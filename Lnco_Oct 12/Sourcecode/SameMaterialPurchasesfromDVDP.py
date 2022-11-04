from string import ascii_uppercase

import pandas as pd
import numpy

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Side, Border
from win32com import client
import pywintypes
from openpyxl.utils import get_column_letter


class BusinessException(Exception):
    pass


def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


def same_mat_dvp(in_config, present_quarter_pd):
    try:
        read_present_quarter_pd = present_quarter_pd
        present_quarter_columns = read_present_quarter_pd.columns
        if in_config["purchase_register_1st_column_name"] in present_quarter_columns and \
                in_config["purchase_register_2nd_column_name"] in present_quarter_columns:
            print("Present Quarter file - The data is starting from first row only")
            pass

        else:
            print("Present Quarter file - The data is not starting from first row ")
            for index, row in read_present_quarter_pd.iterrows():
                if row[0] != in_config["purchase_register_1st_column_name"]:
                    read_present_quarter_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = read_present_quarter_pd.iloc[0]
            read_present_quarter_pd = read_present_quarter_pd[1:]
            read_present_quarter_pd.columns = new_header
            read_present_quarter_pd.reset_index(drop=True, inplace=True)
            read_present_quarter_pd.columns.name = None
        read_present_quarter_pd = read_present_quarter_pd.loc[:, ~read_present_quarter_pd.columns.duplicated(keep='first')]

        input_file = read_present_quarter_pd[read_present_quarter_pd["Material No."].notna()]
        unit_price = read_present_quarter_pd[read_present_quarter_pd["unit price"].notna()]

        if read_present_quarter_pd.shape[0] == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")
        elif len(unit_price) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Unit_price"],
                      body=in_config["Unit_price_body"])
            raise BusinessException("unit price column is empty")
        elif len(input_file) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Document is empty"],
                      body=in_config["Doc_body"])
            raise BusinessException("file is empty")
        else:
            pass
        max_pivot = pd.pivot_table(read_present_quarter_pd, index=["Material No.", "Material Desc", "Vendor Name"],
                                   values=["unit price"], aggfunc=numpy.max, margins=True, margins_name="Grand Total")

        max_pivot = max_pivot.reset_index()
        # print(max_pivot)
        min_pivot = pd.pivot_table(read_present_quarter_pd, index=["Material No.", "Material Desc", "Vendor Name"],
                                   values=["unit price"], aggfunc=numpy.min, margins=True, margins_name="Grand Total")
        min_pivot = min_pivot.reset_index()
        # print(min_pivot)

        com_file = pd.merge(max_pivot, min_pivot, how="outer",
                            on=["Material No.", "Material Desc", "Vendor Name"])
        com_file = com_file.replace(numpy.nan, '', regex=True)

        com_file["Deference"] = com_file["unit price_x"] - com_file["unit price_y"]
        com_file.drop(com_file[com_file["Deference"] <= 1].index, inplace=True)

        com_file = com_file.replace(numpy.nan, 0, regex=True)
        # print(com_file)

        com_file["Variance"] = com_file["Deference"] / com_file["unit price_y"]

        col_name = com_file.columns.values.tolist()

        com_file.sort_values(by=col_name[5], axis=0, ascending=False, inplace=True)
        # com_file[col_name[4]].values[-1] = grand_total
        com_file = com_file.rename(columns={col_name[3]: "Max of uint price"})
        com_file = com_file.rename(columns={col_name[4]: "Min of uint price"})
        with pd.ExcelWriter(in_config["Out_file"], engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            com_file.to_excel(writer, sheet_name=in_config["out_sheet"], index=False, startrow=21)

        wb = load_workbook(in_config["Out_file"])
        ws = wb[in_config["out_sheet"]]

        for cell in ws['D']:
            cell.number_format = "#,###.##"
        for cell in ws['E']:
            cell.number_format = "#,###.##"
        for cell in ws['F']:
            cell.number_format = "#,###.##"
        for cell in ws['G']:
            cell.number_format = "##%"

        ws.delete_rows(idx=23)

        m_row = ws.max_row
        # ws.auto_filter.ref = ws.dimensions

        FullRange = "A22:" + get_column_letter(ws.max_column) \
                   + str(ws.max_row)
        ws.auto_filter.ref = FullRange

        font_style = Font(name="Cambria", size=13, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "22"].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "22"].fill = fill_pattern
            if c == 'G':
                break

        for c in ascii_uppercase:
            ws.column_dimensions[c].width = 35

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=22, min_col=1, max_row=ws.max_row, max_col=7):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        ws.merge_cells('A4:E4')
        ws.merge_cells('A5:E5')
        ws.merge_cells('A6:E6')
        ws.merge_cells('A7:E7')
        ws.merge_cells('A8:E8')
        ws.merge_cells('A9:E9')
        ws.merge_cells('A10:E10')
        ws.merge_cells('A11:E11')
        ws.merge_cells('A12:E12')
        ws.merge_cells('A13:E13')
        ws.merge_cells('A14:E14')

        # Headers implementation
        ws['A1'] = in_config['A1']
        ws['A2'] = in_config['A2']
        ws['A3'] = in_config['A3']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A13'] = in_config['A13']
        ws['A14'] = in_config['A14']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=11, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=13, min_col=1, max_row=13, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=14, min_col=1, max_row=14, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        wb.save(in_config["Out_file"])
        wb.close()

        return com_file
    except FileNotFoundError as f_error:
        print("sent a mail file not found")
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["File_N"],
                  body=in_config["File_N_body"])
        print("Exception: ", f_error)
        return f_error
    except NameError as n_error:
        print("Name Error")
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Name_E"],
                  body=in_config["Name_E_body"])
        print("Exception: ", n_error)
        return n_error
    except KeyError as k_error:
        print("KeyError")
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Key"],
                  body=in_config["Key_body"])
        print("Exception: ", k_error)
        return k_error
    except ValueError as v_error:
        print("ValueError")
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Value_E"],
                  body=in_config["Value_E_body"])
        print("Exception: ", v_error)
        return v_error
    except SyntaxError as s_error:
        print("Exception: ", s_error)
        return s_error


config = {}
present_quarter_pd = pd.DataFrame()
if __name__ == "__main__":
    print(same_mat_dvp(config, present_quarter_pd))


