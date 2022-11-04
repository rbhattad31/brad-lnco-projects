import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from string import ascii_lowercase
from win32com import client
import pywintypes
import os


class BusinessException(Exception):
    pass


def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


def month_wise(in_config, present_quarter_pd):
    try:
        # Read Purchase Register Sheets
        read_present_quarter_pd = present_quarter_pd
        present_quarter_columns = read_present_quarter_pd.columns
        if in_config["purchase_register_1st_column_name"] in present_quarter_columns and \
                in_config["purchase_register_2nd_column_name"] in present_quarter_columns:
            print("Present Quarter file - The data is starting from first row only")
            pass

        else:
            print("Present Quarter file - The data is not starting from first row ")
            for index, row in read_present_quarter_pd.iterrows():
                if row[0] != in_config["purchase_register_1st_column_name"]:
                    read_present_quarter_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = read_present_quarter_pd.iloc[0]
            read_present_quarter_pd = read_present_quarter_pd[1:]
            read_present_quarter_pd.columns = new_header
            read_present_quarter_pd.reset_index(drop=True, inplace=True)
            read_present_quarter_pd.columns.name = None
        read_present_quarter_pd = read_present_quarter_pd.loc[:, ~read_present_quarter_pd.columns.duplicated(keep='first')]

        # Fetch To Address
        to_address = in_config["To_Address"]
        cc_address = in_config["CC_Address"]

        # Check Exception
        if read_present_quarter_pd.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        Present_Quarter_Sheet_col = read_present_quarter_pd.columns.values.tolist()
        for col in ['GR Posting Date', "GR Amt.in loc.cur."]:
            if col not in Present_Quarter_Sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter rows
        date_1 = read_present_quarter_pd[read_present_quarter_pd['GR Posting Date'].notna()]
        gr_amt = read_present_quarter_pd[read_present_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(date_1) == 0:
            subject = in_config["Date_Subject"]
            body = in_config["Date_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Month Column is empty")
        elif len(gr_amt) == 0:
            subject = in_config["GRAmt_Subject"]
            body = in_config["GRAmt_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # Create Month Column
        read_present_quarter_pd['GR Posting Date'] = pd.to_datetime(read_present_quarter_pd['GR Posting Date'])
        read_present_quarter_pd['GR Posting Date'] = read_present_quarter_pd['GR Posting Date'].dt.month_name().str[:3]
        read_present_quarter_pd['Month'] = read_present_quarter_pd['GR Posting Date']

        # Sort based on month
        month_dict = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9,
                      'Oct': 10, 'Nov': 11, 'Dec': 12, 'Grand Total': 13}

        # Create Pivot Table Q3
        pivot_index = ["Month"]
        pivot_values = ["GR Amt.in loc.cur."]
        pivot_present_quarter = pd.pivot_table(read_present_quarter_pd, index=pivot_index, values=pivot_values, aggfunc=numpy.sum, margins=True,
                                               margins_name='Grand Total')

        # Get Pivot Column Names
        col_name = pivot_present_quarter.columns.values.tolist()

        # Rename Column
        pivot_present_quarter = pivot_present_quarter.rename(columns={col_name[0]: in_config["PresentQuarterColumn"]})

        # Remove Index
        pivot_present_quarter = pivot_present_quarter.reset_index()

        # Sort based on month
        pivot_present_quarter = pivot_present_quarter.sort_values('Month', key=lambda x: x.apply(lambda x: month_dict[x]))
        pivot_present_quarter.reset_index(inplace=True, drop=True)

        # Assign Sheets
        pivot_sheet = pivot_present_quarter

        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Delete row of Q4 and Q3 columns values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[1]] == 0)])

        pd.options.mode.chained_assignment = None

        # Get maximum value
        total_value = pivot_sheet.iloc[-1:]
        total_value = total_value.iloc[0, 1]

        # Variance Formula
        variance_list = []
        for index in pivot_sheet.index:
            quarter_value = (pivot_sheet[col_name[1]][index])

            if total_value == 0:
                variance = 1
            else:
                variance = quarter_value / total_value

            variance_list.append(variance)

        # Create Variance Column
        pivot_sheet['Variance'] = variance_list

        # Log Sheet
        with pd.ExcelWriter(in_config["Month_Path"], engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            pivot_sheet.to_excel(writer, sheet_name=in_config["MonthSheet"], index=False,  startrow=16)

        # Check outfile creation
        if os.path.exists(in_config["Month_Path"]):
            print("Month Wise Concentration Logged")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Output file not generated")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(in_config["Month_Path"])
        ws = wb[in_config["MonthSheet"]]

        # Format Q3
        for cell in ws['B']:
            cell.number_format = "#,###,##"

        # Format Variance
        for cell in ws['C']:
            cell.number_format = '0.0%'

        # Format Header
        format_font = Font(name="Calibri", size=11, color="000000", bold=True)
        for c in ascii_lowercase:
            ws[c + "17"].font = format_font

        # Format Footer
        m_row = ws.max_row
        for c in ascii_lowercase:
            ws[c + str(m_row)].font = format_font

        # Header Fill
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "17"].fill = format_fill
            if c == 'c':
                break

        # Footer Fill
        for c in ascii_lowercase:
            ws[c + str(m_row)].fill = format_fill
            if c == 'c':
                break

        # Set Width
        for c in ascii_lowercase:
            ws.column_dimensions[c].width = 15

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=17, min_col=1, max_row=ws.max_row, max_col=3):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        ws.merge_cells('A4:F4')
        ws.merge_cells('A5:F5')
        ws.merge_cells('A6:F6')
        ws.merge_cells('A7:F7')
        ws.merge_cells('A8:F8')
        ws.merge_cells('A9:F9')
        ws.merge_cells('A10:F10')
        ws.merge_cells('A11:F11')
        ws.merge_cells('A12:F12')
        ws.merge_cells('A13:F13')
        ws.merge_cells('A14:F14')

        # Headers implementation
        ws['A1'] = in_config['A1']
        ws['A2'] = in_config['A2']
        ws['A3'] = in_config['A3']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        # Save File
        wb.save(in_config["Month_Path"])
        return ws

    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", file_error)
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Concentration Month Wise Process-", business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-",error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", key_error)
        return key_error


# Read config details and parse to dictionary
config = {}
present_quarter_pd = pd.DataFrame()
if __name__ == "__main__":
    print(month_wise(config, present_quarter_pd))

