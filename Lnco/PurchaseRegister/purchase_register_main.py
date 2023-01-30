import logging

import openpyxl
import xlsxwriter
import pandas as pd
import os.path
import os
import sys
from decouple import Config, RepositoryEnv

from PurchaseRegister.File_Creation_Programs.purchase_present_quarter_file_creation import purchase_present_quarter_file_creation
from PurchaseRegister.File_Creation_Programs.purchase_previous_quarter_file_creation import purchase_previous_quarter_file_creation

import PurchaseRegister.SourceCode.Comparatives.Purchase_type_wise_comparatives as ptcomp
import PurchaseRegister.SourceCode.Comparatives.Month_Wise_comparatives as mwcomp
import PurchaseRegister.SourceCode.Comparatives.Plant_wise_comparatives as pwcomp
import PurchaseRegister.SourceCode.Comparatives.DomImp_wise_sheet_comparatives as dicomp
import PurchaseRegister.SourceCode.Comparatives.VendorTypeWiseComparatives as vtwc

import PurchaseRegister.SourceCode.Concentrations.PurchaseTypeWiseConcentration as ptconc
import PurchaseRegister.SourceCode.Concentrations.MonthWiseConcentration as mwconc
import PurchaseRegister.SourceCode.Concentrations.PlantWiseConcentration as pwconc
import PurchaseRegister.SourceCode.Concentrations.DomAndImportConcentration as diconc
import PurchaseRegister.SourceCode.Concentrations.VendorWiseConcentration as vwc

import PurchaseRegister.SourceCode.DuplicationofVendorNumbers as duplication
import PurchaseRegister.SourceCode.averagedaypurchase as averagedaypurchase
import PurchaseRegister.SourceCode.SameMaterialPurchasesfromDVDP as smpdvdp
import PurchaseRegister.SourceCode.Unit_Price_Comparison as upc
import PurchaseRegister.SourceCode.Inventory_Mapping as im
import PurchaseRegister.SourceCode.SecurityCutoff as sc

from ReusableTasks.send_mail_reusable_task import send_mail

from ReusableTasks.create_sheet_wise_config_dictionary import create_sheet_wise_config_dict


def process_execution(input_files,
                      present_quarter_sheet_name, previous_quarter_sheet_name,
                      vendor_master_sheet_name,
                      mb51_sheet_name, present_quarter_column_name, previous_quarter_column_name,
                      company_name, statutory_audit_quarter, financial_year, config_main, request_id,
                      json_data_list, env_file
                      ):
    print("Starting audit process for the input files")
    logging.info("Starting audit process for the input files")
    # print(input_files)
    mb51_file_path = input_files[0]
    vendor_file_path = input_files[1]
    purchase_register_present_quarter_file_path = input_files[2]
    purchase_register_previous_quarter_file_path = input_files[3]

    config_main['PresentQuarterColumnName'] = present_quarter_column_name
    config_main['PreviousQuarterColumnName'] = previous_quarter_column_name
    config_main['CompanyName'] = company_name
    config_main['StatutoryAuditQuarter'] = statutory_audit_quarter
    config_main['FinancialYear'] = financial_year

    print("*******************************************")
    print("Check if Output file exists")
    project_home_directory = os.getcwd()
    output_file_path = os.path.join(project_home_directory, 'Data', 'Output', 'audit_requests', str(request_id))
    print("Output file folder is : ", output_file_path)
    if not os.path.exists(output_file_path):
        print("Output folder is not exist")
        print("Creating directory: ", output_file_path)
        os.makedirs(output_file_path)
        print("Directory" + output_file_path + " is created")
    output_file_name = company_name.replace(' ', '_') + "_" + str(request_id) + "_Purchase_Audit_Report_Output.xlsx"
    output_file_path = os.path.join(output_file_path, output_file_name)
    print("Output file path is: " + output_file_path)
    config_main['Output_File_Path'] = output_file_path
    if os.path.exists(output_file_path):
        print("Output file exist")
        print("Deleting existing output file")
        os.remove(output_file_path)
        if os.path.exists(output_file_path):
            pass
        else:
            print("existing output file is deleted successfully")
        print("Creating a new output file")
        workbook = xlsxwriter.Workbook(output_file_path)
        workbook.close()
        if os.path.exists(output_file_path):
            print("New output file is created")
        else:
            print("New output file creation is failed")
    else:
        print("Output file not exist")
        print("Creating a new output file")
        workbook = xlsxwriter.Workbook(output_file_path)
        workbook.close()
        if os.path.exists(output_file_path):
            print("New output file is created")
        else:
            print("New output file creation is failed")

    print("*******************************************")

    try:
        print("Reading Purchase registers is started")
        print("Reading present quarter sheet")
        print(purchase_register_present_quarter_file_path)
        read_present_quarter_pd = pd.read_excel(purchase_register_present_quarter_file_path,
                                                present_quarter_sheet_name)
        # print(read_present_quarter_pd.dtypes.to_list)
        read_present_quarter_pd = \
            read_present_quarter_pd.loc[:, ~read_present_quarter_pd.columns.duplicated(keep='first')]

        present_quarter_columns = read_present_quarter_pd.columns
        if config_main["purchase_register_1st_column_name"] in present_quarter_columns and \
                config_main["purchase_register_2nd_column_name"] in present_quarter_columns:
            print("Present Quarter file - The data is starting from first row only")
            pass

        else:
            print("Present Quarter file - The data is not starting from first row ")
            for index, row in read_present_quarter_pd.iterrows():
                if row[0] != config_main["purchase_register_1st_column_name"]:
                    read_present_quarter_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = read_present_quarter_pd.iloc[0]
            read_present_quarter_pd = read_present_quarter_pd[1:]
            read_present_quarter_pd.columns = new_header
            read_present_quarter_pd.reset_index(drop=True, inplace=True)
            read_present_quarter_pd.columns.name = None
        read_present_quarter_pd = \
            read_present_quarter_pd.loc[:, ~read_present_quarter_pd.columns.duplicated(keep='first')]
        present_quarter_all_columns_dataframe = read_present_quarter_pd
        print(
            "Reading purchase register present quarter sheet is complete, creating new input file only with required columns")
        logging.info(
            "Reading purchase register present quarter sheet is complete, creating new input file only with required columns")
        purchase_register_present_quarter_folder_path = os.path.dirname(purchase_register_present_quarter_file_path)
        logging.info("purchase_register_present_quarter_folder_path: \n\t ", purchase_register_present_quarter_folder_path)
        purchase_register_present_quarter_file_name = os.path.basename(
            purchase_register_present_quarter_file_path).lower()
        filtered_purchase_present_file_name = "filtered_" + str(purchase_register_present_quarter_file_name)
        logging.info("filtered_purchase_present_file_name: \n\t", filtered_purchase_present_file_name)
        filtered_purchase_present_file_saving_path = os.path.join(purchase_register_present_quarter_folder_path,
                                                                  filtered_purchase_present_file_name)
        logging.info("filtered_purchase_present_file_saving_path \n\t", filtered_purchase_present_file_saving_path)
        filtered_purchase_present_sheet_name = present_quarter_sheet_name

        purchase_present_quarter_file_creation_output = purchase_present_quarter_file_creation(config_main,
                                                                                               read_present_quarter_pd,
                                                                                               json_data_list,
                                                                                               filtered_purchase_present_file_saving_path,
                                                                                               filtered_purchase_present_sheet_name)

        read_present_quarter_pd = purchase_present_quarter_file_creation_output[0]
        config_main = purchase_present_quarter_file_creation_output[1]

        logging.info("new purchase register present quarter file is created in input folder in request ID folder")

        # reading previous quarter sheet
        print("Reading previous quarter sheet")
        logging.info("Reading previous quarter sheet")
        read_previous_quarter_pd = pd.read_excel(purchase_register_previous_quarter_file_path,
                                                 previous_quarter_sheet_name)
        # print(read_previous_quarter_pd.dtypes.to_list)
        read_previous_quarter_pd = \
            read_previous_quarter_pd.loc[:, ~read_previous_quarter_pd.columns.duplicated(keep='first')]

        previous_quarter_columns = read_previous_quarter_pd.columns
        if config_main["purchase_register_1st_column_name"] in previous_quarter_columns and \
                config_main["purchase_register_2nd_column_name"] in previous_quarter_columns:
            print("Previous Quarter file - The data is starting from first row only")
            pass

        else:
            print("Previous Quarter file - The data is not starting from first row ")
            for index, row in read_previous_quarter_pd.iterrows():
                if row[0] != config_main["purchase_register_1st_column_name"]:
                    read_previous_quarter_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = read_previous_quarter_pd.iloc[0]
            read_previous_quarter_pd = read_previous_quarter_pd[1:]
            read_previous_quarter_pd.columns = new_header
            read_previous_quarter_pd.reset_index(drop=True, inplace=True)
            read_previous_quarter_pd.columns.name = None
        read_previous_quarter_pd = \
            read_previous_quarter_pd.loc[:, ~read_previous_quarter_pd.columns.duplicated(keep='first')]
        print(
            "Reading purchase register previous quarter sheet is complete, creating new input file only with required columns")
        logging.info(
            "Reading purchase register previous quarter sheet is complete, creating new input file only with required columns")
        purchase_register_previous_quarter_folder_path = os.path.dirname(purchase_register_previous_quarter_file_path)
        purchase_register_previous_quarter_file_name = os.path.basename(
            purchase_register_previous_quarter_file_path).lower()
        filtered_purchase_previous_file_name = "filtered_" + str(purchase_register_previous_quarter_file_name)
        filtered_purchase_previous_file_saving_path = os.path.join(purchase_register_previous_quarter_folder_path,
                                                                   filtered_purchase_previous_file_name)
        filtered_purchase_previous_sheet_name = previous_quarter_sheet_name

        read_previous_quarter_pd = purchase_previous_quarter_file_creation(read_previous_quarter_pd, json_data_list,
                                                                           filtered_purchase_previous_file_saving_path,
                                                                           filtered_purchase_previous_sheet_name)
        logging.info("new purchase register previous quarter file is created in input folder in request ID folder")
        print("new purchase register previous quarter file is created in input folder in request ID folder")
    except FileNotFoundError as notfound_error:
        send_mail(to=config_main["To_Mail_Address"], cc=config_main["CC_Mail_Address"],
                  subject=config_main["subject_file_not_found"],
                  body=config_main["body_file_not_found"])
        print(notfound_error)
        logging.error("file not found error occurred: \n\t {}".format(notfound_error))
        logging.exception(notfound_error)
        raise notfound_error
    except ValueError as sheetNotFound_error:
        send_mail(to=config_main["To_Mail_Address"], cc=config_main["CC_Mail_Address"],
                  subject=config_main["subject_sheet_not_found"],
                  body=config_main["body_sheet_not_found"])
        print(sheetNotFound_error)
        logging.error("sheet not found error occurred: \n\t {}".format(sheetNotFound_error))
        logging.exception(sheetNotFound_error)
        raise sheetNotFound_error
    except Exception as exception:
        logging.exception(exception)
        raise exception

    print("*******************************************")
    print("Executing Comparatives Purchase type code")

    try:
        if env_file('comparatives_purchase_wise') == 'YES':
            read_present_quarter_pd_comp_purchase = read_present_quarter_pd
            read_present_quarter_pd_comp_purchase = \
                read_present_quarter_pd_comp_purchase[["Valuation Class", "Valuation Class Text", "GR Amt.in loc.cur.", "Vendor Name"]]
            read_previous_quarter_pd_comp_purchase = read_previous_quarter_pd
            read_previous_quarter_pd_comp_purchase = \
                read_previous_quarter_pd_comp_purchase[
                    ["Valuation Class", "Valuation Class Text", "GR Amt.in loc.cur.", "Vendor Name"]]
            config_purchase_comparatives = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Comparatives_Purchase_sheetname"])
            ptcomp.create_purchase_type_wise(config_main, config_purchase_comparatives,
                                             read_present_quarter_pd_comp_purchase,
                                             read_previous_quarter_pd_comp_purchase)
        elif env_file('comparatives_purchase_wise') == 'NO':
            print("comparatives_purchase_wise process is skipped as per env file")
        else:
            print("select YES/NO for comparatives_purchase_wise process in env file")
            raise Exception("Error in Env file for comparatives_purchase_wise")
    except Exception as e:
        print("Exception caught for Process: Purchase type comparatives ", e)
        logging.error("Exception caught for Process: Purchase type comparatives {0}".format(e))

    print("*******************************************")

    print("Executing Comparatives Month wise code")

    try:
        if env_file('comparatives_month_wise') == 'YES':
            read_present_quarter_pd_comp_month = read_present_quarter_pd
            read_present_quarter_pd_comp_month = read_present_quarter_pd_comp_month[
                ["GR Posting Date", "GR Amt.in loc.cur."]]
            read_previous_quarter_pd_comp_month = read_previous_quarter_pd
            read_previous_quarter_pd_comp_month = read_previous_quarter_pd_comp_month[
                ["GR Posting Date", "GR Amt.in loc.cur."]]
            config_month_comparatives = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Comparatives_Month_sheetname"])
            mwcomp.purchase_month(config_main, config_month_comparatives, read_present_quarter_pd_comp_month,
                                  read_previous_quarter_pd_comp_month)
        elif env_file('comparatives_month_wise') == 'NO':
            print("comparatives_month_wise process is skipped as per env file")
        else:
            print("select YES/NO for comparatives_month_wise process in env file")
            raise Exception("Error in Env file for comparatives_month_wise")
    except Exception as e:
        print("Exception caught for Process: Month wise comparatives ", e)
        logging.error("Exception caught for Process: Month wise comparatives {0}".format(e))
    print("*******************************************")

    print("Executing Comparatives Plant wise code")

    try:
        if env_file('comparatives_plant_wise') == 'YES':
            read_present_quarter_pd_comp_plant = read_present_quarter_pd
            read_present_quarter_pd_comp_plant = read_present_quarter_pd_comp_plant[["Plant", "GR Amt.in loc.cur."]]
            read_previous_quarter_pd_comp_plant = read_previous_quarter_pd
            read_previous_quarter_pd_comp_plant = read_previous_quarter_pd_comp_plant[["Plant", "GR Amt.in loc.cur."]]

            config_plant_comparatives = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Comparatives_Plant_sheetname"])
            pwcomp.create_plant_wise_sheet(config_main, config_plant_comparatives, read_present_quarter_pd_comp_plant,
                                           read_previous_quarter_pd_comp_plant)
        elif env_file('comparatives_plant_wise') == 'NO':
            print("comparatives_plant_wise process is skipped as per env file")
        else:
            print("select YES/NO for comparatives_plant_wise process in env file")
            raise Exception("Error in Env file for comparatives_plant_wise")

    except Exception as e:
        print("Exception caught for Process: Plant wise comparatives ", e)
        logging.error("Exception caught for Process: Plant wise comparatives {0}".format(e))
    print("*******************************************")

    print("Executing Comparatives Domestic and Import wise code")

    try:
        if env_file('comparatives_domestic_import_wise') == 'YES':
            read_present_quarter_pd_comp_domandimp = read_present_quarter_pd
            read_present_quarter_pd_comp_domandimp = \
                read_present_quarter_pd_comp_domandimp[["Currency Key", "GR Amt.in loc.cur."]]
            read_previous_quarter_pd_comp_domandimp = read_previous_quarter_pd
            read_previous_quarter_pd_comp_domandimp = \
                read_previous_quarter_pd_comp_domandimp[["Currency Key", "GR Amt.in loc.cur."]]
            config_domandimp_comparatives = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Comparatives_Dom&Imp_sheetname"])
            dicomp.generate_domestic_and_import_wise(config_main, config_domandimp_comparatives,
                                                     read_present_quarter_pd_comp_domandimp,
                                                     read_previous_quarter_pd_comp_domandimp)
        elif env_file('comparatives_domestic_import_wise') == 'NO':
            print("comparatives_domestic_import_wise process is skipped as per env file")
        else:
            print("select YES/NO for comparatives_domestic_import_wise process in env file")
            raise Exception("Error in Env file for comparatives_domestic_import_wise")

    except Exception as e:
        print("Exception caught for Process: Domestic and import wise comparators ", e)
        logging.error("Exception caught for Process: Domestic and import wise comparators {0}".format(e))

    print("*******************************************")
    print("Executing Vendor Type Wise Comparatives code")

    try:
        if env_file('comparatives_vendor_wise') == 'YES':
            read_present_quarter_pd_comp_vendor = read_present_quarter_pd
            read_present_quarter_pd_comp_vendor = \
                read_present_quarter_pd_comp_vendor[["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]]
            read_previous_quarter_pd_comp_vendor = read_previous_quarter_pd
            read_previous_quarter_pd_comp_vendor = \
                read_previous_quarter_pd_comp_vendor[["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]]
            config_vendor_comp = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Comparatives_Vendor_sheetname"])
            vtwc.create_vendor_wise(config_main, config_vendor_comp, read_present_quarter_pd_comp_vendor,
                                    read_previous_quarter_pd_comp_vendor)
        elif env_file('comparatives_vendor_wise') == 'NO':
            print("comparatives_vendor_wise process is skipped as per env file")
        else:
            print("select YES/NO for comparatives_vendor_wise process in env file")
            raise Exception("Error in Env file for comparatives_vendor_wise")
    except Exception as e:
        print("Exception caught for Process: Vendor type wise comparatives code ", e)
        logging.error("Exception caught for Process: Vendor type wise comparatives code {0}".format(e))

    print("*******************************************")
    print("Executing concentration Purchase type code")

    try:
        if env_file('concentrations_purchase_wise') == 'YES':
            read_present_quarter_pd_purchase = read_present_quarter_pd
            read_present_quarter_pd_purchase = read_present_quarter_pd_purchase[
                ["Valuation Class", "Valuation Class Text", "GR Amt.in loc.cur."]]
            config_concentration_ptw = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Concentrations_Purchase_sheetname"])
            ptconc.purchase_type(config_main, config_concentration_ptw, read_present_quarter_pd_purchase)
        elif env_file('concentrations_purchase_wise') == 'NO':
            print("concentrations_purchase_wise process is skipped as per env file")
        else:
            print("select YES/NO for concentrations_purchase_wise process in env file")
            raise Exception("Error in Env file for concentrations_purchase_wise")

    except Exception as e:
        print("Exception caught for Process: Purchase type concentration ", e)
        logging.error("Exception caught for Process: Purchase type concentration {}".format(e))
    print("*******************************************")

    print("Executing concentration Month wise code")

    try:
        if env_file('concentrations_month_wise') == 'YES':
            read_present_quarter_pd_month = read_present_quarter_pd
            read_present_quarter_pd_month = read_present_quarter_pd_month[["GR Posting Date", "GR Amt.in loc.cur."]]
            config_concentration_mw = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Concentrations_Month_sheetname"])
            mwconc.month_wise(config_main, config_concentration_mw, read_present_quarter_pd_month)
        elif env_file('concentrations_month_wise') == 'NO':
            print("concentrations_month_wise process is skipped as per env file")
        else:
            print("select YES/NO for concentrations_month_wise process in env file")
            raise Exception("Error in Env file for concentrations_month_wise")

    except Exception as e:
        print("Exception caught for Process: Month wise Concentration ", e)
        logging.error("Exception caught for Process: Month wise Concentration {0}".format(e))
    print("*******************************************")

    print("Executing concentration Plant wise code")

    try:
        if env_file('concentrations_plant_wise') == 'YES':
            read_present_quarter_pd_plant = read_present_quarter_pd
            read_present_quarter_pd_plant = read_present_quarter_pd_plant[["Plant", "GR Amt.in loc.cur."]]
            config_concentration_pw = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Concentrations_Plant_sheetname"])
            pwconc.purchase_type(config_main, config_concentration_pw, read_present_quarter_pd_plant)
        elif env_file('concentrations_plant_wise') == 'NO':
            print("concentrations_plant_wise process is skipped as per env file")
        else:
            print("select YES/NO for concentrations_plant_wise process in env file")
            raise Exception("Error in Env file for concentrations_plant_wise")

    except Exception as e:
        print("Exception caught for Process: Plant wise concentration ", e)
        logging.error("Exception caught for Process: Plant wise concentration {0}".format(e))
    print("*******************************************")

    print("Executing concentration Domestic and Import wise code")

    try:
        if env_file('concentrations_domestic_import_wise') == 'YES':
            read_present_quarter_pd_di = read_present_quarter_pd
            read_present_quarter_pd_di = read_present_quarter_pd_di[["Currency Key", "GR Amt.in loc.cur."]]
            config_concentration_di = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Concentrations_Dom&Imp_sheetname"])
            diconc.purchase_type(config_main, config_concentration_di, read_present_quarter_pd_di)
        elif env_file('concentrations_domestic_import_wise') == 'NO':
            print("concentrations_domestic_import_wise process is skipped as per env file")
        else:
            print("select YES/NO for concentrations_domestic_import_wise process in env file")
            raise Exception("Error in Env file for concentrations_domestic_import_wise")
    except Exception as e:
        print("Exception caught for Process: domestic and import wise concentration ", e)
        logging.error("Exception caught for Process: domestic and import wise concentration {0}".format(e))
    print("*******************************************")

    print("Executing Vendor Wise Concentration code")

    try:
        if env_file('concentration_vendor_wise') == 'YES':
            read_present_quarter_pd_vendor = read_present_quarter_pd
            read_present_quarter_pd_vendor = read_present_quarter_pd_vendor[
                ["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]]
            config_vendor_conc = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Concentration_Vendor_sheetname"])
            vwc.con_vendor_wise(config_main, config_vendor_conc, read_present_quarter_pd_vendor)
        elif env_file('concentration_vendor_wise') == 'NO':
            print("concentration_vendor_wise process is skipped as per env file")
        else:
            print("select YES/NO for concentration_vendor_wise process in env file")
            raise Exception("Error in Env file for concentration_vendor_wise")

    except Exception as e:
        print("Exception caught for Process: Vendor wise concentration code", e)
        logging.error("Exception caught for Process: Vendor wise concentration code {0}".format(e))
    print("*******************************************")

    print("Executing Duplication of Vendor number code")

    try:
        if env_file('duplication_of_vendor') == 'YES':
            config_duplication = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Duplication_of_Vendor_sheetname"])
            duplication.vendor_numbers_duplication(config_main,
                                                   config_duplication,
                                                   vendor_file_path,
                                                   vendor_master_sheet_name, json_data_list)
        elif env_file('duplication_of_vendor') == 'NO':
            print("duplication_of_vendor process is skipped as per env file")
        else:
            print("select YES/NO for duplication_of_vendor process in env file")
            raise Exception("Error in Env file for duplication_of_vendor")

    except Exception as e:
        print("Exception caught for Process: duplication of Vendor number ", e)
        logging.error("Exception caught for Process: duplication of Vendor number code {0}".format(e))
    print("*******************************************")

    print("Executing Average Day Purchase code")

    try:
        if env_file('average_day_purchase') == 'YES':
            read_present_quarter_pd_average = read_present_quarter_pd
            read_present_quarter_pd_average = read_present_quarter_pd_average[['GR Amt.in loc.cur.', 'GR Posting Date']]
            config_average_day_purchase = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Average_Day_Purchase_sheetname"])
            averagedaypurchase.average_day_purchase(config_main, config_average_day_purchase,
                                                    read_present_quarter_pd_average)
        elif env_file('average_day_purchase') == 'NO':
            print("average_day_purchase process is skipped as per env file")
        else:
            print("select YES/NO for average_day_purchase process in env file")
            raise Exception("Error in Env file for average_day_purchase")
    except Exception as e:
        print("Exception caught for Process: Average day purchase: ", e)
        logging.error("Exception caught for Process: Average day purchase code {0}".format(e))

    print("*******************************************")

    print("Executing 'Same Material Purchases from Different Vendors & Different prices' code")

    try:
        if env_file('same_material_purchases_DVDP') == 'YES':
            read_present_quarter_pd_smpdvdp = read_present_quarter_pd
            read_present_quarter_pd_smpdvdp = \
                read_present_quarter_pd_smpdvdp[["Material No.", "Material Desc", "Vendor Name", "Unit Price"]]
            config_smpdvdp = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Same_Material_Purchases_DVDP_sheetname"])
            smpdvdp.same_mat_dvp(config_main, config_smpdvdp, read_present_quarter_pd_smpdvdp)
        elif env_file('same_material_purchases_DVDP') == 'NO':
            print("same_material_purchases_DVDP process is skipped as per env file")
        else:
            print("select YES/NO for same_material_purchases_DVDP process in env file")
            raise Exception("Error in Env file for same_material_purchases_DVDP")

    except Exception as e:
        print("Exception caught for Process: Same Material different Vendors ", e)
        logging.error("Exception caught for Process: Same Material different Vendors code {0}".format(e))

    print("*******************************************")

    print("Executing Unit Price Comparison code")

    try:
        if env_file('Unit_Price_Comparison') == 'YES':
            read_present_quarter_pd_unit = read_present_quarter_pd
            read_present_quarter_pd_unit = \
                read_present_quarter_pd_unit[
                    ["GR Amt.in loc.cur.", "GR Qty", "Material No.", "Valuation Class Text", "Vendor Name"]]
            read_previous_quarter_pd_unit = read_previous_quarter_pd
            read_previous_quarter_pd_unit = \
                read_previous_quarter_pd_unit[
                    ["GR Amt.in loc.cur.", "GR Qty", "Material No.", "Valuation Class Text", "Vendor Name"]]
            config_unit_price = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Unit_Price_Comparison_sheetname"])
            upc.create_unit_price(config_main, config_unit_price, read_present_quarter_pd_unit,
                                  read_previous_quarter_pd_unit)
        elif env_file('Unit_Price_Comparison') == 'NO':
            print("Unit_Price_Comparison process is skipped as per env file")
        else:
            print("select YES/NO for Unit_Price_Comparison process in env file")
            raise Exception("Error in Env file for Unit_Price_Comparison")
    except Exception as e:
        print("Exception caught for Process: Unit Price Comparison code ", e)
        logging.error("Exception caught for Process: Unit Price Comparison code {0}".format(e))

    print("*******************************************")

    print("Executing Inventory Mapping code")

    try:
        if env_file('Inventory_Mapping') == 'YES':
            read_present_quarter_pd_inventory = read_present_quarter_pd
            read_present_quarter_pd_inventory = read_present_quarter_pd_inventory[["GR Document Number", "GR Qty"]]
            config_inventory_mapping = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Inventory_Mapping_Sheetname"])
            im.create_inventory_mapping_sheet(config_main, config_inventory_mapping, read_present_quarter_pd_inventory,
                                              mb51_file_path, mb51_sheet_name, json_data_list)
        elif env_file('Inventory_Mapping') == 'NO':
            print("Inventory_Mapping process is skipped as per env file")
        else:
            print("select YES/NO for Inventory_Mapping process in env file")
            raise Exception("Error in Env file for Inventory_Mapping")
    except Exception as e:
        print("Exception caught for Process: Inventory mapping code ", e)
        logging.error("Exception caught for Process: Inventory mapping code {0}".format(e))

    print("*******************************************")

    print("Executing Security Cutoff code")

    try:
        if env_file('Security_Cutoff') == 'YES':
            read_present_quarter_pd_security_cutoff = present_quarter_all_columns_dataframe
            sc.security_cutoff(config_main, read_present_quarter_pd_security_cutoff)

        elif env_file('Inventory_Mapping') == 'NO':
            print("Security Cutoff process is skipped as per env file")
        else:
            print("select YES/NO for Security Cutoff process in env file")
            raise Exception("Error in Env file for Security Cutoff")
    except Exception as e:
        print("Exception caught for Process: Security Cutoff code ", e)
        logging.error("Exception caught for Process: Security Cutoff code {0}".format(e))
    print("*******************************************")
    # ----------------------------------------------------------------------------------
    final_output_file = openpyxl.load_workbook(output_file_path)
    if 'Sheet1' in final_output_file.sheetnames:
        final_output_file.remove(final_output_file['Sheet1'])

    # -----------------------------------------------------------------------------------
    output_sheets_list = final_output_file._sheets

    all_sheets_list = [
        '<Worksheet "Purchase Wise Comparatives">',
        '<Worksheet "Month Wise Comparatives">',
        '<Worksheet "Plant Wise Comparatives">',
        '<Worksheet "DOM&IMP Wise Comparatives">',
        '<Worksheet "Vendor Wise Comparatives">',

        '<Worksheet "Purchase Wise Concentration">',
        '<Worksheet "Month Wise Concentration">',
        '<Worksheet "Plant Wise Concentration">',
        '<Worksheet "Dom&Imp Wise Concentration">',
        '<Worksheet "Vendor Wise Concentration">',

        '<Worksheet "Duplication Of Vendor">',
        '<Worksheet "Average Day Purchase">',
        '<Worksheet "SMP from DVnDP">',
        '<Worksheet "Unit Price Comparison">',
        '<Worksheet "Inventory Mapping">',

        '<Worksheet "Major Vendor Analysis">',

        '<Worksheet "exceptions – COMPARATIVES">',
        '<Worksheet "exceptions – CONCENTRATIONS">',

        '<Worksheet "Unit price as per Unit Price">',
        '<Worksheet "Unit price as per quantity">',
        '<Worksheet "Unit price- change in amount">',

        '<Worksheet "exceptions average day">',

        '<Worksheet "exceptions – inventory mapping">',

        '<Worksheet "Security Cutoff">'
    ]

    sheets_order_list = []
    for sheet in all_sheets_list:
        for output_sheet in output_sheets_list:
            if str(sheet) == str(output_sheet):
                sheets_order_list.append(output_sheet)
                break
    final_output_file._sheets = sheets_order_list
    final_output_file.save(output_file_path)

    # ------------------------------------------------------------------------------------

    config_saving_file_path_in_output = os.path.join(project_home_directory, 'Data', 'Output', 'audit_requests',
                                                     str(request_id), 'config.xlsx')
    config_saving_folder_path_in_output = os.path.join(project_home_directory, 'Data', 'Output', 'audit_requests',
                                                       str(request_id))
    if not os.path.exists(config_saving_folder_path_in_output):
        logging.warning("folder path {0} is not exist".format(config_saving_folder_path_in_output))
        os.makedirs(config_saving_folder_path_in_output)
        logging.warning("created the directory: {0}".format(config_saving_folder_path_in_output))
    print("Creating updated Config file in output folder of Request from Config Dictionary")
    logging.info("Creating updated Config file in output folder of Request from Config Dictionary")
    try:
        new_config_df = pd.DataFrame.from_dict(config_main, orient='index', columns=['Value'])
        new_config_df.index.name = 'Key'
        new_config_df.reset_index(level=0, inplace=True)
        new_config_df.to_excel(config_saving_file_path_in_output, index=False)
        print("Created updated Config file in output folder of Request from Config Dictionary")
        logging.info("Created updated Config file in Output folder of Request from Config Dictionary")
    except Exception as config_file_save_exception:
        logging.warning("Exception occurred while saving updated config file in Output folder of Request Folder")
        logging.exception(config_file_save_exception)
        exception_type, exception_object, exception_traceback = sys.exc_info()
        filename = exception_traceback.tb_frame.f_code.co_filename
        line_number = exception_traceback.tb_lineno
        logging.warning(str(exception_type))
        logging.warning("Exception occurred in file : {} at line number: {}".format(filename, line_number))
    else:
        print("Saved updated config data to an excel file in output folder in the request folder")
    # ------------------------------------------------------------------------------------

    # # Bot success mail notification
    # end_to = config_main['To_Mail_Address']
    # end_cc = config_main['CC_Mail_Address']
    # end_subject = config_main['Success_Mail_Subject']
    # end_body = config_main['Success_Mail_Body']
    # send_mail_with_attachment(to=end_to, cc=end_cc, body=end_body, subject=end_subject,
    #                           attachment_path=output_file_path)
    # print("Process complete mail notification is sent")
    #
    # print("Bot successfully finished Processing of the sheets")
    return output_file_path


if __name__ == '__main__':
    pass
