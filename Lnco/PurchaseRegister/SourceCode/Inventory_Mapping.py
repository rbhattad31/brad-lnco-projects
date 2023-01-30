import pandas as pd
import numpy
import openpyxl
import logging
import os
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase

from ReusableTasks.send_mail_reusable_task import send_mail

from PurchaseRegister.File_Creation_Programs.mb51_file_creation import mb51_file_creation


class BusinessException(Exception):
    pass


def inventory_mapping_business_exception(inventory_mapping_file, main_config):
    inventory_mapping_columns_list = inventory_mapping_file.columns.tolist()
    inventory_mapping_exception_percentage = main_config['Inventory_mapping_exception_percentage'] / 100
    inventory_mapping_business_exception_pd = inventory_mapping_file.loc[
        inventory_mapping_file['Variance'] >= inventory_mapping_exception_percentage, inventory_mapping_columns_list]
    # print(inventory_mapping_business_exception_pd)
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            inventory_mapping_business_exception_pd.to_excel(writer, sheet_name=main_config[
                "Output_inventory_mapping_exceptions_sheetname"], index=False, startrow=1)

    except Exception as File_creation_error:
        logging.error("Exception occurred while creating inventory mapping business exceptions sheet")
        raise File_creation_error

    # Opening and Reading Output File.
    workbook = openpyxl.load_workbook(main_config["Output_File_Path"])
    worksheet = workbook[main_config["Output_inventory_mapping_exceptions_sheetname"]]

    # Adding Background Color
    light_blue_fill = PatternFill(patternType="solid", fgColor="ADD8E6")
    yellow_fill = PatternFill("solid", fgColor="ffff00")
    for j in ascii_uppercase:
        worksheet[j + "2"].fill = light_blue_fill
        if j == 'E':
            break
    worksheet["C2"].fill = yellow_fill

    # # Adding Auto Filter Option
    full_range = "A2:E" + str(worksheet.max_row)
    worksheet.auto_filter.ref = full_range

    # Auto Width Setting
    for c in ascii_uppercase:
        column_length = max(len(str(cell.value)) for cell in worksheet[c])
        worksheet.column_dimensions[c].width = column_length * 1.5
        if c == 'E':
            break

    for cell in worksheet['B']:
        cell.number_format = '#,###.##'
    for cell in worksheet['C']:
        cell.number_format = '#,###.##'
    for cell in worksheet['E']:
        cell.number_format = '0.0%'

    if len(inventory_mapping_business_exception_pd.index) == 0:
        message = "NOTE: No entries found in Inventory Mapping that have Variance > {0}%".format(main_config['Inventory_mapping_exception_percentage'])
        worksheet.merge_cells('A1:E1')
        worksheet['A1'] = message

    # Saving the File
    print(workbook.sheetnames)
    workbook.save(main_config["Output_File_Path"])


def create_inventory_mapping_sheet(main_config, in_config, present_quarter_pd, mb51_file_location, mb51_sheet_name,
                                   json_data_list):
    try:
        # Reading Purchase register File
        # read_excel_data = pd.read_excel(in_config["ExcelPath1"], sheet_name=in_config["Sheet_Name1"], skiprows=6)
        read_excel_data = present_quarter_pd
        read_excel_data = read_excel_data.loc[:, ~read_excel_data.columns.duplicated(keep='first')]

        # print(read_excel_data.head(5))
        # Check Exception
        if read_excel_data.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")

        purchase_sheet_col = read_excel_data.columns.values.tolist()
        for col in ["GR Document Number", "GR Qty"]:
            if col not in purchase_sheet_col:
                subject = in_config["Purchase_ColumnMiss_Subject"]
                body = in_config["Purchase_ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")
        # Filter Rows
        gr_document_number_pd = read_excel_data[read_excel_data['GR Document Number'].notna()]
        gr_qty_pd = read_excel_data[read_excel_data['GR Qty'].notna()]

        if len(gr_document_number_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["GR Document_Number_subject"],
                      body=in_config["GR Document_Number_Body"])
            raise BusinessException("GR Document Number Column is empty")

        elif len(gr_qty_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Gr Qty_Subject"],
                      body=in_config["Gr Qty_Body"])
            raise BusinessException("GR Qty Column is empty")
        else:
            pass

        read_excel_data = read_excel_data[['GR Document Number', 'GR Qty']]

        pivot1_df = pd.pivot_table(read_excel_data, index=["GR Document Number"],
                                   values="GR Qty",
                                   aggfunc=numpy.sum)
        pivot1_df.reset_index()

        # Reading MB 51 File
        mb51_pd = pd.read_excel(mb51_file_location, mb51_sheet_name)
        mb51_pd = mb51_pd.loc[:, ~mb51_pd.columns.duplicated(keep='first')]
        columns = mb51_pd.columns
        if main_config["MB51_first_column"] in columns and \
                main_config["MB51_second_column"] in columns:
            print("MB51 - The data is starting from first row only")
            pass

        else:
            print("MB51 - The data is not starting from first row ")
            for index, row in mb51_pd.iterrows():
                if row[0] != main_config["MB51_first_column"]:
                    mb51_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = mb51_pd.iloc[0]
            mb51_pd = mb51_pd[1:]
            mb51_pd.columns = new_header
            mb51_pd.reset_index(drop=True, inplace=True)
            mb51_pd.columns.name = None
        mb51_pd = mb51_pd.loc[:, ~mb51_pd.columns.duplicated(keep='first')]
        print("MB51 file reading is complete, creating new input file only with required columns")
        logging.info(
            "Reading MB51 sheet is complete, creating new input file only with required columns")
        mb51_folder_path = os.path.dirname(mb51_file_location)
        mb51_file_name = os.path.basename(mb51_file_location).lower()
        filtered_mb51_file_name = "filtered_" + str(mb51_file_name)
        filtered_purchase_present_file_saving_path = os.path.join(mb51_folder_path, filtered_mb51_file_name)
        filtered_purchase_present_sheet_name = mb51_sheet_name
        mb51_pd = mb51_file_creation(mb51_pd, json_data_list, filtered_purchase_present_file_saving_path,
                                     filtered_purchase_present_sheet_name)
        logging.info("new mb51 filtered file is created in input folder in request ID folder")

        # print(mb51_pd.head(5))
        # Check Exception
        if mb51_pd.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")

        mb51_sheet_col = mb51_pd.columns.values.tolist()
        for col in ["Material Document", "Qty in unit of entry", "Movement type"]:
            if col not in mb51_sheet_col:
                subject = in_config["MB51_ColumnMiss_Subject"]
                body = in_config["MB51_ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        material_document_pd = mb51_pd[mb51_pd['Material Document'].notna()]
        qty_unit_of_entry_pd = mb51_pd[mb51_pd['Qty in unit of entry'].notna()]
        movement_type_pd = mb51_pd[mb51_pd['Movement type'].notna()]
        if len(material_document_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Material_Document_subject"],
                      body=in_config["Material_Document_Body"])
            raise BusinessException("Material Document Column is empty")

        elif len(qty_unit_of_entry_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Qty_unit_of_entry_Subject"],
                      body=in_config["Qty_unit_of_entry_Body"])
            raise BusinessException("Qty in unit of entry Column is empty")
        elif len(movement_type_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Movement Type Subject"],
                      body=in_config["Movement Type Body"])
            raise BusinessException("Movement Type Column is empty")
        else:
            pass

        # Taking Required Column to create pivot table
        mb51_pd = mb51_pd[['Material Document', 'Qty in unit of entry', 'Movement type']]
        # print(mb51_pd)
        # write logic to filter only
        movement_types_list = main_config['MB51_Movement_types_list']
        # print(movement_types_list)
        # print(type(movement_types_list))
        # mb51_data_with_movement_type = pd.DataFrame(columns=mb51_pd.columns.values.tolist())
        # for index, row in mb51_pd.iterrows():
        #     if str(row['Movement type']) in movement_types_list:
        #         mb51_data_with_movement_type = mb51_data_with_movement_type.append(row)
        movement_types_list = movement_types_list.strip('][').split(',')
        movement_types_list = [int(item.strip()) for item in movement_types_list]
        # print(movement_types_list)
        mb51_data_with_movement_type = mb51_pd[mb51_pd['Movement type'].isin(movement_types_list)]

        if mb51_data_with_movement_type.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_mail"],
                      body=in_config["Body_mail_MB51_Movement_type_exception"])
            return "Body_mail_MB51_Movement_type_exception"

        # print(mb51_data_with_movement_type)
        print("Creating pivot table on Mb51 after movement types filtered")
        pivot2_df = pd.pivot_table(mb51_data_with_movement_type, index=["Material Document"],
                                   values="Qty in unit of entry",
                                   aggfunc=numpy.sum)
        pivot2_df = pivot2_df.reset_index()
        pivot2_df = pivot2_df.rename(columns={'Material Document': 'GR Document Number'})
        pivot1_df = pivot1_df.reset_index()
        # print(pivot1_df)
        # print(pivot2_df)

        # Merging 2 Pivots
        merge_pd = pd.merge(pivot1_df, pivot2_df, how="outer", on=["GR Document Number"]).fillna(0)
        # print(merge_pd)

        columns_list = merge_pd.columns.values.tolist()

        # create a new column - Success
        merge_pd['Check'] = 0

        # To Remove SettingWithCopyWarning error
        # modifying only one df, so suppressing this warning as it is not affecting
        pd.options.mode.chained_assignment = None

        # variance formula for index
        for index in merge_pd.index:
            present_quarter = merge_pd[columns_list[1]][index]
            mb51 = merge_pd[columns_list[2]][index]

            if round(present_quarter, 2) == round(mb51, 2):
                check = True
            else:
                check = False

            merge_pd['Check'][index] = check

        # Renaming Columns
        inventory_mapping_file = merge_pd.rename(
            columns={columns_list[0]: in_config["Rename_Column1"],
                     columns_list[1]: in_config["Rename_Column2"],
                     columns_list[2]: in_config["Rename_Column3"]})
        # print(inventory_mapping_file)

        purchase_sheet_sum = inventory_mapping_file[in_config["Rename_Column2"]].sum()
        # print(purchase_sheet_sum)
        mb51_sheet_sum = inventory_mapping_file[in_config["Rename_Column3"]].sum()
        # print(mb51_sheet_sum)

        # create variance column
        inventory_mapping_file['Variance'] = 0.0
        columns = inventory_mapping_file.columns.tolist()
        # variance formula for index
        for index in inventory_mapping_file.index:
            present_quarter_row_value = inventory_mapping_file[columns[1]][index]
            mb51_row_value = inventory_mapping_file[columns[2]][index]

            if present_quarter_row_value == 0:
                variance = 0
            else:
                variance = (present_quarter_row_value - mb51_row_value) / present_quarter_row_value
            inventory_mapping_file['Variance'][index] = variance
        # print(inventory_mapping_file)
        # Creating Output File
        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                inventory_mapping_file.to_excel(writer, sheet_name=main_config["Output_Inventory_Mapping_Sheetname"],
                                                index=False, startrow=9)
        except Exception as File_creation_error:
            logging.error("Exception occurred while creating inventory mapping sheet")
            raise File_creation_error

        # Opening and Reading Output File.
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Inventory_Mapping_Sheetname"]]

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        # for i in ascii_uppercase:
        #     ws[i + "1"].font = font_style

        # Adding Background Color
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")

        for j in ascii_uppercase:
            ws[j + "10"].fill = fill_pattern
            if j == 'E':
                break
        ws["C10"].fill = PatternFill("solid", fgColor="ffff00")

        # # Adding Auto Filter Option
        full_range = "A10:E" + str(ws.max_row)
        ws.auto_filter.ref = full_range

        # Auto Width Setting
        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.5
            if c == 'E':
                break

        # Passing Column Values
        ws['B9'] = purchase_sheet_sum
        ws['C9'] = mb51_sheet_sum

        # Number Formatting
        ws['B9'].number_format = '#,###.##'
        ws['C9'].number_format = '#,###.##'

        # Font-style
        ws['B9'].font = font_style
        ws['C9'].font = font_style

        for cell in ws['B']:
            cell.number_format = '#,###.##'
        for cell in ws['C']:
            cell.number_format = '#,###.##'
        for cell in ws['E']:
            cell.number_format = '0.0%'

        # Saving the File
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])

        try:
            inventory_mapping_business_exception(inventory_mapping_file, main_config)
        except Exception as inventory_mapping_business_exception_error:
            print("Exception occurred while creating inventory mapping business exception entries sheet: \n {0}".format(
                inventory_mapping_business_exception_error))

        return create_inventory_mapping_sheet

    # Excepting Errors here
    except FileNotFoundError as notfound_error:
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                  subject=in_config["subject_file_not_found"],
                  body=in_config["body_file_not_found"])
        print("Inventory Mapping Process-", notfound_error)
        logging.exception(notfound_error)

        return notfound_error
    except ValueError as value_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(value_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Inventory Mapping Process-", value_error)
        logging.exception(value_error)
        return value_error
    except BusinessException as business_error:
        print("Inventory Mapping Process-", business_error)
        logging.exception(business_error)
        return business_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Inventory Mapping Process-", type_error)
        logging.exception(type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Inventory Mapping Process-", error)
        logging.exception(error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Inventory Mapping Process-", key_error)
        logging.exception(key_error)
        return key_error
    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Please close the file")
        logging.exception(file_error)
        return file_error


if __name__ == "__main__":
    pass
