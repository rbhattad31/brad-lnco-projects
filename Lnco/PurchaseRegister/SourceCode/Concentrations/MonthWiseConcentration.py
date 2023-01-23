import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from string import ascii_lowercase
import os
import logging
from ReusableTasks.send_mail_reusable_task import send_mail


class BusinessException(Exception):
    pass


def month_concentration_top_weight(month_concentration_dataframe, main_config):
    # save grand total row to delete from datatable to sort
    # print(month_concentration_dataframe)
    grand_total_row = month_concentration_dataframe.tail(1)
    # print(grand_total_row)
    # delete last row from the grand_total_row
    month_concentration_dataframe.drop(month_concentration_dataframe.tail(1).index, inplace=True)
    # print("Deleted Grand total row")
    # sort the dataframe using column name
    month_concentration_dataframe.sort_values(by="Variance", ascending=False, inplace=True)
    # print(month_concentration_dataframe)
    month_concentration_weightage = pd.DataFrame(columns=month_concentration_dataframe.columns)
    # print("empty dataframe is created with columns")
    # print(month_concentration_weightage)
    sum_of_variance = 0
    for index, row in month_concentration_dataframe.iterrows():
        # print(float(row["Variance"]))
        if sum_of_variance < 0.60:
            sum_of_variance = sum_of_variance + float(row["Variance"])
            # print(sum_of_variance)
            month_concentration_weightage = month_concentration_weightage.append(row, ignore_index=True)
            # print("appended row")
        else:
            break
    # print(month_concentration_weightage)
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            month_concentration_weightage.to_excel(writer, sheet_name=main_config[
                "Output_Concentration_Weightage_sheetname"], index=False, startrow=2, startcol=7)
        print("Month wise concentration top weightage entries are logged in the output file")
    except Exception as File_creation_error:

        logging.error("Exception occurred while creating purchase type wise concentration sheet")
        raise File_creation_error

    # Load excel file
    workbook = openpyxl.load_workbook(main_config['Output_File_Path'])

    # Load sheet
    worksheet = workbook[main_config['Output_Concentration_Weightage_sheetname']]

    # Set column widths
    for column_letter in ['h', 'i', 'j']:
        column_length = max(len(str(cell.value)) for cell in worksheet[column_letter])
        worksheet.column_dimensions[column_letter].width = column_length * 1.25

    # row 3 font format, fill color

    calibri_11_black_bold = Font(name="Calibri", size=11, color="000000", bold=True)
    light_blue_solid_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
    thin = Side(border_style="thin", color='b1c5e7')
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cambria_11_black = Font(name='Calibri', size=11, color='000000', bold=False)

    for row in worksheet["H3:J3"]:
        for cell in row:
            cell.fill = light_blue_solid_fill
            cell.font = calibri_11_black_bold

    max_row = len(month_concentration_weightage.index)
    for row in worksheet["H" + str(3 + 1) + ":J" + str(max_row + 3)]:
        for cell in row:
            cell.font = cambria_11_black
            cell.border = thin_border

    # Number format implementation
    for cell in worksheet['I']:
        cell.number_format = "#,###,##"

    # Format Variance
    for cell in worksheet['J']:
        cell.number_format = '0.0%'

    workbook.save(main_config['Output_File_Path'])


def month_wise(main_config, in_config, present_quarter_pd):
    try:
        print(in_config)
        # Read Purchase Register Sheets
        read_present_quarter_pd = present_quarter_pd

        # Create Month Column
        read_present_quarter_pd['GR Posting Date'] = pd.to_datetime(read_present_quarter_pd['GR Posting Date'], errors='coerce')

        read_present_quarter_pd['GR Posting Date'] = read_present_quarter_pd['GR Posting Date'].dt.month_name().str[:3]

        read_present_quarter_pd['Month'] = read_present_quarter_pd['GR Posting Date']

        # Fetch To Address
        to_address = main_config["To_Mail_Address"]
        cc_address = main_config["CC_Mail_Address"]

        # Check Exception
        if read_present_quarter_pd.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        Present_Quarter_Sheet_col = read_present_quarter_pd.columns.values.tolist()
        for col in ['Month', "GR Amt.in loc.cur."]:
            if col not in Present_Quarter_Sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter rows
        date_1 = read_present_quarter_pd[read_present_quarter_pd['Month'].notna()]
        gr_amt = read_present_quarter_pd[read_present_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(date_1) == 0:
            subject = in_config["Date_Subject"]
            body = in_config["Date_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Month Column is empty")
        elif len(gr_amt) == 0:
            subject = in_config["GRAmt_Subject"]
            body = in_config["GRAmt_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # Sort based on month
        month_dict = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9,
                      'Oct': 10, 'Nov': 11, 'Dec': 12, 'Grand Total': 13}

        # Create Pivot Table Q3
        pivot_index = ["Month"]
        pivot_values = ["GR Amt.in loc.cur."]
        pivot_present_quarter = pd.pivot_table(read_present_quarter_pd, index=pivot_index, values=pivot_values, aggfunc=numpy.sum, margins=True,
                                               margins_name='Grand Total')
        # Get Pivot Column Names
        col_name = pivot_present_quarter.columns.values.tolist()

        # Rename Column
        pivot_present_quarter = pivot_present_quarter.rename(columns={col_name[0]: main_config["PresentQuarterColumnName"]})

        # Remove Index
        pivot_present_quarter = pivot_present_quarter.reset_index()

        # Sort based on month
        pivot_present_quarter = pivot_present_quarter.sort_values('Month', key=lambda x: x.apply(lambda x: month_dict[x]))
        pivot_present_quarter.reset_index(inplace=True, drop=True)
        # Assign Sheets
        pivot_sheet = pivot_present_quarter

        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Delete row of Q4 and Q3 columns values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[1]] == 0)])

        pd.options.mode.chained_assignment = None

        # Get maximum value
        total_value = pivot_sheet.iloc[-1:]
        total_value = total_value.iloc[0, 1]

        # Variance Formula
        variance_list = []
        for index in pivot_sheet.index:
            quarter_value = (pivot_sheet[col_name[1]][index])

            if total_value == 0:
                variance = 1
            else:
                variance = quarter_value / total_value

            variance_list.append(variance)

        # Create Variance Column
        pivot_sheet['Variance'] = variance_list

        # Log Sheet
        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                pivot_sheet.to_excel(writer, sheet_name=main_config["Output_Concentrations_Month_sheetname"], index=False,  startrow=16)
        except Exception as File_creation_error:
            logging.error("Exception occurred while creating month wise concentration sheet")
            raise File_creation_error

        # Check outfile creation
        if os.path.exists(main_config["Output_File_Path"]):
            print("Month Wise Concentration Logged")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Output file not generated")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Concentrations_Month_sheetname"]]

        # Format Q3
        for cell in ws['B']:
            cell.number_format = "#,###,##"

        # Format Variance
        for cell in ws['C']:
            cell.number_format = '0.0%'

        # Format Header
        format_font = Font(name="Calibri", size=11, color="000000", bold=True)
        for c in ascii_lowercase:
            ws[c + "17"].font = format_font

        # Format Footer
        m_row = ws.max_row
        for c in ascii_lowercase:
            ws[c + str(m_row)].font = format_font

        # Header Fill
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "17"].fill = format_fill
            if c == 'c':
                break

        # Footer Fill
        for c in ascii_lowercase:
            ws[c + str(m_row)].fill = format_fill
            if c == 'c':
                break

        # Set Width
        for c in ascii_lowercase:
            ws.column_dimensions[c].width = 15

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=17, min_col=1, max_row=ws.max_row, max_col=3):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        ws.merge_cells('A4:F4')
        ws.merge_cells('A5:F5')
        ws.merge_cells('A6:F6')
        ws.merge_cells('A7:F7')
        ws.merge_cells('A8:F8')
        ws.merge_cells('A9:F9')
        ws.merge_cells('A10:F10')
        ws.merge_cells('A11:F11')
        ws.merge_cells('A12:F12')
        ws.merge_cells('A13:F13')
        ws.merge_cells('A14:F14')

        # Headers implementation
        ws['A1'] = main_config['CompanyName']
        ws['A2'] = main_config['StatutoryAuditQuarter']
        ws['A3'] = main_config['FinancialYear']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        # Save File
        wb.save(main_config["Output_File_Path"])

        try:
            month_concentration_top_weight(pivot_sheet, main_config)
        except Exception as month_concentration_top_weight_error:
            logging.error("Exception occurred while creating month wise concentration top weight table")
            raise month_concentration_top_weight_error

        return ws

    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", file_error)
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Concentration Month Wise Process-", business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-",error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", key_error)
        return key_error


# Read config details and parse to dictionary
config = {}
main_config = {}
present_quarter_pd = pd.DataFrame()
if __name__ == "__main__":
    print(month_wise(main_config, config, present_quarter_pd))

