from string import ascii_uppercase
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import PatternFill, Font, Side, Border
from ReusableTasks.send_mail_reusable_task import send_mail
from openpyxl.utils import get_column_letter
import logging


class BusinessException(Exception):
    pass


def vendor_concentration_top_weight(vendor_concentration_dataframe, main_config):
    # save grand total row to delete from datatable to sort
    # print(vendor_concentration_dataframe)
    grand_total_row = vendor_concentration_dataframe.tail(1)
    # print(grand_total_row)
    # delete last row from the grand_total_row
    vendor_concentration_dataframe.drop(vendor_concentration_dataframe.tail(1).index, inplace=True)
    # print("Deleted Grand total row")
    # sort the dataframe using column name
    vendor_concentration_dataframe.sort_values(by="Percentage", ascending=False, inplace=True)
    # print(vendor_concentration_dataframe)
    vendor_concentration_weightage = pd.DataFrame(columns=vendor_concentration_dataframe.columns)
    # print("empty dataframe is created with columns")
    # print(vendor_concentration_weightage)
    sum_of_variance = 0
    for index, row in vendor_concentration_dataframe.iterrows():
        # print(float(row["Percentage"]))
        if sum_of_variance < 0.60:
            sum_of_variance = sum_of_variance + float(row["Percentage"])
            # print(sum_of_variance)
            vendor_concentration_weightage = vendor_concentration_weightage.append(row, ignore_index=True)
            # print("appended row")
        else:
            break
    # print(vendor_concentration_weightage)
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            vendor_concentration_weightage.to_excel(writer, sheet_name=main_config[
                "Output_Concentration_Weightage_sheetname"], index=False, startrow=2, startcol=22)
            print("Vendor wise concentration top weightage entries are logged in the output file")

    except Exception as File_creation_error:
        logging.error("Exception occurred while creating Vendor wise concentration sheet")
        raise File_creation_error

    # Load excel file
    workbook = openpyxl.load_workbook(main_config['Output_File_Path'])

    # Load sheet
    worksheet = workbook[main_config['Output_Concentration_Weightage_sheetname']]

    # Set column widths
    for column_letter in ['w', 'x', 'y', 'z']:
        column_length = max(len(str(cell.value)) for cell in worksheet[column_letter])
        worksheet.column_dimensions[column_letter].width = column_length * 1.25

    # row 3 font format, fill color

    calibri_11_black_bold = Font(name="Calibri", size=11, color="000000", bold=True)
    light_blue_solid_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
    thin = Side(border_style="thin", color='b1c5e7')
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cambria_11_black = Font(name='Calibri', size=11, color='000000', bold=False)

    for row in worksheet["w3:z3"]:
        for cell in row:
            cell.fill = light_blue_solid_fill
            cell.font = calibri_11_black_bold

    max_row = len(vendor_concentration_weightage.index)
    for row in worksheet["W" + str(3 + 1) + ":Z" + str(max_row + 3)]:
        for cell in row:
            cell.font = cambria_11_black
            cell.border = thin_border

    # Number format implementation
    for cell in worksheet['Y']:
        cell.number_format = "#,###,##"

    # Format Variance
    for cell in worksheet['Z']:
        cell.number_format = '0.0%'

    workbook.save(main_config['Output_File_Path'])


def con_vendor_wise(main_config, in_config, present_quarter_pd):
    try:
        read_present_quarter_pd = present_quarter_pd

        read_present_quarter_pd = read_present_quarter_pd[["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]]

        amount = read_present_quarter_pd[read_present_quarter_pd["GR Amt.in loc.cur."].notna()]
        vendor_no = read_present_quarter_pd[read_present_quarter_pd["Vendor No."].notna()]

        if read_present_quarter_pd.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Sourcefile_subject"],
                      body=in_config["Body_mail1"])
            raise BusinessException("Sheet is empty")
        elif len(amount) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Gr_amount"],
                      body=in_config["Gr_amount_body"])
            raise BusinessException("Empty column")

        elif len(vendor_no) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Vendor_No"],
                      body=in_config["Vendor_No_body"])
            raise BusinessException("Vendor No column  missed")

        else:
            pass

        con_vendor_sheet = pd.pivot_table(read_present_quarter_pd, index=["Vendor No.", "Vendor Name"],
                                          values="GR Amt.in loc.cur.", aggfunc=numpy.sum,
                                          margins=True, margins_name="Grand Total", sort=True)
        con_vendor_sheet = con_vendor_sheet.reset_index()
        # print(con_vendor_sheet)

        con_vendor_sheet = con_vendor_sheet.replace(numpy.nan, 0, regex=True)
        name_of_column = con_vendor_sheet.columns.values.tolist()
        con_vendor_sheet.drop(con_vendor_sheet.index[(con_vendor_sheet[name_of_column[2]] <= 0)], inplace=True)
        con_vendor_sheet = con_vendor_sheet.replace(numpy.nan, '', regex=True)

        grand_total = con_vendor_sheet[name_of_column[2]].values[-1]
        con_vendor_sheet[name_of_column[2]].values[-1] = 0
        con_vendor_sheet.sort_values(by=name_of_column[2], axis=0, ascending=False, inplace=True)
        con_vendor_sheet["Percentage"] = ""
        pd.options.mode.chained_assignment = None
        for index in con_vendor_sheet.index:
            vendor_amount = con_vendor_sheet[name_of_column[2]][index]

            if vendor_amount == 0:
                percentage = 1
            else:
                percentage = vendor_amount / grand_total
            con_vendor_sheet["Percentage"][index] = percentage
        con_vendor_sheet[name_of_column[2]].values[-1] = grand_total
        con_vendor_sheet = con_vendor_sheet.rename(columns={name_of_column[2]: main_config["PresentQuarterColumnName"]})

        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                con_vendor_sheet.to_excel(writer, sheet_name=main_config["Output_Concentration_Vendor_sheetname"],
                                          index=False, startrow=17)

        except Exception as File_creation_error:
            logging.error("Exception occurred while creating Vendor wise concentration sheet")
            raise File_creation_error

        try:
            vendor_concentration_top_weight(con_vendor_sheet, main_config)
        except Exception as purchase_concentration_top_weight_error:
            logging.error("Exception occurred while creating purchase type wise concentration top weight table")
            raise purchase_concentration_top_weight_error

        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Concentration_Vendor_sheetname"]]

        for cell in ws["C"]:
            cell.number_format = "#,###.##"
        for cell in ws["D"]:
            cell.number_format = "##.##%"

        full_range = "A18:" + get_column_letter(ws.max_column) \
                     + str(ws.max_row)

        ws.auto_filter.ref = full_range

        font_style = Font(name="Cambria", size=13, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "18"].font = font_style
        for c in ascii_uppercase:
            ws[c + str(ws.max_row)].font = font_style
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "18"].fill = fill_pattern
            if c == "D":
                break

        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 'D':
                break

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=18, min_col=1, max_row=ws.max_row, max_col=4):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Cell merge for headers implementation
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        ws.merge_cells('A4:F4')
        ws.merge_cells('A5:F5')
        ws.merge_cells('A6:F6')
        ws.merge_cells('A7:F7')
        ws.merge_cells('A8:F8')
        ws.merge_cells('A9:F9')
        ws.merge_cells('A10:F10')
        ws.merge_cells('A11:F11')
        ws.merge_cells('A12:F12')
        ws.merge_cells('A13:F13')
        ws.merge_cells('A14:F14')

        # Headers implementation
        ws['A1'] = main_config['CompanyName']
        ws['A2'] = main_config['StatutoryAuditQuarter']
        ws['A3'] = main_config['FinancialYear']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False

        wb.save(main_config["Output_File_Path"])
        wb.close()
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        print(wb.sheetnames)

    except SyntaxError as s_error:
        print("SyntaxError" + str(s_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Syn1"],
                  body=in_config["Synody_1"])
        return s_error
    except FileNotFoundError as f_error:
        print("FileNotFoundError" + str(f_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["File_N1"],
                  body=in_config["File_N1_body"])
        return f_error
    except NameError as n_error:
        print("NameError" + str(n_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Name_E1"],
                  body=in_config["Name_E1_body"])
        return n_error
    except KeyError as k_error:
        print("KeyError" + str(k_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Key1"],
                  body=in_config["Key1_body"])
        return k_error
    except ValueError as v_error:
        print("ValueError" + str(v_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Value_E1"],
                  body=in_config["Value_E1_body"])
        return v_error
    except AttributeError as a_error:
        print("AttributeError" + str(a_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                  subject=in_config["AttributeError1"],
                  body=in_config["AttributeError1_body"])
        return a_error
    except TypeError as t_error:
        print("TypeError" + str(t_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Type_E1"],
                  body=in_config["Type_E1_body"])
        return t_error
    except PermissionError as p_error:
        print("PermissionError" + str(p_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Perm_E1"],
                  body=in_config["Perm_E1_body"])
        return p_error
    except (ImportError, MemoryError, RuntimeError, Exception) as error:
        print("SystemError" + str(error))
        return error


config = {}
main_config = {}
present_quarter_pd = pd.DataFrame()
if __name__ == "__main__":
    print(con_vendor_wise(main_config, config, present_quarter_pd))
