# Importing Libraries

import pandas as pd
import numpy

import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from string import ascii_uppercase
from ReusableTasks.send_mail_reusable_task import send_mail
import logging


class BusinessException(Exception):
    pass


def purchase_comparatives_top_weight(purchase_comparatives_dataframe, main_config):
    # save grand total row to delete from datatable to sort
    # print(purchase_concentration_dataframe)
    grand_total_row = purchase_comparatives_dataframe.tail(1)
    variance = float(grand_total_row['Variance'])
    # print(grand_total_row)
    # delete last row from the grand_total_row
    purchase_comparatives_dataframe.drop(purchase_comparatives_dataframe.tail(1).index, inplace=True)
    # print("Deleted Grand total row")
    # sort the dataframe using column name
    purchase_comparatives_dataframe.sort_values(by="Variance", ascending=False, inplace=True)
    # print(purchase_concentration_dataframe)
    purchase_comparatives_weightage = pd.DataFrame(columns=purchase_comparatives_dataframe.columns)
    # print("empty dataframe is created with columns")
    for index, row in purchase_comparatives_dataframe.iterrows():
        # print(float(row["Variance"]))
        if float(row['Variance']) > variance:
            # print(sum_of_variance)
            purchase_comparatives_weightage = purchase_comparatives_weightage.append(row, ignore_index=True)
            # print("appended row")
        else:
            continue
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            purchase_comparatives_weightage.to_excel(writer, sheet_name=main_config[
                "Output_Comparatives_Weightage_sheetname"], index=False, startrow=2, startcol=1)
            print("purchase type concentration top weightage entries are logged in the output file")

    except Exception as File_creation_error:
        logging.error("Exception occurred while creating purchase type wise concentration sheet: \n {0}".format(
            File_creation_error))
        raise File_creation_error

    # Load excel file
    workbook = openpyxl.load_workbook(main_config['Output_File_Path'])

    # Load sheet
    worksheet = workbook[main_config['Output_Comparatives_Weightage_sheetname']]

    # Set column widths
    for column_letter in ['b', 'c', 'd', 'e', 'f']:
        column_length = max(len(str(cell.value)) for cell in worksheet[column_letter])
        worksheet.column_dimensions[column_letter].width = column_length * 1.25

    # row 3 font format, fill color

    calibri_11_black_bold = Font(name="Calibri", size=11, color="000000", bold=True)
    light_blue_solid_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
    thin = Side(border_style="thin", color='b1c5e7')
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cambria_11_black = Font(name='Calibri', size=11, color='000000', bold=False)

    for row in worksheet["B3:F3"]:
        for cell in row:
            cell.fill = light_blue_solid_fill
            cell.font = calibri_11_black_bold

    max_row = len(purchase_comparatives_weightage.index)
    for row in worksheet["B" + str(3 + 1) + ":F" + str(max_row + 3)]:
        for cell in row:
            cell.font = cambria_11_black
            cell.border = thin_border

    # Number format implementation
    for cell in worksheet['D']:
        cell.number_format = "#,###,##"

    for cell in worksheet['E']:
        cell.number_format = "#,###,##"

    # Format Variance
    for cell in worksheet['F']:
        cell.number_format = '0.0%'

    print(workbook.sheetnames)
    workbook.save(main_config['Output_File_Path'])


def major_vendor_analysis(main_config, present_quarter_final_pivot_pd, present_quarter_pd, previous_quarter_pd):
    present_quarter_final_pivot_pd.drop(present_quarter_final_pivot_pd.tail(1).index, inplace=True)
    # present_quarter_columns_list = present_quarter_pd.columns.tolist()
    # print(present_quarter_columns_list)
    major_vendor_analysis_pd = pd.DataFrame(
        columns=['Valuation Class', 'Valuation Class Text', 'Vendor Name', 'GR Amt.in loc.cur._x',
                 'GR Amt.in loc.cur._y'])
    for index, row in present_quarter_final_pivot_pd.iterrows():
        # print(row[0])
        # print(row[1])
        # print(row[2])
        if row[2] == 0 or row[1] == 0:
            # print("Gr amount or Valuation class text field is empty ")
            continue
        # temp_present_quarter_pd = pd.DataFrame(columns=present_quarter_columns_list)
        # print(temp_present_quarter_pd)
        temp_present_quarter_pd = present_quarter_pd[present_quarter_pd['Valuation Class Text'].isin([row[1]])]
        # print(temp_present_quarter_pd)
        temp_present_quarter_pd = pd.pivot_table(temp_present_quarter_pd,
                                                 index=["Valuation Class", "Valuation Class Text", "Vendor Name"],
                                                 values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=False)
        # print(temp_present_quarter_pd)
        temp_present_quarter_pd.sort_values(by="GR Amt.in loc.cur.", ascending=False, inplace=True)
        # print(temp_present_quarter_pd)
        temp_present_quarter_pd = temp_present_quarter_pd.head(5)
        # print(temp_present_quarter_pd)
        temp_present_quarter_pd.reset_index(inplace=True)
        # print(temp_present_quarter_pd)

        temp_previous_quarter_pd = previous_quarter_pd[previous_quarter_pd['Valuation Class Text'].isin([row[1]])]
        if len(temp_previous_quarter_pd.index) == 0:
            # print("Length of dataframe is zero")
            continue
        # print(temp_previous_quarter_pd)
        temp_previous_quarter_pd = pd.pivot_table(temp_previous_quarter_pd,
                                                  index=["Valuation Class", "Valuation Class Text", "Vendor Name"],
                                                  values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=False)
        # print(temp_previous_quarter_pd)
        temp_merge_pd = pd.merge(temp_present_quarter_pd, temp_previous_quarter_pd, how='left',
                                 on=["Valuation Class", "Valuation Class Text", "Vendor Name"]).fillna(0)
        # print(temp_merge_pd)
        major_vendor_analysis_pd = major_vendor_analysis_pd.append(temp_merge_pd)

    # print(major_vendor_analysis_pd)
    major_vendor_column_names = major_vendor_analysis_pd.columns.values.tolist()
    # print(major_vendor_column_names)

    # create a new column - Success
    major_vendor_analysis_pd['Variance'] = 0
    # print(major_vendor_analysis_pd)
    major_vendor_analysis_pd.reset_index(drop=True, inplace=True)
    # print(major_vendor_analysis_pd)
    pd.options.mode.chained_assignment = None

    # variance formula implementation using index
    for index, row in major_vendor_analysis_pd.iterrows():
        # print(row)
        present_quarter_row_value = major_vendor_analysis_pd[major_vendor_column_names[3]][index]
        # print('present_quarter_row_value')
        # print(present_quarter_row_value)
        previous_quarter_row_value = major_vendor_analysis_pd[major_vendor_column_names[4]][index]
        # print('previous_quarter_row_value')
        # print(previous_quarter_row_value)
        if previous_quarter_row_value == 0:
            variance = 1
        else:
            variance = (present_quarter_row_value - previous_quarter_row_value) / previous_quarter_row_value
        # print(variance)
        major_vendor_analysis_pd['Variance'][index] = variance
        # print(major_vendor_analysis_pd['Variance'][index])
    # print(major_vendor_analysis_pd)
    major_vendor_analysis_pd.drop(['Valuation Class'], axis=1, inplace=True)
    # print(major_vendor_analysis_pd)
    present_quarter_column_name = main_config['PresentQuarterColumnName']
    previous_quarter_column_name = main_config['PreviousQuarterColumnName']

    major_vendor_analysis_pd.set_axis(["Purchase Type", "Major Vendor", "₹ in " + present_quarter_column_name,
                                       "₹ in " + previous_quarter_column_name, "Variance"], axis='columns', inplace=True)
    # print(major_vendor_analysis_pd)
    # major_vendor_analysis_pd.columns = ["Purchase Type", "Major Vendor", "₹ in current quarter",
    #                                     "₹ in Previous quarter", "Variance"]
    major_vendor_column_names = major_vendor_analysis_pd.columns.values.tolist()
    # print(major_vendor_column_names)
    major_vendor_analysis_pd.drop(major_vendor_analysis_pd.loc[(major_vendor_analysis_pd[major_vendor_column_names[2]] == 0) & (major_vendor_analysis_pd[major_vendor_column_names[3]] == 0)].index, inplace=True)
    # print(major_vendor_analysis_pd)
    major_vendor_analysis_pd.fillna(0, inplace=True)
    major_vendor_analysis_pd = pd.DataFrame(major_vendor_analysis_pd).set_index(["Purchase Type", "Major Vendor"])
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            major_vendor_analysis_pd.to_excel(writer, sheet_name=main_config[
                "Output_Major_Vendor_analysis_Sheet_name"], index=True, startrow=1)
        print("Major Vendor Analysis sheet is created in output file")

    except Exception as File_creation_error:
        logging.error(
            "Exception occurred while creating Major Vendor analysis sheet \n\t {0}".format(File_creation_error))
        raise File_creation_error

    # Load excel file
    workbook = openpyxl.load_workbook(main_config['Output_File_Path'])

    # Load sheet
    worksheet = workbook[main_config['Output_Major_Vendor_analysis_Sheet_name']]

    # Set column widths
    for column_letter in ['a', 'b', 'c', 'd', 'e']:
        column_length = max(len(str(cell.value)) for cell in worksheet[column_letter])
        worksheet.column_dimensions[column_letter].width = column_length * 1.25

    # row 3 font format, fill color

    calibri_11_black_bold = Font(name="Calibri", size=11, color="000000", bold=True)
    light_blue_solid_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
    thin = Side(border_style="thin", color='000000')
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cambria_11_black = Font(name='Calibri', size=11, color='000000', bold=False)

    for row in worksheet["A2:E2"]:
        for cell in row:
            cell.fill = light_blue_solid_fill
            cell.font = calibri_11_black_bold

    max_row = len(major_vendor_analysis_pd.index)
    for row in worksheet["A3" + ":E" + str(max_row + 2)]:
        for cell in row:
            cell.font = cambria_11_black
    for row in worksheet["C3" + ":E" + str(max_row + 2)]:
        for cell in row:
            cell.border = thin_border

    # Number format implementation
    for cell in worksheet['C']:
        cell.number_format = "#,###,##"

    for cell in worksheet['D']:
        cell.number_format = "#,###,##"

    # Format Variance
    for cell in worksheet['E']:
        cell.number_format = '0.0%'
    print(workbook.sheetnames)
    workbook.save(main_config['Output_File_Path'])


# Defining a Function
def create_purchase_type_wise(main_config, in_config, present_quarter_pd, previous_quarter_pd):
    try:
        # Read Purchase Register Sheets
        read_present_quarter_pd = present_quarter_pd
        # read_present_quarter_pd = read_present_quarter_pd.loc[:, ~read_present_quarter_pd.columns.duplicated(keep='first')]

        read_previous_quarter_pd = previous_quarter_pd
        # read_previous_quarter_pd = read_previous_quarter_pd.loc[:, ~read_previous_quarter_pd.columns.duplicated(keep='first')]

        # Check Exception
        if read_present_quarter_pd.shape[0] == 0 or read_previous_quarter_pd.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Input Sheet Data is empty")

        previous_quarter_columns_list = read_previous_quarter_pd.columns.values.tolist()
        for col in ["Valuation Class", "Valuation Class Text", "GR Amt.in loc.cur."]:
            if col not in previous_quarter_columns_list:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")

        present_quarter_columns_list = read_present_quarter_pd.columns.values.tolist()
        for col in ["Valuation Class", "Valuation Class Text", "GR Amt.in loc.cur."]:
            if col not in present_quarter_columns_list:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        present_valuation_class_pd = read_present_quarter_pd[read_present_quarter_pd['Valuation Class'].notna()]
        present_valuation_class_text_pd = read_present_quarter_pd[
            read_present_quarter_pd['Valuation Class Text'].notna()]
        present_gr_amount_pd = read_present_quarter_pd[read_present_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(present_valuation_class_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Valuation Class_subject"],
                      body=in_config["Valuation Class_Body"])
            raise BusinessException("Valuation Class Column is empty")

        elif len(present_valuation_class_text_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Valuation Class Text_Subject"],
                      body=in_config["Valuation Class Text_Body"])
            raise BusinessException("Valuation Class Text Column is empty")

        elif len(present_gr_amount_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        previous_valuation_class_pd = read_previous_quarter_pd[read_previous_quarter_pd['Valuation Class'].notna()]
        previous_valuation_class_text_pd = read_previous_quarter_pd[
            read_previous_quarter_pd['Valuation Class Text'].notna()]
        previous_gr_amount_pd = read_previous_quarter_pd[read_previous_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(previous_valuation_class_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Valuation Class_subject"],
                      body=in_config["Valuation Class_Body"])
            raise BusinessException("Valuation Class Column is empty")

        elif len(previous_valuation_class_text_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Valuation Class Text_Subject"],
                      body=in_config["Valuation Class Text_Body"])
            raise BusinessException("Valuation Class Text Column is empty")

        elif len(previous_gr_amount_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # create pivot table
        present_quarter_final_pivot_pd = pd.pivot_table(read_present_quarter_pd,
                                                        index=["Valuation Class", "Valuation Class Text"],
                                                        values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                                        margins_name="Grand Total")

        # reset "indices created during pivot table creation" - for merging
        present_quarter_final_pivot_pd = present_quarter_final_pivot_pd.reset_index()

        # read previous quarters final working file - pd will be replaced with Nan in any blank cells
        previous_quarter_final_pivot_pd = pd.pivot_table(read_previous_quarter_pd,
                                                         index=["Valuation Class", "Valuation Class Text"],
                                                         values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                                         margins_name="Grand Total")

        # replace Nan with blank
        previous_quarter_final_pivot_pd = previous_quarter_final_pivot_pd.replace(numpy.nan, '', regex=True)
        previous_quarter_final_pivot_pd = previous_quarter_final_pivot_pd.reset_index()

        # merging present and previous quarter purchase type wise data -  pd will be replaced with Nan in any blank cells
        merge_pd = pd.merge(present_quarter_final_pivot_pd, previous_quarter_final_pivot_pd, how="outer",
                            on=["Valuation Class", "Valuation Class Text"])

        # print(merge_pd)
        # replacing all Nan's with zeros in Present and previous Quarter's values columns
        merge_pd = merge_pd.replace(numpy.nan, 0, regex=True)
        # print(merge_pd)

        merge_pd_column_list = merge_pd.columns.values.tolist()
        # returns as ['Valuation Class', 'Valuation Class Text', 'GR Amt.in loc.cur.', 'Previous Quarter']

        # dropping columns present and previous quarters both have values as zero
        merge_pd.drop(
            merge_pd.index[(merge_pd[merge_pd_column_list[2]] == 0) & (merge_pd[merge_pd_column_list[3]] == 0)],
            inplace=True)

        # create a new column - Success
        merge_pd['Variance'] = 0
        # print(merge_pd)
        pd.options.mode.chained_assignment = None

        # variance formula implementation using index
        for index in merge_pd.index:
            present_quarter_row_value = merge_pd[merge_pd_column_list[2]][index]
            previous_quarter_row_value = merge_pd[merge_pd_column_list[3]][index]
            if previous_quarter_row_value == 0:
                variance = 1
            else:
                variance = (present_quarter_row_value - previous_quarter_row_value) / previous_quarter_row_value
            merge_pd['Variance'][index] = variance

        # copy present quarter Amount column Grand total, set it as zero, sort the data frame and reassign the value.
        grand_total = merge_pd[merge_pd_column_list[2]].values[-1]
        merge_pd[merge_pd_column_list[2]].values[-1] = 0

        merge_pd[merge_pd_column_list[2]].values[-1] = grand_total

        purchase_type_wise_comparatives_pd = merge_pd.rename(
            columns={merge_pd_column_list[2]: main_config["PresentQuarterColumnName"]})
        purchase_type_wise_comparatives_pd = purchase_type_wise_comparatives_pd.rename(
            columns={merge_pd_column_list[3]: main_config["PreviousQuarterColumnName"]})
        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                purchase_type_wise_comparatives_pd.to_excel(writer, sheet_name=main_config[
                    "Output_Comparatives_Purchase_sheetname"],
                                                            index=False, startrow=16)
        except Exception as File_creation_error:
            logging.error("Exception occurred while creating purchase type wise comparatives sheet")
            raise File_creation_error

        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Comparatives_Purchase_sheetname"]]

        for cell in ws['C']:
            cell.number_format = '#,###.##'
        for cell in ws['D']:
            cell.number_format = '#,###.##'
        for cell in ws['E']:
            cell.number_format = '0.0%'

        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "17"].font = font_style

        m_row = ws.max_row
        for c in ascii_uppercase:
            ws[c + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "17"].fill = fill_pattern
            if c == 'E':
                break

        for c in ascii_uppercase:
            ws.column_dimensions[c].width = 25
        ws.column_dimensions["E"].width = 15

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=17, min_col=1, max_row=ws.max_row, max_col=5):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        ws.merge_cells('A4:E4')
        ws.merge_cells('A5:E5')
        ws.merge_cells('A6:E6')
        ws.merge_cells('A7:E7')
        ws.merge_cells('A8:E8')
        ws.merge_cells('A9:E9')
        ws.merge_cells('A10:E10')
        ws.merge_cells('A11:E11')
        ws.merge_cells('A12:E12')
        ws.merge_cells('A13:E13')
        ws.merge_cells('A14:E14')

        # Headers implementation
        ws['A1'] = main_config['CompanyName']
        ws['A2'] = main_config['StatutoryAuditQuarter']
        ws['A3'] = main_config['FinancialYear']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])

        try:
            purchase_comparatives_top_weight(purchase_type_wise_comparatives_pd, main_config)
        except Exception as major_vendor_analysis_error:
            print("Exception occurred while creating purchase type wise comparatives top weight sheet: \n {0}".format(
                major_vendor_analysis_error))
        try:
            major_vendor_analysis(main_config, present_quarter_final_pivot_pd, present_quarter_pd, previous_quarter_pd)
        except Exception as major_vendor_analysis_error:
            print("Exception occurred while creating major vendor analysis sheet: \n {0}".format(
                major_vendor_analysis_error))

        return purchase_type_wise_comparatives_pd

    # Excepting Errors here
    except FileNotFoundError as notfound_error:
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                  subject=in_config["subject_file_not_found"],
                  body=in_config["body_file_not_found"])
        print("Purchase Type Wise Comparatives Process-", notfound_error)
        return notfound_error
    except ValueError as V_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(V_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", V_error)
        return V_error
    except BusinessException as business_error:
        print("Purchase Type Wise Comparatives Process-", business_error)
        return business_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", key_error)
        return key_error
    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Please close the file")
        return file_error


main_config = {}
config = {}
present_quarter_pd = pd.DataFrame()
previous_quarter_pd = pd.DataFrame()
if __name__ == "__main__":
    print(create_purchase_type_wise(main_config, config, present_quarter_pd, previous_quarter_pd))
