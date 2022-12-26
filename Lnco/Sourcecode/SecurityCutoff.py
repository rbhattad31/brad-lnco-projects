from string import ascii_lowercase

import pandas as pd
import openpyxl
import logging

from openpyxl.styles import PatternFill, Font


# import datetime


def security_cutoff(main_config, security_cutoff_dataframe):
    date_list = main_config['security_cutoff_date_list']
    # print(date_list)
    month_list = ['Mar', 'Jun', 'Sep', 'Dec']
    # sort datatable as per date in ascending order
    gr_posting_date_default_name = main_config['gr_posting_date_default_name']
    gr_posting_date_client_column_name = main_config[gr_posting_date_default_name]
    # print(gr_posting_date_client_column_name)

    security_cutoff_dataframe.sort_values(by=gr_posting_date_client_column_name, inplace=True, ascending=True)
    # print(security_cutoff_dataframe)

    security_cutoff_dataframe[gr_posting_date_client_column_name] = pd.to_datetime(
        security_cutoff_dataframe[gr_posting_date_client_column_name], errors='coerce')

    security_cutoff_dataframe['Month'] = security_cutoff_dataframe[gr_posting_date_client_column_name].dt.month_name().str[:3]
    security_cutoff_dataframe['Day'] = security_cutoff_dataframe[gr_posting_date_client_column_name].dt.strftime("%d")

    security_cutoff_dataframe_output = pd.DataFrame(columns=security_cutoff_dataframe.columns.values.tolist())
    for index, row in security_cutoff_dataframe.iterrows():
        if row['Month'] in month_list and row['Day'] in date_list:
            security_cutoff_dataframe_output = security_cutoff_dataframe_output.append(row)

    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            security_cutoff_dataframe_output.to_excel(writer,
                                                      sheet_name=main_config["Output_Security_Cutoff_Sheet_name"],
                                                      index=False, startrow=1)
    except Exception as File_creation_error:
        logging.error("Exception occurred while creating inventory mapping sheet")
        raise File_creation_error

        # Load Sheet in openpyxl
    wb = openpyxl.load_workbook(main_config["Output_File_Path"])
    ws = wb[main_config["Output_Security_Cutoff_Sheet_name"]]
    max_column = ws.max_column
    # print(max_column)
    # Header Fill
    calibri_11_black_bold = Font(name="Calibri", size=11, color="000000", bold=True)
    format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
    for column in range(1, max_column+1):
        ws.cell(2, column).fill = format_fill
        ws.cell(2, column).font = calibri_11_black_bold
        if column == max_column:
            break
    wb.save(main_config['Output_File_Path'])


if __name__ == "__main__":
    pass
