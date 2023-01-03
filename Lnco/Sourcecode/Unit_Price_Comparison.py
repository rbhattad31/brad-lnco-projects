import logging
from string import ascii_lowercase
from send_mail_reusable_task import send_mail

import numpy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.styles import numbers
import warnings

warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)


class BusinessException(Exception):
    pass


def change_in_unit_price_percentage(main_config, unit_price_present_previous_merge_pd):
    percentage = main_config['unit_price_comparison_unit_price_exception_percentage'] / 100
    unit_price_comparison_columns = unit_price_present_previous_merge_pd.columns.values.tolist()
    unit_price_change_percentage_column_name = unit_price_comparison_columns[14]
    # print(unit_price_change_percentage_column_name)

    change_in_unit_price_percentage_pd = pd.DataFrame(columns=unit_price_comparison_columns)
    for index, row in unit_price_present_previous_merge_pd.iterrows():
        if isinstance(row[unit_price_change_percentage_column_name], float):
            if row[unit_price_change_percentage_column_name] >= percentage:
                change_in_unit_price_percentage_pd = change_in_unit_price_percentage_pd.append(row, ignore_index=True)
            # print("appended row")
        else:
            continue
    # print(change_in_unit_price_percentage_pd)
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            change_in_unit_price_percentage_pd.to_excel(writer, sheet_name=main_config[
                "Output_Unit_Price_Exception_change_in_amount_sheetname"],
                                                        startrow=1, startcol=0, index=False)

    except Exception as File_creation_error:
        logging.error(
            "Exception occurred while creating unit price comparison exception sheet \n\t {0}".format(
                File_creation_error))
        raise File_creation_error

    wb = load_workbook(main_config["Output_File_Path"])
    ws = wb[main_config["Output_Unit_Price_Exception_change_in_amount_sheetname"]]
    present_quarter_column_name = main_config['PresentQuarterColumnName']
    previous_quarter_column_name = main_config['PreviousQuarterColumnName']

    for c in ascii_lowercase:
        column_length = max(len(str(cell.value)) for cell in ws[c])
        ws.column_dimensions[c].width = column_length * 1.25
        if c == 's':
            break

    cell = ws['F1']
    cell.value = present_quarter_column_name
    ws.merge_cells('F1:H1')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell = ws['I1']
    cell.value = previous_quarter_column_name
    ws.merge_cells('I1:K1')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell = ws['L1']
    cell.value = 'Increase/Decrease In Amount'
    ws.merge_cells('L1:S1')
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cambria_12_bold_black = Font(name="Cambria", size=12, bold=True, color="000000")
    ws['F1'].font = cambria_12_bold_black
    ws['I1'].font = cambria_12_bold_black
    ws['L1'].font = cambria_12_bold_black
    fill_solid_light_blue = PatternFill(patternType="solid", fgColor="ADD8E6")
    ws['F1'].fill = fill_solid_light_blue
    ws['L1'].fill = fill_solid_light_blue
    fill_solid_yellow = PatternFill(patternType="solid", fgColor="FFFF00")
    ws['I1'].fill = fill_solid_yellow

    for c in ascii_lowercase:
        ws[c + "2"].fill = fill_solid_light_blue
        if c == 's':
            break

    m_row = ws.max_row
    ws.auto_filter.ref = "A2:S" + str(m_row)
    # print("A2:S" + str(m_row))

    for cell in ws["E"]:
        cell.number_format = "#,###"
    for cell in ws["F"]:
        cell.number_format = "#,###"
    for cell in ws["G"]:
        cell.number_format = "#,###"
    for cell in ws["H"]:
        cell.number_format = "#,###"
    for cell in ws["I"]:
        cell.number_format = "#,###"
    for cell in ws["J"]:
        cell.number_format = "#,###"
    for cell in ws["K"]:
        cell.number_format = "#,###"
    for cell in ws["L"]:
        cell.number_format = "#,###"
    for cell in ws["M"]:
        cell.number_format = "#,###"
    for cell in ws["N"]:
        cell.number_format = "0"
    for cell in ws["O"]:
        cell.number_format = numbers.FORMAT_PERCENTAGE
    for cell in ws["P"]:
        cell.number_format = "0.00"
    for cell in ws["Q"]:
        cell.number_format = "0.00%"
    for cell in ws["R"]:
        cell.number_format = "0.00"
    for cell in ws["S"]:
        cell.number_format = "0.00%"

    print(wb.sheetnames)
    wb.save(main_config['Output_File_Path'])


def amount_quantity_percentage(main_config, unit_price_present_previous_merge_pd):
    percentage = main_config['unit_price_comparison_quantity_exception_percentage'] / 100
    unit_price_comparison_columns = unit_price_present_previous_merge_pd.columns.values.tolist()
    unit_price_change_percentage_column_name = unit_price_comparison_columns[16]
    # print(unit_price_change_percentage_column_name)

    change_in_amount_due_to_quantity_percentage_pd = pd.DataFrame(columns=unit_price_comparison_columns)
    for index, row in unit_price_present_previous_merge_pd.iterrows():
        if isinstance(row[unit_price_change_percentage_column_name], float):
            if row[unit_price_change_percentage_column_name] >= percentage:
                change_in_amount_due_to_quantity_percentage_pd = change_in_amount_due_to_quantity_percentage_pd.append(
                    row, ignore_index=True)
        else:
            continue
    # print(change_in_amount_due_to_quantity_percentage_pd)
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            change_in_amount_due_to_quantity_percentage_pd.to_excel(writer, sheet_name=main_config[
                "Output_Unit_Price_Exception_change_as_per_quantity_sheetname"],
                                                                    startrow=1, startcol=0, index=False)

    except Exception as File_creation_error:
        logging.error(
            "Exception occurred while creating unit price comparison exception sheet \n\t {0}".format(
                File_creation_error))
        raise File_creation_error

    wb = load_workbook(main_config["Output_File_Path"])
    ws = wb[main_config["Output_Unit_Price_Exception_change_as_per_quantity_sheetname"]]
    present_quarter_column_name = main_config['PresentQuarterColumnName']
    previous_quarter_column_name = main_config['PreviousQuarterColumnName']

    for c in ascii_lowercase:
        column_length = max(len(str(cell.value)) for cell in ws[c])
        ws.column_dimensions[c].width = column_length * 1.25
        if c == 's':
            break

    cell = ws['F1']
    cell.value = present_quarter_column_name
    ws.merge_cells('F1:H1')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell = ws['I1']
    cell.value = previous_quarter_column_name
    ws.merge_cells('I1:K1')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell = ws['L1']
    cell.value = 'Increase/Decrease In Amount'
    ws.merge_cells('L1:S1')
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cambria_12_bold_black = Font(name="Cambria", size=12, bold=True, color="000000")
    ws['F1'].font = cambria_12_bold_black
    ws['I1'].font = cambria_12_bold_black
    ws['L1'].font = cambria_12_bold_black
    fill_solid_light_blue = PatternFill(patternType="solid", fgColor="ADD8E6")
    ws['F1'].fill = fill_solid_light_blue
    ws['L1'].fill = fill_solid_light_blue
    fill_solid_yellow = PatternFill(patternType="solid", fgColor="FFFF00")
    ws['I1'].fill = fill_solid_yellow

    for c in ascii_lowercase:
        ws[c + "2"].fill = fill_solid_light_blue
        if c == 's':
            break

    m_row = ws.max_row
    ws.auto_filter.ref = "A2:S" + str(m_row)
    # print("A2:S" + str(m_row))

    for cell in ws["F"]:
        cell.number_format = "#,###"
    for cell in ws["G"]:
        cell.number_format = "#,###"
    for cell in ws["H"]:
        cell.number_format = "#,###"
    for cell in ws["I"]:
        cell.number_format = "#,###"
    for cell in ws["J"]:
        cell.number_format = "#,###"
    for cell in ws["K"]:
        cell.number_format = "#,###"
    for cell in ws["L"]:
        cell.number_format = "#,###"
    for cell in ws["M"]:
        cell.number_format = "#,###"
    for cell in ws["N"]:
        cell.number_format = "0"
    for cell in ws["O"]:
        cell.number_format = numbers.FORMAT_PERCENTAGE
    for cell in ws["P"]:
        cell.number_format = "0.00"
    for cell in ws["Q"]:
        cell.number_format = "0.00%"
    for cell in ws["R"]:
        cell.number_format = "0.00"
    for cell in ws["S"]:
        cell.number_format = "0.00%"

    print(wb.sheetnames)
    wb.save(main_config['Output_File_Path'])


def unit_price_high_low_average(main_config, unit_price_present_previous_merge_pd):
    purchase_remark = main_config["purchase_remark"]
    not_purchase_remark = main_config["not_purchase_remark"]
    unit_price_comparison_columns = unit_price_present_previous_merge_pd.columns.values.tolist()

    # the below line creates dataframe with duplicate columns
    unit_price_high_low_average_pd = unit_price_present_previous_merge_pd[
        [unit_price_comparison_columns[i] for i in range(0, 8)]]

    # remove the duplicate columns and Keep columns comes first
    unit_price_high_low_average_pd = unit_price_high_low_average_pd.loc[:, ~unit_price_high_low_average_pd.columns.duplicated()].copy()

    # remove rows of materials that are not purchased this quarter
    unit_price_high_low_average_pd = unit_price_high_low_average_pd[
        ~unit_price_high_low_average_pd["Remarks"].isin([not_purchase_remark])]

    output_pd = pd.DataFrame(columns=unit_price_high_low_average_pd.columns.values.tolist())
    output_pd['Average'] = 0
    unit_price_column_name = unit_price_high_low_average_pd.columns.values.tolist()[7]
    for index, row in unit_price_high_low_average_pd.iterrows():
        material_number = row["Material No."]
        # print(row["Material No."])
        if row["Material No."] in output_pd["Material No."].values:
            print("Material number is already evaluated")
            continue
        else:
            # filter rows with material number
            temp_df = unit_price_high_low_average_pd[unit_price_high_low_average_pd["Material No."] == material_number]
            temp_df['Average'] = 0
            # print(temp_df)
            unit_price_column_total = temp_df[unit_price_column_name].sum()
            # print(unit_price_column_total)
            number_of_entries_found = len(temp_df.index)
            # print(number_of_entries_found)
            unit_price_average = unit_price_column_total / number_of_entries_found
            # print(unit_price_average)
            temp_df.sort_values(by=unit_price_column_name, axis=0, ascending=False, inplace=True)
            # print(temp_df)

            max_row = temp_df.head(1)
            # print(max_row)
            max_row['Average'] = unit_price_average
            # print(max_row)
            output_pd = output_pd.append(max_row)

            if number_of_entries_found == 1:
                continue

            min_row = temp_df.tail(1)
            # print(min_row)
            min_row['Average'] = unit_price_average
            # print(min_row)
            output_pd = output_pd.append(min_row)

    print(output_pd)
    output_pd.drop(['Concat'], axis=1, inplace=True)
    print(output_pd)
    output_pd = pd.DataFrame(output_pd).set_index(["Material No.", "Valuation Class Text"])
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            output_pd.to_excel(writer, sheet_name=main_config["Output_Unit_Price_Exception_Unit_price_wise"],
                               startrow=1, startcol=0, index=True)

    except Exception as File_creation_error:
        logging.error(
            "Exception occurred while creating unit price comparison exception sheet as per unit price wise \n\t {0}".format(
                File_creation_error))
        raise File_creation_error

    wb = load_workbook(main_config["Output_File_Path"])
    ws = wb[main_config["Output_Unit_Price_Exception_Unit_price_wise"]]

    for c in ascii_lowercase:
        column_length = max(len(str(cell.value)) for cell in ws[c])
        ws.column_dimensions[c].width = column_length * 1.25
        if c == 'G':
            break

    fill_solid_light_blue = PatternFill(patternType="solid", fgColor="ADD8E6")

    for c in ascii_lowercase:
        ws[c + "2"].fill = fill_solid_light_blue
        if c == 'G':
            break

    m_row = ws.max_row
    ws.auto_filter.ref = "A2:G" + str(m_row + 2)
    # print("A2:S" + str(m_row))

    for cell in ws["F"]:
        cell.number_format = "#,###"
    for cell in ws["G"]:
        cell.number_format = "#,###"
    for cell in ws["H"]:
        cell.number_format = "#,###"
    for cell in ws["I"]:
        cell.number_format = "#,###"

    thin = Side(border_style="thin", color='000000')
    cambria_11_black = Font(name='Calibri', size=11, color='000000', bold=False)
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws["C3" + ":H" + str(m_row + 2)]:
        for cell in row:
            cell.border = thin_border
    for row in ws["A3" + ":H" + str(m_row + 2)]:
        for cell in row:
            cell.font = cambria_11_black

    print(wb.sheetnames)
    wb.save(main_config['Output_File_Path'])


def create_unit_price(main_config, in_config, present_quarter_pd, previous_quarter_pd):
    try:

        # Fetch To Address
        to_address = main_config["To_Mail_Address"]
        cc_address = main_config["CC_Mail_Address"]

        # Check Exception
        if present_quarter_pd.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        present_quarter_columns = present_quarter_pd.columns.values.tolist()
        for col in ["GR Amt.in loc.cur.", "GR Qty", "Material No.", "Valuation Class Text", "Vendor Name"]:
            if col not in present_quarter_columns:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        present_quarter_pivot = pd.pivot_table(present_quarter_pd,
                                               index=["Material No.", "Valuation Class Text", "Vendor Name"],
                                               values=["GR Amt.in loc.cur.", "GR Qty"], aggfunc=numpy.sum, margins=True,
                                               margins_name="Grand Total", sort=True)
        # Drop last column of a dataframe
        present_quarter_pivot = present_quarter_pivot[:-1]

        present_quarter_pivot = present_quarter_pivot.reset_index()
        present_quarter_pivot = present_quarter_pivot.replace(numpy.nan, 0, regex=True)
        columns = present_quarter_pivot.columns.values.tolist()
        numpy.seterr(divide='ignore')
        present_quarter_pivot['Unit Price'] = ""
        pd.options.mode.chained_assignment = None

        for index in present_quarter_pivot.index:
            gr_amount = present_quarter_pivot[columns[3]][index]
            gr_quantity = present_quarter_pivot[columns[4]][index]
            if gr_quantity != 0:
                unit_price = gr_amount / gr_quantity
                present_quarter_pivot['Unit Price'][index] = unit_price

        columns = present_quarter_pivot.columns.values.tolist()
        present_quarter_pivot = present_quarter_pivot.rename(
            columns={columns[3]: "GR Amt.in loc.cur.1"})
        present_quarter_pivot = present_quarter_pivot.rename(
            columns={columns[4]: "GR Qty1"})
        present_quarter_pivot = present_quarter_pivot.rename(
            columns={columns[5]: "Unit Price1"})

        present_quarter_pivot['Concat'] = ""
        present_quarter_pivot["Concat"] = present_quarter_pivot["Material No."].astype(str) + present_quarter_pivot[
            "Valuation Class Text"].astype(str) + present_quarter_pivot["Vendor Name"].astype(str)

        present_quarter_pivot = present_quarter_pivot[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.1",
             "GR Qty1", "Unit Price1"]]

        # Check Exception
        if previous_quarter_pd.shape[0] == 0:
            subject = in_config["EmptyInput_Subject1"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        previous_quarter_columns = previous_quarter_pd.columns.values.tolist()
        for col in ["GR Amt.in loc.cur.", "GR Qty", "Material No.", "Valuation Class Text", "Vendor Name"]:
            if col not in previous_quarter_columns:
                subject = in_config["ColumnMiss_Subject1"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        previous_quarter_pivot = pd.pivot_table(previous_quarter_pd, values=["GR Amt.in loc.cur.", "GR Qty"],
                                                index=["Material No.", "Valuation Class Text", "Vendor Name"],
                                                aggfunc={"GR Amt.in loc.cur.": numpy.sum, 'GR Qty': numpy.sum},
                                                margins=True)

        previous_quarter_pivot = previous_quarter_pivot.reset_index()
        previous_quarter_pivot = previous_quarter_pivot.replace(numpy.nan, 0, regex=True)

        # Drop last column of a dataframe
        previous_quarter_pivot = previous_quarter_pivot[:-1]
        columns = previous_quarter_pivot.columns.values.tolist()

        previous_quarter_pivot['Unit Price'] = ""
        pd.options.mode.chained_assignment = None
        for index in previous_quarter_pivot.index:
            gr_amount = previous_quarter_pivot[columns[3]][index]
            gr_quantity = previous_quarter_pivot[columns[4]][index]
            if gr_quantity != 0 and type(gr_amount) != 'str' and type(gr_quantity) != 'str':
                unit_price = gr_amount / gr_quantity
                previous_quarter_pivot['Unit Price'][index] = unit_price

        columns = previous_quarter_pivot.columns.values.tolist()
        previous_quarter_pivot = previous_quarter_pivot.rename(
            columns={columns[3]: "GR Amt.in loc.cur.2"})
        previous_quarter_pivot = previous_quarter_pivot.rename(
            columns={columns[4]: "GR Qty2"})
        previous_quarter_pivot = previous_quarter_pivot.rename(
            columns={columns[5]: "Unit Price2"})

        previous_quarter_pivot['Concat'] = ""
        previous_quarter_pivot["Concat"] = previous_quarter_pivot["Material No."].astype(str) + previous_quarter_pivot[
            "Valuation Class Text"].astype(str) + previous_quarter_pivot["Vendor Name"].astype(str)

        previous_quarter_pivot = previous_quarter_pivot[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.2", "GR Qty2",
             "Unit Price2"]]

        unit_price_present_quarter = pd.merge(present_quarter_pivot, previous_quarter_pivot, how="left",
                                              on=["Material No.", "Valuation Class Text", "Vendor Name", "Concat"],
                                              copy=False)
        unit_price_present_quarter = unit_price_present_quarter[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.1", "GR Qty1",
             "Unit Price1"]]
        unit_price_previous_quarter = pd.merge(previous_quarter_pivot, present_quarter_pivot, how="left",
                                               on=["Material No.", "Valuation Class Text", "Vendor Name", "Concat"],
                                               copy=False)
        unit_price_previous_quarter = unit_price_previous_quarter[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.2", "GR Qty2",
             "Unit Price2"]]
        # columns = unit_price_previous_quarter.columns.values.tolist()

        unit_price_present_previous_merge = pd.merge(unit_price_present_quarter, unit_price_previous_quarter,
                                                     how="outer",
                                                     on=["Material No.", "Valuation Class Text", "Vendor Name",
                                                         "Concat"], copy=False)

        unit_price_present_previous_merge = unit_price_present_previous_merge[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.1", "GR Qty1",
             "Unit Price1", "GR Amt.in loc.cur.2", "GR Qty2", "Unit Price2"]]
        unit_price_present_previous_merge = unit_price_present_previous_merge.reset_index()

        unit_price_present_previous_merge = unit_price_present_previous_merge.replace(numpy.nan, 0, regex=True)
        unit_price_present_previous_merge["Remarks"] = " "

        columns = unit_price_present_previous_merge.columns.values.tolist()
        columns.remove("Remarks")
        columns.insert(4, "Remarks")
        # Re-order columns
        unit_price_present_previous_merge = unit_price_present_previous_merge[columns]

        # Unit_Price_Comparison.sort_values
        columns = unit_price_present_previous_merge.columns.values.tolist()
        pd.options.mode.chained_assignment = None
        main_config["purchase_remark"] = "Purchased in the current quarter"
        main_config["not_purchase_remark"] = "Not purchased in the current quarter"
        for index in unit_price_present_previous_merge.index:
            if unit_price_present_previous_merge[columns[6]][index] == 0:
                unit_price_present_previous_merge['Remarks'][index] = "Not purchased in the current quarter"

            elif unit_price_present_previous_merge[columns[9]][index] == 0:
                unit_price_present_previous_merge['Remarks'][index] = "Purchased in the current quarter"
            else:
                pass

        unit_price_present_previous_merge['Increase/decrease in Amount'] = ""
        pd.options.mode.chained_assignment = None
        for index in unit_price_present_previous_merge.index:
            gr_amount = unit_price_present_previous_merge[columns[6]][index]
            gr_quantity = unit_price_present_previous_merge[columns[9]][index]
            unit_price = gr_amount - gr_quantity
            unit_price_present_previous_merge['Increase/decrease in Amount'][index] = unit_price

        unit_price_present_previous_merge['Increase/decrease in Quantity'] = ""
        pd.options.mode.chained_assignment = None
        for index in unit_price_present_previous_merge.index:
            gr_amount = unit_price_present_previous_merge[columns[7]][index]
            gr_quantity = unit_price_present_previous_merge[columns[10]][index]
            unit_price = gr_amount - gr_quantity
            unit_price_present_previous_merge['Increase/decrease in Quantity'][index] = unit_price

        unit_price_present_previous_merge['Increase/decrease in Unit Price'] = ""
        pd.options.mode.chained_assignment = None
        for index in unit_price_present_previous_merge.index:
            gr_amount = unit_price_present_previous_merge[columns[8]][index]
            gr_quantity = unit_price_present_previous_merge[columns[11]][index]
            if isinstance(gr_amount, str) or isinstance(gr_quantity, str):
                continue
            unit_price = gr_amount - gr_quantity
            unit_price_present_previous_merge['Increase/decrease in Unit Price'][index] = unit_price

        columns = unit_price_present_previous_merge.columns.values.tolist()
        unit_price_present_previous_merge['Increase/decrease in unit price (%)'] = ""
        pd.options.mode.chained_assignment = None
        for index in unit_price_present_previous_merge.index:
            gr_amount = unit_price_present_previous_merge[columns[14]][index]
            gr_quantity = unit_price_present_previous_merge[columns[11]][index]
            if isinstance(gr_amount, str) or isinstance(gr_quantity, str):
                continue
            elif gr_quantity != 0:
                unit_price = gr_amount / gr_quantity
                unit_price_present_previous_merge['Increase/decrease in unit price (%)'][index] = unit_price

        columns = unit_price_present_previous_merge.columns.values.tolist()
        unit_price_present_previous_merge['In amount due to Qty'] = ""
        pd.options.mode.chained_assignment = None
        for index in unit_price_present_previous_merge.index:
            gr_amount = unit_price_present_previous_merge[columns[13]][index]
            gr_quantity = unit_price_present_previous_merge[columns[11]][index]
            if isinstance(gr_amount, str) or isinstance(gr_quantity, str):
                continue
            else:
                unit_price = gr_amount * gr_quantity
                unit_price_present_previous_merge['In amount due to Qty'][index] = unit_price
        columns = unit_price_present_previous_merge.columns.values.tolist()
        unit_price_present_previous_merge['In amount due to Qty (%)'] = ""
        pd.options.mode.chained_assignment = None
        for index in unit_price_present_previous_merge.index:
            gr_amount = unit_price_present_previous_merge[columns[16]][index]
            gr_quantity = unit_price_present_previous_merge[columns[12]][index]

            if isinstance(gr_amount, str) or isinstance(gr_quantity, str):
                continue
            elif gr_quantity != 0:
                unit_price = gr_amount / gr_quantity
                unit_price_present_previous_merge['In amount due to Qty (%)'][index] = unit_price

        columns = unit_price_present_previous_merge.columns.values.tolist()
        unit_price_present_previous_merge['In amount due to price'] = ""
        pd.options.mode.chained_assignment = None

        for index in unit_price_present_previous_merge.index:
            gr_amount = unit_price_present_previous_merge[columns[7]][index]
            gr_quantity = unit_price_present_previous_merge[columns[14]][index]
            if isinstance(gr_amount, str) or isinstance(gr_quantity, str):
                continue
            else:
                unit_price = gr_amount * gr_quantity
                unit_price_present_previous_merge['In amount due to price'][index] = unit_price

        columns = unit_price_present_previous_merge.columns.values.tolist()
        unit_price_present_previous_merge['In amount due to unit price (%)'] = ""
        pd.options.mode.chained_assignment = None
        for index in unit_price_present_previous_merge.index:
            gr_amount = unit_price_present_previous_merge[columns[18]][index]
            gr_quantity = unit_price_present_previous_merge[columns[12]][index]
            if isinstance(gr_amount, str) or isinstance(gr_quantity, str):
                continue
            elif gr_quantity != 0:
                unit_price = gr_amount / gr_quantity
                unit_price_present_previous_merge['In amount due to unit price (%)'][index] = unit_price

        #  Rename Columns
        unit_price_present_previous_merge = unit_price_present_previous_merge.rename(
            columns={columns[6]: "GR Amt.in loc.cur."})
        unit_price_present_previous_merge = unit_price_present_previous_merge.rename(
            columns={columns[7]: "GR Qty"})
        unit_price_present_previous_merge = unit_price_present_previous_merge.rename(
            columns={columns[8]: "Unit Price"})
        unit_price_present_previous_merge = unit_price_present_previous_merge.rename(
            columns={columns[9]: "GR Amt.in loc.cur."})
        unit_price_present_previous_merge = unit_price_present_previous_merge.rename(
            columns={columns[10]: "GR Qty"})
        unit_price_present_previous_merge = unit_price_present_previous_merge.rename(
            columns={columns[11]: "Unit Price"})
        unit_price_present_previous_merge = unit_price_present_previous_merge.drop(columns=["index"])
        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                unit_price_present_previous_merge.to_excel(writer, sheet_name=main_config[
                    "Output_Unit_Price_Comparison_sheetname"], startrow=24,
                                                           index=False)
        except Exception as File_creation_error:
            logging.error(
                "Exception occurred while creating unit price comparison sheet \n\t {0}".format(File_creation_error))
            raise File_creation_error

        wb = load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Unit_Price_Comparison_sheetname"]]
        present_quarter_column_name = main_config['PresentQuarterColumnName']
        previous_quarter_column_name = main_config['PreviousQuarterColumnName']
        cell = ws['F24']
        cell.value = present_quarter_column_name
        ws.merge_cells('F24:H24')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = ws['I24']
        cell.value = previous_quarter_column_name
        ws.merge_cells('I24:K24')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = ws['L24']
        cell.value = 'Increase/Decrease In Amount'
        ws.merge_cells('L24:S24')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        ws['F24'].font = font_style
        ws['I24'].font = font_style
        ws['L24'].font = font_style
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        ws['F24'].fill = fill_pattern
        ws['L24'].fill = fill_pattern
        fill_pattern = PatternFill(patternType="solid", fgColor="FFFF00")
        ws['I24'].fill = fill_pattern

        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "25"].fill = format_fill
            if c == 's':
                break
        m_row = ws.max_row

        ws.auto_filter.ref = "A25:S" + str(m_row)

        for cell in ws["E"]:
            cell.number_format = "#,###"
        for cell in ws["F"]:
            cell.number_format = "#,###"
        for cell in ws["G"]:
            cell.number_format = "#,###"
        for cell in ws["H"]:
            cell.number_format = "#,###"
        for cell in ws["I"]:
            cell.number_format = "#,###"
        for cell in ws["J"]:
            cell.number_format = "#,###"
        for cell in ws["K"]:
            cell.number_format = "#,###"
        for cell in ws["L"]:
            cell.number_format = "#,###"
        for cell in ws["M"]:
            cell.number_format = "#,###"
        for cell in ws["N"]:
            cell.number_format = "0"
        for cell in ws["O"]:
            cell.number_format = numbers.FORMAT_PERCENTAGE
        for cell in ws["P"]:
            cell.number_format = "0.00"
        for cell in ws["Q"]:
            cell.number_format = "0.00%"
        for cell in ws["R"]:
            cell.number_format = "0.00"
        for cell in ws["S"]:
            cell.number_format = "0.00%"

        m_row = ws.max_row

        ws['F23'] = '=SUBTOTAL(9,F26:F' + str(m_row) + ')'
        ws['G23'] = '=SUBTOTAL(9,G26:G' + str(m_row) + ')'
        ws['I23'] = '=SUBTOTAL(9,I26:I' + str(m_row) + ')'
        ws['J23'] = '=SUBTOTAL(9,J26:J' + str(m_row) + ')'
        ws['M23'] = '=SUBTOTAL(9,M26:M' + str(m_row) + ')'
        ws['L23'] = '=SUBTOTAL(9,L26:L' + str(m_row) + ')'
        ws['P23'] = '=SUBTOTAL(9,P26:P' + str(m_row) + ')'
        ws['R23'] = '=SUBTOTAL(9,R26:R' + str(m_row) + ')'

        # Auto-fit column width
        for c in ascii_lowercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 's':
                break

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=26, min_col=1, max_row=ws.max_row, max_col=19):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        ws.merge_cells('A4:E4')
        ws.merge_cells('A5:E5')
        ws.merge_cells('A6:E6')
        ws.merge_cells('A7:E7')
        ws.merge_cells('A8:E8')
        ws.merge_cells('A9:E9')
        ws.merge_cells('A10:E10')
        ws.merge_cells('A11:E11')
        ws.merge_cells('A12:E12')
        ws.merge_cells('A13:E13')
        ws.merge_cells('A14:E14')

        # Headers implementation
        ws['A1'] = main_config['CompanyName']
        ws['A2'] = main_config['StatutoryAuditQuarter']
        ws['A3'] = main_config['FinancialYear']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])
        wb.close()

        try:
            change_in_unit_price_percentage(main_config, unit_price_present_previous_merge)
        except Exception as File_creation_error:
            logging.error(
                "Exception occurred while creating unit price comparison business exception sheet as per 'increase/decrease in unit price %' column: \n\t {0}".format(
                    File_creation_error))
            logging.exception(File_creation_error)
            print(File_creation_error)
        try:
            amount_quantity_percentage(main_config, unit_price_present_previous_merge)
        except Exception as File_creation_error:
            logging.error(
                "Exception occurred while creating unit price comparison business exception sheet as per 'In amount due to Qty (%)' column: \n\t {0}".format(
                    File_creation_error))
            logging.exception(File_creation_error)
            print(File_creation_error)

        try:
            unit_price_high_low_average(main_config, unit_price_present_previous_merge)
        except Exception as File_creation_error:
            logging.error(
                "Exception occurred while creating unit price comparison business exception sheet as per 'Unit Price' column: \n\t {0}".format(
                    File_creation_error))
            logging.exception(File_creation_error)
            print(File_creation_error)

    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Unit Price Comparison Process-", business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", key_error)
        return key_error


if __name__ == "__main__":
    pass
