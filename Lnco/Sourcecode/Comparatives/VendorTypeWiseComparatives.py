import warnings
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from string import ascii_uppercase
import logging
from purchase_send_mail_reusable_task import send_mail

warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)


class BusinessException(Exception):
    pass


def vendor_comparatives_top_weight(vendor_comparatives_dataframe, main_config, percentage):
    # save grand total row to delete from datatable to sort
    # grand_total_row = vendor_comparatives_dataframe.tail(1)
    # print(grand_total_row)
    # delete last row from the grand_total_row
    vendor_comparatives_dataframe.drop(vendor_comparatives_dataframe.tail(1).index, inplace=True)
    # print("Deleted Grand total row")
    # sort the dataframe using column name
    col_list = vendor_comparatives_dataframe.columns.values.tolist()
    vendor_comparatives_dataframe.sort_values(by=col_list[2], ascending=False, inplace=True)
    vendor_comparatives_weightage = pd.DataFrame(columns=vendor_comparatives_dataframe.columns)
    # print("empty dataframe is created with columns")
    for index, row in vendor_comparatives_dataframe.iterrows():
        if float(row['Percentage']) > percentage:
            vendor_comparatives_weightage = vendor_comparatives_weightage.append(row, ignore_index=True)
        else:
            continue
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            vendor_comparatives_weightage.to_excel(writer, sheet_name=main_config[
                "Output_Comparatives_Weightage_sheetname"], index=False, startrow=2, startcol=23)
            print("vendor concentration top weightage entries are logged in the output file")

    except Exception as File_creation_error:
        logging.error("Exception occurred while creating vendor type wise concentration sheet: \n {0}".format(
            File_creation_error))
        raise File_creation_error

    # Load excel file
    workbook = openpyxl.load_workbook(main_config['Output_File_Path'])

    # Load sheet
    worksheet = workbook[main_config['Output_Comparatives_Weightage_sheetname']]

    # Set column widths
    for column_letter in ['x', 'y', 'z', 'aa', 'ab', 'ac']:
        column_length = max(len(str(cell.value)) for cell in worksheet[column_letter])
        worksheet.column_dimensions[column_letter].width = column_length * 1.25

    # row 3 font format, fill color

    calibri_11_black_bold = Font(name="Calibri", size=11, color="000000", bold=True)
    light_blue_solid_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
    thin = Side(border_style="thin", color='b1c5e7')
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cambria_11_black = Font(name='Calibri', size=11, color='000000', bold=False)

    for row in worksheet["X3:AC3"]:
        for cell in row:
            cell.fill = light_blue_solid_fill
            cell.font = calibri_11_black_bold

    max_row = len(vendor_comparatives_weightage.index)
    for row in worksheet["X" + str(3 + 1) + ":AC" + str(max_row + 3)]:
        for cell in row:
            cell.font = cambria_11_black
            cell.border = thin_border

    # Number format implementation
    for cell in worksheet['Z']:
        cell.number_format = "#,###,##"
        if cell.value == 0:
            cell.value = '-'
            cell.alignment = Alignment(horizontal='center')

    for cell in worksheet['AA']:
        cell.number_format = "#,###,##"
        if cell.value == 0:
            cell.value = '-'
            cell.alignment = Alignment(horizontal='center')

    for cell in worksheet['AB']:
        cell.number_format = "#,###,##"
        if cell.value == 0:
            cell.value = '-'
            cell.alignment = Alignment(horizontal='center')

    # Format Variance
    for cell in worksheet['AC']:
        cell.number_format = '0.0%'

    workbook.save(main_config['Output_File_Path'])


# Defining a Function
def create_vendor_wise(main_config, in_config, present_quarter_pd, previous_quarter_pd):
    try:
        read_present_quarter_pd = present_quarter_pd
        read_previous_quarter_pd = previous_quarter_pd
        # Checking Exception starts here
        # present quarter
        if read_present_quarter_pd.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Input Sheet Data is empty")

        present_quarter_columns_list = read_present_quarter_pd.columns.values.tolist()
        for col in ["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]:
            if col not in present_quarter_columns_list:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        vendor_no_pd = read_present_quarter_pd[read_present_quarter_pd['Vendor No.'].notna()]
        vendor_name_pd = read_present_quarter_pd[read_present_quarter_pd['Vendor Name'].notna()]
        gr_amt_pd = read_present_quarter_pd[read_present_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(vendor_no_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Vendor No._subject"],
                      body=in_config["Vendor No._Body"])
            raise BusinessException("Vendor No. Column is empty")
        elif len(vendor_name_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Vendor Name_Subject"],
                      body=in_config["Vendor Name_Body"])
            raise BusinessException("Vendor Name Column is empty")
        elif len(gr_amt_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # present quarter exceptions ends here
        # previous quarter exceptions starts here
        if read_previous_quarter_pd.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Input Sheet Data is empty")

        previous_quarter_columns_list = read_previous_quarter_pd.columns.values.tolist()
        for col in ["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]:
            if col not in previous_quarter_columns_list:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows

        vendor_no_pd = read_previous_quarter_pd[read_previous_quarter_pd['Vendor No.'].notna()]
        vendor_name_pd = read_previous_quarter_pd[read_previous_quarter_pd['Vendor Name'].notna()]
        gr_amt_pd = read_previous_quarter_pd[read_previous_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(vendor_no_pd) == 0:

            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Vendor No._subject"],
                      body=in_config["Vendor No._Body"])
            raise BusinessException("Vendor No. Column is empty")

        elif len(vendor_name_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Vendor Name_Subject"],
                      body=in_config["Vendor Name_Body"])
            raise BusinessException("Vendor Name Column is empty")

        elif len(gr_amt_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # exception ends here

        # create pivot table
        # print(read_present_quarter_pd)

        present_quarter_final_file_pd = pd.pivot_table(read_present_quarter_pd, index=["Vendor No.", "Vendor Name"],
                                                       values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                                       margins_name="Grand Total")
        # print(present_quarter_final_file_pd)
        # pd[:-1] will select all the rows but not last row, omitting Grand total row
        present_quarter_final_file_pd = present_quarter_final_file_pd[:-1]
        # print(present_quarter_final_file_pd)
        # reset "indices created during pivot table creation" - for merging
        present_quarter_final_file_pd = present_quarter_final_file_pd.reset_index()
        # print(present_quarter_final_file_pd)

        # read previous quarters final working file - pd will be replaced with Nan in any blank cells
        previous_quarter_final_file_pd = pd.pivot_table(read_previous_quarter_pd,
                                                        index=['Vendor No.', 'Vendor Name'],
                                                        values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                                        margins_name="Grand Total")
        # print(previous_quarter_final_file_pd)
        previous_quarter_final_file_pd = previous_quarter_final_file_pd[:-1]
        # print(previous_quarter_final_file_pd)
        previous_quarter_final_file_pd = previous_quarter_final_file_pd.reset_index()
        # print(previous_quarter_final_file_pd)

        # replace Nan with blank
        present_quarter_final_file_pd = present_quarter_final_file_pd.replace(numpy.nan, 0, regex=True)
        previous_quarter_final_file_pd = previous_quarter_final_file_pd.replace(numpy.nan, 0, regex=True)

        # print(present_quarter_final_file_pd)
        # print(previous_quarter_final_file_pd)
        # merging present and previous quarter vendor wise data -  pd will be replaced with Nan in any blank cells
        merge_pd = pd.merge(present_quarter_final_file_pd, previous_quarter_final_file_pd, how="outer",
                            on=["Vendor No.", "Vendor Name"])

        # replacing all Nan's with zeros in Present and previous Quarter's values columns
        merge_pd = merge_pd.replace(numpy.nan, 0, regex=True)

        col_list = merge_pd.columns.values.tolist()
        # returns as ['Valuation Class', 'Valuation Class Text', 'GR Amt.in loc.cur.', 'Previous Quarter']

        # dropping columns present and previous quarters both have values as zero
        merge_pd.drop(merge_pd.index[(merge_pd[col_list[2]] == 0) & (merge_pd[col_list[3]] == 0)],
                      inplace=True)
        merge_pd.sort_values(by=col_list[2], axis=0, ascending=False, inplace=True)

        # create a new column - Success
        merge_pd['Variance'] = 0

        pd.options.mode.chained_assignment = None

        # variance formula implementation using index
        for index in merge_pd.index:
            present_quarter_row_value = merge_pd[col_list[2]][index]
            previous_quarter_row_value = merge_pd[col_list[3]][index]
            variance = present_quarter_row_value - previous_quarter_row_value
            merge_pd['Variance'][index] = variance

        col_list = merge_pd.columns.values.tolist()
        merge_pd.drop(merge_pd.index[(merge_pd[col_list[3]] == 0) & (merge_pd[col_list[4]] == 0)],
                      inplace=True)
        merge_pd['Percentage'] = ''
        pd.options.mode.chained_assignment = None
        # variance formula implementation using index
        for index in merge_pd.index:
            previous_quarter_row_value = merge_pd[col_list[3]][index]
            variance_row_value = merge_pd[col_list[4]][index]
            if previous_quarter_row_value == 0:
                percentage = 1
            else:
                percentage = variance_row_value / previous_quarter_row_value
            merge_pd['Percentage'][index] = percentage
        vendor_wise_comparatives_pd = merge_pd.rename(
            columns={col_list[2]: main_config["PresentQuarterColumnName"]})
        vendor_wise_comparatives_pd = vendor_wise_comparatives_pd.rename(
            columns={col_list[3]: main_config["PreviousQuarterColumnName"]})

        present_quarter_subtotal = vendor_wise_comparatives_pd[main_config["PresentQuarterColumnName"]].sum()
        # print(present_quarter_subtotal)
        previous_quarter_subtotal = vendor_wise_comparatives_pd[main_config["PreviousQuarterColumnName"]].sum()
        # print(previous_quarter_subtotal)
        variance_subtotal = present_quarter_subtotal - previous_quarter_subtotal
        # print(variance_subtotal)
        percentage_overall = variance_subtotal / previous_quarter_subtotal
        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                vendor_wise_comparatives_pd.to_excel(writer,
                                                     sheet_name=main_config[
                                                         "Output_Comparatives_Vendor_sheetname"], index=False, startrow=17)
        except Exception as File_creation_error:
            logging.error("Exception occurred while creating vendor wise concentration sheet")
            raise File_creation_error
        try:
            vendor_comparatives_top_weight(vendor_wise_comparatives_pd, main_config, percentage_overall)
        except Exception as vendor_comparatives_top_weight_error:
            print("Exception occurred while creating vendor wise concentration sheet: \n {0}".format(
                vendor_comparatives_top_weight_error))

        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Comparatives_Vendor_sheetname"]]

        ws['C17'] = present_quarter_subtotal
        ws['D17'] = previous_quarter_subtotal
        ws['E17'] = variance_subtotal
        ws['F17'] = percentage_overall

        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for char in ascii_uppercase:
            ws[char + "17"].font = font_style
            if char == 'F':
                break

        # max_row_number = ws.max_row
        # for char in ascii_uppercase:
        #     ws[char + str(max_row_number)].font = font_style
        #     if char == 'F':
        #         break

        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for char in ascii_uppercase:
            ws[char + "18"].fill = fill_pattern
            if char == 'F':
                break

        max_row_number = ws.max_row

        ws.auto_filter.ref = "A18:F" + str(max_row_number)
        for c in ascii_uppercase:
            ws.column_dimensions[c].width = 20
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["B"].width = 35
        # ws.delete_rows(max_row_number)
        # ws.delete_rows(max_row_number - 1)

        for cell in ws["C"]:
            cell.number_format = "#,##,###"
            if cell.value == 0:
                cell.value = '-'
                cell.alignment = Alignment(horizontal='center')

        for cell in ws["D"]:
            cell.number_format = "#,##,###"
            if cell.value == 0:
                cell.value = '-'
                cell.alignment = Alignment(horizontal='center')
        for cell in ws['E']:
            cell.number_format = '##,##'
            if cell.value == 0:
                cell.value = '-'
                cell.alignment = Alignment(horizontal='center')
        for cell in ws['F']:
            cell.number_format = '0.0%'

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=18, min_col=1, max_row=ws.max_row, max_col=6):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Cell merge for headers implementation
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        ws.merge_cells('A4:F4')
        ws.merge_cells('A5:F5')
        ws.merge_cells('A6:F6')
        ws.merge_cells('A7:F7')
        ws.merge_cells('A8:F8')
        ws.merge_cells('A9:F9')
        ws.merge_cells('A10:F10')
        ws.merge_cells('A11:F11')
        ws.merge_cells('A12:F12')
        ws.merge_cells('A13:F13')
        ws.merge_cells('A14:F14')

        # Headers implementation
        ws['A1'] = main_config['CompanyName']
        ws['A2'] = main_config['StatutoryAuditQuarter']
        ws['A3'] = main_config['FinancialYear']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])

        return vendor_wise_comparatives_pd

    # Excepting Errors here
    except PermissionError as file_error:
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                  subject=in_config["SystemE_Subject"],
                  body=in_config["SystemE_Body"])
        print("Please close the file")
        print("Exception: ", file_error)
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Vendor Wise comparatives Process", notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Vendor Wise comparatives Process-", business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Vendor Wise comparatives Process-", value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Vendor Wise comparatives Process-", type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Vendor Wise comparatives Process-", error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Vendor Wise comparatives Process-", key_error)
        return key_error


config = {}
main_config = {}
present_quarter_pd = pd.DataFrame()
previous_quarter_pd = pd.DataFrame()
if __name__ == "__main__":
    print(create_vendor_wise(main_config, config, present_quarter_pd, previous_quarter_pd))
