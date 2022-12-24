import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from string import ascii_lowercase
from send_mail_reusable_task import send_mail
import logging


class BusinessException(Exception):
    pass


def month_comparatives_top_weight(month_comparatives_dataframe, main_config):
    # save grand total row to delete from datatable to sort
    grand_total_row = month_comparatives_dataframe.tail(1)
    variance = float(grand_total_row['Variance'])
    # print(grand_total_row)
    # delete last row from the grand_total_row
    month_comparatives_dataframe.drop(month_comparatives_dataframe.tail(1).index, inplace=True)
    # print("Deleted Grand total row")
    # sort the dataframe using column name
    month_comparatives_dataframe.sort_values(by="Variance", ascending=False, inplace=True)
    month_comparatives_weightage = pd.DataFrame(columns=month_comparatives_dataframe.columns)
    # print("empty dataframe is created with columns")
    for index, row in month_comparatives_dataframe.iterrows():
        # print(float(row["Variance"]))
        if float(row['Variance']) > variance:
            # print(sum_of_variance)
            month_comparatives_weightage = month_comparatives_weightage.append(row, ignore_index=True)
            # print("appended row")
        else:
            continue
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            month_comparatives_weightage.to_excel(writer, sheet_name=main_config[
                "Output_Comparatives_Weightage_sheetname"], index=False, startrow=2, startcol=7)
            print("month wise concentration top weightage entries are logged in the output file")

    except Exception as File_creation_error:
        logging.error("Exception occurred while creating month wise concentration sheet: \n {0}".format(
            File_creation_error))
        raise File_creation_error

    # Load excel file
    workbook = openpyxl.load_workbook(main_config['Output_File_Path'])

    # Load sheet
    worksheet = workbook[main_config['Output_Comparatives_Weightage_sheetname']]

    # Set column widths
    for column_letter in ['h', 'i', 'j', 'k', 'l']:
        column_length = max(len(str(cell.value)) for cell in worksheet[column_letter])
        worksheet.column_dimensions[column_letter].width = column_length * 1.25

    # row 3 font format, fill color

    calibri_11_black_bold = Font(name="Calibri", size=11, color="000000", bold=True)
    light_blue_solid_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
    thin = Side(border_style="thin", color='b1c5e7')
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cambria_11_black = Font(name='Calibri', size=11, color='000000', bold=False)

    for row in worksheet["H3:L3"]:
        for cell in row:
            cell.fill = light_blue_solid_fill
            cell.font = calibri_11_black_bold

    max_row = len(month_comparatives_weightage.index)
    for row in worksheet["H" + str(3 + 1) + ":L" + str(max_row + 3)]:
        for cell in row:
            cell.font = cambria_11_black
            cell.border = thin_border

    # Number format implementation
    for cell in worksheet['I']:
        cell.number_format = "#,###,##"

    for cell in worksheet['K']:
        cell.number_format = "#,###,##"

    # Format Variance
    for cell in worksheet['L']:
        cell.number_format = '0.0%'

    workbook.save(main_config['Output_File_Path'])


def purchase_month(main_config, in_config, present_quarter_pd, previous_quarter_pd):
    try:
        read_present_quarter_pd = present_quarter_pd

        # Create Month Column
        read_present_quarter_pd['GR Posting Date'] = pd.to_datetime(read_present_quarter_pd['GR Posting Date'], errors='coerce')

        read_present_quarter_pd['GR Posting Date'] = read_present_quarter_pd['GR Posting Date'].dt.month_name().str[:3]

        read_present_quarter_pd['Month'] = read_present_quarter_pd['GR Posting Date']

        read_previous_quarter_pd = previous_quarter_pd

        # Create Month Column
        read_previous_quarter_pd['GR Posting Date'] = pd.to_datetime(read_previous_quarter_pd['GR Posting Date'], errors='coerce')
        read_previous_quarter_pd['GR Posting Date'] = read_previous_quarter_pd['GR Posting Date'].dt.month_name().str[:3]
        read_previous_quarter_pd['Month'] = read_previous_quarter_pd['GR Posting Date']

        # Check Exception
        if read_present_quarter_pd.shape[0] == 0 or read_previous_quarter_pd.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")

        previous_quarter_sheet_col = read_previous_quarter_pd.columns.values.tolist()
        for col in ["Month", "GR Amt.in loc.cur."]:
            if col not in previous_quarter_sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        present_quarter_sheet_columns = read_present_quarter_pd.columns.values.tolist()
        for col in ["Month", "GR Amt.in loc.cur."]:
            if col not in present_quarter_sheet_columns:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        month_pd = read_present_quarter_pd[read_present_quarter_pd['Month'].notna()]
        gr_amt_pd = read_present_quarter_pd[read_present_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(month_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Month_subject"],
                      body=in_config["Month_Body"])
            raise BusinessException("Month Column is empty")

        elif len(gr_amt_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        month_pd_2 = read_previous_quarter_pd[read_previous_quarter_pd['Month'].notna()]
        gr_amt_pd_2 = read_previous_quarter_pd[read_previous_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(month_pd_2) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Month_subject"],
                      body=in_config["Month_Body"])
            raise BusinessException("Month Column is empty")

        elif len(gr_amt_pd_2) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # create Pivot Table Q4
        pivot_index = ["Month"]
        pivot_values = ["GR Amt.in loc.cur."]
        pivot_present_quarter = pd.pivot_table(read_present_quarter_pd, index=pivot_index, values=pivot_values, aggfunc=numpy.sum, margins=True,
                                               margins_name='Grand Total')

        # Get Pivot Column Names
        col_name = pivot_present_quarter.columns.values.tolist()

        # Rename Column
        pivot_present_quarter = pivot_present_quarter.rename(columns={col_name[0]: main_config["PresentQuarterColumnName"]})
        pivot_present_quarter = pivot_present_quarter.reset_index()

        # Sort based on month
        month_dict = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9,
                      'Oct': 10, 'Nov': 11, 'Dec': 12, 'Grand Total': 13}

        pivot_present_quarter = pivot_present_quarter.sort_values('Month', key=lambda x: x.apply(lambda y: month_dict[y]))
        pivot_present_quarter.reset_index(inplace=True, drop=True)

        # Create Pivot Table Q3
        pivot_index = ["Month"]
        pivot_values = ["GR Amt.in loc.cur."]
        pivot_previous_quarter = pd.pivot_table(read_previous_quarter_pd, index=pivot_index, values=pivot_values, aggfunc=numpy.sum, margins=True,
                                                margins_name='Grand Total')

        # Get Pivot Column Names
        col_name = pivot_previous_quarter.columns.values.tolist()

        # Rename Column
        pivot_previous_quarter = pivot_previous_quarter.rename(columns={col_name[0]: main_config["PreviousQuarterColumnName"]})

        # Remove Index
        pivot_previous_quarter = pivot_previous_quarter.reset_index()

        # Sort based on month
        pivot_previous_quarter = pivot_previous_quarter.sort_values('Month', key=lambda x: x.apply(lambda a: month_dict[a]))
        pivot_previous_quarter.reset_index(inplace=True, drop=True)

        # Merge Pivot Sheets
        month_comparatives_pd = pd.concat([pivot_present_quarter, pivot_previous_quarter], axis=1, sort=False)
        # Remove Empty Rows
        month_comparatives_pd = month_comparatives_pd.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = month_comparatives_pd.columns.values.tolist()

        # Delete row of Q4 and Q3 columns values as zero
        month_comparatives_pd.drop(month_comparatives_pd.index[(month_comparatives_pd[col_name[1]] == 0) & (month_comparatives_pd[col_name[3]] == 0)])

        pd.options.mode.chained_assignment = None

        # Variance Formula
        variance_list = []
        for index in month_comparatives_pd.index:
            present_quarter_row_value = (month_comparatives_pd[col_name[1]][index])
            previous_quarter_row_value = (month_comparatives_pd[col_name[3]][index])

            if previous_quarter_row_value == 0:
                variance = 1
            else:
                variance = (present_quarter_row_value - previous_quarter_row_value) / previous_quarter_row_value

            variance_list.append(variance)

        # Create Variance Column
        month_comparatives_pd['Variance'] = variance_list
        try:
            # Log Sheet
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                month_comparatives_pd.to_excel(writer, sheet_name=main_config["Output_Comparatives_Month_sheetname"], index=False, startrow=16)
        except Exception as File_creation_error:
            logging.error("Exception occurred while creating month wise comparatives sheet")
            raise File_creation_error

        try:
            month_comparatives_top_weight(month_comparatives_pd, main_config)
        except Exception as month_comparatives_top_weight_error:
            print("Exception occurred while creating month type wise concentration sheet: \n {0}".format(
                month_comparatives_top_weight_error))

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Comparatives_Month_sheetname"]]

        # Format Q4 & Q3
        for col in ['B', 'D']:
            for cell in ws[col]:
                cell.number_format = "#,###,##.##"

        # Format Variance
        for cell in ws['E']:
            cell.number_format = '0.0%'

        # Format Header
        format_font = Font(name="Calibri", size=11, color="000000", bold=True)
        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        for c in ascii_lowercase:
            ws[c + "17"].font = format_font

        # Format Footer
        m_row = ws.max_row
        for c in ascii_lowercase:
            ws[c + str(m_row)].font = format_font

        # Header Fill
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "17"].fill = format_fill
            if c == 'e':
                break

        # Set Width
        for c in ascii_lowercase:
            ws.column_dimensions[c].width = 20

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=17, min_col=1, max_row=ws.max_row, max_col=5):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Cell merge for headers implementation
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        ws.merge_cells('A4:E4')
        ws.merge_cells('A5:E5')
        ws.merge_cells('A6:E6')
        ws.merge_cells('A7:E7')
        ws.merge_cells('A8:E8')
        ws.merge_cells('A9:E9')
        ws.merge_cells('A10:E10')
        ws.merge_cells('A11:E11')
        ws.merge_cells('A12:E12')
        ws.merge_cells('A13:E13')
        ws.merge_cells('A14:E14')

        # Headers implementation
        ws['A1'] = main_config['CompanyName']
        ws['A2'] = main_config['StatutoryAuditQuarter']
        ws['A3'] = main_config['FinancialYear']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        # Save File
        wb.save(main_config["Output_File_Path"])
        return ws

    # Excepting Errors here
    except FileNotFoundError as notfound_error:
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["subject_file_not_found"],
                  body=in_config["body_file_not_found"])
        print("Month Type Wise Comparatives Process-", notfound_error)
        return notfound_error
    except ValueError as V_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(V_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", V_error)
        return V_error
    except BusinessException as business_error:
        print("Month Type Wise Comparatives Process-", business_error)
        return business_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Month Type Wise Comparatives Process-", type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Month Type Wise Comparatives Process-", error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Month Type Wise Comparatives Process-", key_error)
        return key_error
    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Please close the file")
        return file_error


# Read config details and parse to dictionary
config = {}
main_config = {}
present_quarter_pd = pd.DataFrame()
previous_quarter_pd = pd.DataFrame()
if __name__ == "__main__":
    print(purchase_month(main_config, config, present_quarter_pd, previous_quarter_pd))
