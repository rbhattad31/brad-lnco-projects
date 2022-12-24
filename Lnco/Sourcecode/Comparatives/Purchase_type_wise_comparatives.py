# Importing Libraries

import pandas as pd
import numpy

import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from string import ascii_uppercase
from send_mail_reusable_task import send_mail
import logging


class BusinessException(Exception):
    pass


def purchase_comparatives_top_weight(purchase_comparatives_dataframe, main_config):
    # save grand total row to delete from datatable to sort
    # print(purchase_concentration_dataframe)
    grand_total_row = purchase_comparatives_dataframe.tail(1)
    variance = float(grand_total_row['Variance'])
    # print(grand_total_row)
    # delete last row from the grand_total_row
    purchase_comparatives_dataframe.drop(purchase_comparatives_dataframe.tail(1).index, inplace=True)
    # print("Deleted Grand total row")
    # sort the dataframe using column name
    purchase_comparatives_dataframe.sort_values(by="Variance", ascending=False, inplace=True)
    # print(purchase_concentration_dataframe)
    purchase_comparatives_weightage = pd.DataFrame(columns=purchase_comparatives_dataframe.columns)
    # print("empty dataframe is created with columns")
    for index, row in purchase_comparatives_dataframe.iterrows():
        # print(float(row["Variance"]))
        if float(row['Variance']) > variance:
            # print(sum_of_variance)
            purchase_comparatives_weightage = purchase_comparatives_weightage.append(row, ignore_index=True)
            # print("appended row")
        else:
            continue
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            purchase_comparatives_weightage.to_excel(writer, sheet_name=main_config[
                "Output_Comparatives_Weightage_sheetname"], index=False, startrow=2, startcol=1)
            print("purchase type concentration top weightage entries are logged in the output file")

    except Exception as File_creation_error:
        logging.error("Exception occurred while creating purchase type wise concentration sheet: \n {0}".format(
            File_creation_error))
        raise File_creation_error

    # Load excel file
    workbook = openpyxl.load_workbook(main_config['Output_File_Path'])

    # Load sheet
    worksheet = workbook[main_config['Output_Comparatives_Weightage_sheetname']]

    # Set column widths
    for column_letter in ['b', 'c', 'd', 'e', 'f']:
        column_length = max(len(str(cell.value)) for cell in worksheet[column_letter])
        worksheet.column_dimensions[column_letter].width = column_length * 1.25

    # row 3 font format, fill color

    calibri_11_black_bold = Font(name="Calibri", size=11, color="000000", bold=True)
    light_blue_solid_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
    thin = Side(border_style="thin", color='b1c5e7')
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cambria_11_black = Font(name='Calibri', size=11, color='000000', bold=False)

    for row in worksheet["B3:F3"]:
        for cell in row:
            cell.fill = light_blue_solid_fill
            cell.font = calibri_11_black_bold

    max_row = len(purchase_comparatives_weightage.index)
    for row in worksheet["B" + str(3 + 1) + ":F" + str(max_row + 3)]:
        for cell in row:
            cell.font = cambria_11_black
            cell.border = thin_border

    # Number format implementation
    for cell in worksheet['D']:
        cell.number_format = "#,###,##"

    for cell in worksheet['E']:
        cell.number_format = "#,###,##"

    # Format Variance
    for cell in worksheet['F']:
        cell.number_format = '0.0%'

    workbook.save(main_config['Output_File_Path'])


# Defining a Function
def create_purchase_type_wise(main_config, in_config, present_quarter_pd, previous_quarter_pd):
    try:
        # Read Purchase Register Sheets
        read_present_quarter_pd = present_quarter_pd
        read_present_quarter_pd = read_present_quarter_pd.loc[:,
                                  ~read_present_quarter_pd.columns.duplicated(keep='first')]

        read_previous_quarter_pd = previous_quarter_pd
        read_previous_quarter_pd = read_previous_quarter_pd.loc[:,
                                   ~read_previous_quarter_pd.columns.duplicated(keep='first')]

        # Check Exception
        if read_present_quarter_pd.shape[0] == 0 or read_previous_quarter_pd.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Input Sheet Data is empty")

        PreviousQuarterSheet_col = read_previous_quarter_pd.columns.values.tolist()
        for col in ["Valuation Class", "Valuation Class Text", "GR Amt.in loc.cur."]:
            if col not in PreviousQuarterSheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")

        PresentQuarterSheetColumns = read_present_quarter_pd.columns.values.tolist()
        for col in ["Valuation Class", "Valuation Class Text", "GR Amt.in loc.cur."]:
            if col not in PresentQuarterSheetColumns:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        Valuation_pd = read_present_quarter_pd[read_present_quarter_pd['Valuation Class'].notna()]
        Valuation_Text_pd = read_present_quarter_pd[read_present_quarter_pd['Valuation Class Text'].notna()]
        Gr_Amt_pd = read_present_quarter_pd[read_present_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(Valuation_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Valuation Class_subject"],
                      body=in_config["Valuation Class_Body"])
            raise BusinessException("Valuation Class Column is empty")

        elif len(Valuation_Text_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Valuation Class Text_Subject"],
                      body=in_config["Valuation Class Text_Body"])
            raise BusinessException("Valuation Class Text Column is empty")

        elif len(Gr_Amt_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        Valuation_pd_2 = read_previous_quarter_pd[read_previous_quarter_pd['Valuation Class'].notna()]
        Valuation_Text_pd_2 = read_previous_quarter_pd[read_previous_quarter_pd['Valuation Class Text'].notna()]
        Gr_Amt_pd_2 = read_previous_quarter_pd[read_previous_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(Valuation_pd_2) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Valuation Class_subject"],
                      body=in_config["Valuation Class_Body"])
            raise BusinessException("Valuation Class Column is empty")

        elif len(Valuation_Text_pd_2) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Valuation Class Text_Subject"],
                      body=in_config["Valuation Class Text_Body"])
            raise BusinessException("Valuation Class Text Column is empty")

        elif len(Gr_Amt_pd_2) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # create pivot table
        purchase_type_wise_pd = pd.pivot_table(read_present_quarter_pd,
                                               index=["Valuation Class", "Valuation Class Text"],
                                               values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                               margins_name="Grand Total")

        # reset "indices created during pivot table creation" - for merging
        purchase_type_wise_pd = purchase_type_wise_pd.reset_index()
        # print(purchase_type_wise_pd)
        # print(purchase_type_wise_pd.dtypes)
        # read previous quarters final working file - pd will be replaced with Nan in any blank cells
        previous_quarter_final_file_pd = pd.pivot_table(read_previous_quarter_pd,
                                                        index=["Valuation Class", "Valuation Class Text"],
                                                        values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                                        margins_name="Grand Total")

        # replace Nan with blank
        previous_quarter_final_file_pd = previous_quarter_final_file_pd.replace(numpy.nan, '', regex=True)
        previous_quarter_final_file_pd = previous_quarter_final_file_pd.reset_index()
        # print(previous_quarter_final_file_pd)
        # print(previous_quarter_final_file_pd.dtypes)
        # merging present and previous quarter purchase type wise data -  pd will be replaced with Nan in any blank cells
        merge_pd = pd.merge(purchase_type_wise_pd, previous_quarter_final_file_pd, how="outer",
                            on=["Valuation Class", "Valuation Class Text"])

        # print(merge_pd)
        # replacing all Nan's with zeros in Present and previous Quarter's values columns
        merge_pd = merge_pd.replace(numpy.nan, 0, regex=True)
        # print(merge_pd)

        Col_List = merge_pd.columns.values.tolist()
        # returns as ['Valuation Class', 'Valuation Class Text', 'GR Amt.in loc.cur.', 'Previous Quarter']

        # dropping columns present and previous quarters both have values as zero
        # print(merge_pd[[Col_List[2], Col_List[3]]])
        merge_pd.drop(merge_pd.index[(merge_pd[Col_List[2]] == 0) & (merge_pd[Col_List[3]] == 0)],
                      inplace=True)
        # print(merge_pd[[Col_List[2], Col_List[3]]])
        # create a new column - Success
        merge_pd['Variance'] = 0
        # print(merge_pd)
        pd.options.mode.chained_assignment = None

        # variance formula implementation using index
        for index in merge_pd.index:
            present_quarter_row_value = merge_pd[Col_List[2]][index]
            previous_quarter_row_value = merge_pd[Col_List[3]][index]
            if previous_quarter_row_value == 0:
                variance = 1
            else:
                variance = (present_quarter_row_value - previous_quarter_row_value) / previous_quarter_row_value
            merge_pd['Variance'][index] = variance

        # copy present quarter Amount column Grand total, set it as zero, sort the data frame and reassign the value.
        grand_total = merge_pd[Col_List[2]].values[-1]
        merge_pd[Col_List[2]].values[-1] = 0

        merge_pd[Col_List[2]].values[-1] = grand_total

        purchase_type_wise_comparatives_pd = merge_pd.rename(
            columns={Col_List[2]: main_config["PresentQuarterColumnName"]})
        purchase_type_wise_comparatives_pd = purchase_type_wise_comparatives_pd.rename(
            columns={Col_List[3]: main_config["PreviousQuarterColumnName"]})
        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                purchase_type_wise_comparatives_pd.to_excel(writer,sheet_name=main_config[
                                                                "Output_Comparatives_Purchase_sheetname"],
                                                            index=False, startrow=16)
        except Exception as File_creation_error:
            logging.error("Exception occurred while creating purchase type wise comparatives sheet")
            raise File_creation_error

        try:
            purchase_comparatives_top_weight(purchase_type_wise_comparatives_pd, main_config)
        except Exception as purchase_comparatives_top_weight_error:
            print("Exception occurred while creating purchase type wise concentration sheet: \n {0}".format(
                purchase_comparatives_top_weight_error))

        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Comparatives_Purchase_sheetname"]]

        for cell in ws['C']:
            cell.number_format = '#,###.##'
        for cell in ws['D']:
            cell.number_format = '#,###.##'
        for cell in ws['E']:
            cell.number_format = '0.0%'

        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "17"].font = font_style

        m_row = ws.max_row
        for c in ascii_uppercase:
            ws[c + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "17"].fill = fill_pattern
            if c == 'E':
                break

        for c in ascii_uppercase:
            ws.column_dimensions[c].width = 25
        ws.column_dimensions["E"].width = 15

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=17, min_col=1, max_row=ws.max_row, max_col=5):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        ws.merge_cells('A4:E4')
        ws.merge_cells('A5:E5')
        ws.merge_cells('A6:E6')
        ws.merge_cells('A7:E7')
        ws.merge_cells('A8:E8')
        ws.merge_cells('A9:E9')
        ws.merge_cells('A10:E10')
        ws.merge_cells('A11:E11')
        ws.merge_cells('A12:E12')
        ws.merge_cells('A13:E13')
        ws.merge_cells('A14:E14')

        # Headers implementation
        ws['A1'] = main_config['CompanyName']
        ws['A2'] = main_config['StatutoryAuditQuarter']
        ws['A3'] = main_config['FinancialYear']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])

        return purchase_type_wise_comparatives_pd

    # Excepting Errors here
    except FileNotFoundError as notfound_error:
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                  subject=in_config["subject_file_not_found"],
                  body=in_config["body_file_not_found"])
        print("Purchase Type Wise Comparatives Process-", notfound_error)
        return notfound_error
    except ValueError as V_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(V_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", V_error)
        return V_error
    except BusinessException as business_error:
        print("Purchase Type Wise Comparatives Process-", business_error)
        return business_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", key_error)
        return key_error
    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Please close the file")
        return file_error


main_config = {}
config = {}
present_quarter_pd = pd.DataFrame()
previous_quarter_pd = pd.DataFrame()
if __name__ == "__main__":
    print(create_purchase_type_wise(main_config, config, present_quarter_pd, previous_quarter_pd))
