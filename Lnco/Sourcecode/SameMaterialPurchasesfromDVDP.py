from string import ascii_uppercase

import pandas as pd
import numpy

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Side, Border
from send_mail_reusable_task import send_mail
from openpyxl.utils import get_column_letter


class BusinessException(Exception):
    pass


def same_mat_dvp(main_config, in_config, present_quarter_pd):
    try:
        read_present_quarter_pd = present_quarter_pd
        input_file = read_present_quarter_pd[read_present_quarter_pd["Material No."].notna()]
        unit_price = read_present_quarter_pd[read_present_quarter_pd["Unit Price"].notna()]

        if read_present_quarter_pd.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")
        elif len(unit_price) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Unit_price"],
                      body=in_config["Unit_price_body"])
            raise BusinessException("unit price column is empty")
        elif len(input_file) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Document is empty"],
                      body=in_config["Doc_body"])
            raise BusinessException("file is empty")
        else:
            pass
        max_pivot = pd.pivot_table(read_present_quarter_pd, index=["Material No.", "Material Desc", "Vendor Name"],
                                   values=["Unit Price"], aggfunc=numpy.max, margins=True, margins_name="Grand Total")

        max_pivot = max_pivot.reset_index()

        min_pivot = pd.pivot_table(read_present_quarter_pd, index=["Material No.", "Material Desc", "Vendor Name"],
                                   values=["Unit Price"], aggfunc=numpy.min, margins=True, margins_name="Grand Total")
        min_pivot = min_pivot.reset_index()

        com_file = pd.merge(max_pivot, min_pivot, how="outer",
                            on=["Material No.", "Material Desc", "Vendor Name"])
        com_file = com_file.replace(numpy.nan, '', regex=True)

        com_file["Deference"] = com_file["Unit Price_x"] - com_file["Unit Price_y"]
        com_file.drop(com_file[com_file["Deference"] <= 1].index, inplace=True)

        com_file = com_file.replace(numpy.nan, 0, regex=True)

        com_file["Percentage"] = com_file["Deference"] / com_file["Unit Price_y"]

        col_name = com_file.columns.values.tolist()

        com_file.sort_values(by=col_name[5], axis=0, ascending=False, inplace=True)
        # com_file[col_name[4]].values[-1] = grand_total
        com_file = com_file.rename(columns={col_name[3]: "Max of unit price"})
        com_file = com_file.rename(columns={col_name[4]: "Min of unit price"})
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            com_file.to_excel(writer, sheet_name=main_config["Output_Same_Material_Purchases_DVDP_sheetname"], index=False, startrow=21)

        wb = load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Same_Material_Purchases_DVDP_sheetname"]]

        for cell in ws['D']:
            cell.number_format = "#,###.##"
        for cell in ws['E']:
            cell.number_format = "#,###.##"
        for cell in ws['F']:
            cell.number_format = "#,###.##"
        for cell in ws['G']:
            cell.number_format = "##%"

        ws.delete_rows(idx=23)

        m_row = ws.max_row
        # ws.auto_filter.ref = ws.dimensions

        FullRange = "A22:" + get_column_letter(ws.max_column) \
                   + str(ws.max_row)
        ws.auto_filter.ref = FullRange

        font_style = Font(name="Cambria", size=13, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "22"].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "22"].fill = fill_pattern
            if c == 'G':
                break

        for c in ascii_uppercase:
            ws.column_dimensions[c].width = 35

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=22, min_col=1, max_row=ws.max_row, max_col=7):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        ws.merge_cells('A4:E4')
        ws.merge_cells('A5:E5')
        ws.merge_cells('A6:E6')
        ws.merge_cells('A7:E7')
        ws.merge_cells('A8:E8')
        ws.merge_cells('A9:E9')
        ws.merge_cells('A10:E10')
        ws.merge_cells('A11:E11')
        ws.merge_cells('A12:E12')
        ws.merge_cells('A13:E13')
        ws.merge_cells('A14:E14')

        # Headers implementation
        ws['A1'] = main_config['CompanyName']
        ws['A2'] = main_config['StatutoryAuditQuarter']
        ws['A3'] = main_config['FinancialYear']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A13'] = in_config['A13']
        ws['A14'] = in_config['A14']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=11, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=13, min_col=1, max_row=13, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=14, min_col=1, max_row=14, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])
        wb.close()

        return com_file
    except FileNotFoundError as f_error:
        print("sent a mail file not found")
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["File_N"],
                  body=in_config["File_N_body"])
        print("Exception: ", f_error)
        return f_error
    except NameError as n_error:
        print("Name Error")
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Name_E"],
                  body=in_config["Name_E_body"])
        print("Exception: ", n_error)
        return n_error
    except KeyError as k_error:
        print("KeyError")
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Key"],
                  body=in_config["Key_body"])
        print("Exception: ", k_error)
        return k_error
    except ValueError as v_error:
        print("ValueError")
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=in_config["Value_E"],
                  body=in_config["Value_E_body"])
        print("Exception: ", v_error)
        return v_error
    except SyntaxError as s_error:
        print("Exception: ", s_error)
        return s_error


config = {}
main_config = {}
present_quarter_pd = pd.DataFrame()
if __name__ == "__main__":
    print(same_mat_dvp(main_config, config, present_quarter_pd))


