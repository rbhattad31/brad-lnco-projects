import pandas as pd
import numpy
import openpyxl
import logging
import os
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase

from send_mail_reusable_task import send_mail

from AWS_and_SQL_programs.mb51_file_creation import mb51_file_creation


class BusinessException(Exception):
    pass


def create_Inventory_Mapping_sheet(main_config, in_config, present_quarter_pd, mb51_file_location, mb51_sheet_name,
                                   json_data_list):
    try:
        # Reading Purchase register File
        # read_excel_data = pd.read_excel(in_config["ExcelPath1"], sheet_name=in_config["Sheet_Name1"], skiprows=6)
        read_excel_data = present_quarter_pd
        read_excel_data = read_excel_data.loc[:, ~read_excel_data.columns.duplicated(keep='first')]

        # print(read_excel_data.head(5))
        # Check Exception
        if read_excel_data.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")

        Purchase_Sheet_col = read_excel_data.columns.values.tolist()
        for col in ["GR Document Number", "GR Qty"]:
            if col not in Purchase_Sheet_col:
                subject = in_config["Purchase_ColumnMiss_Subject"]
                body = in_config["Purchase_ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")
        # Filter Rows
        GR_Document_Number_pd = read_excel_data[read_excel_data['GR Document Number'].notna()]
        Gr_Qty_pd = read_excel_data[read_excel_data['GR Qty'].notna()]

        if len(GR_Document_Number_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["GR Document_Number_subject"],
                      body=in_config["GR Document_Number_Body"])
            raise BusinessException("GR Document Number Column is empty")

        elif len(Gr_Qty_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Gr Qty_Subject"],
                      body=in_config["Gr Qty_Body"])
            raise BusinessException("GR Qty Column is empty")
        else:
            pass

        read_excel_data = read_excel_data[['GR Document Number', 'GR Qty']]
        pivot1_df = pd.pivot_table(read_excel_data, index=["GR Document Number"],
                                   values="GR Qty",
                                   aggfunc=numpy.sum)
        pivot1_df.reset_index()

        # Reading MB 51 File
        mb51_pd = pd.read_excel(mb51_file_location, mb51_sheet_name)
        mb51_pd = mb51_pd.loc[:,
                  ~mb51_pd.columns.duplicated(keep='first')]
        columns = mb51_pd.columns
        if main_config["MB51_first_column"] in columns and \
                main_config["MB51_second_column"] in columns:
            print("MB51 - The data is starting from first row only")
            pass

        else:
            print("MB51 - The data is not starting from first row ")
            for index, row in mb51_pd.iterrows():
                if row[0] != main_config["MB51_first_column"]:
                    mb51_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = mb51_pd.iloc[0]
            mb51_pd = mb51_pd[1:]
            mb51_pd.columns = new_header
            mb51_pd.reset_index(drop=True, inplace=True)
            mb51_pd.columns.name = None
        mb51_pd = mb51_pd.loc[:,
                  ~mb51_pd.columns.duplicated(keep='first')]
        print("MB51 file reading is complete, creating new input file only with required columns")
        logging.info(
            "Reading purchase register present quarter sheet is complete, creating new input file only with required columns")
        mb51_folder_path = os.path.dirname(mb51_file_location)
        mb51_file_name = os.path.basename(mb51_file_location).lower()
        filtered_mb51_file_name = "filtered_" + str(mb51_file_name)
        filtered_purchase_present_file_saving_path = os.path.join(mb51_folder_path,filtered_mb51_file_name)
        filtered_purchase_present_sheet_name = mb51_sheet_name
        mb51_pd = mb51_file_creation(mb51_pd, json_data_list, filtered_purchase_present_file_saving_path, filtered_purchase_present_sheet_name)
        logging.info("new mb51 filtered file is created in input folder in request ID folder")
        mb51_pd.columns.values[0:2] = ["Material Document", "Qty in unit of entry"]
        # print(mb51_pd.head(5))
        # Check Exception
        if mb51_pd.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")

        MB51_Sheet_col = mb51_pd.columns.values.tolist()
        for col in ["Material Document", "Qty in unit of entry"]:
            if col not in MB51_Sheet_col:
                subject = in_config["MB51_ColumnMiss_Subject"]
                body = in_config["MB51_ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        Material_Document_pd = mb51_pd[mb51_pd['Material Document'].notna()]
        Qty_unit_of_entry_pd = mb51_pd[mb51_pd['Qty in unit of entry'].notna()]

        if len(Material_Document_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Material_Document_subject"],
                      body=in_config["Material_Document_Body"])
            raise BusinessException("Material Document Column is empty")

        elif len(Qty_unit_of_entry_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Qty_unit_of_entry_Subject"],
                      body=in_config["Qty_unit_of_entry_Body"])
            raise BusinessException("Qty in unit of entry Column is empty")
        else:
            pass

        # Taking Required Column to create pivot table
        mb51_pd = mb51_pd[['Material Document', 'Qty in unit of entry']]

        pivot2_df = pd.pivot_table(mb51_pd, index=["Material Document"],
                                   values="Qty in unit of entry",
                                   aggfunc=numpy.sum)
        pivot2_df = pivot2_df.reset_index()
        pivot2_df = pivot2_df.rename(columns={'Material Document': 'GR Document Number'})
        pivot1_df = pivot1_df.reset_index()
        # print(pivot1_df)
        # print(pivot2_df)

        # Merging 2 Pivots
        merge_pd = pd.merge(pivot1_df, pivot2_df, how="outer", on=["GR Document Number"])
        # print("at merging")
        # print(merge_pd.head(10))
        columns_list = merge_pd.columns.values.tolist()

        # create a new column - Success
        merge_pd['Check'] = 0

        # To Remove SettingWithCopyWarning error
        # modifying only one df, so suppressing this warning as it is not affecting
        pd.options.mode.chained_assignment = None

        # variance formula for index
        for index in merge_pd.index:
            Q4 = merge_pd[columns_list[1]][index]
            Q3 = merge_pd[columns_list[2]][index]

            if Q4 == Q3:
                CHECK = True
            else:
                CHECK = False

            merge_pd['Check'][index] = CHECK

        # Renaming Columns
        Inventory_Mapping_file = merge_pd.rename(
            columns={columns_list[0]: in_config["Rename_Column1"],
                     columns_list[1]: in_config["Rename_Column2"],
                     columns_list[2]: in_config["Rename_Column3"]})
        # print(Inventory_Mapping_file)

        Purchase_Sheet_Sum = Inventory_Mapping_file[in_config["Rename_Column2"]].sum()
        # print(Purchase_Sheet_Sum)
        MB51_Sheet_Sum = Inventory_Mapping_file[in_config["Rename_Column3"]].sum()
        # print(MB51_Sheet_Sum)

        # Creating Output File
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            Inventory_Mapping_file.to_excel(writer, sheet_name=main_config["Output_Inventory_Mapping_Sheetname"],
                                            index=False, startrow=9)

        # Opening and Reading Output File.
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Inventory_Mapping_Sheetname"]]

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for i in ascii_uppercase:
            ws[i + "1"].font = font_style

        # Adding Background Color
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")

        for j in ascii_uppercase:
            ws[j + "10"].fill = fill_pattern
            if j == 'D':
                break
        ws["C10"].fill = PatternFill("solid", fgColor="ffff00")

        # # Adding Auto Filter Option
        FullRange = "A10:D" + str(ws.max_row)
        ws.auto_filter.ref = FullRange

        # Auto Width Setting
        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.5
            if c == 'D':
                break

        # Passing Column Values
        ws['B9'] = Purchase_Sheet_Sum
        ws['C9'] = MB51_Sheet_Sum

        # Number Formatting
        ws['B9'].number_format = '#,###.##'
        ws['C9'].number_format = '#,###.##'

        # Font-style
        ws['B9'].font = font_style
        ws['C9'].font = font_style

        # Saving the File
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])

        return create_Inventory_Mapping_sheet

    # Excepting Errors here
    except FileNotFoundError as notfound_error:
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                  subject=in_config["subject_file_not_found"],
                  body=in_config["body_file_not_found"])
        print("Inventory Mapping Process-", notfound_error)
        return notfound_error
    except ValueError as V_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(V_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Inventory Mapping Process-", V_error)
        return V_error
    except BusinessException as business_error:
        print("Inventory Mapping Process-", business_error)
        return business_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Inventory Mapping Process-", type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Inventory Mapping Process-", error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Inventory Mapping Process-", key_error)
        return key_error
    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Please close the file")
        return file_error


if __name__ == "__main__":
    pass
