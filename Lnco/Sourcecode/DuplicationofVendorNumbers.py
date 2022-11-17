from string import ascii_lowercase
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import PatternFill, Side, Border
from send_mail_reusable_task import send_mail
import os


class BusinessException(Exception):
    pass


def vendor_numbers_duplication(main_config, in_config):
    try:
        # Read Purchase Register Sheets
        vendor_data = pd.read_excel(main_config["VendorMasterFilePath"], sheet_name=main_config["VendorMasterSheetName"])

        # Fetch To Address
        to_address = main_config["To_Mail_Address"]
        cc_address = main_config["CC_Mail_Address"]

        # Check data in input sheet
        if vendor_data.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        vendor_col = vendor_data.columns.values.tolist()
        for col in ["Vendor", "Name 1", "Tax Number"]:
            if col not in vendor_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter rows
        vendor_name = vendor_data[vendor_data['Name 1'].notna()]
        vendor_no = vendor_data[vendor_data['Vendor'].notna()]
        vendor_tax = vendor_data[vendor_data['Tax Number'].notna()]

        # Check Exception
        if len(vendor_no) == 0:
            subject = in_config["EmptyVendorNo_Subject"]
            body = in_config["EmptyVendorNo_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Vendor Number Column is empty")
        elif len(vendor_name) == 0:
            subject = in_config["EmptyVendorName_Subject"]
            body = in_config["EmptyVendorName_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Vendor Name Column is empty")
        elif len(vendor_tax) == 0:
            subject = in_config["EmptyTax_Subject"]
            body = in_config["EmptyTax_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Tax Number Column is empty")
        else:
            pass

        # Mark Empty rows
        vendor_data = vendor_data.replace(numpy.nan, "Empty", regex=True)

        # create Pivot Table
        pivot_index = ["Vendor", "Name 1", "Tax Number"]
        pivot_data = pd.pivot_table(vendor_data, index=pivot_index, sort=True)


        # Remove Index
        pivot_data = pivot_data.reset_index()

        # Assign Sheet
        pivot_sheet = pivot_data[["Vendor", "Name 1", "Tax Number"]]

        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)
        pivot_sheet = pivot_sheet.replace("Empty", '', regex=True)

        # Create Duplicate Column
        pivot_sheet['Duplicate'] = ""
        pivot_sheet['Lower case'] = pivot_sheet["Name 1"].str.lower()

        # Map Duplicate Rows
        pivot_sheet['Duplicate'] = pivot_sheet.duplicated(subset=["Lower case"], keep=False) \
            .map({True: 'Yes', False: 'No'})

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Assign Sheet
        pivot_sheet = pivot_sheet[["Vendor", "Name 1", "Tax Number", 'Duplicate']]

        # Delete row where vendor number columns values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[0]] == 0)], inplace=True)

        # Sort Table
        pivot_sheet = pivot_sheet.sort_values(by='Vendor', ascending=True)

        # Log Sheet
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            pivot_sheet.to_excel(writer, sheet_name=main_config["Output_Duplication_of_Vendor_sheetname"], index=False)

        # Check outfile creation
        if os.path.exists(main_config["Output_File_Path"]):
            print("Duplication of vendor numbers logged successfully")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Output file not generated")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Duplication_of_Vendor_sheetname"]]

        # Header Fill
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "1"].fill = format_fill
            if c == 'c':
                break

        # Highlight row
        format_fill = PatternFill(patternType='solid', fgColor='FFFF00')
        for cell in ws['d']:
            if cell.value == 'Yes':
                ws['B' + str(cell.row)].fill = format_fill

        # Set Width
        for c in ascii_lowercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 'c':
                break

        # Delete Duplicate Column
        ws.delete_cols(idx=4)

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        ws.sheet_view.showGridLines = False

        # Save File
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])
        return ws

    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Duplication Process-", end="")
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Duplication Process-", end="")
        return notfound_error
    except BusinessException as business_error:
        print("Duplication Process-", end="")
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Duplication Process-", end="")
        return value_error
    except TypeError as type_error:
        print("Duplication Process-", end="")
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Duplication Process-", end="")
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Duplication Process-", end="")
        print("Please check the given keyword is correct")
        return key_error


# Read config details and parse to dictionary
config = {}
main_config = {}

if __name__ == "__main__":
    print(vendor_numbers_duplication(main_config, config))

