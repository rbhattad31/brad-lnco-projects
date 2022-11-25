from string import ascii_lowercase
from send_mail_reusable_task import send_mail

import numpy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.styles import numbers
import warnings
warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)

class BusinessException(Exception):
    pass

def create_unit_price(main_config, in_config, present_quarter_pd, previous_quarter_pd):
    try:

        Excel_data = present_quarter_pd
        # Fetch To Address
        to_address = main_config["To_Mail_Address"]
        cc_address = main_config["CC_Mail_Address"]

        # Check Exception
        if Excel_data.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        Q4Sheet_col = Excel_data.columns.values.tolist()
        for col in ["GR Amt.in loc.cur.", "GR Qty", "Material No.", "Valuation Class Text", "Vendor Name"]:
            if col not in Q4Sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        pivot_Q4 = pd.pivot_table(Excel_data, index=["Material No.", "Valuation Class Text", "Vendor Name"],
                                  values=["GR Amt.in loc.cur.", "GR Qty"], aggfunc=numpy.sum, margins=True,
                                  margins_name="Grand Total", sort=True)
        # Drop last column of a dataframe
        pivot_Q4 = pivot_Q4[:-1]

        pivot_Q4 = pivot_Q4.reset_index()
        pivot_Q4 = pivot_Q4.replace(numpy.nan, 0, regex=True)
        columns = pivot_Q4.columns.values.tolist()
        numpy.seterr(divide='ignore')
        pivot_Q4['Unit Price'] = ""
        pd.options.mode.chained_assignment = None

        for index in pivot_Q4.index:
            GR_amt = pivot_Q4[columns[3]][index]
            GR_qty = pivot_Q4[columns[4]][index]
            if GR_qty != 0:
                Unit_price = GR_amt / GR_qty
                pivot_Q4['Unit Price'][index] = Unit_price

        columns = pivot_Q4.columns.values.tolist()
        pivot_Q4 = pivot_Q4.rename(
            columns={columns[3]: "GR Amt.in loc.cur.1"})
        pivot_Q4 = pivot_Q4.rename(
            columns={columns[4]: "GR Qty1"})
        pivot_Q4 = pivot_Q4.rename(
            columns={columns[5]: "Unit Price1"})

        pivot_Q4['Concat'] = ""
        pivot_Q4["Concat"] = pivot_Q4["Material No."].astype(str) + pivot_Q4["Valuation Class Text"].astype(str) + pivot_Q4[
            "Vendor Name"].astype(str)
        pivot_Q4 = pivot_Q4[["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.1",
                             "GR Qty1", "Unit Price1"]]

        # Q3 Pivot
        # Excel_data = pd.read_excel(main_config["InputFilePath"], sheet_name=main_config["PreviousQuarterSheetName"],skiprows=in_config["Skiprow_Q3"])

        Excel_data = previous_quarter_pd
        # Check Exception
        if Excel_data.shape[0] == 0:
            subject = in_config["EmptyInput_Subject1"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        Q3Sheet_col = Excel_data.columns.values.tolist()
        for col in ["GR Amt.in loc.cur.", "GR Qty", "Material No.", "Valuation Class Text", "Vendor Name"]:
            if col not in Q3Sheet_col:
                subject = in_config["ColumnMiss_Subject1"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        pivot_Q3_1 = pd.pivot_table(Excel_data, values=["GR Amt.in loc.cur.", "GR Qty"],
                                    index=["Material No.", "Valuation Class Text", "Vendor Name"],
                                    aggfunc={"GR Amt.in loc.cur.": numpy.sum, 'GR Qty': numpy.sum}, margins=True)

        pivot_Q3 = pivot_Q3_1
        pivot_Q3 = pivot_Q3.reset_index()
        pivot_Q3 = pivot_Q3.replace(numpy.nan, 0, regex=True)

        # Drop last column of a dataframe
        pivot_Q3 = pivot_Q3[:-1]
        columns = pivot_Q3.columns.values.tolist()

        pivot_Q3['Unit Price'] = ""
        pd.options.mode.chained_assignment = None
        for index in pivot_Q3.index:
            GR_amt = pivot_Q3[columns[3]][index]
            GR_qty = pivot_Q3[columns[4]][index]
            if GR_qty != 0 and type(GR_amt) != 'str' and type(GR_qty) != 'str':
                Unit_price = GR_amt / GR_qty
                pivot_Q3['Unit Price'][index] = Unit_price

        columns = pivot_Q3.columns.values.tolist()
        pivot_Q3 = pivot_Q3.rename(
            columns={columns[3]: "GR Amt.in loc.cur.2"})
        pivot_Q3 = pivot_Q3.rename(
            columns={columns[4]: "GR Qty2"})
        pivot_Q3 = pivot_Q3.rename(
            columns={columns[5]: "Unit Price2"})

        pivot_Q3['Concat'] = ""
        pivot_Q3["Concat"] = pivot_Q3["Material No."].astype(str) + pivot_Q3["Valuation Class Text"].astype(str) + pivot_Q3[
            "Vendor Name"].astype(str)
        pivot_Q3 = pivot_Q3[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.2", "GR Qty2",
             "Unit Price2"]]

        Unit_Price_Q4 = pd.merge(pivot_Q4, pivot_Q3, how="left",
                                 on=["Material No.", "Valuation Class Text", "Vendor Name", "Concat"], copy=False)
        Unit_Price_Q4 = Unit_Price_Q4[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.1", "GR Qty1",
             "Unit Price1"]]
        Unit_Price_Q3 = pd.merge(pivot_Q3, pivot_Q4, how="left",
                                 on=["Material No.", "Valuation Class Text", "Vendor Name", "Concat"], copy=False)
        Unit_Price_Q3 = Unit_Price_Q3[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.2", "GR Qty2",
             "Unit Price2"]]
        columns = Unit_Price_Q3.columns.values.tolist()


        Unit_Price = pd.merge(Unit_Price_Q4, Unit_Price_Q3, how="outer",
                              on=["Material No.", "Valuation Class Text", "Vendor Name", "Concat"], copy=False)
        # Unit_Price = pd.concat([Unit_Price_Q4, Unit_Price_Q3], ignore_index=True, sort=True)
        Unit_Price = Unit_Price[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.1", "GR Qty1",
             "Unit Price1", "GR Amt.in loc.cur.2", "GR Qty2", "Unit Price2"]]
        Unit_Price = Unit_Price.reset_index()

        Unit_Price = Unit_Price.replace(numpy.nan, 0, regex=True)
        Unit_Price["Remarks"] = " "


        columns = Unit_Price.columns.values.tolist()
        columns.remove("Remarks")
        columns.insert(4, "Remarks")
        # Re-order columns
        Unit_Price = Unit_Price[columns]

        # Unit_Price_Comparison.sort_values
        columns = Unit_Price.columns.values.tolist()
        pd.options.mode.chained_assignment = None

        for index in Unit_Price.index:
            if Unit_Price[columns[6]][index] == 0:
                Unit_Price['Remarks'][index] = "Not purchased in the current quarter"
            elif Unit_Price[columns[9]][index] == 0:
                Unit_Price['Remarks'][index] = "Purchased in the current quarter"
            else:
                pass

        Unit_Price['Increase/decrease in Amount'] = ""
        pd.options.mode.chained_assignment = None
        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[6]][index]
            GR_qty = Unit_Price[columns[9]][index]
            Unit_price = GR_amt - GR_qty
            Unit_Price['Increase/decrease in Amount'][index] = Unit_price

        Unit_Price['Increase/decrease in Quantity'] = ""
        pd.options.mode.chained_assignment = None
        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[7]][index]
            GR_qty = Unit_Price[columns[10]][index]
            Unit_price = GR_amt - GR_qty
            Unit_Price['Increase/decrease in Quantity'][index] = Unit_price

        Unit_Price['Increase/decrease in Unit Price'] = ""
        pd.options.mode.chained_assignment = None
        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[8]][index]
            GR_qty = Unit_Price[columns[11]][index]
            if isinstance(GR_amt, str) or isinstance(GR_qty, str):
                continue
            Unit_price = GR_amt - GR_qty
            Unit_Price['Increase/decrease in Unit Price'][index] = Unit_price

        columns = Unit_Price.columns.values.tolist()
        Unit_Price['Increase/decrease in unit price (%)'] = ""
        pd.options.mode.chained_assignment = None
        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[14]][index]
            GR_qty = Unit_Price[columns[11]][index]
            if isinstance(GR_amt, str) or isinstance(GR_qty, str):
                continue
            elif GR_qty != 0:
                Unit_price = GR_amt / GR_qty
                Unit_Price['Increase/decrease in unit price (%)'][index] = Unit_price

        columns = Unit_Price.columns.values.tolist()
        Unit_Price['In amount due to Qty'] = ""
        pd.options.mode.chained_assignment = None
        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[13]][index]
            GR_qty = Unit_Price[columns[11]][index]
            if isinstance(GR_amt, str) or isinstance(GR_qty, str):
                continue
            else:
                Unit_price = GR_amt * GR_qty
                Unit_Price['In amount due to Qty'][index] = Unit_price
        columns = Unit_Price.columns.values.tolist()
        Unit_Price['In amount due to Qty (%)'] = ""
        pd.options.mode.chained_assignment = None
        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[16]][index]
            GR_qty = Unit_Price[columns[12]][index]

            if isinstance(GR_amt, str) or isinstance(GR_qty, str):
                continue
            elif GR_qty != 0:
                Unit_price = GR_amt / GR_qty
                Unit_Price['In amount due to Qty (%)'][index] = Unit_price

        columns = Unit_Price.columns.values.tolist()
        Unit_Price['In amount due to price'] = ""
        pd.options.mode.chained_assignment = None

        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[7]][index]
            GR_qty = Unit_Price[columns[14]][index]
            if isinstance(GR_amt, str) or isinstance(GR_qty, str):
                continue
            else:
                Unit_price = GR_amt * GR_qty
                Unit_Price['In amount due to price'][index] = Unit_price

        columns = Unit_Price.columns.values.tolist()
        Unit_Price['In amount due to unit price (%)'] = ""
        pd.options.mode.chained_assignment = None
        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[18]][index]
            GR_qty = Unit_Price[columns[12]][index]
            if isinstance(GR_amt, str) or isinstance(GR_qty, str):
                continue
            elif GR_qty != 0:
                Unit_price = GR_amt / GR_qty
                Unit_Price['In amount due to unit price (%)'][index] = Unit_price

        #  Rename Columns
        Unit_Price = Unit_Price.rename(
            columns={columns[6]: "GR Amt.in loc.cur."})
        Unit_Price = Unit_Price.rename(
            columns={columns[7]: "GR Qty"})
        Unit_Price = Unit_Price.rename(
            columns={columns[8]: "Unit Price"})
        Unit_Price = Unit_Price.rename(
            columns={columns[9]: "GR Amt.in loc.cur."})
        Unit_Price = Unit_Price.rename(
            columns={columns[10]: "GR Qty"})
        Unit_Price = Unit_Price.rename(
            columns={columns[11]: "Unit Price"})
        Unit_Price = Unit_Price.drop(columns=["index"])

        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            Unit_Price.to_excel(writer, sheet_name=main_config["Output_Unit_Price_Comparison_sheetname"], startrow=24, index=False)

        wb = load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Unit_Price_Comparison_sheetname"]]

        cell = ws['F24']
        cell.value = 'Current Quarter Q4'
        ws.merge_cells('F24:H24')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = ws['I24']
        cell.value = 'Previous Quarter Q3'
        ws.merge_cells('I24:K24')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = ws['L24']
        cell.value = 'Increase/Decrease In Amount'
        ws.merge_cells('L24:S24')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        ws['F24'].font = font_style
        ws['I24'].font = font_style
        ws['L24'].font = font_style
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        ws['F24'].fill = fill_pattern
        ws['L24'].fill = fill_pattern
        fill_pattern = PatternFill(patternType="solid", fgColor="FFFF00")
        ws['I24'].fill = fill_pattern

        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "25"].fill = format_fill
            if c == 's':
                break
        m_row = ws.max_row

        ws.auto_filter.ref = "A25:S" + str(m_row)

        for cell in ws["E"]:
            cell.number_format = "#,###"
        for cell in ws["F"]:
            cell.number_format = "#,###"
        for cell in ws["G"]:
            cell.number_format = "#,###"
        for cell in ws["H"]:
            cell.number_format = "#,###"
        for cell in ws["I"]:
            cell.number_format = "#,###"
        for cell in ws["J"]:
            cell.number_format = "#,###"
        for cell in ws["K"]:
            cell.number_format = "#,###"
        for cell in ws["L"]:
            cell.number_format = "#,###"
        for cell in ws["M"]:
            cell.number_format = "#,###"
        for cell in ws["N"]:
            cell.number_format = "0"
        for cell in ws["O"]:
            cell.number_format = numbers.FORMAT_PERCENTAGE
        for cell in ws["P"]:
            cell.number_format = "0.00"
        for cell in ws["Q"]:
            cell.number_format = "0.00"
        for cell in ws["R"]:
            cell.number_format = "0.00"
        for cell in ws["S"]:
            cell.number_format = "0.00%"
        ws.delete_rows(m_row)
        ws.delete_rows(m_row - 1)

        m_row = ws.max_row

        ws['F23'] = '=SUBTOTAL(9,F26:F' + str(m_row) + ')'
        ws['G23'] = '=SUBTOTAL(9,G26:G' + str(m_row) + ')'
        ws['I23'] = '=SUBTOTAL(9,I26:I' + str(m_row) + ')'
        ws['J23'] = '=SUBTOTAL(9,J26:J' + str(m_row) + ')'
        ws['M23'] = '=SUBTOTAL(9,M26:M' + str(m_row) + ')'
        ws['L23'] = '=SUBTOTAL(9,L26:L' + str(m_row) + ')'
        ws['P23'] = '=SUBTOTAL(9,P26:P' + str(m_row) + ')'
        ws['R23'] = '=SUBTOTAL(9,R26:R' + str(m_row) + ')'

        # Auto-fit column width
        for c in ascii_lowercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 's':
                break

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=26, min_col=1, max_row=ws.max_row, max_col=19):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        ws.merge_cells('A4:E4')
        ws.merge_cells('A5:E5')
        ws.merge_cells('A6:E6')
        ws.merge_cells('A7:E7')
        ws.merge_cells('A8:E8')
        ws.merge_cells('A9:E9')
        ws.merge_cells('A10:E10')
        ws.merge_cells('A11:E11')
        ws.merge_cells('A12:E12')
        ws.merge_cells('A13:E13')
        ws.merge_cells('A14:E14')

        # Headers implementation
        ws['A1'] = main_config['CompanyName']
        ws['A2'] = main_config['StatutoryAuditQuarter']
        ws['A3'] = main_config['FinancialYear']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])
        wb.close()

        return Unit_Price

    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Unit Price Comparison Process-", business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", key_error)
        return key_error


config = {}
main_config = {}
present_quarter_pd = pd.DataFrame()
previous_quarter_pd = pd.DataFrame()
if __name__ == "__main__":
    print(create_unit_price(main_config, config, present_quarter_pd, previous_quarter_pd))
