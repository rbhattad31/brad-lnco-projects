import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Side,  Border
from string import ascii_lowercase
from openpyxl.styles import Alignment
from send_mail_reusable_task import send_mail
import os
import logging


class BusinessException(Exception):
    pass


def plant_concentration_top_weight(plant_concentration_dataframe, main_config):
    # save grand total row to delete from datatable to sort
    # print(plant_concentration_dataframe)
    grand_total_row = plant_concentration_dataframe.tail(1)
    # print(grand_total_row)
    # delete last row from the grand_total_row
    plant_concentration_dataframe.drop(plant_concentration_dataframe.tail(1).index, inplace=True)
    # print("Deleted Grand total row")
    # sort the dataframe using column name
    plant_concentration_dataframe.sort_values(by="Variance", ascending=False, inplace=True)
    # print(plant_concentration_dataframe)
    plant_concentration_weightage = pd.DataFrame(columns=plant_concentration_dataframe.columns)
    # print("empty dataframe is created with columns")
    # print(plant_concentration_weightage)
    sum_of_variance = 0
    for index, row in plant_concentration_dataframe.iterrows():
        # print(float(row["Variance"]))
        if sum_of_variance < 0.60:
            sum_of_variance = sum_of_variance + float(row["Variance"])
            # print(sum_of_variance)
            plant_concentration_weightage = plant_concentration_weightage.append(row, ignore_index=True)
            # print("appended row")
        else:
            break
    # print(plant_concentration_weightage)
    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            plant_concentration_weightage.to_excel(writer, sheet_name=main_config[
                "Output_Concentration_Weightage_sheetname"], index=False, startrow=2, startcol=12)
        print("plant type concentration top weightage entries are logged in the output file")
    except Exception as File_creation_error:
        logging.error("Exception occurred while creating purchase type wise concentration sheet")
        raise File_creation_error
    # Load excel file
    workbook = openpyxl.load_workbook(main_config['Output_File_Path'])

    # Load sheet
    worksheet = workbook[main_config['Output_Concentration_Weightage_sheetname']]

    # Set column widths
    for column_letter in ['m', 'n', 'o']:
        column_length = max(len(str(cell.value)) for cell in worksheet[column_letter])
        worksheet.column_dimensions[column_letter].width = column_length * 1.25

    # row 3 font format, fill color

    calibri_11_black_bold = Font(name="Calibri", size=11, color="000000", bold=True)
    light_blue_solid_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
    thin = Side(border_style="thin", color='b1c5e7')
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cambria_11_black = Font(name='Calibri', size=11, color='000000', bold=False)

    for row in worksheet["m3:o3"]:
        for cell in row:
            cell.fill = light_blue_solid_fill
            cell.font = calibri_11_black_bold

    max_row = len(plant_concentration_weightage.index)
    for row in worksheet["M" + str(3 + 1) + ":O" + str(max_row + 3)]:
        for cell in row:
            cell.font = cambria_11_black
            cell.border = thin_border

    # Number format implementation
    for cell in worksheet['N']:
        cell.number_format = "#,###,##"

    # Format Variance
    for cell in worksheet['O']:
        cell.number_format = '0.0%'

    workbook.save(main_config['Output_File_Path'])


def purchase_type(main_config, in_config, present_quarter_pd):
    try:
        logging.info("Starting plant wise concentration code execution")
        # Read Purchase Register Sheets
        read_present_quarter_pd = present_quarter_pd

        # Fetch To Address
        to_address = main_config["To_Mail_Address"]
        cc_address = main_config["CC_Mail_Address"]
        # Check Exception
        if read_present_quarter_pd.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Empty present quarter purchase register found")
            raise BusinessException("Sheet is empty")

        # Check Column Present
        present_quarter_Sheet_col = read_present_quarter_pd.columns.values.tolist()
        for col in ['Plant', "GR Amt.in loc.cur."]:
            if col not in present_quarter_Sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                logging.error("{} Column is missing".format(col))
                raise BusinessException(col + " Column is missing")

        # Filter rows
        plant = read_present_quarter_pd[read_present_quarter_pd['Plant'].notna()]
        gr_amt = read_present_quarter_pd[read_present_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(plant) == 0:
            subject = in_config["Plant_Subject"]
            body = in_config["Plant_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Plant Column is empty")
            raise BusinessException("Plant Column is empty")
        elif len(gr_amt) == 0:
            subject = in_config["GRAmt_Subject"]
            body = in_config["GRAmt_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("GR Amt Column is empty")
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # Create Pivot Table Q4
        pivot_index = ["Plant"]
        pivot_values = ["GR Amt.in loc.cur."]
        pivot_present_quarter = pd.pivot_table(read_present_quarter_pd, index=pivot_index, values=pivot_values, aggfunc=numpy.sum, margins=True,
                                               margins_name='Grand Total')
        logging.info("Present quarter pivot table is created")
        # Remove Index
        pivot_present_quarter = pivot_present_quarter.reset_index()

        # Assign Pivot Sheets
        pivot_sheet = pivot_present_quarter

        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()
        # Delete row of Q4 and Q3 columns values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[1]] == 0)], inplace=True)

        # Create Variance Column
        pivot_sheet['Variance'] = ""

        pd.options.mode.chained_assignment = None

        # Get maximum value
        total_value = pivot_sheet.iloc[-1:]
        total_value = total_value.iloc[0, 1]

        # Variance Formula
        for index in pivot_sheet.index:
            quarter_value = pivot_sheet[col_name[1]][index]

            if total_value == 0:
                variance = 1
            else:
                variance = quarter_value / total_value

            pivot_sheet['Variance'][index] = variance

        # Change Column names of Q4
        pivot_sheet = pivot_sheet.rename(columns={col_name[1]: main_config["PresentQuarterColumnName"]})

        # Log Sheet
        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                pivot_sheet.to_excel(writer, sheet_name=main_config["Output_Concentrations_Plant_sheetname"], index=False,  startrow=16)
        except Exception as File_creation_error:
            logging.error("Exception occurred while creating plant wise concentration sheet")
            raise File_creation_error
        try:
            plant_concentration_top_weight(pivot_sheet, main_config)
        except Exception as plant_concentration_top_weight_error:
            logging.error("Exception occurred while creating plant wise concentration top weight table")
            raise plant_concentration_top_weight_error

        # Check outfile creation
        if os.path.exists(main_config["Output_File_Path"]):
            print("Plant Wise Concentration sheet is created")
            logging.info("Plant wise concentration sheet is created")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.warning("Plant Wise Concentration sheet is not created")
            raise BusinessException("Output file not generated")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Concentrations_Plant_sheetname"]]

        # Format Q4 & Q3
        for cell in ws['B']:
            cell.number_format = "#,###,##"

        # Format Variance
        for cell in ws['C']:
            cell.number_format = '0.0%'

        # Format Header
        format_font = Font(name="Calibri", size=11, color="000000", bold=True)
        for c in ascii_lowercase:
            ws[c + "17"].font = format_font

        # Format Footer
        m_row = ws.max_row
        for c in ascii_lowercase:
            ws[c + str(m_row)].font = format_font

        # Header Fill
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "17"].fill = format_fill
            if c == 'c':
                break

        # Footer Fill
        for c in ascii_lowercase:
            ws[c + str(m_row)].fill = format_fill
            if c == 'c':
                break

        # Alignment
        cell = ws['A' + str(m_row)]
        cell.alignment = Alignment(horizontal='right', vertical='center')

        # Set Width
        for c in ascii_lowercase:
            ws.column_dimensions[c].width = 20

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=17, min_col=1, max_row=ws.max_row, max_col=3):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        ws.merge_cells('A4:E4')
        ws.merge_cells('A5:E5')
        ws.merge_cells('A6:E6')
        ws.merge_cells('A7:E7')
        ws.merge_cells('A8:E8')
        ws.merge_cells('A9:E9')
        ws.merge_cells('A10:E10')
        ws.merge_cells('A11:E11')
        ws.merge_cells('A12:E12')
        ws.merge_cells('A13:E13')
        ws.merge_cells('A14:E14')

        # Headers implementation
        ws['A1'] = main_config['CompanyName']
        ws['A2'] = main_config['StatutoryAuditQuarter']
        ws['A3'] = main_config['FinancialYear']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False
        print(wb.sheetnames)
        # Save File
        wb.save(main_config["Output_File_Path"])
        logging.info("Completed plant wise concentration code execution")
        return ws

    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(file_error))
        print("Please close the file")
        logging.exception(file_error)
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(notfound_error))
        logging.exception(notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Concentration Plant Wise Process-", str(business_error))
        logging.exception(business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(value_error))
        logging.exception(value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(type_error))
        logging.exception(type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(error))
        logging.exception(error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(key_error))
        logging.exception(key_error)
        return key_error


config = {}
main_config = {}
present_quarter_pd = pd.DataFrame()

if __name__ == "__main__":
    purchase_type(main_config, config, present_quarter_pd)

