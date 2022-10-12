import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Side,  Border
from string import ascii_lowercase
from openpyxl.styles import Alignment
from win32com import client
import pywintypes
import os


class BusinessException(Exception):
    pass


def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


def purchase_type(in_config, present_quarter_pd):

    try:

        # Read Purchase Register Sheets
        read_present_quarter_pd = present_quarter_pd
        present_quarter_columns = read_present_quarter_pd.columns
        if in_config["purchase_register_1st_column_name"] in present_quarter_columns and \
                in_config["purchase_register_2nd_column_name"] in present_quarter_columns:
            print("Present Quarter file - The data is starting from first row only")
            pass

        else:
            print("Present Quarter file - The data is not starting from first row ")
            for index, row in read_present_quarter_pd.iterrows():
                if row[0] != in_config["purchase_register_1st_column_name"]:
                    read_present_quarter_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = read_present_quarter_pd.iloc[0]
            read_present_quarter_pd = read_present_quarter_pd[1:]
            read_present_quarter_pd.columns = new_header
            read_present_quarter_pd.reset_index(drop=True, inplace=True)
            read_present_quarter_pd.columns.name = None
        read_present_quarter_pd = read_present_quarter_pd.loc[:, ~read_present_quarter_pd.columns.duplicated(keep='first')]

        # Fetch To Address
        to_address = in_config["To_Address"]
        cc_address = in_config["CC_Address"]

        # Check Exception
        if read_present_quarter_pd.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        PresentQuarterSheetColumns = read_present_quarter_pd.columns.values.tolist()
        for col in ["Valuation Class", "Valuation Class Text", "GR Amt.in loc.cur."]:
            if col not in PresentQuarterSheetColumns:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter rows
        val_class = read_present_quarter_pd[read_present_quarter_pd['Valuation Class'].notna()]
        val_txt = read_present_quarter_pd[read_present_quarter_pd['Valuation Class Text'].notna()]
        gr_amt = read_present_quarter_pd[read_present_quarter_pd['GR Amt.in loc.cur.'].notna()]

        # Check Exception
        if len(val_class) == 0:
            subject = in_config["EmptyValClass_Subject"]
            body = in_config["EmptyValClass_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Valuation Class Column is empty")
        elif len(val_txt) == 0:
            subject = in_config["EmptyValClassTxt_Subject"]
            body = in_config["EmptyValClassTxt_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Valuation Class Text Column is empty")
        elif len(gr_amt) == 0:
            subject = in_config["GRAmt_Subject"]
            body = in_config["GRAmt_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # create Pivot Table Q4
        pivot_index = ["Valuation Class", "Valuation Class Text"]
        pivot_values = ["GR Amt.in loc.cur."]
        pivot_present_quarter = pd.pivot_table(read_present_quarter_pd, index=pivot_index, values=pivot_values, aggfunc=numpy.sum, margins=True,
                                               margins_name='Grand Total', sort=True)

        # Remove Index
        pivot_present_quarter = pivot_present_quarter.reset_index()

        # Assign Sheet
        pivot_sheet = pivot_present_quarter

        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Delete row of Q4 and Q3 columns values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[2]] == 0)], inplace=True)

        # Create Variance Column
        pivot_sheet['Variance'] = ""

        pd.options.mode.chained_assignment = None

        # Get maximum value
        total_value = pivot_sheet.iloc[-1:]
        total_value = total_value.iloc[0, 2]

        # Variance Formula
        for index in pivot_sheet.index:
            quarter_value = pivot_sheet[col_name[2]][index]
            if total_value == 0:
                variance = 1
            else:
                variance = quarter_value / total_value

            pivot_sheet['Variance'][index] = variance

        # Change Column names of Q4 and Q3
        pivot_sheet = pivot_sheet.rename(columns={col_name[2]: in_config["PresentQuarterColumn"]})

        # Log Sheet
        with pd.ExcelWriter(in_config["Type_Path"], engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            pivot_sheet.to_excel(writer, sheet_name=in_config["TypeSheet"], index=False,  startrow=16)
        # Check outfile creation
        if os.path.exists(in_config["Type_Path"]):
            print("Type Wise Comparatives Logged")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Output file not generated")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(in_config["Type_Path"])
        ws = wb[in_config["TypeSheet"]]

        # Format Q3
        for cell in ws['C']:
            cell.number_format = "#,###,##"

        # Format Variance
        for cell in ws['D']:
            cell.number_format = '0.0%'

        # Format Header
        format_font = Font(name="Calibri", size=11, color="000000", bold=True)
        for c in ascii_lowercase:
            ws[c + "17"].font = format_font

        # Format Footer
        m_row = ws.max_row
        for c in ascii_lowercase:
            ws[c + str(m_row)].font = format_font

        # Header Fill
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "17"].fill = format_fill
            if c == 'd':
                break

        # Footer Fill
        for c in ascii_lowercase:
            ws[c + str(m_row)].fill = format_fill
            if c == 'd':
                break

        # Merge Grand Total
        ws.merge_cells('A'+str(m_row)+':B'+str(m_row))
        cell = ws['A'+str(m_row)]
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set Width
        for c in ascii_lowercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 'd':
                break

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=17, min_col=1, max_row=ws.max_row, max_col=4):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        ws.merge_cells('A4:E4')
        ws.merge_cells('A5:E5')
        ws.merge_cells('A6:E6')
        ws.merge_cells('A7:E7')
        ws.merge_cells('A8:E8')
        ws.merge_cells('A9:E9')
        ws.merge_cells('A10:E10')
        ws.merge_cells('A11:E11')
        ws.merge_cells('A12:E12')
        ws.merge_cells('A13:E13')
        ws.merge_cells('A14:E14')

        # Headers implementation
        ws['A1'] = in_config['A1']
        ws['A2'] = in_config['A2']
        ws['A3'] = in_config['A3']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False

        # Save File
        wb.save(in_config["Type_Path"])
        wb = openpyxl.load_workbook(in_config["Type_Path"])
        print(wb.sheetnames)
        wb.save(in_config["Type_Path"])
        return ws

    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Type Wise Process-", file_error)
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Type Wise Process-", notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Concentration Type Wise Process-", business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Type Wise Process-", value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Type Wise Process-", type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Type Wise Process-", error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Type Wise Process-", key_error)
        return key_error


# Read config details and parse to dictionary
config = {}
present_quarter_pd = pd.DataFrame()

if __name__ == "__main__":
    print(purchase_type(config, present_quarter_pd))
