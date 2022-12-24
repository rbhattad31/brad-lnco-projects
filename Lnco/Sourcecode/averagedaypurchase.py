import os
from string import ascii_lowercase
import xlrd_compdoc_commented as xlrd
import openpyxl
import pandas as pd
from send_mail_reusable_task import send_mail
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
import logging

#  Function produce average day purchase output sheet


def average_day_purchase_weightage(pivot_present_quarter, main_config):
    average_day_purchase_weightage = pd.DataFrame(columns=pivot_present_quarter.columns)

    lowest_negative_row = pivot_present_quarter.tail(1)
    print(lowest_negative_row)
    lowest_negative_percentage = float(lowest_negative_row['Percentage'])
    print(lowest_negative_percentage)
    highest_positive_row = pivot_present_quarter.head(1)
    print(highest_positive_row)
    highest_positive_percentage = float(highest_positive_row['Percentage'])
    print(highest_positive_percentage)
    lowest_positive_percentage = highest_positive_percentage
    print(lowest_positive_percentage)
    highest_negative_percentage = lowest_negative_percentage
    print(highest_negative_percentage)
    lowest_positive_row = highest_positive_row
    highest_negative_row = lowest_negative_row
    for index, row in pivot_present_quarter.iterrows():
        percentage = float(row['Percentage'])
        if 0 <= percentage < lowest_positive_percentage:
            lowest_positive_percentage = percentage
            lowest_positive_row = row
        if 0 > percentage > highest_negative_percentage:
            highest_negative_percentage = percentage
            highest_negative_row = row
    average_day_purchase_weightage = average_day_purchase_weightage.append(highest_positive_row, ignore_index=True)
    average_day_purchase_weightage = average_day_purchase_weightage.append(lowest_positive_row, ignore_index=True)
    average_day_purchase_weightage = average_day_purchase_weightage.append(highest_negative_row, ignore_index=True)
    average_day_purchase_weightage = average_day_purchase_weightage.append(lowest_negative_row, ignore_index=True)
    # print(average_day_purchase_weightage)

    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            average_day_purchase_weightage.to_excel(writer, sheet_name=main_config[
                "Output_Average_Day_Weightage_Sheetname"], index=False, startrow=0, startcol=0)
            print("Average day purchase highest and lowest entries are logged in the output file")

    except Exception as File_creation_error:
        logging.error("Exception occurred while creating Average day purchase highest and lowest entries sheet: \n {0}".format(
            File_creation_error))
        raise File_creation_error

    # Load excel file
    workbook = openpyxl.load_workbook(main_config['Output_File_Path'])

    # Load sheet
    worksheet = workbook[main_config['Output_Average_Day_Weightage_Sheetname']]

    # Assign max row value to variable
    m_row = worksheet.max_row

    # Set column widths
    for c in ascii_lowercase:
        column_length = max(len(str(cell.value)) for cell in worksheet[c])
        worksheet.column_dimensions[c].width = column_length * 1.25
        if c == 'e':
            break

    # row 3 font format, fill color
    fill_hawkes_blue_color = PatternFill(fgColor='d9e1f2', fill_type="solid")
    cambria_11_sapphire_bold = Font(name='Cambria', size=11, color='002060', bold=True)
    cambria_11_sapphire = Font(name='Cambria', size=11, color='002060', bold=False)

    for row in worksheet["A1:E1"]:
        for cell in row:
            cell.fill = fill_hawkes_blue_color
            cell.font = cambria_11_sapphire_bold

    # Cell border implementation for the table
    thin = Side(border_style="thin", color='b1c5e7')

    for row in worksheet.iter_rows(min_row=1, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # font format
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.font = cambria_11_sapphire

    # Number format implementation
    for cell in worksheet['A']:
        cell.number_format = 'dd-mm-yyyy'

    for cell in worksheet['B']:
        cell.number_format = '#,###'

    for cell in worksheet['C']:
        cell.number_format = '#,###'

    for cell in worksheet['D']:
        cell.number_format = '#,###'

    for cell in worksheet['E']:
        cell.number_format = '000%'

    workbook.save(main_config['Output_File_Path'])


def average_day_purchase(main_config, in_config, present_quarter_pd):
    to_ = main_config["To_Mail_Address"]
    cc_ = main_config["CC_Mail_Address"]

    try:

        # Reading the data from excel
        read_present_quarter_pd = present_quarter_pd

        # Empty data check
        if read_present_quarter_pd.empty:
            print('Empty DataFrame')
            raise ValueError

        # Column existence check
        if pd.Series(['GR Amt.in loc.cur.', 'GR Posting Date']).isin(read_present_quarter_pd.columns).all():
            # print("Columns exist")
            pass
        else:
            print("One or more columns doesnt exist in the read_present_quarter_pd")
            raise KeyError

        # Empty column data check
        if read_present_quarter_pd['GR Amt.in loc.cur.'].isnull().all():
            print('Column GR Amt.in loc.cur. is empty')
            raise ValueError

        if read_present_quarter_pd['GR Posting Date'].isnull().all():
            print('Column GR Posting Date is empty')
            raise ValueError

        # Generating pivot table
        pivot_present_quarter = pd.pivot_table(read_present_quarter_pd, values='GR Amt.in loc.cur.',
                                               index='GR Posting Date',
                                               observed=True, sort=True, aggfunc='sum', margins=True,
                                               margins_name="Grand Total")

        # Sort pivot table
        pivot_present_quarter = pivot_present_quarter.reset_index()

        # Store the last grand total value to variable
        present_total_gr_amt = (pivot_present_quarter.iloc[-1]['GR Amt.in loc.cur.'])

        # Remove last row from pivot table
        pivot_present_quarter = pivot_present_quarter.iloc[:-1]

        # Arithmetic operations

        pivot_present_quarter['Average'] = present_total_gr_amt / (len(pivot_present_quarter))
        pivot_present_quarter['Variance'] = pivot_present_quarter['Average'] - pivot_present_quarter['GR Amt.in loc.cur.']
        pivot_present_quarter['Percentage'] = pivot_present_quarter['Variance'] / pivot_present_quarter['Average']
        pivot_present_quarter.columns = ['GR Posting Date', 'Amount as per purchase register', 'Average purchases',
                                         'Variance','Percentage']

        # Sort pivot table based of percentage column in ascending order
        pivot_present_quarter.sort_values(by='Percentage', ascending=False, inplace=True)

        # Save pivot table to Excel file
        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                pivot_present_quarter.to_excel(writer, sheet_name=main_config['Output_Average_Day_Purchase_sheetname'],
                                               startrow=24, index=False)
        except Exception as File_creation_error:
            logging.error("Exception occurred while creating average day purchases sheet")
            raise File_creation_error
        try:
            average_day_purchase_weightage(pivot_present_quarter, main_config)
        except Exception as average_day_high_low_entries_error:
            print("Exception occurred while creating average day Highest and Least Values Table in excel sheet: \n {0}".format(
                average_day_high_low_entries_error))

        #   Formatting and styling the Excel data
        # Load excel file
        workbook = openpyxl.load_workbook(main_config['Output_File_Path'])

        # Load sheet
        worksheet = workbook[main_config['Output_Average_Day_Purchase_sheetname']]

        # Assign max row value to variable
        m_row = worksheet.max_row

        # Set column widths
        for c in ascii_lowercase:
            column_length = max(len(str(cell.value)) for cell in worksheet[c])
            worksheet.column_dimensions[c].width = column_length * 1.25
            if c == 'e':
                break

        # Implement subtotal formula for max row values
        worksheet['B24'] = '=SUBTOTAL(9,B26:B' + str(m_row) + ')'
        worksheet['C24'] = '=SUBTOTAL(9,C26:C' + str(m_row) + ')'

        # Set font style variable configuration
        cambria_11_sapphire_bold = Font(name='Cambria', size=11, color='002060', bold=True)
        cambria_12_sapphire = Font(name='Cambria', size=12, color='002060', bold=False)
        cambria_12_sapphire_bold_underline = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        cambria_14_sapphire_bold = Font(name='Cambria', size=14, color='002060', bold=True)

        # Implement the configuration to appropriate rows
        # Set Cambria to Total values above the table
        worksheet['B24'].font = cambria_11_sapphire_bold
        worksheet['C24'].font = cambria_11_sapphire_bold

        # Number format implementation
        for cell in worksheet['A']:
            cell.number_format = 'dd-mm-yyyy'

        for cell in worksheet['B']:
            cell.number_format = '#,###'

        for cell in worksheet['C']:
            cell.number_format = '#,###'

        for cell in worksheet['D']:
            cell.number_format = '#,###'

        for cell in worksheet['E']:
            cell.number_format = '000%'

        # Cell color variable assignment
        fill_hawkes_blue_color = PatternFill(fgColor='d9e1f2', fill_type="solid")

        # Cell color implementation for appropriate rows
        for row in worksheet["A25:E25"]:
            for cell in row:
                cell.fill = fill_hawkes_blue_color
                cell.font = cambria_11_sapphire_bold

        # Cell border implementation
        thin = Side(border_style="thin", color='b1c5e7')

        for row in worksheet.iter_rows(min_row=25, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for row in worksheet.iter_rows(min_row=26, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.font = Font(name='Cambria', size=11, color='002060')

        for row in worksheet.iter_rows(min_row=24, min_col=1, max_row=24, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')

        # Cell merge for headers implementation
        worksheet.merge_cells('A1:E1')
        worksheet.merge_cells('A2:E2')
        worksheet.merge_cells('A3:E3')
        worksheet.merge_cells('A4:E4')
        worksheet.merge_cells('A5:E5')
        worksheet.merge_cells('A6:E6')
        worksheet.merge_cells('A7:E7')
        worksheet.merge_cells('A8:E8')
        worksheet.merge_cells('A9:E9')
        worksheet.merge_cells('A10:E10')
        worksheet.merge_cells('A11:E11')
        worksheet.merge_cells('A12:E12')
        worksheet.merge_cells('A13:E13')
        worksheet.merge_cells('A14:E14')
        worksheet.merge_cells('A15:E15')
        worksheet.merge_cells('A16:E16')
        worksheet.merge_cells('A17:E17')
        worksheet.merge_cells('A18:E18')
        worksheet.merge_cells('A19:E19')
        worksheet.merge_cells('A20:E20')
        worksheet.merge_cells('A21:E21')
        worksheet.merge_cells('A22:E22')

        # Headers implementation
        worksheet['A1'] = main_config['CompanyName']
        worksheet['A2'] = main_config['StatutoryAuditQuarter']
        worksheet['A3'] = main_config['FinancialYear']
        worksheet['A4'] = in_config['A4']
        worksheet['A5'] = in_config['A5']
        worksheet['A7'] = in_config['A7']
        worksheet['A8'] = in_config['A8']
        worksheet['A10'] = in_config['A11']
        worksheet['A11'] = in_config['A12']
        worksheet['A13'] = in_config['A13']
        worksheet['A14'] = in_config['A14']
        worksheet['A15'] = in_config['A15']
        worksheet['A16'] = in_config['A16']

        # Headers formatting and styling
        for row in worksheet.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = cambria_14_sapphire_bold

        for row in worksheet.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = cambria_12_sapphire_bold_underline

        for row in worksheet.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = cambria_12_sapphire_bold_underline

        for row in worksheet.iter_rows(min_row=13, min_col=1, max_row=13, max_col=1):
            for cell in row:
                cell.font = cambria_12_sapphire_bold_underline

        for row in worksheet.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = cambria_12_sapphire

        for row in worksheet.iter_rows(min_row=11, min_col=1, max_row=11, max_col=1):
            for cell in row:
                cell.font = cambria_12_sapphire

        for row in worksheet.iter_rows(min_row=14, min_col=1, max_row=16, max_col=1):
            for cell in row:
                cell.font = cambria_12_sapphire

        # Remove Gridlines
        worksheet.sheet_view.showGridLines = False

        worksheet.auto_filter.ref = "A25:E" + str(m_row)

        workbook.save(main_config['Output_File_Path'])
        # Empty cell check
        wb = xlrd.open_workbook(main_config['Output_File_Path'])
        wb_sheet = wb.sheet_by_index(0)

        for row in range(25, wb_sheet.nrows):
            for column in range(0, wb_sheet.ncols):
                if wb_sheet.cell_value(row, column) == "":
                    print('row', row + 1, 'col', column + 1, 'is empty')
                    raise RuntimeError

        wb = openpyxl.load_workbook(main_config['Output_File_Path'])
        print(wb.sheetnames)
        wb.save(main_config['Output_File_Path'])

    #  Exceptions handling
    except FileNotFoundError:
        print("Check with the file paths")
        print("Sending notification through outlook mail...")

        subject_ = 'Required file not found'
        body_ = in_config['FileNotFoundError']
        send_mail(to_, cc_, subject_, body_)
        print("exception: ", FileNotFoundError)
    except PermissionError:
        print("Close the excel file before execution")
        os.system('TASKKILL /F /IM excel.exe')
        print("Re executing...")
        average_day_purchase(main_config, main_config, present_quarter_pd)
        print("exception: ", PermissionError)
    except KeyError:
        print("Check with the input column names provided from the source file and program")
        print("Sending notification through outlook mail...")
        subject_ = 'Required column not found'
        body_ = in_config['KeyError']
        send_mail(to_, cc_, subject_, body_)
        print("exception: ", KeyError)
    except ValueError:
        print("Check if the column data values are empty and also"
              "the excel file could have no data or one of the column data is empty")
        print("Sending notification through outlook mail...")
        subject_ = 'Excel file has empty values'
        body_ = in_config['ValueError']
        send_mail(to_, cc_, subject_, body_)
        print("exception: ", ValueError)
    except TypeError:
        print('Type error occurred')
        print("Sending notification through outlook mail...")
        subject_ = 'Incorrect column format'
        body_ = in_config['TypeError']
        send_mail(to_, cc_, subject_, body_)
        print("exception: ", TypeError)
    except RuntimeError:
        print('Run time error occurred')
        print("Sending notification through outlook mail...")
        subject_ = 'Empty cell found'
        body_ = in_config['RuntimeError']
        send_mail(to_, cc_, subject_, body_)
        print("exception: ", RuntimeError)
    except FileExistsError:
        print("Directory not found")
        print("Sending notification through outlook mail...")
        subject_ = 'Directory not found'
        body_ = in_config['FileExistsError']
        send_mail(to_, cc_, subject_, body_)
        print("exception: ", FileExistsError)
    except SystemError:
        print("System Exception occurred")
        print("exception: ", SystemError)
    except ArithmeticError:
        print("Look for overflow, zero division and floating point errors causing formula and values")
        print("exception: ", ArithmeticError)
    except IndexError:
        print("Look for incorrect index values")
        print("exception: ", IndexError)
    except SyntaxError:
        print("Look for incorrect syntax input")
        print("exception: ", SyntaxError)
    finally:
        print("Process is over")


if __name__ == "__main__":
    pass
