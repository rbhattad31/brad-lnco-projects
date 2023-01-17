import openpyxl
from ReusableTasks.send_mail_reusable_task import send_mail
import os


# function "reading_sheet_config_data_to_dict" reads sheet wise config file and creates sheet specific config dictionary
def create_sheet_wise_config_dict(sheet_name):
    try:
        config = {}
        present_working_directory = os.getcwd()
        config_file_path = os.path.join(os.path.dirname(present_working_directory), 'Input', 'Config.xlsx')
        work_book = openpyxl.load_workbook(config_file_path)
        work_sheet = work_book[sheet_name]
        maximum_row = work_sheet.max_row
        maximum_col = work_sheet.max_column

        for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            cell_name = config_details[0].value
            cell_value = config_details[1].value
            config[cell_name] = cell_value

        return config

    except Exception as config_error:
        print("Failed to load config file for sheet:", sheet_name)
        print(config_error)
        to = "kalyan.gundu@bradsol.com"
        cc = "kalyan.gundu@bradsol.com"
        subject = "Config reading is failed for sheet: " + sheet_name
        body = '''
Hello,

Config file is failed to load. Continuing with next process.

Thanks & Regards,
L & Co  

'''
        send_mail(to=to, cc=cc, subject=subject, body=body)
        raise Exception
