import openpyxl


def create_main_config_dictionary(path, sheet_name):
    try:
        dict_main_config = {}
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook[sheet_name]
        maximum_row = worksheet.max_row
        maximum_col = worksheet.max_column

        for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            key = row[0].value
            value = row[1].value
            dict_main_config[key] = value
        workbook.close()
        return dict_main_config

    except Exception as config_error:
        # print("failed to load main config file. Hence, stopping the BOT")
        # print(config_error)
        raise config_error


if __name__ == '__main__':
    pass
