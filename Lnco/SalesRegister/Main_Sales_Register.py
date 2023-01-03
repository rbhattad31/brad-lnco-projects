import logging

import openpyxl
import xlsxwriter
import pandas as pd
import os.path
import os
import sys
from decouple import Config, RepositoryEnv

from sales_send_mail_reusable_task import send_mail, send_mail_with_attachment


def reading_sheets_names_from_config_main_sheet(path, sheet_name):
    try:
        config_sheets = {}
        work_book = openpyxl.load_workbook(path)
        work_sheet = work_book[sheet_name]
        maximum_row = work_sheet.max_row
        maximum_col = work_sheet.max_column

        for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            cell_name = config_details[0].value
            cell_value = config_details[1].value
            config_sheets[cell_name] = cell_value
        work_book.save(path)
        return config_sheets

    except Exception as config_error:
        # print("failed to load main config file. Hence, stopping the BOT")
        # print(config_error)
        raise config_error


# function "reading_sheet_config_data_to_dict" reads sheet wise config file and creates sheet specific config dictionary
def reading_sheet_config_data_to_dict(sheet_name):
    try:
        config = {}
        work_book = openpyxl.load_workbook("Input/Config.xlsx")
        work_sheet = work_book[sheet_name]
        maximum_row = work_sheet.max_row
        maximum_col = work_sheet.max_column

        for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            cell_name = config_details[0].value
            cell_value = config_details[1].value
            config[cell_name] = cell_value

        return config

    except Exception as config_error:
        print("Failed to load config file for sheet:", sheet_name)
        print(config_error)
        to = "kalyan.gundu@bradsol.com"
        cc = "kalyan.gundu@bradsol.com"
        subject = "Config reading is failed for sheet: " + sheet_name
        body = '''
Hello,

Config file is failed to load. Continuing with next process.

Thanks & Regards,
L & Co  

'''
        send_mail(to=to, cc=cc, subject=subject, body=body)
        raise Exception


def process_execution(input_files,
                      present_quarter_sheet_name, previous_quarter_sheet_name,
                      present_quarter_column_name, previous_quarter_column_name,
                      company_name, statutory_audit_quarter, financial_year, config_main, request_id,
                      ):
    print("Starting audit process for the input files")
    logging.info("Starting audit process for the input files")
    print(input_files)
    sales_register_present_quarter_file_path = input_files[0]
    sales_register_previous_quarter_file_path = input_files[1]

    config_main['PresentQuarterColumnName'] = present_quarter_column_name
    config_main['PreviousQuarterColumnName'] = previous_quarter_column_name
    config_main['CompanyName'] = company_name
    config_main['StatutoryAuditQuarter'] = statutory_audit_quarter
    config_main['FinancialYear'] = financial_year

    # reading env file
    env_file = 'envfiles/sales_register_quality.env'
    print("ENV_FILE: ", env_file)

    env_file = Config(RepositoryEnv(env_file))

    print("*******************************************")
    # send Bot starting mail
    start_to = config_main['To_Mail_Address']
    start_cc = config_main['CC_Mail_Address']
    start_subject = config_main['Start_Mail_Subject']
    start_body = config_main['Start_Mail_Body']
    send_mail(to=start_to, cc=start_cc, body=start_body, subject=start_subject)
    print("Process start mail notification is sent")

    print("*******************************************")
    print("Check if Output file exists")
    # output_file_path = config_main["Output_File_Path"]
    # output_file_path = "Output/Output.xlsx"
    project_home_directory = os.getcwd()
    output_file_path = os.path.join(project_home_directory, 'Data', 'Output', 'audit_requests', str(request_id))
    print("Output file folder is : ", output_file_path)
    if not os.path.exists(output_file_path):
        print("Output folder is not exist")
        print("Creating directory: ", output_file_path)
        os.makedirs(output_file_path)
        print("Directory" + output_file_path + " is created")
    output_file_name = company_name.replace(' ', '_') + "_" + str(request_id) + "_Purchase_Register_Output.xlsx"
    output_file_path = os.path.join(output_file_path, output_file_name)
    print("Output file path is: " + output_file_path)
    config_main['Output_File_Path'] = output_file_path
    if os.path.exists(output_file_path):
        print("Output file exist")
        print("Deleting existing output file")
        os.remove(output_file_path)
        if os.path.exists(output_file_path):
            pass
        else:
            print("existing output file is deleted successfully")
        print("Creating a new output file")
        workbook = xlsxwriter.Workbook(output_file_path)
        workbook.close()
        if os.path.exists(output_file_path):
            print("New output file is created")
        else:
            print("New output file creation is failed")
    else:
        print("Output file not exist")
        print("Creating a new output file")
        workbook = xlsxwriter.Workbook(output_file_path)
        workbook.close()
        if os.path.exists(output_file_path):
            print("New output file is created")
        else:
            print("New output file creation is failed")

    print("*******************************************")

    try:
        print("Reading Purchase registers is started")
        print("Reading present quarter sheet")
        print(sales_register_present_quarter_file_path)
        read_present_quarter_pd = pd.read_excel(sales_register_present_quarter_file_path,
                                                present_quarter_sheet_name)
        # print(read_present_quarter_pd.dtypes.to_list)
        read_present_quarter_pd = \
            read_present_quarter_pd.loc[:, ~read_present_quarter_pd.columns.duplicated(keep='first')]

        present_quarter_columns = read_present_quarter_pd.columns
        if config_main["sales_register_1st_column_name"] in present_quarter_columns and \
                config_main["sales_register_2nd_column_name"] in present_quarter_columns:
            print("Present Quarter file - The data is starting from first row only")
            pass

        else:
            print("Present Quarter file - The data is not starting from first row ")
            for index, row in read_present_quarter_pd.iterrows():
                if row[0] != config_main["sales_register_1st_column_name"]:
                    read_present_quarter_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = read_present_quarter_pd.iloc[0]
            read_present_quarter_pd = read_present_quarter_pd[1:]
            read_present_quarter_pd.columns = new_header
            read_present_quarter_pd.reset_index(drop=True, inplace=True)
            read_present_quarter_pd.columns.name = None
        read_present_quarter_pd = \
            read_present_quarter_pd.loc[:, ~read_present_quarter_pd.columns.duplicated(keep='first')]
        present_quarter_all_columns_dataframe = read_present_quarter_pd
        # print(
        #     "Reading purchase register present quarter sheet is complete, creating new input file only with required columns")
        # logging.info(
        #     "Reading purchase register present quarter sheet is complete, creating new input file only with required columns")
        # purchase_register_present_quarter_folder_path = os.path.dirname(purchase_register_present_quarter_file_path)
        # purchase_register_present_quarter_file_name = os.path.basename(
        #     purchase_register_present_quarter_file_path).lower()
        # filtered_purchase_present_file_name = "filtered_" + str(purchase_register_present_quarter_file_name)
        # filtered_purchase_present_file_saving_path = os.path.join(purchase_register_present_quarter_folder_path,
        #                                                           filtered_purchase_present_file_name)
        # filtered_purchase_present_sheet_name = present_quarter_sheet_name

        # reading previous quarter sheet
        # print("Reading previous quarter sheet")
        # logging.info("Reading previous quarter sheet")
        # read_previous_quarter_pd = pd.read_excel(sales_register_previous_quarter_file_path,
        #                                          previous_quarter_sheet_name)
        # # print(read_previous_quarter_pd.dtypes.to_list)
        # read_previous_quarter_pd = \
        #     read_previous_quarter_pd.loc[:, ~read_previous_quarter_pd.columns.duplicated(keep='first')]
        #
        # previous_quarter_columns = read_previous_quarter_pd.columns
        # if config_main["sales_register_1st_column_name"] in previous_quarter_columns and \
        #         config_main["sales_register_2nd_column_name"] in previous_quarter_columns:
        #     print("Previous Quarter file - The data is starting from first row only")
        #     pass
        #
        # else:
        #     print("Previous Quarter file - The data is not starting from first row ")
        #     for index, row in read_previous_quarter_pd.iterrows():
        #         if row[0] != config_main["purchase_register_1st_column_name"]:
        #             read_previous_quarter_pd.drop(index, axis=0, inplace=True)
        #         else:
        #             break
        #     new_header = read_previous_quarter_pd.iloc[0]
        #     read_previous_quarter_pd = read_previous_quarter_pd[1:]
        #     read_previous_quarter_pd.columns = new_header
        #     read_previous_quarter_pd.reset_index(drop=True, inplace=True)
        #     read_previous_quarter_pd.columns.name = None
        # read_previous_quarter_pd = \
        #     read_previous_quarter_pd.loc[:, ~read_previous_quarter_pd.columns.duplicated(keep='first')]
        # print(
        #     "Reading purchase register previous quarter sheet is complete, creating new input file only with required columns")
        # logging.info(
        #     "Reading purchase register previous quarter sheet is complete, creating new input file only with required columns")
        # purchase_register_previous_quarter_folder_path = os.path.dirname(sales_register_previous_quarter_file_path)
        # purchase_register_previous_quarter_file_name = os.path.basename(
        #     sales_register_previous_quarter_file_path).lower()
        # filtered_purchase_previous_file_name = "filtered_" + str(purchase_register_previous_quarter_file_name)
        # filtered_purchase_previous_file_saving_path = os.path.join(purchase_register_previous_quarter_folder_path,
        #                                                            filtered_purchase_previous_file_name)
        # filtered_purchase_previous_sheet_name = previous_quarter_sheet_name
        #
        # read_previous_quarter_pd = purchase_previous_quarter_file_creation(read_previous_quarter_pd, json_data_list,
        #                                                                    filtered_purchase_previous_file_saving_path,
        #                                                                    filtered_purchase_previous_sheet_name)
        # logging.info("new purchase register previous quarter file is created in input folder in request ID folder")
        # print("new purchase register previous quarter file is created in input folder in request ID folder")
    except FileNotFoundError as notfound_error:
        send_mail(to=config_main["To_Mail_Address"], cc=config_main["CC_Mail_Address"],
                  subject=config_main["subject_file_not_found"],
                  body=config_main["body_file_not_found"])
        print(notfound_error)
        logging.error("file not found error occurred: \n\t {}".format(notfound_error))
        raise notfound_error
    except ValueError as sheetNotFound_error:
        send_mail(to=config_main["To_Mail_Address"], cc=config_main["CC_Mail_Address"],
                  subject=config_main["subject_sheet_not_found"],
                  body=config_main["body_sheet_not_found"])
        print(sheetNotFound_error)
        logging.error("sheet not found error occurred: \n\t {}".format(sheetNotFound_error))
        raise sheetNotFound_error

    print("*******************************************")
    print("Executing Comparatives Purchase type code")

    try:
        if env_file('GST Rate Check') == 'YES':
            sales_present_quarter_pd = read_present_quarter_pd

            config_gst_rate_check = reading_sheet_config_data_to_dict(
                sheet_name=config_main["Config_GST_Rate_Check_sheet_name"])
            ptcomp.create_purchase_type_wise(config_main, config_gst_rate_check, sales_present_quarter_pd)

        elif env_file('GST Rate Check') == 'NO':
            print("GST Rate Check process is skipped as per env file")
        else:
            print("select YES/NO for GST Rate Check process in env file")
            raise Exception("Error in Env file for 'GST Rate Check' sheet")
    except Exception as e:
        print("Exception caught for Process: 'GST Rate Check' Sheet: ", e)

    print("*******************************************")

    final_output_file = openpyxl.load_workbook(output_file_path)
    if 'Sheet1' in final_output_file.sheetnames:
        final_output_file.remove(final_output_file['Sheet1'])
    final_output_file.save(output_file_path)

    # ------------------------------------------------------------------------------------

    config_saving_file_path_in_output = os.path.join(project_home_directory, 'Data', 'Output', 'audit_requests',
                                                     str(request_id), 'config.xlsx')
    config_saving_folder_path_in_output = os.path.join(project_home_directory, 'Data', 'Output', 'audit_requests',
                                                       str(request_id))
    if not os.path.exists(config_saving_folder_path_in_output):
        logging.warning("folder path {0} is not exist".format(config_saving_folder_path_in_output))
        os.makedirs(config_saving_folder_path_in_output)
        logging.warning("created the directory: {0}".format(config_saving_folder_path_in_output))
    print("Creating new Config file in output folder of Request from Config Dictionary")
    logging.info("Creating new Config file in output folder of Request from Config Dictionary")
    try:
        new_config_df = pd.DataFrame.from_dict(config_main, orient='index', columns=['Value'])
        new_config_df.index.name = 'Key'
        new_config_df.reset_index(level=0, inplace=True)
        new_config_df.to_excel(config_saving_file_path_in_output, index=False)
        print("Created new Config file in output folder of Request from Config Dictionary")
        logging.info("Created new Config file in Output folder of Request from Config Dictionary")
    except Exception as config_file_save_exception:
        logging.warning("Exception occurred while saving config file in Output folder of Request Folder directory")
        logging.exception(config_file_save_exception)
        exception_type, exception_object, exception_traceback = sys.exc_info()
        filename = exception_traceback.tb_frame.f_code.co_filename
        line_number = exception_traceback.tb_lineno
        logging.warning(str(exception_type))
        logging.warning("Exception occurred in file : {} at line number: {}".format(filename, line_number))
    # ------------------------------------------------------------------------------------
    print("Saved config data to an excel file")

    # Bot success mail notification
    end_to = config_main['To_Mail_Address']
    end_cc = config_main['CC_Mail_Address']
    end_subject = config_main['Success_Mail_Subject']
    end_body = config_main['Success_Mail_Body']
    send_mail_with_attachment(to=end_to, cc=end_cc, body=end_body, subject=end_subject,
                              attachment_path=output_file_path)
    print("Process complete mail notification is sent")

    print("Bot successfully finished Processing of the sheets")
    return output_file_path


if __name__ == '__main__':
    sales_present_quarter_file_path = 'sales register q1.xlsx'
    sales_previous_quarter_file_path = ''
    input_files = [sales_present_quarter_file_path, sales_previous_quarter_file_path]
    present_quarter_sheet_name = 'Sheet1'
    previous_quarter_sheet_name = ''
    present_quarter_column_name = 'Q1 FY 2022-23'
    previous_quarter_column_name = 'Q4 FY 2021-22'
    company_name = 'Pitti Engineering Limited'
    statutory_audit_quarter = ''
    financial_year = '2022-23'

    path = 'Input/Config.xlsx'
    config_sheet_name = 'Main'
    config_main = reading_sheets_names_from_config_main_sheet(path, config_sheet_name)

    request_id = 20
    process_execution(input_files,
                      present_quarter_sheet_name, previous_quarter_sheet_name,
                      present_quarter_column_name, previous_quarter_column_name,
                      company_name, statutory_audit_quarter, financial_year, config_main, request_id,
                      )
