import logging
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase
from openpyxl.utils import get_column_letter
from Lnco.ReusableTasks.send_mail_reusable_task import send_mail


class BusinessException(Exception):
    pass


def same_material_domestic(dict_main_config, dict_in_config, sales_present_quarter_pd):
    try:
        logging.info("Starting Same Material domestic code execution")

        # Fetch To Address
        str_to_address = dict_main_config["To_Mail_Address"]
        str_cc_address = dict_main_config["CC_Mail_Address"]

        # Check Exception
        if sales_present_quarter_pd.shape[0] == 0:
            str_subject = dict_in_config["EmptyInput_Subject"]
            str_body = dict_in_config["EmptyInput_Body"]
            send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
            logging.error("Empty present quarter Sales Register found")
            raise BusinessException("Sheet is empty")

        # Check Column Present
        sales_present_quarter_column_list = sales_present_quarter_pd.columns.values.tolist()
        for col in ["Material No.", "Material Description", "Doc. Type Text", "So Unit Price"]:
            if col not in sales_present_quarter_column_list:
                str_subject = dict_in_config["ColumnMiss_Subject"]
                str_body = dict_in_config["ColumnMiss_Body"]
                str_body = str_body.replace("ColumnName +", col)
                send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
                logging.error("{} Column is missing".format(col))
                raise BusinessException(col + " Column is missing")

        # Filter rows
        pd_material_no = sales_present_quarter_pd[sales_present_quarter_pd['Material No.'].notna()]
        pd_material_description = sales_present_quarter_pd[sales_present_quarter_pd['Material Description'].notna()]
        pd_doc_type_text = sales_present_quarter_pd[sales_present_quarter_pd['Doc. Type Text'].notna()]
        pd_so_unit_price = sales_present_quarter_pd[sales_present_quarter_pd['So Unit Price'].notna()]

        if len(pd_material_no) == 0:
            str_subject = dict_in_config["material_no_Subject"]
            str_body = dict_in_config["material_no_Body"]
            send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
            logging.error("Material No. Column is empty")
            raise BusinessException("Material No. Column is empty")
        elif len(pd_material_description) == 0:
            str_subject = dict_in_config["material_description_Subject"]
            str_body = dict_in_config["material_description_Body"]
            send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
            logging.error("Material Description Column is empty")
            raise BusinessException("Material Description Column is empty")
        elif len(pd_doc_type_text) == 0:
            str_subject = dict_in_config["doc_type_text_Subject"]
            str_body = dict_in_config["doc_type_text_Body"]
            send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
            logging.error("Doc type text Column is empty")
            raise BusinessException("Doc type text Column is empty")
        elif len(pd_so_unit_price) == 0:
            str_subject = dict_in_config["so_unit_price_Subject"]
            str_body = dict_in_config["so_unit_price_Body"]
            send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
            logging.error("So unit price Column is empty")
            raise BusinessException("So unit price Column is empty")
        else:
            pass

        # Create Pivot Table Sales Register
        try:
            list_pivot_index = ["Material No.", "Material Description", "Doc. Type Text"]
            str_pivot_values = "So Unit Price"
            sales_register_max_pivot_df = pd.pivot_table(sales_present_quarter_pd, index=list_pivot_index,
                                                         values=str_pivot_values,
                                                         aggfunc=numpy.max,
                                                         margins=False,
                                                         margins_name="Grand Total")
            sales_register_max_pivot_df = sales_register_max_pivot_df.reset_index()
            print("Same Material Domestic max Pivot table is Created")
            logging.info("Same Material Domestic max Pivot table is Created")
        except Exception as create_pivot_table:
            str_subject = dict_in_config["subject_max_pivot_table"]
            str_body = dict_in_config["body_max_pivot_table"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Same Material Domestic Wise Process-", str(create_pivot_table))
            logging.error("Same Material Domestic max pivot table is not created")
            raise create_pivot_table
        try:
            list_pivot_index = ["Material No.", "Material Description", "Doc. Type Text"]
            str_pivot_values = "So Unit Price"
            sales_register_min_pivot_df = pd.pivot_table(sales_present_quarter_pd, index=list_pivot_index,
                                                         values=str_pivot_values,
                                                         aggfunc=numpy.min,
                                                         margins=False,
                                                         margins_name="Grand Total")
            sales_register_min_pivot_df = sales_register_min_pivot_df.reset_index()
            print("Same Material Domestic min Pivot table is Created")
            logging.info("Same Material Domestic min Pivot table is Created")
        except Exception as create_pivot_table:
            str_subject = dict_in_config["subject_min_pivot_table"]
            str_body = dict_in_config["body_min_pivot_table"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Same Material Domestic Wise Process-", str(create_pivot_table))
            logging.error("Same Material Domestic min pivot table is not created")
            logging.exception(create_pivot_table)
            raise create_pivot_table

        same_material_domestic_df = pd.merge(sales_register_min_pivot_df, sales_register_max_pivot_df, how="outer",
                                             on=["Material No.", "Material Description", "Doc. Type Text"])

        # comment
        same_material_domestic_df = same_material_domestic_df.replace(numpy.nan, 0, regex=True)
        col_name = same_material_domestic_df.columns.values.tolist()

        # comment
        same_material_domestic_df.drop(
            same_material_domestic_df.index[(
                    same_material_domestic_df[col_name[2]] == 'PLL Credit Memo Req')], inplace=True)

        same_material_domestic_df.drop(
            same_material_domestic_df.index[(
                    same_material_domestic_df[col_name[2]] == 'Export Ordr w/o Duty')],
            inplace=True)

        same_material_domestic_df.drop(
            same_material_domestic_df.index[(
                    same_material_domestic_df[col_name[2]] == 'Export Order')],
            inplace=True)

        same_material_domestic_df.drop(
            same_material_domestic_df.index[(
                    same_material_domestic_df[col_name[2]] == 'Scrap Order')],
            inplace=True)
        same_material_domestic_df.drop(
            same_material_domestic_df.index[(
                    same_material_domestic_df[col_name[2]] == 'Debit Memo Request')],
            inplace=True)
        same_material_domestic_df.drop(
            same_material_domestic_df.index[(
                    same_material_domestic_df[col_name[2]] == 'Returns')],
            inplace=True)
        same_material_domestic_df.drop(
            same_material_domestic_df.index[(
                    same_material_domestic_df[col_name[2]] == 'Asset Sale Order')],
            inplace=True)

        same_material_domestic_df['Variance'] = ''
        pd.options.mode.chained_assignment = None

        # variance formula for index
        for index in same_material_domestic_df.index:
            min_unit_price = same_material_domestic_df[col_name[3]][index]
            max_unit_price = same_material_domestic_df[col_name[4]][index]

            float_variance = max_unit_price - min_unit_price

            same_material_domestic_df['Variance'][index] = float_variance

        same_material_domestic_df['Variance %'] = ''
        col_name = same_material_domestic_df.columns.values.tolist()
        for index in same_material_domestic_df.index:
            float_variance = same_material_domestic_df['Variance'][index]
            min_unit_price = same_material_domestic_df[col_name[3]][index]

            float_variance = float_variance / min_unit_price

            same_material_domestic_df['Variance %'][index] = float_variance

        # col_name = same_material_domestic_df.columns.values.tolist()
        # same_material_domestic_df['Concentration'] = ''
        # pd.options.mode.chained_assignment = None
        # total_variance = same_material_domestic_df[col_name[5]].sum()
        # # print(total_variance)
        # for index in same_material_domestic_df.index:
        #     float_variance = same_material_domestic_df[col_name[5]][index]
        #     if float_variance == 0:
        #         concentration = 0
        #     else:
        #         concentration = float_variance / total_variance
        #
        #     same_material_domestic_df['Concentration'][index] = concentration

        col_name = same_material_domestic_df.columns.values.tolist()
        same_material_domestic_df.sort_values(by=col_name[6], axis=0, ascending=False, inplace=True)

        same_material_domestic_df['Remarks'] = ''
        # print(same_material_domestic_df)
        same_material_domestic_df.reset_index(inplace=True)
        # print(same_material_domestic_df)
        pd.options.mode.chained_assignment = None
        for index, row in same_material_domestic_df.iterrows():
            if index == 10:
                # print(index)
                # print("breaking the loop")
                break
            else:
                same_material_domestic_df.loc[index, 'Remarks'] = 'Major'
                # print(row['Remarks'])
                # print(index)
        same_material_domestic_df = same_material_domestic_df.drop(columns=["index"])
        # Change Column names
        same_material_domestic_df = same_material_domestic_df.rename(columns={col_name[3]: "Min of So Unit Price"})
        same_material_domestic_df = same_material_domestic_df.rename(columns={col_name[4]: "Max of So Unit Price"})
        # print(same_material_domestic_df)
        try:
            # Log Sheet
            with pd.ExcelWriter(dict_main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                same_material_domestic_df.to_excel(writer,
                                                   sheet_name=dict_main_config["Output_SameMaterialDomestic_sheetname"],
                                                   index=False, startrow=2)
            print("Same Material Domestic sheet Out file is saved")
        except Exception as save_output_file:
            str_subject = dict_in_config["subject_save_output_file"]
            str_body = dict_in_config["body_save_output_file"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Same Material Domestic Wise Process-", str(save_output_file))
            logging.critical("Same Material Domestic sheet Out file is not saved")
            return save_output_file

        # Load Sheet in openpyxl
        try:
            workbook = openpyxl.load_workbook(dict_main_config["Output_File_Path"])
            print("Same Material Domestic Work Book is loaded")
            logging.info("Same Material Domestic Work Book is loaded")
        except Exception as load_work_book:
            str_subject = dict_in_config["load_work_book_subject"]
            str_body = dict_in_config["load_work_book_body"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Same Material Domestic Wise Process-", str(load_work_book))
            logging.critical("Same Material Domestic work book is not loaded")
            return load_work_book
        try:
            worksheet = workbook[dict_main_config["Output_SameMaterialDomestic_sheetname"]]
            print("Same Material Domestic Work Sheet is loaded")
            logging.info("Same Material Domestic Work Sheet is loaded")
        except Exception as load_work_sheet:
            str_subject = dict_in_config["load_work_sheet_subject"]
            str_body = dict_in_config["load_work_sheet_body"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Same Material Domestic Wise Process-", str(load_work_sheet))
            logging.critical("Same Material Domestic work sheet is not loaded")
            return load_work_sheet
        for cell in worksheet["G"]:
            cell.number_format = "0%"
        # for cell in worksheet["B"]:
        #    cell.number_format = "#,##"

        full_range = "A3:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
        worksheet.auto_filter.ref = full_range
        cambria_11_bold_black = Font(name="Cambria", size=11, bold=True, color="000000")
        for c in ascii_uppercase:
            worksheet[c + "3"].font = cambria_11_bold_black
        # for c in ascii_uppercase:
        # worksheet[c + str(worksheet.max_row)].font = font_style
        solid_light_blue_fill = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            worksheet[c + "3"].fill = solid_light_blue_fill
            if c == "H":
                break

        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in worksheet[c])
            worksheet.column_dimensions[c].width = column_length * 1.25
            if c == 'H':
                break
        worksheet['D2'] = '=SUBTOTAL(9,D4:D' + str(worksheet.max_row) + ')'
        worksheet['E2'] = '=SUBTOTAL(9,E4:E' + str(worksheet.max_row) + ')'
        worksheet['F2'] = '=SUBTOTAL(9,F4:F' + str(worksheet.max_row) + ')'
        # Save File
        try:
            print(workbook.sheetnames)
            workbook.save(dict_main_config["Output_File_Path"])
            print("Same Material Domestic Work Sheet file is closed")
            logging.info("Same Material Domestic Work Sheet file is Closed")
        except Exception as close_file_exception:
            str_subject = dict_in_config["close_work_sheet_file_subject"]
            str_body = dict_in_config["close_work_sheet_file_body"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Same Material Domestic Process-", str(close_file_exception))
            logging.error("Same Material Domestic work sheet file is not closed")
            logging.exception(close_file_exception)
            return close_file_exception
        logging.info("Completed Same Material Domestic code execution")
        return same_material_domestic

    except PermissionError as file_error:
        str_subject = dict_in_config["Permission_Error_Subject"]
        str_body = dict_in_config["Permission_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Same Material Domestic Process-", str(file_error))
        print("Please close the file")
        logging.exception(file_error)
        return file_error
    except FileNotFoundError as notfound_error:
        str_subject = dict_in_config["FileNotFound_Subject"]
        str_body = dict_in_config["FileNotFound_Body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Same Material Domestic Process-", str(notfound_error))
        logging.exception(notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Same Material Domestic Process-", str(business_error))
        logging.exception(business_error)
        return business_error
    except ValueError as value_error:
        str_subject = dict_in_config["Value_Error"]
        str_body = dict_in_config["Value_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Same Material Domestic Process-", str(value_error))
        logging.exception(value_error)
        return value_error
    except TypeError as type_error:
        str_subject = dict_in_config["Type_Error"]
        str_body = dict_in_config["Type_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Same Material Domestic Process-", str(type_error))
        logging.exception(type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        str_subject = dict_in_config["SystemError_Subject"]
        str_body = dict_in_config["SystemError_Body"]
        str_body = str_body.replace("SystemError +", str(error))
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Same Material Domestic Process-", str(error))
        logging.exception(error)
        return error
    except KeyError as key_error:
        str_subject = dict_in_config["Name_Error"]
        str_body = dict_in_config["Name_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Same Material Domestic Process-", str(key_error))
        logging.exception(key_error)
        return key_error
    except NameError as nameError:
        str_subject = dict_in_config["Key_Error"]
        str_body = dict_in_config["Key_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Same Material Domestic Process-", str(nameError))
        logging.exception(nameError)
        return nameError
    except AttributeError as attributeError:
        str_subject = dict_in_config["Attribute_Error"]
        str_body = dict_in_config["Attribute_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Same Material Domestic Process-", str(attributeError))
        logging.exception(attributeError)
        return attributeError


if __name__ == "__main__":
    pass
