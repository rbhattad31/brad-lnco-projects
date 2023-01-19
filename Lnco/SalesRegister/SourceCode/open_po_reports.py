import logging
from string import ascii_lowercase
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Side, Border


class open_po_Exception(Exception):
    pass


def open_po_reports(open_po_df, main_config):
    # print(open_po_df)
    print(list(open_po_df.columns.values.tolist()))

    try:
        open_po_df["Order Date"] = pd.to_datetime(open_po_df["Order Date"], errors='coerce')
        open_po_df["PO Date"] = pd.to_datetime(open_po_df["PO Date"], errors='coerce')
    except Exception as changing_date_format_exception:
        exception_message = changing_date_format_exception
        logging.error("Exception occurred while changing the date format")
        raise open_po_Exception(exception_message)
    try:
        recent_audit_date = open_po_df["Order Date"].max()  # error
        # print(recent_audit_date)
    except Exception as recent_audit_date_exception:
        exception_message = recent_audit_date_exception
        logging.error("Exception occurred while recent audit date exception the date format")
        raise open_po_Exception(exception_message)
    open_po_df["no of days"] = 0
    output_file_path = main_config['Output_File_path']
    print(output_file_path)
    output_sheet_name = main_config['Output_Open_PO_sheetname']
    print(output_sheet_name)

    for index in open_po_df.index:
        po_date = open_po_df.loc[index]["PO Date"]
        no_of_days = recent_audit_date - po_date
        # print(no_of_days)
        open_po_df.at[index, 'no of days'] = no_of_days

    try:
        open_po_df["PO Date"] = pd.to_datetime(open_po_df["PO Date"]).dt.strftime("%d-%m-%Y")
        open_po_df["Order Date"] = pd.to_datetime(open_po_df["Order Date"]).dt.strftime("%d-%m-%Y")
        # print(open_po_df)
    except Exception as changing_date_format_exception:
        exception_message = changing_date_format_exception
        logging.error("Below exception occurred while changing the date format")
        raise open_po_Exception(exception_message)

    try:
        # print(open_po_df)
        print(open_po_df.columns.values.tolist())
        # open_po_df.to_excel(output_file_path, sheet_name=output_sheet_name, startrow=1, index=False)
        with pd.ExcelWriter(output_file_path, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            open_po_df.to_excel(writer, sheet_name=output_sheet_name, index=False)
    except Exception as file_saving_exception:
        exception_message = file_saving_exception
        logging.error("Below exception occurred while saving the file ")
        raise open_po_Exception(exception_message)

#   load in openpyxl
    workbook = openpyxl.load_workbook(output_file_path)
    worksheet = workbook[output_sheet_name]
    worksheet['P1'] = recent_audit_date  # date format change
    print(recent_audit_date)
    # value = datetime.datetime.strptime(recent_audit_date, "%Y-%m-%d %H:%M:%S")
    # cell = worksheet['P1']
    # cell.value = value
    # cell.number_format = '%d-%m-%Y'
    # Format Header
    calibri_11_black_bold_font = Font(name="Calibri", size=11, color="000000", bold=True)
    # print(ascii_lowercase)
    for c in ascii_lowercase:
        # print(c)
        worksheet[c + "2"].font = calibri_11_black_bold_font

    # Header Fill
    light_blue_fill = PatternFill(patternType='solid', fgColor='87ceeb')
    for c in ascii_lowercase:
        worksheet[c + "2"].fill = light_blue_fill
        if c == 'p':
            break

    thin = Side(border_style="thin", color='000000')
    for row in worksheet.iter_rows(min_row=3, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Set Width
    # for c in ascii_lowercase:
        # worksheet.column_dimensions[c].width = 30

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                print(e)

            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
    print(workbook.sheetnames)
    workbook.save(output_file_path)
    return output_file_path


if __name__ == '__main__':
    pass
