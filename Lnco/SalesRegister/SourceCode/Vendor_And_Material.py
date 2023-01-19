import logging
from msilib.schema import Font
from string import ascii_uppercase
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from ReusableTasks.send_mail_reusable_task import send_mail
import numpy
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

# import warnings

# warnings.filterwarnings("ignore", category=RuntimeWarning)
# warnings.simplefilter(action='ignore', category=FutureWarning)


class BusinessException(Exception):
    pass


def Vendor_And_Material(main_config, in_config, present_quarter_pd, previous_quarter_path):
    try:

        sales_register_data = present_quarter_pd
        # print(Excel_data)
        # Fetch To Address
        to_address = main_config["To_Mail_Address"]
        cc_address = main_config["CC_Mail_Address"]

        # Check Exception
        if sales_register_data.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        sales_register_col = sales_register_data.columns.values.tolist()
        for col in ["Payer", "Payer Name", "Material Description", "Billing Qty.", "Base Price in INR"]:
            if col not in sales_register_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        Payer = sales_register_data[sales_register_data['Payer'].notna()]
        Payer_Name = sales_register_data[sales_register_data['Payer Name'].notna()]
        Material_Dec = sales_register_data[sales_register_data['Material Description'].notna()]
        quantity = sales_register_data[sales_register_data['Billing Qty.'].notna()]
        price_q1 = sales_register_data[sales_register_data['Base Price in INR'].notna()]

        if len(Payer) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Payer_q1_subject"],
                      body=in_config["Payer_q1_Body"])
            logging.error("Payer q1 Column is empty")
            raise BusinessException("Payer q1 Column is empty")

        elif len(Payer_Name) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Payer_Name_q1_subject"],
                      body=in_config["Payer_Name_q1_Body"])
            logging.error("Payer Name q1 Column is empty")
            raise BusinessException("Payer Name q1 Column is empty")

        elif len(Material_Dec) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Material_Dec_q1_Subject"],
                      body=in_config["Material_Dec_q1_Body"])
            logging.error("Material Description q1 Column is empty")
            raise BusinessException("Material Description q1 Column is empty")

        elif len(quantity) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Billing_Qty_q1_Subject"],
                      body=in_config["Billing_Qty_q1_Body"])
            logging.error("Billing Qty q1 Column is empty")
            raise BusinessException("Billing Qty q1 Column is empty")
        elif len(price_q1) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["price_INR_q1_Subject"],
                      body=in_config["price_INR_q1_Body"])
            logging.error("Base price in INR q1 Column is empty")
            raise BusinessException("Base Price in INR q1 Column is empty")

        else:
            pass

        try:
            pivot_Q1 = pd.pivot_table(sales_register_data, index=["Payer", "Payer Name", "Material Description"],
                                      values=["Billing Qty.", "Base Price in INR"], aggfunc=numpy.sum, margins=False,
                                      margins_name="Grand Total", sort=True)
            print("vendor and material Comparison q1 Pivot table is Created")
            logging.info("vendor and material Comparison q1 Pivot table is Created")
        except Exception as create_pivot_table:
            subject = in_config["subject_pivot_q1_table"]
            body = in_config["body_pivot_q1_table"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("vendor and material Comparison Wise Process-", str(create_pivot_table))
            logging.info("vendor and material Comparison q1 pivot table is not created")
            raise create_pivot_table

        # print(pivot_Q1)
        # Drop last column of a dataframe
        # pivot_Q1 = pivot_Q1[:-1]
        pivot_Q1 = pivot_Q1.reset_index()
        pivot_Q1[["Payer"]] = pivot_Q1[["Payer"]].fillna('').astype(int, errors='ignore')
        columns = pivot_Q1.columns.values.tolist()
        pivot_Q1 = pivot_Q1.rename(columns={columns[1]: "Payer Name"})
        # print(pivot_Q1)
        pivot_Q1 = pivot_Q1.replace(numpy.nan, 0, regex=True)
        columns = pivot_Q1.columns.values.tolist()
        # numpy.seterr(divide='ignore')
        pivot_Q1['Unit Price'] = ""
        pd.options.mode.chained_assignment = None

        for index in pivot_Q1.index:
            quantity = pivot_Q1[columns[3]][index]
            # print(GR_amt)
            qty_q4 = pivot_Q1[columns[4]][index]
            # print(GR_qty)
            if qty_q4 == 0:
                Unit_price = 0
            else:
                Unit_price = quantity / qty_q4
            # print(Unit_price)
            pivot_Q1['Unit Price'][index] = Unit_price

            # print(Unit_price)

        columns = pivot_Q1.columns.values.tolist()
        pivot_Q1 = pivot_Q1.rename(
            columns={columns[4]: "Billing Qty.1"})
        pivot_Q1 = pivot_Q1.rename(
            columns={columns[3]: "Base Price in INR1"})
        pivot_Q1 = pivot_Q1.rename(
            columns={columns[5]: "Unit Price1"})

        # pivot_Q4['Concat'] = ""
        # pivot_Q4["Concat"] = pivot_Q4["Material No."].astype(str) + pivot_Q4["Valuation Class Text"].astype(str) + \
        # pivot_Q4["Vendor Name"].astype(str)
        pivot_Q1 = pivot_Q1[
            ["Payer", "Payer Name", "Material Description", "Billing Qty.1", "Base Price in INR1", "Unit Price1"]]
        # print(pivot_Q1)

        # Q3 Pivot Excel_data = pd.read_excel(main_config["InputFilePath"], sheet_name=main_config[
        # "PreviousQuarterSheetName"], skiprows=in_config["Skiprow_Q3"])

        sales_register_data = pd.read_excel(previous_quarter_path)
        # print(Excel_data)
        sales_register_data = sales_register_data.loc[:, ~sales_register_data.columns.duplicated(keep='first')]
        columns = sales_register_data.columns
        if main_config["sales_register_1st_column_name"] in columns and \
                main_config["sales_register_2nd_column_name"] in columns:
            print("Previous quarter q4 - The data is starting from first row only")
            pass

        else:
            print("Previous quarter q4 - The data is not starting from first row ")
            for index, row in sales_register_data.iterrows():
                if row[0] != main_config["sales_register_1st_column_name"]:
                    sales_register_data.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = sales_register_data.iloc[0]
            sales_register_data = sales_register_data[1:]
            sales_register_data.columns = new_header
            sales_register_data.reset_index(drop=True, inplace=True)
            sales_register_data.columns.name = None
        sales_register_data = sales_register_data.loc[:, ~sales_register_data.columns.duplicated(keep='first')]
        # Check Exception
        if sales_register_data.shape[0] == 0:
            subject = in_config["EmptyInput_Subject1"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        Q3Sheet_col = sales_register_data.columns.values.tolist()
        for col in ["Payer", "Payer Name", "Material Description", "Billing Qty.", "Base Price in INR"]:
            if col not in Q3Sheet_col:
                subject = in_config["ColumnMiss_Subject1"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        payer_q4 = sales_register_data[sales_register_data['Payer'].notna()]
        payer_Name_q4 = sales_register_data[sales_register_data['Payer Name'].notna()]
        material_dec_q4 = sales_register_data[sales_register_data['Material Description'].notna()]
        qty_q4 = sales_register_data[sales_register_data['Billing Qty.'].notna()]
        price_q4 = sales_register_data[sales_register_data['Base Price in INR'].notna()]

        if len(payer_q4) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Payer_q4_subject"],
                      body=in_config["Payer_q4_Body"])
            logging.error("Payer q4 Column is empty")
            raise BusinessException("Payer q4 Column is empty")

        elif len(payer_Name_q4) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Payer_Name_q4_subject"],
                      body=in_config["Payer_Name_q4_Body"])
            logging.error("Payer Name q4 Column is empty")
            raise BusinessException("Payer Name q4 Column is empty")

        elif len(material_dec_q4) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Material_Dec_q4_Subject"],
                      body=in_config["Material_Dec_q4_Body"])
            logging.error("Material Description q4 Column is empty")
            raise BusinessException("Material Description q4 Column is empty")

        elif len(qty_q4) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Billing_Qty_q4_Subject"],
                      body=in_config["Billing_Qty_q4_Body"])
            logging.error("Billing Qty q4 Column is empty")
            raise BusinessException("Billing Qty q4 Column is empty")
        elif len(price_q4) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["price_INR_q4_Subject"],
                      body=in_config["price_INR_q4_Body"])
            logging.error("Base price in INR q4 Column is empty")
            raise BusinessException("Base Price in q4 INR Column is empty")

        else:
            pass

        sales_register_data = sales_register_data[["Payer", "Payer Name", "Material Description", "Billing Qty.", "Base Price in INR"]]
        # print(Excel_data)
        try:
            pivot_Q4 = pd.pivot_table(sales_register_data, index=["Payer", "Payer Name", "Material Description"],
                                      values=["Billing Qty.", "Base Price in INR"], aggfunc=numpy.sum,
                                      margins=False,
                                      margins_name="Grand Total", sort=True)
            print("vendor and material Comparison previous quarter Pivot table is Created")
            logging.info("vendor and material Comparison previous quarter Pivot table is Created")
        except Exception as create_pivot_table:
            subject = in_config["subject_pivot_q4_table"]
            body = in_config["body_pivot_q4_table"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("vendor and material Comparison previous quarter Wise Process-", str(create_pivot_table))
            logging.error("vendor and material Comparison previous quarter pivot table is not created")
            raise create_pivot_table

        pivot_Q4 = pivot_Q4.reset_index()

        pivot_Q4 = pivot_Q4.replace(numpy.nan, 0, regex=True)

        # Drop last column of a dataframe
        # pivot_Q4 = pivot_Q4[:-1]
        columns = pivot_Q4.columns.values.tolist()

        pivot_Q4['Unit Price'] = ""
        pd.options.mode.chained_assignment = None

        for index in pivot_Q4.index:
            quantity = pivot_Q4[columns[3]][index]
            qty_q4 = pivot_Q4[columns[4]][index]
            if qty_q4 == 0:
                Unit_price = 0
            else:
                Unit_price = quantity / qty_q4
            # print(Unit_price)
            pivot_Q4['Unit Price'][index] = Unit_price

        columns = pivot_Q4.columns.values.tolist()
        pivot_Q4 = pivot_Q4.rename(
            columns={columns[4]: "Billing Qty.2"})
        pivot_Q4 = pivot_Q4.rename(
            columns={columns[3]: "Base Price in INR2"})
        pivot_Q4 = pivot_Q4.rename(
            columns={columns[5]: "Unit Price2"})
        # print(columns)
        pivot_Q4 = pivot_Q4[
            ["Payer", "Payer Name", "Material Description", "Billing Qty.2", "Base Price in INR2", "Unit Price2"]]
        # print(pivot_Q4)
        vendor_and_material_Q1 = pd.merge(pivot_Q1, pivot_Q4, how="left",
                                          on=["Payer", "Payer Name", "Material Description"], copy=False)
        vendor_and_material_Q1 = vendor_and_material_Q1[
            ["Payer", "Payer Name", "Material Description", "Billing Qty.1",
             "Base Price in INR1", "Unit Price1"]]
        vendor_and_material_Q4 = pd.merge(pivot_Q4, pivot_Q1, how="left",
                                          on=["Payer", "Payer Name", "Material Description"], copy=False)
        vendor_and_material_Q4 = vendor_and_material_Q4[
            ["Payer", "Payer Name", "Material Description", "Billing Qty.2", "Base Price in INR2", "Unit Price2"]]
        # columns = Unit_Price_Q3.columns.values.tolist()

        vendor_and_material = pd.merge(vendor_and_material_Q1, vendor_and_material_Q4, how="outer",
                                       on=["Payer", "Payer Name", "Material Description"], copy=False)
        # Unit_Price = pd.concat([Unit_Price_Q4, Unit_Price_Q3], ignore_index=True, sort=True)
        vendor_and_material = vendor_and_material[
            ["Payer", "Payer Name", "Material Description", "Billing Qty.1", "Base Price in INR1",
             "Unit Price1", "Billing Qty.2", "Base Price in INR2", "Unit Price2"]]
        # print(vendor_and_material)
        vendor_and_material = vendor_and_material.reset_index()
        # print(vendor_and_material)
        vendor_and_material = vendor_and_material.replace(numpy.nan, 0, regex=True)

        columns = vendor_and_material.columns.values.tolist()
        vendor_and_material = vendor_and_material.drop(columns=["index"])
        vendor_and_material.sort_values(by=columns[1], axis=0, ascending=True, inplace=False)

        vendor_and_material['Nature of Product'] = ""
        pd.options.mode.chained_assignment = None
        for index in vendor_and_material.index:
            if (vendor_and_material[columns[4]][index] == 0) & (vendor_and_material[columns[7]][index] != 0):
                vendor_and_material['Nature of Product'][index] = "No values in Q1"
            elif (vendor_and_material[columns[4]][index] != 0) & (vendor_and_material[columns[7]][index] == 0):
                vendor_and_material['Nature of Product'][index] = "New product"
            elif (vendor_and_material[columns[4]][index] != 0) & (vendor_and_material[columns[7]][index] != 0):
                vendor_and_material['Nature of Product'][index] = "Common product"
            elif (vendor_and_material[columns[4]][index] == 0) & (vendor_and_material[columns[7]][index] == 0):
                vendor_and_material['Nature of Product'][index] = "No data in Q1 & Q4"
            else:
                pass

        vendor_and_material['Amount1'] = ""
        pd.options.mode.chained_assignment = None

        for index in vendor_and_material.index:
            quantity = vendor_and_material[columns[4]][index]
            # print(qty_q1)
            qty_q4 = vendor_and_material[columns[7]][index]
            # print(qty_q4)
            unit_price_q4 = vendor_and_material[columns[9]][index]
            # print(unit_price_q4)
            Amount = (+(quantity - qty_q4)) * unit_price_q4
            # print(Amount)
            vendor_and_material['Amount1'][index] = Amount

        columns = vendor_and_material.columns.values.tolist()
        vendor_and_material['%1'] = ""
        pd.options.mode.chained_assignment = None
        for index in vendor_and_material.index:
            qty_amount = vendor_and_material[columns[10]][index]
            # print(qty_amount)
            price_INR_q4 = vendor_and_material[columns[7]][index]
            # print(price_INR_q4)
            if price_INR_q4 == 0:
                percentage = 0

            else:
                percentage = (qty_amount / price_INR_q4)
            vendor_and_material['%1'][index] = percentage

        vendor_and_material['Amount'] = ""
        pd.options.mode.chained_assignment = None
        for index in vendor_and_material.index:
            unit_price_q1 = vendor_and_material[columns[5]][index]

            # print(unit_price_q1)
            unit_price_q4 = vendor_and_material[columns[8]][index]
            # print(unit_price_q4)
            quantity = vendor_and_material[columns[3]][index]
            # print(qty_q1)

            Amount = (+(unit_price_q1 - unit_price_q4)) * quantity
            # print(Amount)
            vendor_and_material['Amount'][index] = Amount

        columns = vendor_and_material.columns.values.tolist()
        vendor_and_material['%'] = ""
        pd.options.mode.chained_assignment = None
        for index in vendor_and_material.index:
            amount_q1 = vendor_and_material[columns[12]][index]
            # print(amount_q1)
            price_INR_q4 = vendor_and_material[columns[7]][index]
            # print(price_INR_q4)
            if price_INR_q4 == 0:
                percentage = 0
            else:
                percentage = amount_q1 / price_INR_q4
            vendor_and_material['%'][index] = percentage

        # print(vendor_and_material)

        #  Rename Columns

        # print(vendor_and_material)
        vendor_and_material = vendor_and_material.rename(
            columns={columns[3]: "Sum of Billing Qty."})
        # print(vendor_and_material)
        vendor_and_material = vendor_and_material.rename(
            columns={columns[4]: "Sum of Base Price in INR"})
        vendor_and_material = vendor_and_material.rename(
            columns={columns[5]: "Unit Price"})
        vendor_and_material = vendor_and_material.rename(
            columns={columns[6]: "Sum of Billing Qty."})
        # print(vendor_and_material)
        vendor_and_material = vendor_and_material.rename(
            columns={columns[7]: "Sum of Base Price in INR"})
        vendor_and_material = vendor_and_material.rename(
            columns={columns[8]: "Unit Price"})
        vendor_and_material = vendor_and_material.rename(
            columns={columns[10]: "Amount"})
        vendor_and_material = vendor_and_material.rename(
            columns={columns[11]: "%"})

        # print(vendor_and_material)
        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                vendor_and_material.to_excel(writer,
                                             sheet_name=main_config["Output_VendorAndMaterial_Comparison_sheetname"],
                                             startrow=2, index=False)
            print("vendor and material Comparison sheet Out file is saved")
        except Exception as save_output_file:
            subject = in_config["subject_save_output_file"]
            body = in_config["body_save_output_file"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("vendor and material Comparison Wise Process-", str(save_output_file))
            logging.info("vendor and material Comparison sheet Out file is not saved")
            return save_output_file
            # Load Sheet in openpyxl
        wb = load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_VendorAndMaterial_Comparison_sheetname"]]

        cell = ws['D2']
        cell.value = 'Current Quarter Q1'
        ws.merge_cells('D2:F2')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = ws['G2']
        cell.value = 'Previous Quarter Q4'
        ws.merge_cells('G2:I2')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = ws['K2']
        cell.value = 'Qty variance'
        ws.merge_cells('K2:L2')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = ws['M2']
        cell.value = 'Price variance'
        ws.merge_cells('M2:N2')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cambria_12_bold_black_font = Font(name="Cambria", size=12, bold=True, color="000000")
        ws['D2'].font = cambria_12_bold_black_font
        ws['G2'].font = cambria_12_bold_black_font
        ws['K2'].font = cambria_12_bold_black_font
        ws['M2'].font = cambria_12_bold_black_font
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        ws['D2'].fill = fill_pattern
        ws['G2'].fill = fill_pattern

        fill_pattern = PatternFill(patternType="solid", fgColor="FFFF00")
        ws['K2'].fill = fill_pattern
        ws['M2'].fill = fill_pattern

        for cell in ws["D"]:
            cell.number_format = "#,##"
        for cell in ws["E"]:
            cell.number_format = "#,##"
        for cell in ws["F"]:
            cell.number_format = "#,##"
        for cell in ws["G"]:
            cell.number_format = "#,##"
        for cell in ws["H"]:
            cell.number_format = "#,##"
        for cell in ws["I"]:
            cell.number_format = "#,##"
        for cell in ws["K"]:
            cell.number_format = "#,##"
        for cell in ws["L"]:
            cell.number_format = "0%"
        for cell in ws["M"]:
            cell.number_format = "#,##"
        for cell in ws["N"]:
            cell.number_format = "0%"

        Full_range = "A3:" + get_column_letter(ws.max_column) + str(ws.max_row)
        ws.auto_filter.ref = Full_range
        cambria_12_bold_black_font = Font(name="Cambria", size=11, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "3"].font = cambria_12_bold_black_font
            # for c in ascii_uppercase:
            # ws[c + str(ws.max_row)].font = font_style
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "3"].fill = fill_pattern
            if c == "N":
                break

        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 'N':
                break
        ws['D1'] = '=SUBTOTAL(9,D4:D' + str(ws.max_row) + ')'
        ws['E1'] = '=SUBTOTAL(9,E4:E' + str(ws.max_row) + ')'
        ws['G1'] = '=SUBTOTAL(9,G4:G' + str(ws.max_row) + ')'
        ws['H1'] = '=SUBTOTAL(9,H4:H' + str(ws.max_row) + ')'
        ws['K1'] = '=SUBTOTAL(9,K4:K' + str(ws.max_row) + ')'
        ws['M1'] = '=SUBTOTAL(9,M4:M' + str(ws.max_row) + ')'
        # ws['L1'] = ws['K1'] / ws['H1']
        ws['L1'] = '=K1/H1'
        ws['N1'] = "=M1/H1"
        # sheet['D3'] = '=C4/3'
        # ws['L1'] = '=SUBTOTAL(9,k1:H1' + str(ws.max_column) + ')'
        # ws['N1'] = '=SUBTOTAL(9,N4:N' + str(ws.max_row) + ')'

        # Save File
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])
        logging.info("Completed Vendor and Material wise Comparison code execution")
        return Vendor_And_Material

    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Please close the file")
        logging.exception(PermissionError)
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("vendor and material Comparison Process-", notfound_error)
        logging.exception(FileNotFoundError)
        return notfound_error
    except BusinessException as business_error:
        print("vendor and material Comparison Process-", business_error)
        logging.exception(business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Vendor and Material Wise Comparison Process-", value_error)
        logging.exception(value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("vendor and material wise Comparison Process-", type_error)
        logging.exception(type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Vendor and Material wise Comparison Process-", error)
        logging.exception(error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Vendor and Material wise Comparison Process-", key_error)
        logging.exception(key_error)
        return key_error


if __name__ == "__main__":
    pass
