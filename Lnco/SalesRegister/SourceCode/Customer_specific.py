from string import ascii_lowercase
from openpyxl.styles import Font, PatternFill, Side, Border
import openpyxl
import logging
import pandas as pd


class customer_specific_exception(Exception):
    pass


def customer_specific(sales_register_df, dict_config_main):
    try:
        pd.options.mode.chained_assignment = None
        sales_register_filtered_df = sales_register_df[["Material No.", "Material Description", "Payer Name"]]
        # print(sales_register_filtered_df)
    except Exception as filtered_column_name_error:
        str_exception_message = filtered_column_name_error
        logging.error("Exception occurred while specified filtered was not found in the input list")
        raise customer_specific_exception(str_exception_message)

    try:
        # # Keep only duplicate rows
        sales_register_filtered_df.sort_values("Material Description", inplace=True)

        # drop all the entries which are having unique material description
        duplicate_series = sales_register_filtered_df['Material Description'].duplicated(keep=False)
        sr_material_description_duplicates_pd = sales_register_filtered_df[duplicate_series]

        # drop all the duplicates of 3 columns
        sr_material_description_duplicates_pd["Payer Name"] = sr_material_description_duplicates_pd["Payer Name"].str.title()
        sr_material_description_duplicates_pd = sr_material_description_duplicates_pd.drop_duplicates(
            subset=["Material No.", "Material Description", "Payer Name"], keep='first')

        # # Keep the entries which are having same material number and material description
        # Get unique values in materials description
        list_material_description_values = sr_material_description_duplicates_pd['Material Description'].tolist()
        list_material_description_unique_values = list(set(list_material_description_values))

        # for each material description, get the count of unique values in material No. Column
        # if the count is not equal to one, delete all entries with the material description
        for material_description in list_material_description_unique_values:
            material_description_filtered_pd = sr_material_description_duplicates_pd[sr_material_description_duplicates_pd['Material Description'] == material_description]
            list_material_number_values = material_description_filtered_pd['Material No.'].tolist()
            list_material_number_unique_values = list(set(list_material_number_values))
            # delete all entries of material description if rows are having multiple material numbers or single entry
            if len(list_material_number_unique_values) != 1 or len(material_description_filtered_pd.index) == 1:
                sr_material_description_duplicates_pd = sr_material_description_duplicates_pd[sr_material_description_duplicates_pd['Material Description'] != material_description]
                continue

        # sr_duplicates_filtered_df = sales_register_filtered_df.drop_duplicates(keep='first')
        # duplicate = sr_duplicates_filtered_df[
        #     sr_duplicates_filtered_df.duplicated(subset=["Material Description"], keep=False)]
        # # print(duplicate)
    except KeyError:
        str_exception_message = "material Description column was not found"
        logging.error("Exception occurred while specified Material description was not found in the input list")
        raise customer_specific_exception(str_exception_message)
    try:
        output_file_path = dict_config_main["Output_File_Path"]
        output_sheet_name = dict_config_main["Output_Customer_Specific_sheetname"]

        # duplicate.to_excel(output_file_path, sheet_name=output_sheet_name, index=False)
        with pd.ExcelWriter(output_file_path, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            sr_material_description_duplicates_pd.to_excel(writer, sheet_name=output_sheet_name, index=False)

        workbook = openpyxl.load_workbook(output_file_path)
        worksheet = workbook[output_sheet_name]

        format_font = Font(name="Calibri", size=11, color="000000", bold=True)
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')

        for c in ascii_lowercase:
            worksheet[c + "1"].fill = format_fill
            worksheet[c + "1"].font = format_font
            if c == 'c':
                break

        thin = Side(border_style="thin", color='000000')
        for row in worksheet.iter_rows(min_row=1, min_col=1, max_row=worksheet.max_row, max_col=3):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for c in ascii_lowercase:
            worksheet.column_dimensions[c].width = 45
        print(workbook.sheetnames)
        workbook.save(output_file_path)
        print(output_file_path)
        return output_file_path

    except Exception as exception:
        print(exception)
        pass


if __name__ == '__main__':
    pass
