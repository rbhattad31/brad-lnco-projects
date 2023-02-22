import logging
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from string import ascii_uppercase
from openpyxl.utils import get_column_letter
import os
from Lnco.ReusableTasks.send_mail_reusable_task import send_mail


class BusinessException(Exception):
    pass


def product_mix_comparison(main_config, in_config, present_quarter_pd, previous_quarter_pd):
    try:
        logging.info("Starting product mix comparison code execution")
        # Read Purchase Register Sheets
        read_present_quarter_pd = present_quarter_pd
        read_previous_quarter_pd = previous_quarter_pd
        # Fetch To Address
        to_address = main_config["To_Mail_Address"]
        cc_address = main_config["CC_Mail_Address"]

        # Check Exception
        if read_present_quarter_pd.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Empty present quarter Sales Register found")
            raise BusinessException("Sheet is empty")

        if read_previous_quarter_pd.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Empty previous quarter Sales Register found")
            raise BusinessException("Sheet is empty")

        # Check Column exist for Present quarter
        present_quarter_columns = read_present_quarter_pd.columns.values.tolist()
        for col in ["Material Type Descri", "Billing Qty.", "Base Price in INR"]:
            if col not in present_quarter_columns:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                logging.error("{} Column is missing".format(col))
                raise BusinessException(col + " Column is missing")

        # Check Column exist for Previous quarter
        previous_quarter_columns = read_previous_quarter_pd.columns.values.tolist()
        for col in ["Material Type Descri", "Billing Qty.", "Base Price in INR"]:
            if col not in previous_quarter_columns:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                logging.error("{} Column is missing".format(col))
                raise BusinessException(col + " Column is missing")

        # Filter rows
        present_quarter_product_type_descp = read_present_quarter_pd[
            read_present_quarter_pd['Material Type Descri'].notna()]
        present_quarter_billing_qty = read_present_quarter_pd[read_present_quarter_pd['Billing Qty.'].notna()]
        present_quarter_price_inr = read_present_quarter_pd[read_present_quarter_pd['Base Price in INR'].notna()]

        if len(present_quarter_product_type_descp) == 0:
            subject = in_config["Product_type_descp_Subject"]
            body = in_config["Product_type_descp_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Product type descp Column is empty")
            raise BusinessException("Product type descp Column is empty")
        elif len(present_quarter_billing_qty) == 0:
            subject = in_config["Billing_Qty_Subject"]
            body = in_config["Billing_Qty_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Billing Qty Column is empty")
            raise BusinessException("Billing Qty Column is empty")
        elif len(present_quarter_price_inr) == 0:
            subject = in_config["Base_Price_INR_Subject"]
            body = in_config["Base_Price_INR_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Base Price in INR Column is empty")
            raise BusinessException("Base Price in INR Column is empty")
        else:
            pass

        # Filter rows
        previous_quarter_product_type_descp = read_present_quarter_pd[
            read_present_quarter_pd['Material Type Descri'].notna()]
        previous_quarter_billing_qty = read_present_quarter_pd[read_present_quarter_pd['Billing Qty.'].notna()]
        previous_quarter_price_inr = read_present_quarter_pd[read_present_quarter_pd['Base Price in INR'].notna()]

        if len(previous_quarter_product_type_descp) == 0:
            subject = in_config["Product_type_descp_Subject"]
            body = in_config["Product_type_descp_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Product type descp Column is empty")
            raise BusinessException("Product type descp Column is empty")
        elif len(previous_quarter_billing_qty) == 0:
            subject = in_config["Billing_Qty_Subject"]
            body = in_config["Billing_Qty_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Billing Qty Column is empty")
            raise BusinessException("Billing Qty Column is empty")
        elif len(previous_quarter_price_inr) == 0:
            subject = in_config["Base_Price_INR_Subject"]
            body = in_config["Base_Price_INR_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Base Price in INR Column is empty")
            raise BusinessException("Base Price in INR Column is empty")
        else:
            pass

        # Create Pivot Table Sales Register for present quarter
        try:
            read_present_quarter_pd[["Material Type Descri"]] = read_present_quarter_pd[
                ["Material Type Descri"]].fillna('')
            pivot_index = ["Material Type Descri"]
            pivot_values = ["Base Price in INR", "Billing Qty."]
            present_quarter_product_mix_pivot_df = pd.pivot_table(read_present_quarter_pd, index=pivot_index,
                                                                  values=pivot_values,
                                                                  aggfunc=numpy.sum,
                                                                  margins=False,
                                                                  margins_name="Grand Total")

            print(" Product mix Comparison Pivot table is Created for present quarter")
            logging.info("Product mix Comparison Pivot table is Created for present quarter")
        except Exception as create_pivot_table:
            subject = in_config["subject_pivot_table"]
            body = in_config["body_pivot_table"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("Product mix Comparison Wise Process-", str(create_pivot_table))
            logging.info("Product mix Comparison pivot table is not created for present quarter")
            raise create_pivot_table

        # Create Pivot Table Sales Register for previous quarter
        try:
            read_previous_quarter_pd[["Material Type Descri"]] = read_previous_quarter_pd[
                ["Material Type Descri"]].fillna('')
            pivot_index = ["Material Type Descri"]
            pivot_values = ["Base Price in INR", "Billing Qty."]
            previous_quarter_product_mix_pivot_df = pd.pivot_table(read_previous_quarter_pd, index=pivot_index,
                                                                   values=pivot_values,
                                                                   aggfunc=numpy.sum,
                                                                   margins=False,
                                                                   margins_name="Grand Total")

            print(" Product mix Comparison Pivot table is Created for previous quarter")
            logging.info("Product mix Comparison Pivot table is Created for previous quarter")
        except Exception as create_pivot_table:
            subject = in_config["subject_pivot_table"]
            body = in_config["body_pivot_table"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("Product mix Comparison Wise Process-", str(create_pivot_table))
            logging.info("Product mix Comparison pivot table is not created")
            raise create_pivot_table

        # merge two pivots
        # print(present_quarter_product_mix_pivot_df)
        # print(previous_quarter_product_mix_pivot_df)
        product_mix_output_pd = pd.merge(present_quarter_product_mix_pivot_df, previous_quarter_product_mix_pivot_df,
                                         how="outer",
                                         on=["Material Type Descri"])
        # print(product_mix_output_pd)
        product_mix_output_pd.reset_index(inplace=True)
        # print(product_mix_output_pd)
        # Formulas
        product_mix_output_pd['Quantity Variance %'] = 0
        pd.options.mode.chained_assignment = None
        product_mix_output_pd_columns_list = product_mix_output_pd.columns.values.tolist()
        # variance formula implementation using index
        for index in product_mix_output_pd.index:
            present_quarter_quantity = product_mix_output_pd[product_mix_output_pd_columns_list[2]][index]
            previous_quarter_quantity = product_mix_output_pd[product_mix_output_pd_columns_list[4]][index]
            if previous_quarter_quantity == 0:
                variance_percentage = 1
            else:
                variance_percentage = (present_quarter_quantity - previous_quarter_quantity) / previous_quarter_quantity
            product_mix_output_pd['Quantity Variance %'][index] = variance_percentage

        product_mix_output_pd['amount Variance %'] = 0
        product_mix_output_pd_columns_list = product_mix_output_pd.columns.values.tolist()
        # variance formula implementation using index
        for index in product_mix_output_pd.index:
            present_quarter_amount = product_mix_output_pd[product_mix_output_pd_columns_list[1]][index]
            previous_quarter_amount = product_mix_output_pd[product_mix_output_pd_columns_list[3]][index]
            if present_quarter_amount == 0:
                variance_percentage = 1
            else:
                variance_percentage = (present_quarter_amount - previous_quarter_amount) / previous_quarter_amount
            product_mix_output_pd['amount Variance %'][index] = variance_percentage

        # Get Pivot Column Names
        col_name = product_mix_output_pd.columns.values.tolist()

        # Change Column names
        product_mix_output_pd = product_mix_output_pd.rename(columns={col_name[1]: "Sum of Base price in INR"})
        product_mix_output_pd = product_mix_output_pd.rename(columns={col_name[2]: "Sum Of Billing Qty"})
        product_mix_output_pd = product_mix_output_pd.rename(columns={col_name[3]: "Sum of Base price in INR"})
        product_mix_output_pd = product_mix_output_pd.rename(columns={col_name[4]: "Sum Of Billing Qty"})

        # Saving
        try:
            # Log Sheet
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                product_mix_output_pd.to_excel(writer,
                                               sheet_name=main_config["Output_Product_Mix_Comparison_sheetname"],
                                               index=False, startrow=2)
            print("Product mix Comparison sheet Out file is saved")
        except Exception as save_output_file:
            subject = in_config["subject_save_output_file"]
            body = in_config["body_save_output_file"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("Product mix Comparison Wise Process-", str(save_output_file))
            logging.info("Product mix Comparison sheet Out file is not saved")
            return save_output_file

        # Check outfile creation
        if os.path.exists(main_config["Output_File_Path"]):
            print("Product mix Comparison Logged")
            logging.info("Product mix Comparison sheet is created")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.warning("Product mix Comparison sheet is not created")
            raise BusinessException("Output file not generated")
        # Formatting
        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Product_Mix_Comparison_sheetname"]]

        for cell in ws["B"]:
            cell.number_format = "#,##"
        for cell in ws["C"]:
            cell.number_format = "#,##"
        for cell in ws["D"]:
            cell.number_format = "#,##"
        for cell in ws["E"]:
            cell.number_format = "#,##"
        for cell in ws["F"]:
            cell.number_format = "0.0%"
        for cell in ws["G"]:
            cell.number_format = "0.0%"

        full_range = "A3:" + get_column_letter(ws.max_column) + str(ws.max_row)
        ws.auto_filter.ref = full_range
        font_style = Font(name="Cambria", size=11, bold=True, color="000000")
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "3"].font = font_style
            ws[c + "3"].fill = fill_pattern
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == "G":
                break

        ws['B1'] = '=SUBTOTAL(9,B4:B' + str(ws.max_row) + ')'
        ws['C1'] = '=SUBTOTAL(9,C4:C' + str(ws.max_row) + ')'
        ws['D1'] = '=SUBTOTAL(9,B4:B' + str(ws.max_row) + ')'
        ws['E1'] = '=SUBTOTAL(9,C4:C' + str(ws.max_row) + ')'
        ws.merge_cells('B2:C2')
        ws.merge_cells('D2:E2')
        ws['B2'].value = 'Present Quarter'
        ws['D2'].value = 'Previous Quarter'
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D2'].alignment = Alignment(horizontal='center', vertical='center')

        ws['B2'].fill = fill_pattern
        ws['D2'].fill = fill_pattern

        cambria_12_bold_black_font = Font(name="Cambria", size=12, bold=True, color="000000")
        ws['B2'].font = cambria_12_bold_black_font
        ws['D2'].font = cambria_12_bold_black_font
        # Save File
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])
        logging.info("Completed Product mix Comparison code execution")
        return ws

    except PermissionError as file_error:
        subject = in_config["Permission_Error_Subject"]
        body = in_config["Permission_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Product mix Comparison Process-", str(file_error))
        print("Please close the file")
        logging.exception(file_error)
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Product mix Comparison Process-", str(notfound_error))
        logging.exception(notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Product mix Comparison Process-", str(business_error))
        logging.exception(business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["Value_Error"]
        body = in_config["Value_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Product mix Comparison Process-", str(value_error))
        logging.exception(value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["Type_Error"]
        body = in_config["Type_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Product mix Comparison Process-", str(type_error))
        logging.exception(type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Product mix Comparison Process-", str(error))
        logging.exception(error)
        return error
    except KeyError as key_error:
        subject = in_config["Name_Error"]
        body = in_config["Name_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Product mix Comparison Process-", str(key_error))
        logging.exception(key_error)
        return key_error
    except NameError as nameError:
        subject = in_config["Key_Error"]
        body = in_config["Key_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Product mix Comparison Process-", str(nameError))
        logging.exception(nameError)
        return nameError
    except AttributeError as attributeError:
        subject = in_config["Attribute_Error"]
        body = in_config["Attribute_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Product mix Comparison Process-", str(attributeError))
        logging.exception(attributeError)
        return attributeError


if __name__ == "__main__":
    pass
