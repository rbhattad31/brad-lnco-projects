from string import ascii_uppercase
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from ReusableTasks.send_mail_reusable_task import send_mail
import os
import logging


class BusinessException(Exception):
    pass


def customer_wise_concentration(main_config, in_config, present_quarter_pd):
    try:
        logging.info("Starting Customer wise concentration code execution")
        # Read Sales Register Sheets
        read_present_quarter_pd = present_quarter_pd

        # Fetch To Address
        to_address = main_config["To_Mail_Address"]
        cc_address = main_config["CC_Mail_Address"]

        # Check Exception
        if read_present_quarter_pd.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Empty present quarter Sales Register found")
            raise BusinessException("Sheet is empty")

        # Check Column Present
        present_quarter_col = read_present_quarter_pd.columns.values.tolist()
        for col in ['Payer Name(Customer Name)', "Base Price in INR"]:
            if col not in present_quarter_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                logging.error("{} Column is missing".format(col))
                raise BusinessException(col + " Column is missing")

        # Filter rows
        payer_name = read_present_quarter_pd[read_present_quarter_pd['Payer Name(Customer Name)'].notna()]
        price_inr = read_present_quarter_pd[read_present_quarter_pd['Base Price in INR'].notna()]

        if len(payer_name) == 0:
            subject = in_config["Payer_Name_Subject"]
            body = in_config["Payer_Name_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Payer Name Column is empty")
            raise BusinessException("Payer Name Column is empty")
        elif len(price_inr) == 0:
            subject = in_config["Base_Price_INR_Subject"]
            body = in_config["Base_Price_INR_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Base Price in INR Column is empty")
            raise BusinessException("Base Price in INR Column is empty")
        else:
            pass

        # Create Pivot Table
        try:
            pivot_index = ["Payer Name(Customer Name)"]
            pivot_values = ["Base Price in INR"]
            pivot_sales = pd.pivot_table(read_present_quarter_pd, index=pivot_index, values=pivot_values,
                                         aggfunc=numpy.sum,
                                         margins=True,
                                         margins_name='Grand Total')
            print("Customer wise Concentration sheet pivot table is created")
            logging.info("Customer wise Concentration pivot table is created")
        except Exception as create_pivot_table:
            subject = in_config["subject_pivot_table"]
            body = in_config["body_pivot_table"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("Concentration Customer Wise Process-", str(create_pivot_table))
            logging.info("Customer Wise Concentration pivot table is not created")
            raise create_pivot_table

        # Remove Index
        pivot_sales = pivot_sales.reset_index()

        # Assign Pivot Sheets
        pivot_sheet = pivot_sales

        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Delete row Base Price in INR column values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[1]] == 0)], inplace=False)

        # Create Concentration Column
        pivot_sheet['Concentration'] = ""

        pd.options.mode.chained_assignment = None

        # Get maximum value
        total_value = pivot_sheet.iloc[-1:]
        total_value = total_value.iloc[0, 1]

        # Concentration Formula
        for index in pivot_sheet.index:
            quarter_value = pivot_sheet[col_name[1]][index]

            if total_value == 0:
                concentration = 1
            else:
                concentration = quarter_value / total_value

            pivot_sheet['Concentration'][index] = concentration

        # Change Column names of Base Price in INR
        pivot_sheet = pivot_sheet.rename(columns={col_name[1]: main_config["PresentQuarterColumnName"]})
        pivot_sheet = pivot_sheet.rename(columns={col_name[0]: in_config["PresentQuarterColumnName1"]})
        try:
            # Log Sheet
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                pivot_sheet.to_excel(writer, sheet_name=main_config["Output_Concentration_Customer_sheetname"],
                                     index=False)
                print("Customer Wise Concentration Output file is saved")
                logging.info("Customer Wise Concentration Output file is saved")
        except Exception as saving_output_file:
            subject = in_config["subject_save_output_file"]
            body = in_config["body_save_output_file"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("Concentration Customer Wise Process-", str(saving_output_file))
            logging.info("Customer Wise Concentration Output file is not saved")
            return saving_output_file

        # Check outfile creation
        if os.path.exists(main_config["Output_File_Path"]):
            print("Customer Wise Concentration Logged")
            logging.info("Customer wise concentration sheet is created")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.warning("Customer Wise Concentration sheet is not created")
            raise BusinessException("Output file not generated")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Concentration_Customer_sheetname"]]

        for cell in ws["C"]:
            cell.number_format = "0%"

        full_range = "A1:" + get_column_letter(ws.max_column) \
                     + str(ws.max_row)
        ws.auto_filter.ref = full_range
        font_style = Font(name="Cambria", size=11, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "1"].font = font_style
        for c in ascii_uppercase:
            ws[c + str(ws.max_row)].font = font_style
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "1"].fill = fill_pattern
            if c == "C":
                break

        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 'C':
                break

        # Save File
        wb.save(main_config["Output_File_Path"])
        logging.info("Completed Customer wise concentration code execution")
        return ws

    except PermissionError as file_error:
        subject = in_config["Permission_Error_Subject"]
        body = in_config["Permission_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Customer Wise Process-", str(file_error))
        print("Please close the file")
        logging.exception(file_error)
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Customer Wise Process-", str(notfound_error))
        logging.exception(notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Concentration Customer Wise Process-", str(business_error))
        logging.exception(business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["Value_Error"]
        body = in_config["Value_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Customer Wise Process-", str(value_error))
        logging.exception(value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["Type_Error"]
        body = in_config["Type_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Customer Wise Process-", str(type_error))
        logging.exception(type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Customer Wise Process-", str(error))
        logging.exception(error)
        return error
    except KeyError as key_error:
        subject = in_config["Name_Error"]
        body = in_config["Name_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Customer Wise Process-", str(key_error))
        logging.exception(key_error)
        return key_error
    except NameError as nameError:
        subject = in_config["Key_Error"]
        body = in_config["Key_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Customer Wise Process-", end="")
        return nameError
    except AttributeError as attributeError:
        subject = in_config["Attribute_Error"]
        body = in_config["Attribute_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Customer Wise Process-", str(attributeError))
        logging.exception(attributeError)
        return attributeError
    except IndexError as indexError:
        subject = in_config["ColumnMiss_Subject"]
        body = in_config["ColumnMiss_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Customer Wise Process-", str(indexError))
        logging.exception(indexError)
        return indexError


if __name__ == "__main__":
    pass
