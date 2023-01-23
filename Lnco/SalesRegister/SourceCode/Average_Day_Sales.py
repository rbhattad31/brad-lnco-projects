import os
from string import ascii_lowercase
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
import logging


class average_day_sales_Exception(Exception):
    pass


def create_config_dict_from_config_file(path, sheet_name):
    try:

        dict_config_main = {}

        if not os.path.exists(path):
            raise average_day_sales_Exception("Config file is not exist in the path provided")
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as config_file_error:
            exception_message = "Below error is occurred while loading Config file from path {0}. \n\t {1}".format(path,
                                                                                                                   config_file_error)
            logging.error("Exception occurred while loading Config file from path")

            raise average_day_sales_Exception(exception_message)

        try:
            worksheet = workbook[sheet_name]
        except Exception as work_sheet_exception:
            exception_message = "Below error is occurred while loading Config Excel file sheet {0}. \n\t {1}".format(
                sheet_name, work_sheet_exception)
            logging.error("Exception occurred while loading Config Excel file sheet")
            raise average_day_sales_Exception(exception_message)

        maximum_row = worksheet.max_row
        maximum_col = worksheet.max_column

        for config_details in worksheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            key = config_details[0].value
            value = config_details[1].value
            dict_config_main[key] = value

        try:
            workbook.save(path)
        except Exception as config_save_exception:
            exception_message = "Below error is occurred while saving Config file in path {0} \n\t {1}".format(path,
                                                                                                               config_save_exception)
            logging.error("Exception occurred while saving Config file from path")
            raise average_day_sales_Exception(exception_message)

        return dict_config_main

    except Exception as config_file_read_error:
        exception_message = "Below error is occurred while reading Config file in path {0} \n\t {1}".format(path,
                                                                                                            config_file_read_error)
        logging.error("Exception occurred while reading Config file in path")
        raise average_day_sales_Exception(exception_message)


def leap_year(year):
    try:
        if type(year) != int:
            raise Exception
    except Exception:
        exception_message = "year value {0} passed to the function is not a number".format(year)
        logging.error("year value{0} passed to the function is not a number")
        raise average_day_sales_Exception(exception_message)

    # divided by 100 means century year (ending with 00)
    # century year divided by 400 is leap year
    if (year % 400 == 0) and (year % 100 == 0):
        print("{0} is a leap year".format(year))
        bool_leap_year = True
    # not divided by 100 means not a century year
    #  divided by 4 is a leap year
    elif (year % 4 == 0) and (year % 100 != 0):
        print("{0} is a leap year".format(year))
        bool_leap_year = True
    # if not divided by both 400 (century year) and 4 (not century year)
    # year is not leap year
    else:
        print("{0} is not a leap year".format(year))
        bool_leap_year = False

    return bool_leap_year


# def average_day_sales(str_sales_register_file_path, str_sales_register_file_sheet_name, main_config):
def average_day_sales(sales_register_df, main_config):
    # creating a pivot table
    sales_register_columns_list = sales_register_df.columns.values.tolist()
    columns_not_found_list = []
    for column in ["Billing Date", "Base Price in INR"]:
        if column not in sales_register_columns_list:
            columns_not_found_list.append(column)
    if len(columns_not_found_list) != 0:
        print("below columns are missing in input sales register file, hence stopping the program. \n\t {0}".format(
            columns_not_found_list))
        exception_message = "Exception occurred: below columns are missing in input sales register file, hence stopping the program. \n\t {0}".format(
            columns_not_found_list)
        raise average_day_sales_Exception(exception_message)
    try:
        average_day_sales_pivot_df = pd.pivot_table(sales_register_df, index="Billing Date", values="Base Price in INR",
                                                    aggfunc=np.sum)
    except Exception as pivot_table_exception:
        exception_message = "Below exception occurred while creating the pivot table for average day sales report,Hence stopping the program.\n\t {0}".format(
            pivot_table_exception)
        logging.error("Exception occurred while creating the pivot table for average day sales report")
        raise average_day_sales_Exception(exception_message)

        # reset index for having proper index for the dataframe and unsetting index column making usual column
    average_day_sales_pivot_df = average_day_sales_pivot_df.reset_index()
    try:
        float_base_price_column_sum = average_day_sales_pivot_df["Base Price in INR"].sum()
        print("float_base_price_column_sum:", float_base_price_column_sum)
    except Exception as base_price_sum_exception:
        exception_message = "Exception occurred while calculating sum of values in base price column"
        print(base_price_sum_exception)
        logging.error("Exception occurred while calculating sum of values in base price column")
        raise average_day_sales_Exception(exception_message)

    try:
        average_day_sales_pivot_df["month"] = average_day_sales_pivot_df["Billing Date"].dt.month_name().str[:3]
    except Exception as month_column_finding_exception:
        exception_message = "Exception occurred while creating month column from billing date that be due to values other than dates"
        print(month_column_finding_exception)
        logging.error(
            "Exception occurred while creating month column from billing date that be due to values other than dates")
        raise average_day_sales_Exception(exception_message)

    month_unique_values_list = average_day_sales_pivot_df["month"].unique().tolist()
    print(month_unique_values_list)
    try:
        average_day_sales_pivot_df["year"] = average_day_sales_pivot_df["Billing Date"].dt.year
    except Exception as year_column_finding_exception:
        exception_message = "Below Exception occurred while creating year column from billing date that be due to values other than dates. \n \t {0}".format(
            year_column_finding_exception)
        logging.error(
            "Exception occurred while creating year column from billing date that be due to values other than dates")
        raise average_day_sales_Exception(exception_message)

    year_unique_values_list = average_day_sales_pivot_df["year"].unique().tolist()
    # print(year_unique_values_list)
    year = year_unique_values_list[0]
    if len(year_unique_values_list) != 1:
        print("sales register contains data of more than 1 year. hence stopping the bot")
        exception_message = "sales register contains data of more than 1 year. hence stopping the bot"
        logging.error("sales register contains data of more than 1 year. hence stopping the bot")
        raise average_day_sales_Exception(exception_message)

    print(year)
    quarter = None
    q1_months = ['Apr', 'May', 'Jun']
    q2_months = ['Jul', 'Aug', 'Sep']
    q3_months = ['Oct', 'Nov', 'Dec']
    q4_months = ['Jan', 'Feb', 'Mar']

    if len(month_unique_values_list) == 3:
        for month in q1_months:
            if month in month_unique_values_list:
                quarter = q1_months
        for month in q2_months:
            if month in month_unique_values_list:
                quarter = q2_months
        for month in q3_months:
            if month in month_unique_values_list:
                quarter = q3_months

        for month in q4_months:
            if month in month_unique_values_list:
                quarter = q4_months
        print(quarter)
    else:
        print("sales register contains data of more than 3 months for a quarter. hence stopping the bot")
        exception_message = "sales register contains data of more than 3 months for a quarter. hence stopping the bot"
        logging.error("sales register contains data of more than 3 months for a quarter. hence stopping the bot")
        raise average_day_sales_Exception(exception_message)

    if quarter == q4_months:
        # find year whether it is leap or not
        bool_leap_year = leap_year(year)
        if bool_leap_year:
            int_days = 91
        else:
            int_days = 90
    elif quarter == q1_months:
        int_days = 91
    elif quarter == q2_months or quarter == q3_months:
        int_days = 92
    else:
        print(
            "The program could not find financial timeline information such as quarter, months and year of sales register")
        exception_message = "The program could not find financial timeline information such as quarter, months and year of sales register"
        logging.error(
            "The program could not find financial timeline information such as quarter, months and year of sales register")
        raise average_day_sales_Exception(exception_message)

    print(int_days)
    float_average_sales_for_day = float_base_price_column_sum / int_days
    print("float_average_sales_for_day:", float_average_sales_for_day)

    average_day_sales_pivot_df["average sales for day"] = float_average_sales_for_day
    try:
        float_average_sales_for_day_sum = average_day_sales_pivot_df["average sales for day"].sum()
        # print(float_average_sales_for_day_sum)
    except Exception as average_sales_per_day_calculation_exception:
        exception_message = "Below exception occurred while calculating sum of 'average sales for day' column \n\t {0}".format(
            average_sales_per_day_calculation_exception)
        logging.error("Below exception occurred while calculating sum of 'average sales for day' column")
        raise average_day_sales_Exception(exception_message)

    average_day_sales_pivot_df["Variance"] = average_day_sales_pivot_df[
                                                 "Base Price in INR"] - float_average_sales_for_day  # average
    try:
        float_variance_sum = average_day_sales_pivot_df["Variance"].sum()
        # print(float_variance_sum)
    except Exception as variance_sum_exception:
        exception_message = "Below Exception occurred while creating year column from billing date that ay be due to values other than dates. \n \t {0}".format(
            variance_sum_exception)
        logging.error(
            "Below exception occurred while creating year column from billing date that ay be due to values other than dates")
        raise average_day_sales_Exception(exception_message)

    average_day_sales_pivot_df["Concentration"] = average_day_sales_pivot_df["Variance"] / float_variance_sum
    # print(average_day_sales_pivot_df)

    # sorting the Concentration in descending order
    # average_day_sales_pivot_df = average_day_sales_pivot_df.sort_values(by=["Concentration"],
    #                                                                     ascending=False)  # descending order
    # creating a column as remarks

    # And for those concentrations greater than 25% and -25% name them as major in remarks column
    positive_threshold_percentage = main_config['average_day_sales_threshold_percentage'] / 100
    negative_threshold_percentage = main_config['average_day_sales_threshold_percentage'] / (-100)
    remarks = main_config['average_day_sales_remarks_keyword']
    average_day_sales_pivot_df["Remarks"] = ''
    for index, row in average_day_sales_pivot_df.iterrows():
        # print(row['Concentration'])
        if row['Concentration'] >= positive_threshold_percentage or row['Concentration'] <= \
                negative_threshold_percentage:
            average_day_sales_pivot_df.at[index, 'Remarks'] = remarks

    average_day_sales_pivot_df.drop("month", inplace=True, axis=1)
    average_day_sales_pivot_df.drop("year", inplace=True, axis=1)
    #  date time
    average_day_sales_pivot_df["Billing Date"] = pd.to_datetime(average_day_sales_pivot_df["Billing Date"]).dt.strftime(
        "%d-%b-%Y")

    output_file_path = main_config['Output_File_Path']
    print(output_file_path)
    output_sheet_name = main_config['Output_Average_Day_Sales_sheetname']
    print(output_sheet_name)
    try:
        with pd.ExcelWriter(output_file_path, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            average_day_sales_pivot_df.to_excel(writer, sheet_name=output_sheet_name,
                                                index=False, startrow=2)
    except Exception as file_saving_exception:
        exception_message = file_saving_exception
        logging.error("Below exception occurred while saving the file ")
        raise average_day_sales_Exception(exception_message)

    # Load Sheet in openpyxl
    workbook = openpyxl.load_workbook(output_file_path)
    worksheet = workbook[output_sheet_name]

    print("workbook", workbook)
    worksheet['B2'] = float_base_price_column_sum  # "=sum(B6:B93)
    worksheet['C2'] = float_average_sales_for_day_sum  # sum of average  sales for day
    worksheet['D2'] = float_variance_sum

    # Number format implementation
    for cell in worksheet['B']:
        cell.number_format = "#,###,##"
    for cell in worksheet['C']:
        cell.number_format = "#,###,##"
    for cell in worksheet['D']:
        cell.number_format = "#,###,##"

    # Format Variance
    for cell in worksheet['E']:
        cell.number_format = '0.0%'

    # Format Header
    cambria_11_black_bold_font = Font(name="Cambria", size=11, color="000000", bold=True)
    # print(ascii_lowercase)
    for c in ascii_lowercase:
        worksheet[c + "3"].font = cambria_11_black_bold_font

    # Header Fill
    fill_solid_light_blue = PatternFill(patternType='solid', fgColor='ADD8E6')
    for f in ascii_lowercase:
        worksheet[f + "3"].fill = fill_solid_light_blue
        if f == 'f':
            break

    thin = Side(border_style="thin", color='000000')
    for row in worksheet.iter_rows(min_row=3, min_col=1, max_row=worksheet.max_row, max_col=6):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Set Width

    # worksheet.column_dimensions[c].width = 25
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 16
    worksheet.column_dimensions['C'].width = 19
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 13
    worksheet.column_dimensions['F'].width = 8

    print(workbook.sheetnames)
    workbook.save(output_file_path)
    return output_file_path


if __name__ == '__main__':
    pass
