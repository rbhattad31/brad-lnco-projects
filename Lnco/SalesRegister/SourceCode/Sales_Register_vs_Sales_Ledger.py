import logging
from string import ascii_lowercase
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Side, Border, Font


class sr_vs_sl_Exception(Exception):
    pass


def sales_ledger_pivot_table(sales_ledger_df):
    # knowing the cols
    print(list(sales_ledger_df.columns.values.tolist()))
    # getting index & values of register
    values_of_ledger = ["Credit", "Debit"]
    sales_ledger_df_pivot_table = pd.pivot_table(sales_ledger_df, index="G/L Acct Long Text", values=values_of_ledger, aggfunc=np.sum)
    # print(sales_ledger_df_pivot_table)
    # resetting index
    sales_ledger_df_pivot_table.reset_index(inplace=True)
    # creating new col net_amount
    sales_ledger_df_pivot_table['net_amount'] = sales_ledger_df_pivot_table['Credit'] - sales_ledger_df_pivot_table['Debit']
    return sales_ledger_df_pivot_table


def sr_vs_sl(data_of_sales_register, main_config, sales_ledger_new_dataframe):
    out_put_file_path = main_config["Output_File_Path"]
    out_put_sheet_name = main_config["Output_Sales_Register_vs_Ledger_sheetname"]

    # print(data_of_sales_register)
    print(data_of_sales_register.columns.values.tolist())
    # getting index & values of register
    index_values_of_sale_register = ["Doc. Type Text", "Type of sale"]
    try:
        dataframe_to_pivot_table_of_register = pd.pivot_table(data_of_sales_register, index=index_values_of_sale_register, values="Base Price in INR", aggfunc=np.sum)
        # print(dataframe_to_pivot_table_of_register)
    except Exception as pivot_table_exception:
        exception_message = "exception occurred while creating a pivot table for sr vs sl report, Hence stopping the bot".format(pivot_table_exception)
        logging.error(" exception occurred while creating a pivot table for sr vs sl report, Hence stopping the bot.")
        raise sr_vs_sl_Exception(exception_message)

    dataframe_to_pivot_table_of_register = dataframe_to_pivot_table_of_register.reset_index()

    try:
        df_sales_ledger = sales_ledger_new_dataframe
        # print(df_sales_ledger)
        dataframe_to_pivot_table_of_ledger = sales_ledger_pivot_table(df_sales_ledger)
    except Exception as sales_ledger_exception:
        exception_message = " exception occurred while reading the file or there is no file_path: \n\t {0}".format(sales_ledger_exception)
        logging.error(exception_message)
        logging.exception(sales_ledger_exception)
        raise sr_vs_sl_Exception(exception_message)

    pd_columns_register = dataframe_to_pivot_table_of_register.columns.values.tolist()
    print(pd_columns_register)
    # print("mmm")
    # pd_columns_register.columns = pd_columns_register.columns.str.replace('Base Price in INR', 'amount as per sr')
    try:
        dataframe_to_pivot_table_of_register = dataframe_to_pivot_table_of_register.rename(columns={pd_columns_register[2]: "amount as per sr"})
        # print(dataframe_to_pivot_table_of_register)
    except Exception as rename_exception:
        exception_message = "exception occurred while renaming a pivot table for sr vs sl report, Hence stopping the bot".format(rename_exception)
        logging.error(" exception occurred while renaming a pivot table for sr vs sl report, Hence stopping the bot.")
        raise sr_vs_sl_Exception(exception_message)

    try:
        sub_total_of_register = dataframe_to_pivot_table_of_register['amount as per sr'].sum()
        # print(sub_total_of_register)
    except Exception as sub_total_exception:
        exception_message = " exception occurred while calculating sub total of register, Hence stopping the bot".format(sub_total_exception)
        logging.error(" exception occurred while calculating sub total of register, Hence stopping the bot.")
        raise sr_vs_sl_Exception(exception_message)

    # to show in tabular form
    try:
        data_of_sales_ledger_table = pd.DataFrame(dataframe_to_pivot_table_of_register)
        # print(data_of_sales_ledger_table)
    except Exception as ledger_table_exception:
        exception_message = " exception occurred while creating a ledger table for sr vs sl report, Hence stopping the bot".format(ledger_table_exception)
        logging.error(" exception occurred while creating a pivot ledger for sr vs sl report, Hence stopping the bot.")
        raise sr_vs_sl_Exception(exception_message)

    # 1
    float_net_amount_sale_scrap = dataframe_to_pivot_table_of_ledger.loc[
        dataframe_to_pivot_table_of_ledger['G/L Acct Long Text'] == 'Sale of  Scrap', 'net_amount'].iloc[0]
    # print(float_net_amount_sale_scrap)
    dataframe_to_pivot_table_of_register.loc[dataframe_to_pivot_table_of_register[
                                                 'Doc. Type Text'] == 'Scrap Order', 'amount as per Sales Ledger'] = float_net_amount_sale_scrap

    # 2
    float_net_amount_sale_export = dataframe_to_pivot_table_of_ledger.loc[
        dataframe_to_pivot_table_of_ledger['G/L Acct Long Text'] == 'Sales Export', 'net_amount'].iloc[0]
    # print(float_net_amount_sale_export)
    dataframe_to_pivot_table_of_register.loc[dataframe_to_pivot_table_of_register[
                                                 'Doc. Type Text'] == 'Export Order', 'amount as per Sales Ledger'] = float_net_amount_sale_export

    # 3
    float_net_amount_sale_job_work_charges = dataframe_to_pivot_table_of_ledger.loc[
        dataframe_to_pivot_table_of_ledger['G/L Acct Long Text'] == 'Sales Job Work Charges', 'net_amount'].iloc[0]
    # print(float_net_amount_sale_job_work_charges)
    dataframe_to_pivot_table_of_register.loc[dataframe_to_pivot_table_of_register[
                                                 'Doc. Type Text'] == 'Service Order', 'amount as per Sales Ledger'] = float_net_amount_sale_job_work_charges

    # 4
    float_net_amount_sales_domestic = dataframe_to_pivot_table_of_ledger.loc[
        dataframe_to_pivot_table_of_ledger['G/L Acct Long Text'] == 'Sales Domestic', 'net_amount'].iloc[0]
    # print(float_net_amount_sales_domestic)
    dataframe_to_pivot_table_of_register.loc[dataframe_to_pivot_table_of_register[
                                                 'Doc. Type Text'] == 'Standard Order', 'amount as per Sales Ledger'] = float_net_amount_sales_domestic

    # 5
    float_net_amount_income_clearing = dataframe_to_pivot_table_of_ledger.loc[
        dataframe_to_pivot_table_of_ledger['G/L Acct Long Text'] == 'Inter Unit Income clearing', 'net_amount'].iloc[0]
    # print(float_net_amount_income_clearing)

    # 6
    float_net_amount_job_work_charges = dataframe_to_pivot_table_of_ledger.loc[
        dataframe_to_pivot_table_of_ledger['G/L Acct Long Text'] == 'Inter Unit Job work Charges', 'net_amount'].iloc[
        0]
    # print(float_net_amount_job_work_charges)

    # 7
    float_net_amount_job_work_income = dataframe_to_pivot_table_of_ledger.loc[
        dataframe_to_pivot_table_of_ledger['G/L Acct Long Text'] == 'Inter Unit Job work Income', 'net_amount'].iloc[0]
    # print(float_net_amount_job_work_income)

    # sum three is in ledger
    sum_of_three_i_s = float_net_amount_income_clearing + float_net_amount_job_work_charges + float_net_amount_job_work_income
    # print(sum_of_three_i_s)
    # 4a
    dataframe_to_pivot_table_of_register.loc[
        dataframe_to_pivot_table_of_register[
            'Doc. Type Text'] == 'INTER PLANT SERVICES', 'amount as per Sales Ledger'] = sum_of_three_i_s

    total_amount_as_per_sr = dataframe_to_pivot_table_of_register['amount as per sr'].sum()
    total_amount_as_per_sl = dataframe_to_pivot_table_of_register['amount as per Sales Ledger'].sum()
    # print(total_amount_as_per_sr)
    # print("Amount of ledger")
    # print(total_amount_as_per_sl)

    dataframe_to_pivot_table_of_register["amount as per Sales Ledger"].fillna(0, inplace=True)

    # difference sr - sl
    amount_as_per_sr = 'amount as per sr'
    amount_as_per_sales_ledger = 'amount as per Sales Ledger'

    dataframe_to_pivot_table_of_register['Difference'] = dataframe_to_pivot_table_of_register.apply(lambda row: row[amount_as_per_sr] - row[amount_as_per_sales_ledger], axis=1)
    total_difference = dataframe_to_pivot_table_of_register['Difference'].sum()
    # print(total_difference)
    try:
        # dataframe_to_pivot_table_of_register.to_excel(out_put_file_path, sheet_name=out_put_sheet_name, index=False, startrow=4)
        with pd.ExcelWriter(out_put_file_path, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            dataframe_to_pivot_table_of_register.to_excel(writer, sheet_name=out_put_sheet_name, index=False, startrow=4)
    except Exception as saving_excel_file:
        exception_message = " exception occurred while saving file, Hence stopping the bot. \n\t {0}".format(saving_excel_file)
        logging.error(exception_message)
        raise sr_vs_sl_Exception(exception_message)

    # showing sum value top of the table
    workbook = openpyxl.load_workbook(out_put_file_path)
    worksheet = workbook[out_put_sheet_name]
    worksheet['C4'].value = total_amount_as_per_sr
    # for sum of sales ledger
    total_amount_as_per_sl = float_net_amount_sale_scrap + float_net_amount_sale_export + float_net_amount_sale_job_work_charges + float_net_amount_sales_domestic + sum_of_three_i_s
    # print(total_amount_as_per_sl)

    worksheet['D4'].value = total_amount_as_per_sl
    worksheet['E4'].value = total_difference

    # Format
    for cell in worksheet['C']:
        cell.number_format = "#,###,##"
    # Format
    for cell in worksheet['D']:
        cell.number_format = "#,###,##"
    # Format
    for cell in worksheet['E']:
        cell.number_format = "#,###,##"

    # Format Header
    calibri_11_black_bold_font = Font(name="Calibri", size=11, color="000000", bold=True)
    # print(ascii_lowercase)
    for c in ascii_lowercase:
        # print(c)
        worksheet[c + "3"].font = calibri_11_black_bold_font

    # Header Fill
    solid_yellow_fill = PatternFill(patternType='solid', fgColor='FFFF00')
    for e in ascii_lowercase:
        worksheet[e + "5"].fill = solid_yellow_fill
        if e == 'e':
            break

        thin = Side(border_style="thin", color='000000')
        for row in worksheet.iter_rows(min_row=5, min_col=1, max_row=worksheet.max_row, max_col=5):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Set Width
        for c in ascii_lowercase:
            worksheet.column_dimensions[c].width = 25
    print(workbook.sheetnames)
    workbook.save(out_put_file_path)


if __name__ == '__main__':
    pass
