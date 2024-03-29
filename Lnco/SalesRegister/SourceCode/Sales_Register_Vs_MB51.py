import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase
import logging
from ReusableTasks.send_mail_reusable_task import send_mail


class BusinessException(Exception):
    pass


def create_sales_register_vs_mb51_sheet(main_config, in_config, sales_present_quarter_pd, mb51_pd):
    try:
        logging.info("Starting Sales Register Vs MB51 code execution")
        read_excel_data = sales_present_quarter_pd
        read_excel_data = read_excel_data.loc[:, ~read_excel_data.columns.duplicated(keep='first')]

        if read_excel_data.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_q1_mail"],
                      body=in_config["Body_q1_mail"])
            logging.error("Empty present quarter Sales Register found")
            raise BusinessException("Sheet is empty")

        sales_sheet_col = read_excel_data.columns.values.tolist()
        for col in ["Material No.", "Material Description", "Billing Qty."]:
            if col not in sales_sheet_col:
                subject = in_config["Sales_ColumnMiss_Subject"]
                body = in_config["Sales_ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                logging.error("{} Column is missing".format(col))
                raise BusinessException(col + " Column is missing")
        # Filter Rows
        material_no_pd = read_excel_data[read_excel_data['Material No.'].notna()]
        material_desc = read_excel_data[read_excel_data['Material Description'].notna()]
        billing_qty = read_excel_data[read_excel_data['Billing Qty.'].notna()]

        if len(material_no_pd) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Material_Number_subject"],
                      body=in_config["Material_Number_Body"])
            logging.error("Material NO. Column is empty")
            raise BusinessException("Material No. Column is empty")

        elif len(material_desc) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Material_Dec_Subject"],
                      body=in_config["Material_Dec_Body"])
            logging.error("Material Description Column is empty")
            raise BusinessException("Material Description Column is empty")

        elif len(billing_qty) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Billing_Qty_Subject"],
                      body=in_config["Billing_Qty_Body"])
            logging.error("Billing Qty Column is empty")
            raise BusinessException("Billing Qty Column is empty")
        else:
            pass

        # read_excel_data = present_quarter_pd
        try:
            read_excel_data = read_excel_data[["Material No.", "Material Description", "Billing Qty."]]
            pivot1_df = pd.pivot_table(read_excel_data, index=["Material No.", "Material Description"],
                                       values="Billing Qty.",
                                       aggfunc=numpy.sum)
            print("Sales Wise Pivot Table is created")
            logging.info("Sales Wise Process Pivot Table is created")
        except Exception as create_pivot_table:
            subject = in_config["subject_pivot_table"]
            body = in_config["body_pivot_table"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("Sales Wise Process-", str(create_pivot_table))
            logging.error("Sales Wise pivot table is not created")
            raise create_pivot_table
        # print(pivot1_df.index)
        # print(pivot1_df)
        pivot1_df = pivot1_df.reset_index()
        # print(pivot1_df)
        pivot1_df[["Material No."]] = pivot1_df[["Material No."]].fillna('').astype(str, errors='ignore')

        # Reading MB 51 File
        read_mb51_excel_data = mb51_pd
        # print(read_excel_data_2.head(5))
        # Check Exception
        if read_mb51_excel_data.shape[0] == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")

        mb51_sheet_col = read_mb51_excel_data.columns.values.tolist()
        for col in ["Material", "Material description", "Quantity"]:
            if col not in mb51_sheet_col:
                subject = in_config["MB51_ColumnMiss_Subject"]
                body = in_config["MB51_ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject,
                          body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        material = read_mb51_excel_data[read_mb51_excel_data['Material'].notna()]
        material_desc = read_mb51_excel_data[read_mb51_excel_data['Material description'].notna()]
        quantity = read_mb51_excel_data[read_mb51_excel_data['Quantity'].notna()]

        if len(material) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Material_subject"],
                      body=in_config["Material_Body"])
            logging.error("Material Column is empty")
            raise BusinessException("Material Column is empty")

        elif len(material_desc) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Material_Dec_Subject"],
                      body=in_config["Material_Dec_Body"])
            logging.error("Material Description column is empty")
            raise BusinessException("Material Description Column is empty")
        elif len(quantity) == 0:
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                      subject=in_config["Quantity_Subject"],
                      body=in_config["Quantity_Body"])
            logging.error("Quantity column is empty")
            raise BusinessException("Quantity Column is empty")
        else:
            pass

        # Taking Required Column to create pivot table
        try:
            read_mb51_excel_data = read_mb51_excel_data[["Material", "Material description", "Quantity"]]
            mb51_pivot_df = pd.pivot_table(read_mb51_excel_data, index=["Material", "Material description"],
                                           values="Quantity",
                                           aggfunc=numpy.sum)
            print("MB51 Pivot Table is created")
        except Exception as create_pivot_table:
            subject = in_config["subject_pivot_table"]
            body = in_config["body_pivot_table"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("MB51 Process-", str(create_pivot_table))
            logging.error("MB51 pivot table is not created")
            raise create_pivot_table
        mb51_pivot_df = mb51_pivot_df.reset_index()

        columns = mb51_pivot_df.columns.values.tolist()
        mb51_pivot_df = mb51_pivot_df.rename(columns={columns[0]: "Material No.",
                                                      columns[1]: "Material Description"})

        # print(pivot2_df.dtypes.tolist())
        mb51_pivot_df[["Material No."]] = mb51_pivot_df[["Material No."]].fillna('').astype(str, errors='ignore')

        merge_pd = pd.merge(pivot1_df, mb51_pivot_df, how="outer", on=["Material No.", "Material Description"],
                            copy=False)
        # print(merge_pd)
        columns_list = merge_pd.columns.values.tolist()

        # create a new column - Success
        merge_pd['Difference'] = ''
        merge_pd[[columns_list[2], columns_list[3]]] = merge_pd[[columns_list[2], columns_list[3]]].fillna(0)
        # To Remove SettingWithCopyWarning error
        # modifying only one df, so suppressing this warning as it is not affecting
        pd.options.mode.chained_assignment = None

        # variance formula for index
        for index in merge_pd.index:
            sales = merge_pd[columns_list[2]][index]
            mb51 = merge_pd[columns_list[3]][index]

            difference = sales + mb51

            merge_pd['Difference'][index] = difference

        try:
            merge_pd.rename(columns={'Billing Qty.': 'Sales Register Billing Quantity', 'Quantity': 'MB51 Billing Quantity'}, inplace=True)
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                merge_pd.to_excel(writer, sheet_name=main_config[
                    "Output_Concentration_SalesRegister_Vs_MB51_sheetname"], index=False)
            print("Sales Register vs MB51 sheet Out file is saved")
        except Exception as save_output_file:
            subject = in_config["subject_save_output_file"]
            body = in_config["body_save_output_file"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("Sales Register Vs MB51 Process-", str(save_output_file))
            logging.error(" Sales Register Vs MB51 sheet Out file is not saved")
            return save_output_file

        # Opening and Reading Output File.

        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Concentration_SalesRegister_Vs_MB51_sheetname"]]

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for i in ascii_uppercase:
            ws[i + "1"].font = font_style

        # Adding Background Color
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")

        for j in ascii_uppercase:
            ws[j + "1"].fill = fill_pattern
            if j == 'E':
                break
        # ws["C1"].fill = PatternFill("solid", fgColor="ffff00")

        # Adding Auto Filter Option
        full_range = "A1:E" + str(ws.max_row)
        ws.auto_filter.ref = full_range

        # Auto Width Setting
        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.5
            if c == 'E':
                break

        # Saving the File
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])
        logging.info("Completed the Sales Register Vs MB51")

        return create_sales_register_vs_mb51_sheet

    # Excepting Errors here
    except FileNotFoundError as notfound_error:
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"],
                  subject=in_config["subject_file_not_found"],
                  body=in_config["body_file_not_found"])
        print("Sales Register Vs MB51 Process-", notfound_error)
        logging.exception(notfound_error)
        return notfound_error
    except ValueError as V_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(V_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Sales Register Vs MB51 Process-", V_error)
        logging.exception(V_error)
        return V_error
    except BusinessException as business_error:
        print("Sales Register Vs MB51 Process-", business_error)
        logging.exception(business_error)
        return business_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Sales Register Vs MB51 Process-", type_error)
        logging.exception(type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Sales Register Vs MB51 Process-", error)
        logging.exception(error)
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Sales Register Vs MB51 Process-", key_error)
        logging.exception(key_error)
        return key_error
    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Please close the file", file_error)
        logging.exception(file_error)
        return file_error


if __name__ == "__main__":
    pass
