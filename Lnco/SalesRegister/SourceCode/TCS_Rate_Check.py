import logging

import numpy
import pandas as pd
# import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase
from openpyxl.utils import get_column_letter
import os
from ReusableTasks.send_mail_reusable_task import send_mail


class BusinessException(Exception):
    pass


def tcs_rate_check(main_config, in_config, present_quarter_pd):
    try:
        logging.info("Starting TCS rate check code execution")
        # Read Sales Register Sheets
        read_excel_data = present_quarter_pd
        read_excel_data['Billing Date'] = pd.to_datetime(read_excel_data['Billing Date'], errors='coerce')

        read_excel_data['Month'] = read_excel_data['Billing Date'].dt.month_name().str[:3]
        # print(read_excel_data)
        # read_excel_data['Month'] = read_excel_data['Billing Date']
        # Fetch To Address
        to_address = main_config["To_Mail_Address"]
        cc_address = main_config["CC_Mail_Address"]

        # Check Exception
        if read_excel_data.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Empty present quarter Sales Register found")
            raise BusinessException("Sheet is empty")

        # Check Column Present
        present_quarter_col = read_excel_data.columns.values.tolist()
        for col in ['Plant', 'Ref.Doc.No.', 'Billing Date', 'Month', 'Payer Name', 'Material No.',
                    'Sales Order', 'Delivery No.',
                    'Billing No.', 'PO. No.', 'PO Date', 'Material Description', 'Billing Qty.', 'Base Price in INR',
                    'CGST Value', 'SGST Value', 'IGST Value', 'JTCS Value', 'Grand Total Value(IN', 'Doc. Type Text']:
            if col not in present_quarter_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                logging.error("{} Column is missing".format(col))
                raise BusinessException(col + " Column is missing")

        # Filter rows
        plant = read_excel_data[read_excel_data['Plant'].notna()]
        ref_doc_no = read_excel_data[read_excel_data['Ref.Doc.No.'].notna()]
        billing_date = read_excel_data[read_excel_data['Billing Date'].notna()]
        month = read_excel_data[read_excel_data['Month'].notna()]
        payer_name = read_excel_data[read_excel_data['Payer Name'].notna()]
        material_no = read_excel_data[read_excel_data['Material No.'].notna()]
        sales_order = read_excel_data[read_excel_data['Sales Order'].notna()]
        delivery_no = read_excel_data[read_excel_data['Delivery No.'].notna()]
        billing_no = read_excel_data[read_excel_data['Billing No.'].notna()]
        po_no = read_excel_data[read_excel_data['PO. No.'].notna()]
        po_date = read_excel_data[read_excel_data['PO Date'].notna()]
        material_description = read_excel_data[read_excel_data['Material Description'].notna()]
        billing_qty = read_excel_data[read_excel_data['Billing Qty.'].notna()]
        base_price_inr = read_excel_data[read_excel_data['Base Price in INR'].notna()]
        cgst_value = read_excel_data[read_excel_data['CGST Value'].notna()]
        sgst_value = read_excel_data[read_excel_data['SGST Value'].notna()]
        igst_value = read_excel_data[read_excel_data['IGST Value'].notna()]
        jtcs_value = read_excel_data[read_excel_data['JTCS Value'].notna()]
        grand_total_value = read_excel_data[read_excel_data['Grand Total Value(IN'].notna()]
        doc_type_text = read_excel_data[read_excel_data['Doc. Type Text'].notna()]
        # type_of_Sale = read_excel_data[read_excel_data['Type of sale'].notna()]

        if len(plant) == 0:
            subject = in_config["Plant_Subject"]
            body = in_config["Plant_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Plant Column is empty")
            raise BusinessException("Plant Column is empty")
        elif len(ref_doc_no) == 0:
            subject = in_config["ref_doc_no_Subject"]
            body = in_config["ref_doc_no_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("ref doc no Column is empty")
            raise BusinessException("ref doc no Column is empty")
        elif len(billing_date) == 0:
            subject = in_config["billing_date_Subject"]
            body = in_config["billing_date_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("billing date Column is empty")
            raise BusinessException("billing date Column is empty")
        elif len(month) == 0:
            subject = in_config["month_Subject"]
            body = in_config["month_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("month Column is empty")
            raise BusinessException("month Column is empty")
        elif len(payer_name) == 0:
            subject = in_config["payer_name_Subject"]
            body = in_config["payer_name_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("payer name Column is empty")
            raise BusinessException("payer name Column is empty")
        elif len(material_no) == 0:
            subject = in_config["material_no_Subject"]
            body = in_config["material_no_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("material no Column is empty")
            raise BusinessException("material no Column is empty")
        elif len(sales_order) == 0:
            subject = in_config["sales_order_Subject"]
            body = in_config["sales_order_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("sales order Column is empty")
            raise BusinessException("sales order Column is empty")
        elif len(delivery_no) == 0:
            subject = in_config["delivery_no_Subject"]
            body = in_config["delivery_no_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("delivery no Column is empty")
            raise BusinessException("delivery no Column is empty")
        elif len(billing_no) == 0:
            subject = in_config["Base_Price_INR_Subject"]
            body = in_config["Base_Price_INR_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Base Price in INR Column is empty")
            raise BusinessException("Base Price in INR Column is empty")
        elif len(po_no) == 0:
            subject = in_config["po_no_Subject"]
            body = in_config["po_no_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("po no Column is empty")
            raise BusinessException("po no Column is empty")
        elif len(po_date) == 0:
            subject = in_config["po_date_Subject"]
            body = in_config["po_date_INR_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("po date Column is empty")
            raise BusinessException(" po date Column is empty")
        elif len(material_description) == 0:
            subject = in_config["material_description_Subject"]
            body = in_config["material_description_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("material description Column is empty")
            raise BusinessException("material description Column is empty")
        elif len(billing_qty) == 0:
            subject = in_config["billing_qty_Subject"]
            body = in_config["billing_qty_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("billing qty Column is empty")
            raise BusinessException("billing qty Column is empty")
        elif len(base_price_inr) == 0:
            subject = in_config["base_price_inr_Subject"]
            body = in_config["base_price_inr_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Base Price in INR Column is empty")
            raise BusinessException("Base Price in INR Column is empty")
        elif len(cgst_value) == 0:
            subject = in_config["cgst_value_Subject"]
            body = in_config["cgst_value_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("CGST Value Column is empty")
            raise BusinessException("CGST Value Column is empty")
        elif len(sgst_value) == 0:
            subject = in_config["sgst_value_Subject"]
            body = in_config["sgst_value_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("SGST Value Column is empty")
            raise BusinessException("SGST Value Column is empty")
        elif len(igst_value) == 0:
            subject = in_config["igst_value_Subject"]
            body = in_config["igst_value_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("IGST Value Column is empty")
            raise BusinessException("IGST Value Column is empty")
        elif len(jtcs_value) == 0:
            subject = in_config["jtcs_value_Subject"]
            body = in_config["jtcs_value_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("JTCS Value Column is empty")
            raise BusinessException("JTCS Value Column is empty")
        elif len(grand_total_value) == 0:
            subject = in_config["grand_total_value_Subject"]
            body = in_config["grand_total_value_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Grand Total Value Column is empty")
            raise BusinessException("Grand Total Value Column is empty")
        elif len(doc_type_text) == 0:
            subject = in_config["doc_type_text_Subject"]
            body = in_config["doc_type_text_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Doc Type Text Column is empty")
            raise BusinessException("Doc Type Text Column is empty")
        else:
            pass

        # Create Pivot Table TCS Rate Check
        try:
            read_tcs_rate_check = read_excel_data[
                ['Plant', 'Ref.Doc.No.', 'Billing Date', 'Month', 'Payer Name', 'Material No.',
                 'Sales Order', 'Delivery No.', 'Billing No.', 'PO. No.', 'PO Date', 'Material Description',
                 'Billing Qty.', 'Base Price in INR', 'CGST Value', 'SGST Value', 'IGST Value', 'JTCS Value',
                 'Grand Total Value(IN', 'Doc. Type Text']]
            # print(read_tcs_rate_check)
            print("TCS Rate Check Pivot table is created")
            logging.info("TCS Rate Check Pivot table is created")
        except Exception as create_pivot_table:
            subject = in_config["subject_pivot_table"]
            body = in_config["body_pivot_table"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("TCS Rate Check Wise Process-", str(create_pivot_table))
            logging.error("TCS Rate Check pivot table is not created")
            raise create_pivot_table

        month_dict = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9,
                      'Oct': 10, 'Nov': 11, 'Dec': 12, 'Grand Total': 13}
        read_tcs_rate_check = read_tcs_rate_check.sort_values('Month', key=lambda x: x.apply(lambda y: month_dict[y]))

        tcs_rate_check = read_tcs_rate_check
        # Remove Empty Rows
        tcs_rate_check = tcs_rate_check.replace(numpy.nan, ' ', regex=True)

        tcs_rate_check['Type of Sale'] = ""
        col_name = tcs_rate_check.columns.values.tolist()
        pd.options.mode.chained_assignment = None

        for index in tcs_rate_check.index:
            # doc_type_text = tcs_rate_check[col_name[19]][index]
            # print(doc_type_text)
            if tcs_rate_check[col_name[19]][index] == 'Scrap Order':
                tcs_rate_check['Type of Sale'][index] = 'Scrap Sales'
            elif tcs_rate_check[col_name[19]][index] == 'Export Ordr w/o Duty':
                tcs_rate_check['Type of Sale'][index] = 'Export Sales'
            elif tcs_rate_check[col_name[19]][index] == 'Export Order':
                tcs_rate_check['Type of Sale'][index] = 'Export Sales'
            elif tcs_rate_check[col_name[19]][index] == 'Trade Order':
                tcs_rate_check['Type of Sale'][index] = 'Domestic Sales'
            elif tcs_rate_check[col_name[19]][index] == 'Standard Order':
                tcs_rate_check['Type of Sale'][index] = 'Domestic Sales'
            elif tcs_rate_check[col_name[19]][index] == 'Service Order':
                tcs_rate_check['Type of Sale'][index] = 'Domestic Sales'
            elif tcs_rate_check[col_name[19]][index] == 'SEZ Sales order':
                tcs_rate_check['Type of Sale'][index] = 'Domestic Sales'
            elif tcs_rate_check[col_name[19]][index] == 'Asset Sale Order':
                tcs_rate_check['Type of Sale'][index] = 'Sale of Asset'
            elif tcs_rate_check[col_name[19]][index] == 'INTER PLANT SERVICES':
                tcs_rate_check['Type of Sale'][index] = 'Job work services'
            elif tcs_rate_check[col_name[19]][index] == 'PLL Credit Memo Req':
                tcs_rate_check['Type of Sale'][index] = 'Sales return'
            elif tcs_rate_check[col_name[19]][index] == 'Returns':
                tcs_rate_check['Type of Sale'][index] = 'Sales return'
            elif tcs_rate_check[col_name[19]][index] == 'Debit Memo Request':
                tcs_rate_check['Type of Sale'][index] = 'Debit Memo'
            else:
                pass
        tcs_rate_check.drop(tcs_rate_check.index[(tcs_rate_check[col_name[20]] != 'Scrap Sales')], inplace=True)
        # Get Pivot Column Names
        tcs_rate_check['applicability'] = ""
        col_name = tcs_rate_check.columns.values.tolist()

        col_name.remove("applicability")
        col_name.insert(12, "applicability")

        tcs_rate_check = tcs_rate_check[col_name]
        col_name = tcs_rate_check.columns.values.tolist()
        pd.options.mode.chained_assignment = None

        for index in tcs_rate_check.index:
            # doc_type_text = tcs_rate_check[col_name[11]][index]
            # print(doc_type_text)

            if tcs_rate_check[col_name[11]][index] == 'WASTE PAPER & GARBAGE SCRAP':
                tcs_rate_check['applicability'][index] = "Not applicable"
            elif "Slit Coil" in tcs_rate_check[col_name[11]][index]:
                tcs_rate_check['applicability'][index] = "Not applicable"
            elif "SIDE TRIMMING COIL" in tcs_rate_check[col_name[11]][index]:
                tcs_rate_check['applicability'][index] = "Not applicable"
            elif tcs_rate_check[col_name[11]][index] == 'WOODEN SCRAP':
                tcs_rate_check['applicability'][index] = "Not applicable"
            elif tcs_rate_check[col_name[11]][index] == 'Corrugated Side angles Scrap':
                tcs_rate_check['applicability'][index] = "Not applicable"
            elif tcs_rate_check[col_name[11]][index] == 'FIRE WOOD SCRAP':
                tcs_rate_check['applicability'][index] = "Not applicable"
            elif tcs_rate_check[col_name[11]][index] != ['FIRE WOOD SCRAP', 'WASTE PAPER & GARBAGE SCRAP',
                                                         'WOODEN SCRAP', 'Corrugated Side angles Scrap']:
                tcs_rate_check['applicability'][index] = "applicable"
            else:
                pass

        tcs_rate_check['TCS as per lnco'] = ""
        col_name = tcs_rate_check.columns.values.tolist()

        col_name.remove("TCS as per lnco")
        col_name.insert(19, "TCS as per lnco")

        tcs_rate_check = tcs_rate_check[col_name]
        col_name = tcs_rate_check.columns.values.tolist()
        pd.options.mode.chained_assignment = None

        for index in tcs_rate_check.index:
            grand_total_value = tcs_rate_check[col_name[20]][index]
            # print(grand_total_value)
            applicable = tcs_rate_check[col_name[12]][index]
            # print(applicable)
            if applicable == 'Not applicable':
                continue
            # continue
            # read applica\bilty cel value
            # check value if it is applicable
            # if it is not applicable, skip the calculation
            # If it is applicable calcualte
            else:
                tcs_lnco = (grand_total_value * 1) / 100
            tcs_rate_check['TCS as per lnco'][index] = tcs_lnco

        tcs_rate_check['Difference'] = ""
        pd.options.mode.chained_assignment = None
        col_name = tcs_rate_check.columns.values.tolist()
        for index in tcs_rate_check.index:
            tcs_lnco = tcs_rate_check[col_name[19]][index]
            JTCS = tcs_rate_check[col_name[18]][index]
            applicable = tcs_rate_check[col_name[12]][index]
            # print(applicable)
            if applicable == 'Not applicable':
                difference = 0

            # print(tcs_lnco)
            else:
                difference = tcs_lnco - JTCS
            # print(Amount)
            tcs_rate_check['Difference'][index] = difference

        tcs_rate_check = tcs_rate_check.rename(columns={col_name[4]: "Payer Name"})
        # print(tcs_rate_check)
        try:
            # Log Sheet
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                tcs_rate_check.to_excel(writer, sheet_name=main_config["Output_TCS_Rate_Check_sheetname"],
                                        index=False, startrow=2)
                print("TCS Rate Check Output file is saved")
                logging.info("TCS Rate Check Output file is saved")
        except Exception as saving_output_file:
            subject = in_config["subject_save_output_file"]
            body = in_config["body_save_output_file"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("TCS Rate Check Wise Process-", str(saving_output_file))
            logging.error("TCS Rate Check Output file is not Saved")
            return saving_output_file

        # Check outfile creation
        if os.path.exists(main_config["Output_File_Path"]):
            print("TCS Rate Check Wise Logged")
            logging.info("TCS Rate Check sheet is created")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.warning("TCS Rate Check sheet is not created")
            raise BusinessException("Output file not generated")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_TCS_Rate_Check_sheetname"]]

        for cell in ws["X"]:
            cell.number_format = "#,##"

        full_range = "A3:" + get_column_letter(ws.max_column) + str(ws.max_row)
        ws.auto_filter.ref = full_range
        font_style = Font(name="Cambria", size=11, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "3"].font = font_style
        # for c in ascii_uppercase:
        #    ws[c + str(ws.max_row)].font = font_style
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "3"].fill = fill_pattern
            if c == "X":
                break

        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 'X':
                break
        ws['S2'] = '=SUBTOTAL(9,S3:K' + str(ws.max_row) + ')'
        ws['T2'] = '=SUBTOTAL(9,T3:M' + str(ws.max_row) + ')'
        ws['X2'] = '=SUBTOTAL(9,X3:K' + str(ws.max_row) + ')'

        # Save File
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])
        logging.info("Completed TCS Rate Check code execution")
        return ws

    except PermissionError as file_error:
        subject = in_config["Permission_Error_Subject"]
        body = in_config["Permission_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("TCS Rate Check wise Process-", str(file_error))
        logging.exception(file_error)
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("TCS Rate Check Wise Process-", str(notfound_error))
        logging.exception(notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("TCS Rate Check Wise Process-", str(business_error))
        logging.exception(business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["Value_Error"]
        body = in_config["Value_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("TCS Rate Check Wise Process-", str(value_error))
        logging.exception(value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["Type_Error"]
        body = in_config["Type_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("TCS Rate Check Wise Process-", str(type_error))
        logging.exception(type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("TCS Rate Check Wise Process-", str(error))
        logging.exception(error)
        return error
    except KeyError as key_error:
        subject = in_config["Name_Error"]
        body = in_config["Name_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("TCS Rate Check Wise Process-", str(key_error))
        logging.exception(key_error)
        return key_error
    except NameError as nameError:
        subject = in_config["Key_Error"]
        body = in_config["Key_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("TCS Rate Check Wise Process-", str(nameError))
        logging.exception(nameError)
        return nameError
    except AttributeError as attributeError:
        subject = in_config["Attribute_Error"]
        body = in_config["Attribute_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("TCS Rate Check Wise Process-", str(attributeError))
        print(attributeError)
        return attributeError


if __name__ == "__main__":
    pass
