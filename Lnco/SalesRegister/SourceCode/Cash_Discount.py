import logging
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from string import ascii_uppercase
from openpyxl.utils import get_column_letter
from ReusableTasks.send_mail_reusable_task import send_mail


class BusinessException(Exception):
    pass


def cash_discount(dict_main_config, dict_in_config, sales_present_quarter_pd):
    try:
        pd.options.mode.chained_assignment = None
        logging.info("Starting Same Material Scrap code execution")
        # Read sales Register Sheets
        sales_present_quarter_pd['Billing Date'] = pd.to_datetime(sales_present_quarter_pd['Billing Date'],
                                                                  errors='coerce')
        sales_present_quarter_pd['Month'] = sales_present_quarter_pd['Billing Date'].dt.month_name().str[:3]
        # Fetch To Address
        str_to_address = dict_main_config["To_Mail_Address"]
        str_cc_address = dict_main_config["CC_Mail_Address"]

        # Check Exception
        if sales_present_quarter_pd.shape[0] == 0:
            str_subject = dict_in_config["EmptyInput_Subject"]
            str_body = dict_in_config["EmptyInput_Body"]
            send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
            logging.error("Empty present quarter Sales Register found")
            raise BusinessException("Sheet is empty")

        # Check Column Present
        sales_present_quarter_columns_list = sales_present_quarter_pd.columns.values.tolist()
        for col in ["Payer Name", "Month", "Base Price in INR"]:
            if col not in sales_present_quarter_columns_list:
                str_subject = dict_in_config["ColumnMiss_Subject"]
                str_body = dict_in_config["ColumnMiss_Body"]
                str_body = str_body.replace("ColumnName +", col)
                send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
                logging.error("{} Column is missing".format(col))
                raise BusinessException(col + " Column is missing")

        # Filter rows
        pd_payer_name = sales_present_quarter_pd[sales_present_quarter_pd['Payer Name'].notna()]
        pd_month = sales_present_quarter_pd[sales_present_quarter_pd['Month'].notna()]
        pd_base_price_inr = sales_present_quarter_pd[sales_present_quarter_pd['Base Price in INR'].notna()]

        if len(pd_payer_name) == 0:
            str_subject = dict_in_config["payer_name_Subject"]
            str_body = dict_in_config["payer_name_Body"]
            send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
            logging.error("Payer Name Column is empty")
            raise BusinessException("Material No. Column is empty")
        elif len(pd_month) == 0:
            str_subject = dict_in_config["month_Subject"]
            str_body = dict_in_config["month_Body"]
            send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
            logging.error("Month Column is empty")
            raise BusinessException("Month Column is empty")
        elif len(pd_base_price_inr) == 0:
            str_subject = dict_in_config["base_price_inr_Subject"]
            str_body = dict_in_config["base_price_inr_Body"]
            send_mail(to=str_to_address, cc=str_cc_address, subject=str_subject, body=str_body)
            logging.error("Base Price in INR Column is empty")
            raise BusinessException("Base Price in INR Column is empty")
        else:
            pass

        # Create Pivot Table cash discount
        try:

            str_pivot_index = "Payer Name"
            str_pivot_columns = "Month"
            str_pivot_values = "Base Price in INR"
            cash_discount_df = pd.pivot_table(sales_present_quarter_pd, index=str_pivot_index,
                                              columns=str_pivot_columns,
                                              values=str_pivot_values,
                                              aggfunc=numpy.sum,
                                              margins=True,
                                              margins_name="Grand Total")
            print("Cash Discount Pivot table is Created")
            logging.info("Cash Discount Pivot table is Created")
        except Exception as create_pivot_table:
            str_subject = dict_in_config["subject_pivot_table"]
            str_body = dict_in_config["body_pivot_table"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Cash Discount Wise Process-", str(create_pivot_table))
            logging.critical("Cash Discount pivot table is not created")
            raise create_pivot_table

        cash_discount_df = cash_discount_df.reset_index()
        cash_discount_columns = cash_discount_df.columns.values.tolist()
        # print(cash_discount_columns)
        month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Oct', 'Nov',
                       'Dec']
        new_month_order = []
        for month in month_order:
            if month in cash_discount_columns:
                new_month_order.append(month)
                # print(new_month_order)
        new_month_order.insert(0, cash_discount_columns[0])
        new_month_order.append(cash_discount_columns[-1])
        # print(new_month_order)
        cash_discount_df = cash_discount_df[new_month_order]

        new_cash_discount_df = cash_discount_df
        cash_discount_df = pd.merge(cash_discount_df, new_cash_discount_df, how="outer",
                                    on=["Payer Name"],
                                    copy=False)
        # print(cash_discount_df)
        print(cash_discount_df.columns.values.tolist())

        values = ["ABB India Limited -", "ABB India Limited - Vadodara", "ABB India Limited-Faridabad",
                  "Dol Motors Pvt Limited", "Maharashtra Electro Mechanical Work",
                  "Locomotive Manufacturing and Servic", "Grand Total"]
        cash_discount_df = cash_discount_df[cash_discount_df["Payer Name"].isin(values)]
        # print(cash_discount_df)
        # Using df.loc
        cash_discount_df = cash_discount_df.drop(columns="Grand Total_y")

        columns = cash_discount_df.columns.values.tolist()
        cash_discount_df = cash_discount_df.rename(columns={columns[1]: columns[1].replace('_x', '')})
        cash_discount_df = cash_discount_df.rename(columns={columns[2]: columns[2].replace('_x', '')})
        cash_discount_df = cash_discount_df.rename(columns={columns[3]: columns[3].replace('_x', '')})
        cash_discount_df = cash_discount_df.rename(columns={columns[4]: columns[4].replace('_x', '')})
        cash_discount_df = cash_discount_df.rename(columns={columns[5]: columns[5].replace('_y', '')})
        cash_discount_df = cash_discount_df.rename(columns={columns[6]: columns[6].replace('_y', '')})
        cash_discount_df = cash_discount_df.rename(columns={columns[7]: columns[7].replace('_y', '')})

        try:
            # Log Sheet
            with pd.ExcelWriter(dict_main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                cash_discount_df.to_excel(writer,
                                          sheet_name=dict_main_config["Output_Cash_Discount_sheetname"],
                                          index=False, startrow=2)
            print("Cash Discount sheet Out file is saved")
            logging.info("Cash Discount sheet Out file is saved")
        except Exception as save_output_file:
            str_subject = dict_in_config["subject_save_output_file"]
            str_body = dict_in_config["body_save_output_file"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Cash Discount Wise Process-", str(save_output_file))
            logging.critical("Cash Discount sheet Out file is not saved")
            return save_output_file

        # Load Sheet in openpyxl
        try:
            workbook = openpyxl.load_workbook(dict_main_config["Output_File_Path"])
            print("Cash Discount Work Book is loaded")
            logging.info("Cash Discount Work Book is loaded")
        except Exception as load_work_book:
            str_subject = dict_in_config["load_work_book_subject"]
            str_body = dict_in_config["load_work_book_body"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Cash Discount Wise Process-", str(load_work_book))
            logging.critical("Cash Discount work book is not loaded")
            return load_work_book
        try:
            worksheet = workbook[dict_main_config["Output_Cash_Discount_sheetname"]]
            print("Cash Discount Work Sheet is loaded")
            logging.info("Cash Discount Work Sheet is loaded")
        except Exception as load_work_sheet:
            str_subject = dict_in_config["load_work_sheet_subject"]
            str_body = dict_in_config["load_work_sheet_body"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Cash Discount Wise Process-", str(load_work_sheet))
            logging.critical("Cash Discount work sheet is not loaded")
            return load_work_sheet

        full_range = "A3:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
        worksheet.auto_filter.ref = full_range
        cambria_11_bold_black = Font(name="Cambria", size=11, bold=True, color="000000")
        for c in ascii_uppercase:
            worksheet[c + "3"].font = cambria_11_bold_black
        for c in ascii_uppercase:
            worksheet[c + str(worksheet.max_row)].font = cambria_11_bold_black
        solid_light_blue_fill = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            worksheet[c + "3"].fill = solid_light_blue_fill
            if c == "H":
                break

        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in worksheet[c])
            worksheet.column_dimensions[c].width = column_length * 1.25
            if c == 'H':
                break

        worksheet['B10'] = '=SUM(B3:B9)'
        worksheet['C10'] = '=SUM(C3:C9)'
        worksheet['D10'] = '=SUM(D3:D9)'
        worksheet['E10'] = '=SUM(E3:E9)'

        worksheet['F4'] = '=(B4*6.39%) *180/365'
        worksheet['F5'] = '=(B5*6.39%) *180/365'
        worksheet['F6'] = '=(B6*6.39%) *180/365'
        worksheet['F7'] = '=(B7*9%) *90/365'
        worksheet['F8'] = '=(B8*2%)'
        worksheet['F9'] = '=(B9*9%) *90/365'
        worksheet['F10'] = '=SUM(F3:F9)'

        worksheet['G4'] = '=(C4*6.39%) *180/365'
        worksheet['G5'] = '=(C5*6.39%) *180/365'
        worksheet['G6'] = '=(C6*6.39%) *180/365'
        worksheet['G7'] = '=(C7*9%) *90/365'
        worksheet['G8'] = '=(C8*2%)'
        worksheet['G9'] = '=(C9*9%) *90/365'
        worksheet['G10'] = '=SUM(G3:G9)'

        worksheet['H4'] = '=(D4*6.39%) *180/365'
        worksheet['H5'] = '=(D5*6.39%) *180/365'
        worksheet['H6'] = '=(D6*6.39%) *180/365'
        worksheet['H7'] = '=(D7*9%) *90/365'
        worksheet['H8'] = '=(D8*2%)'
        worksheet['H9'] = '=(D9*9%) *90/365'
        worksheet['H10'] = '=SUM(H3:H9)'

        cell = worksheet['B2']
        cell.value = 'SALES AMOUNT AS PER SR'
        worksheet.merge_cells('B2:E2')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = worksheet['F2']
        cell.value = 'DISCOUNT AS PER LNCO'
        worksheet.merge_cells('F2:H2')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cambria_12_bold_black_font = Font(name="Cambria", size=12, bold=True, color="000000")
        worksheet['B2'].font = cambria_12_bold_black_font
        worksheet['F2'].font = cambria_12_bold_black_font

        solid_light_blue_fill = PatternFill(patternType="solid", fgColor="ADD8E6")
        worksheet['B2'].fill = solid_light_blue_fill
        worksheet['F2'].fill = solid_light_blue_fill
        # Save File
        try:
            print(workbook.sheetnames)
            workbook.save(dict_main_config["Output_File_Path"])
            print("Cash Discount Work Sheet file is closed")
            logging.info("Cash Discount Work Sheet file is Closed")
        except Exception as close_file:
            str_subject = dict_in_config["close_work_sheet_file_subject"]
            str_body = dict_in_config["close_work_sheet_file_body"]
            send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"],
                      subject=str_subject,
                      body=str_body)
            print("Cash Discount Process-", str(close_file))
            logging.critical("Cash Discount work sheet file is not closed")
            return close_file
        logging.info("Completed Cash Discount code execution")
        return cash_discount

    except PermissionError as file_error:
        str_subject = dict_in_config["Permission_Error_Subject"]
        str_body = dict_in_config["Permission_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Cash Discount Process-", str(file_error))
        print("Please close the file")
        logging.exception(file_error)
        return file_error
    except FileNotFoundError as notfound_error:
        str_subject = dict_in_config["FileNotFound_Subject"]
        str_body = dict_in_config["FileNotFound_Body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Cash Discount Process-", str(notfound_error))
        logging.exception(notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Cash Discount Process-", str(business_error))
        logging.exception(business_error)
        return business_error
    except ValueError as value_error:
        str_subject = dict_in_config["Value_Error"]
        str_body = dict_in_config["Value_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Cash Discount Process-", str(value_error))
        logging.exception(value_error)
        return value_error
    except TypeError as type_error:
        str_subject = dict_in_config["Type_Error"]
        str_body = dict_in_config["Type_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Cash Discount Process-", str(type_error))
        logging.exception(type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        str_subject = dict_in_config["SystemError_Subject"]
        str_body = dict_in_config["SystemError_Body"]
        str_body = str_body.replace("SystemError +", str(error))
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Cash Discount Process-", str(error))
        logging.exception(error)
        return error
    except KeyError as key_error:
        str_subject = dict_in_config["Name_Error"]
        str_body = dict_in_config["Name_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Cash Discount Process-", str(key_error))
        logging.exception(key_error)
        return key_error
    except NameError as nameError:
        str_subject = dict_in_config["Key_Error"]
        str_body = dict_in_config["Key_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Cash Discount Process-", str(nameError))
        logging.exception(nameError)
        return nameError
    except AttributeError as attributeError:
        str_subject = dict_in_config["Attribute_Error"]
        str_body = dict_in_config["Attribute_Error_body"]
        send_mail(to=dict_main_config["To_Mail_Address"], cc=dict_main_config["CC_Mail_Address"], subject=str_subject,
                  body=str_body)
        print("Cash Discount Process-", str(attributeError))
        logging.exception(attributeError)
        return attributeError


if __name__ == "__main__":
    pass
