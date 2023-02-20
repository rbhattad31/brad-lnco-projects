import logging
from string import ascii_lowercase
from openpyxl.styles import Font
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side
import openpyxl


class RPT_Exception(Exception):
    pass


def related_parties_transaction(rpt_df, dict_main_config):
    str_output_file_path = dict_main_config["Output_File_Path"]
    str_output_sheet_name = dict_main_config["Output_Related_Party_Transactions_sheetname"]
    str_payer_list = dict_main_config["related_party_transaction_payer_list"]
    # filtering
    rpt_payers = str_payer_list
    # print(type(rpt_payers))
    # convert string to list
    list_rpt_payers = rpt_payers.split(",")
    # print(type(list_rpt_payers))
    # print(list_rpt_payers)
    # print(len(list_rpt_payers))
    # for x1 in range(len(list_rpt_payers)):
    #     print(list_rpt_payers[x1])
    logging.info(list_rpt_payers)
    # selecting rows based on condition
    # print(rpt_df)
    rpt_df = rpt_df.loc[rpt_df['Payer Name'].isin(list_rpt_payers)]
    # print('\nResult dataframe :\n', rpt_df)

    # creating pivot table sales register
    try:
        rpt_pivot_df = pd.pivot_table(rpt_df, index="Payer Name", values="So Unit Price")
        # print(rpt_pivot_df)
    except Exception as Sales_register_file_read_exception:
        str_exception_message = "Below exception occurred while creating pivot table: \n\t {0}".format(
            Sales_register_file_read_exception)
        print(Sales_register_file_read_exception)
        logging.error(str_exception_message)
        raise RPT_Exception(str_exception_message)

    # indexing pivot table
    rpt_pivot_df = rpt_pivot_df.reset_index()
    # print(rpt_pivot_df)

    # creating columns in excel
    result_df = pd.DataFrame(
        columns=['Payer Name', 'Min of So Unit Price', 'Max of So unit price', 'Variance', 'Variance %','Remarks'])
    try:

        for client in list_rpt_payers:
            clients = client.split(",")
            # print(clients)
            # print(type(clients))
            df_rpt_payer = rpt_df.loc[rpt_df['Payer Name'].isin(clients)]
            # print('\nResult dataframe :\n', df_rpt_payer)
            # calculating first minimum value  in column
            df_min = df_rpt_payer['So Unit Price'].min()
            # print("df_min: ", df_min)
            # calculating first maximum value  in column
            df_max = df_rpt_payer['So Unit Price'].max()
            # print("df_max: ", df_max)
            # creating first variance formula
            df_variance = df_max - df_min
            # df_Remarks = "NA", "Minor", "Major"
            df_variance_percentage = df_variance / df_min
            result_df.loc[len(result_df.index)] = [client, df_min, df_max, df_variance, df_variance_percentage, '']
            # result_df = pd.DataFrame(list(zip(clients, df_min, df_max, df_variance)), columns=['Payer Name', 'Min of So Unit Price', 'Max of So unit price', 'Variance', 'Remarks'])
            # print(result_df)
            # print(type(result_df))
            logging.debug("Dataframe created for all payers - as in below line".format(clients))
            logging.debug(result_df)
    except Exception as Sales_register_file_read_exception:
        str_exception_message = "Below exception occurred while filtering payer names  :\n\t {0}".format(
            Sales_register_file_read_exception)
        logging.error(Sales_register_file_read_exception)
        raise RPT_Exception(str_exception_message)

    # creating remarks column and giving comments
    list_of_variance = result_df['Variance'].tolist()
    print(list_of_variance)
    list_of_variance.remove(0)
    print(list_of_variance)
    highest_negative = max(list_of_variance)
    least_negative = min(list_of_variance)
    for index, row in result_df.iterrows():
        if row['Variance'] == least_negative:
            result_df.at[index, 'Remarks'] = 'Minor'
        if row['Variance'] == highest_negative:
            result_df.at[index, 'Remarks'] = 'Major'
        if row['Variance'] == 0:
            result_df.at[index, 'Remarks'] = 'NA'
            result_df_ = result_df.head()
            # print(result_df_)
        # print(row)
    # saving data to excel
    try:
        # result_df.to_excel(str_output_file_path, sheet_name=str_output_sheet_name, index=False)
        with pd.ExcelWriter(str_output_file_path, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            result_df.to_excel(writer, sheet_name=str_output_sheet_name,
                               index=False)
    except Exception as saving_data_to_excel:
        str_exception_message = "Below exception occurred while saving file :\n\t {0}".format(saving_data_to_excel)
        logging.error(saving_data_to_excel)
        raise RPT_Exception(str_exception_message)

    # Load Sheet in openpyxl
    try:
        workbook = openpyxl.load_workbook(str_output_file_path)
        worksheet = workbook[str_output_sheet_name]
    except Exception as loading_sheet_in_openpyxl:
        str_exception_message = "Below exception occurred while file opening sheet in openpyxl :\n\t {0}".format(
            loading_sheet_in_openpyxl)
        logging.error(loading_sheet_in_openpyxl)
        raise RPT_Exception(str_exception_message)
    # Format
    for cell in worksheet['B']:
        cell.number_format = "#,###,##"
    for cell in worksheet['C']:
        cell.number_format = "#,###,##"
    for cell in worksheet['D']:
        cell.number_format = "#,###,##"
    for cell in worksheet['E']:
        cell.number_format = "0%"

    # Format Header
    calibre_11_black_bold = Font(name="Calibre", size=11, color="000000", bold=True)

    # Header Fill
    solid_light_blue_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
    for c in ascii_lowercase:
        worksheet[c + "1"].font = calibre_11_black_bold
        worksheet[c + "1"].fill = solid_light_blue_fill
        worksheet.column_dimensions[c].width = 35
        if c == 'f':
            break

    # applying border
    black_thin_border = Side(border_style="thin", color='000000')
    for row in worksheet.iter_rows(min_row=1, min_col=1, max_row=worksheet.max_row, max_col=6):
        for cell in row:
            cell.border = Border(top=black_thin_border, left=black_thin_border, right=black_thin_border,
                                 bottom=black_thin_border)
    # to save in openpyxl
    try:
        print(workbook.sheetnames)
        workbook.save(str_output_file_path)

    except Exception as saving_data_in_openpyxl:
        str_exception_message = "Below exception occurred while saving file in openpyxl :\n\t {0}".format(
            saving_data_in_openpyxl)
        logging.error("Below exception occurred while saving file in openpyxl")
        raise RPT_Exception(str_exception_message)


if __name__ == '__main__':
    pass
