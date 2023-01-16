import os
from string import ascii_lowercase
import xlrd_compdoc_commented as xlrd
import openpyxl
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
import logging
import numpy as np


def gst_rate_check(main_config, in_config, sales_present_quarter_pd, hsn_pd):
    sales_present_quarter_pd = sales_present_quarter_pd[['Plant', 'Ref.Doc.No.', 'Payer Name(Customer Name)',
                                                         'Base Price in INR', 'CGST Value', 'SGST Value', 'IGST Value',
                                                         'JTCS Value', 'Grand Total Value(IN', 'HSN Code']]

    pd.options.mode.chained_assignment = None
    sales_present_quarter_pd['Total GST as per client'] = 0

    # Variance Formula
    for index in sales_present_quarter_pd.index:
        try:
            total_gst_as_per_client = sales_present_quarter_pd['CGST Value'][index] + \
                                      sales_present_quarter_pd['SGST Value'][index] + \
                                      sales_present_quarter_pd['IGST Value'][index]
            sales_present_quarter_pd['Total GST as per client'][index] = total_gst_as_per_client
        except Exception as error:
            # print('%s (%s-%s): %s' % (error.filename, error.lineno, error.offset, error.text))
            print(error)
    print(sales_present_quarter_pd)
    print("Total GST as per client")
    sales_present_quarter_pd['GST as per client'] = 0
    for index in sales_present_quarter_pd.index:
        try:
            # gst_as_per_client = sales_present_quarter_pd['Total GST as per client'][index] / sales_present_quarter_pd['Base Price in INR'][index]
            if sales_present_quarter_pd['Base Price in INR'][index] == 0:
                sales_present_quarter_pd['GST as per client'][index] = main_config[
                    'default_gst_rate_for_base_price_0_value_entries']
                continue
            gst_as_per_client = np.divide(sales_present_quarter_pd['Total GST as per client'][index],
                                          sales_present_quarter_pd['Base Price in INR'][index])
            sales_present_quarter_pd['GST as per client'][index] = gst_as_per_client
        except Exception as error:
            print("error")
            print(error)

    print(sales_present_quarter_pd)
    print("Columns are created")
    hsn_pd.rename(columns={'HSN Codes': 'HSN Code'}, inplace=True)
    # hsn_pd.columns.values = ['HSN Code', 'GST Rate', 'Description']
    print(hsn_pd.columns.values.tolist())
    gst_rate_check_pd = pd.merge(sales_present_quarter_pd, hsn_pd, on='HSN Code', how='left')
    print(gst_rate_check_pd)
    print(gst_rate_check_pd.columns.values.tolist())
    gst_rate_check_pd = gst_rate_check_pd[
        ['Plant', 'Ref.Doc.No.', 'Payer Name(Customer Name)', 'Base Price in INR', 'CGST Value', 'SGST Value',
         'IGST Value', 'JTCS Value', 'Grand Total Value(IN', 'HSN Code', 'Total GST as per client', 'GST as per client',
         'GST Rate']]
    gst_rate_check_pd['Difference'] = 0

    for index in gst_rate_check_pd.index:
        try:
            difference = gst_rate_check_pd['GST as per client'][index] - gst_rate_check_pd['GST Rate'][index]
            gst_rate_check_pd['Difference'][index] = difference
        except Exception as error:
            # print('%s (%s-%s): %s' % (error.filename, error.lineno, error.offset, error.text))
            print(error)

    try:
        with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            gst_rate_check_pd.to_excel(writer, sheet_name=main_config["Output_GST_Rate_Check_sheet_name"], index=False,
                                       startrow=4)
    except Exception as File_creation_error:
        logging.error("Exception occurred while creating gst rate check sheet")
        raise File_creation_error
    print('Gst rate Check output is saved in output file')

    wb = openpyxl.load_workbook(main_config["Output_File_Path"])
    ws = wb[main_config["Output_GST_Rate_Check_sheet_name"]]

    # Format Q4 & Q3
    for column in ['D', 'E', 'F', 'G', 'H', 'I', 'K']:
        for cell in ws[column]:
            cell.number_format = "#,###,##"

    # Format Variance
    for column in ['L', 'M', 'N']:
        for cell in ws[column]:
            cell.number_format = '0.0%'

    fill_solid_light_blue = PatternFill(patternType='solid', fgColor='ADD8E6')
    for c in ascii_lowercase:
        ws[c + "5"].fill = fill_solid_light_blue
        if c == 'n':
            break
    calibri_11_black_bold = Font(name="Calibri", size=11, color="000000", bold=True)
    for c in ascii_lowercase:
        ws[c + "5"].font = calibri_11_black_bold
        if c == 'n':
            break
    thin = Side(border_style="thin", color='b1c5e7')

    for row in ws.iter_rows(min_row=6, min_col=1, max_row=ws.max_row, max_col=14):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Set Width
    for c in ascii_lowercase:
        max_cell_length = max(len(str(cell.value)) for cell in ws[c])
        ws.column_dimensions[c].width = max_cell_length * 1.25
        if c == 'n':
            break

    fill_solid_orange = PatternFill(patternType='solid', fgColor='FFA500')
    fill_red_orange = PatternFill(patternType='solid', fgColor='FF0000')
    for i in range(6, ws.max_row+1):
        try:
            if round(ws.cell(i, 14).value, 2) != 0:
                ws.cell(i, 14).fill = fill_solid_orange
        except Exception as round_error:
            ws.cell(i, 14).fill = fill_red_orange

    ws.sheet_view.showGridLines = False
    wb.save(main_config["Output_File_Path"])


if __name__ == '__main__':
    pass