import logging
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase
from openpyxl.utils import get_column_letter
import os
from ReusableTasks.send_mail_reusable_task import send_mail


class BusinessException(Exception):
    pass


def plant_wise_concentration(main_config, in_config, present_quarter_pd):
    try:
        logging.info("Starting plant wise concentration code execution")
        # Read Purchase Register Sheets
        read_present_quarter_pd = present_quarter_pd

        # Fetch To Address
        to_address = main_config["To_Mail_Address"]
        cc_address = main_config["CC_Mail_Address"]

        # Check Exception
        if read_present_quarter_pd.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Empty present quarter Sales Register found")
            raise BusinessException("Sheet is empty")

        # Check Column Present
        present_quarter_col = read_present_quarter_pd.columns.values.tolist()
        for col in ['Plant', "Base Price in INR"]:
            if col not in present_quarter_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                logging.error("{} Column is missing".format(col))
                raise BusinessException(col + " Column is missing")

        # Filter rows
        plant = read_present_quarter_pd[read_present_quarter_pd['Plant'].notna()]
        price_inr = read_present_quarter_pd[read_present_quarter_pd['Base Price in INR'].notna()]

        if len(plant) == 0:
            subject = in_config["Plant_Subject"]
            body = in_config["Plant_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Plant Column is empty")
            raise BusinessException("Plant Column is empty")
        elif len(price_inr) == 0:
            subject = in_config["Base_Price_INR_Subject"]
            body = in_config["Base_Price_INR_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.error("Base Price in INR Column is empty")
            raise BusinessException("Base Price in INR Column is empty")
        else:
            pass

        # Create Pivot Table Q4
        try:
            pivot_index = ["Plant"]
            pivot_values = ["Base Price in INR"]
            pivot_sales = pd.pivot_table(read_present_quarter_pd, index=pivot_index, values=pivot_values,
                                         aggfunc=numpy.sum,
                                         margins=True,
                                         margins_name='Grand Total')
            print("Plant Wise Concentration Pivot table is created")
            logging.info("Plant Wise Concentration Pivot table is created")
        except Exception as create_pivot_table:
            subject = in_config["subject_pivot_table"]
            body = in_config["body_pivot_table"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("Concentration Plant Wise Process-", str(create_pivot_table))
            logging.info("Plant Wise Concentration pivot table is not created")
            raise create_pivot_table

        # Remove Index
        pivot_sales = pivot_sales.reset_index()
        # Assign Pivot Sheets
        pivot_sheet = pivot_sales

        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Delete row of Base price in INR column values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[1]] == 0)], inplace=True)

        # Create Concentration Column
        pivot_sheet['Concentration'] = ""

        pd.options.mode.chained_assignment = None

        # Get maximum value
        total_value = pivot_sheet.iloc[-1:]
        total_value = total_value.iloc[0, 1]

        # Variance Formula
        for index in pivot_sheet.index:
            quarter_value = pivot_sheet[col_name[1]][index]

            if total_value == 0:
                concentration = 1
            else:
                concentration = quarter_value / total_value

            pivot_sheet['Concentration'][index] = concentration

        # Change Column names of Base Price in INR
        pivot_sheet = pivot_sheet.rename(columns={col_name[1]: main_config["PresentQuarterColumnName"]})
        try:
            # Log Sheet
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                pivot_sheet.to_excel(writer, sheet_name=main_config["Output_Concentration_Plant_sheetname"], index=False)
                print("Plant Wise Concentration Output file is saved")
                logging.info("Plant Wise Concentration Output file is saved")
        except Exception as saving_output_file:
            subject = in_config["subject_save_output_file"]
            body = in_config["body_save_output_file"]
            send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
            print("Concentration Plant Wise Process-", str(saving_output_file))
            logging.info("Plant Wise Concentration Output file is not Saved")
            return saving_output_file

        # Check outfile creation
        if os.path.exists(main_config["Output_File_Path"]):
            print("Plant Wise Concentration Logged")
            logging.info("Plant wise concentration sheet is created")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            logging.warning("Plant Wise Concentration sheet is not created")
            raise BusinessException("Output file not generated")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(main_config["Output_File_Path"])
        ws = wb[main_config["Output_Concentration_Plant_sheetname"]]

        for cell in ws["C"]:
            cell.number_format = "0%"

        full_range = "A1:" + get_column_letter(ws.max_column) + str(ws.max_row)
        ws.auto_filter.ref = full_range
        font_style = Font(name="Cambria", size=11, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "1"].font = font_style
        for c in ascii_uppercase:
            ws[c + str(ws.max_row)].font = font_style
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "1"].fill = fill_pattern
            if c == "C":
                break

        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 'C':
                break

        # Save File
        print(wb.sheetnames)
        wb.save(main_config["Output_File_Path"])
        logging.info("Completed Plant wise concentration code execution")
        return ws

    except PermissionError as file_error:
        subject = in_config["Permission_Error_Subject"]
        body = in_config["Permission_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(file_error))
        logging.exception(file_error)
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(notfound_error))
        logging.exception(notfound_error)
        return notfound_error
    except BusinessException as business_error:
        print("Concentration Plant Wise Process-", str(business_error))
        logging.exception(business_error)
        return business_error
    except ValueError as value_error:
        subject = in_config["Value_Error"]
        body = in_config["Value_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(value_error))
        logging.exception(value_error)
        return value_error
    except TypeError as type_error:
        subject = in_config["Type_Error"]
        body = in_config["Type_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(type_error))
        logging.exception(type_error)
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(error))
        logging.exception(error)
        return error
    except KeyError as key_error:
        subject = in_config["Name_Error"]
        body = in_config["Name_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(key_error))
        logging.exception(key_error)
        return key_error
    except NameError as nameError:
        subject = in_config["Key_Error"]
        body = in_config["Key_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(nameError))
        logging.exception(nameError)
        return nameError
    except AttributeError as attributeError:
        subject = in_config["Attribute_Error"]
        body = in_config["Attribute_Error_body"]
        send_mail(to=main_config["To_Mail_Address"], cc=main_config["CC_Mail_Address"], subject=subject, body=body)
        print("Concentration Plant Wise Process-", str(attributeError))
        logging.exception(attributeError)
        return attributeError


if __name__ == "__main__":
    pass
