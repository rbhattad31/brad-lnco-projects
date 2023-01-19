
import openpyxl
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill
import logging
import re
from openpyxl.utils import get_column_letter


def sequence_check(main_config, in_config, sales_present_quarter_pd):
    sales_present_quarter_pd = sales_present_quarter_pd[
        ['Plant', 'Month', 'Billing Date', 'Type of sale', 'Ref.Doc.No.']]
    # print(sales_present_quarter_pd)
    sales_present_quarter_pd = sales_present_quarter_pd.drop_duplicates(keep='first')
    # print(sales_present_quarter_pd)
    sales_columns_list = sales_present_quarter_pd.columns.values.tolist()
    # pd.options.mode.chained_assignment = None

    # get unique values in a column
    list_plants = sales_present_quarter_pd['Plant'].unique().tolist()
    # print(list_plants)
    count = 0
    for plant in list_plants:
        # plant_as_list = list(plant)
        temp_df = pd.DataFrame(columns=sales_columns_list)

        temp_df = sales_present_quarter_pd.drop(sales_present_quarter_pd[sales_present_quarter_pd['Plant'] != plant].index)
        temp_df.sort_values(['Type of sale', 'Ref.Doc.No.'], ascending=[True, True], inplace=True)
        temp_df.reset_index(drop=True, inplace=True)

        temp_df['Ref.Doc.No.letters'] = ''
        temp_df['Ref.Doc.No.Numbers'] = 0
        temp_df['Difference'] = 0

        # print(temp_df)

        temp_df[["Ref.Doc.No."]] = temp_df[["Ref.Doc.No."]].fillna('').astype(str, errors='ignore')
        temp_df['Billing Date'] = pd.to_datetime(temp_df['Billing Date'], errors='coerce').dt.strftime("%d-%m-%Y")

        for index in temp_df.index:

            str_ref_doc_number = temp_df.loc[index]['Ref.Doc.No.']

            str_number = re.findall(r'\d+', str_ref_doc_number)[0]
            int_number = int(str_number)

            str_letters = re.findall(r'[a-zA-Z]+', str_ref_doc_number)[0]

            # temp_df.loc[index]['Ref.Doc.No.letters'] = str_ref_doc_number.replace('([A-Z]+)', '')
            temp_df.at[index, 'Ref.Doc.No.letters'] = str_letters
            # temp_df.loc[index]['Ref.Doc.No.Numbers'] = str_ref_doc_number.extract('([A-Z]+)')
            temp_df.at[index, 'Ref.Doc.No.Numbers'] = int_number

            if index == 0:
                temp_df.at[index, 'Difference'] = 0
                continue

            if temp_df.loc[index]['Type of sale'] != temp_df.loc[index-1]['Type of sale']:
                temp_df.at[index, 'Difference'] = 0
                continue

            temp_df.at[index, 'Difference'] = temp_df.loc[index]['Ref.Doc.No.Numbers'] - temp_df.loc[
                index - 1]['Ref.Doc.No.Numbers']
        # print(temp_df)

        try:
            with pd.ExcelWriter(main_config["Output_File_Path"], engine="openpyxl", mode="a",
                                if_sheet_exists="overlay") as writer:
                temp_df.to_excel(writer, sheet_name=main_config["Config_Sequence_Check_sheet_name"],
                                 index=False,
                                 startrow=4, startcol=count * 9 + 1)
                count = count + 1
        except Exception as File_creation_error:
            logging.error("Exception occurred while creating gst rate check sheet")
            raise File_creation_error
    print('Sequence Check output is saved in output file')

    workbook = openpyxl.load_workbook(main_config["Output_File_Path"])
    worksheet = workbook[main_config["Config_Sequence_Check_sheet_name"]]

    # color fill for header
    fill_solid_light_blue = PatternFill(patternType='solid', fgColor='ADD8E6')
    max_column = worksheet.max_column
    max_row = worksheet.max_row

    for column_number in range(max_column):
        column_letter = get_column_letter(column_number+1)
        if worksheet[column_letter + '5'].value is not None:

            worksheet[column_letter + "5"].fill = fill_solid_light_blue

    # border
    blue_thin_border = Side(border_style="thin", color='b1c5e7')
    for count in range(len(list_plants)):
        for row in worksheet.iter_rows(min_row=6, min_col=1, max_row=max_row, max_col=max_column):
            for cell in row:
                if cell.value is None:
                    continue
                cell.border = Border(top=blue_thin_border, left=blue_thin_border, right=blue_thin_border, bottom=blue_thin_border)

    # Set Width
    for column_number in range(max_column):
        column_letter = get_column_letter(column_number+1)
        max_cell_length = max(len(str(cell.value)) for cell in worksheet[column_letter])
        worksheet.column_dimensions[column_letter].width = max_cell_length * 1.25

    worksheet.sheet_view.showGridLines = False
    print(workbook.sheetnames)
    workbook.save(main_config["Output_File_Path"])


if __name__ == '__main__':
    pass
