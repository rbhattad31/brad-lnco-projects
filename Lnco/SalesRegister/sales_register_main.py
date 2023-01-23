import openpyxl
import xlsxwriter
import pandas as pd
import os.path
import os
import sys
import logging

from ReusableTasks.send_mail_reusable_task import send_mail, send_mail_with_attachment

from SalesRegister.SourceCode.Customer_Wise_Concentration import customer_wise_concentration
from SalesRegister.SourceCode.Month_Wise_Concentration import month_wise_concentration
from SalesRegister.SourceCode.Plant_Wise_Concentration import plant_wise_concentration
from SalesRegister.SourceCode.Type_of_Sales_Wise_Concentration import type_of_sale_wise_concentration
from SalesRegister.SourceCode.Sales_Register_Vs_MB51 import create_sales_register_vs_mb51_sheet
from SalesRegister.SourceCode.Product_Mix_Comparison import product_mix_comparison
from SalesRegister.SourceCode.Vendor_And_Material import Vendor_And_Material

from SalesRegister.SourceCode.gst_rate_check import gst_rate_check
from SalesRegister.SourceCode.TCS_Rate_Check import tcs_rate_check
from SalesRegister.SourceCode.Average_Day_Sales import average_day_sales
from SalesRegister.SourceCode.Related_parties_transaction import related_parties_transaction

from SalesRegister.SourceCode.Same_Material_Scrap import same_material_scrap
from SalesRegister.SourceCode.Same_Material_Domestic import same_material_domestic

from SalesRegister.SourceCode.Customer_specific import customer_specific
from SalesRegister.SourceCode.Sales_Register_vs_Sales_Ledger import sr_vs_sl
from SalesRegister.SourceCode.open_po_reports import open_po_reports

from SalesRegister.SourceCode.sequence_check import sequence_check
from SalesRegister.SourceCode.Cash_Discount import cash_discount

from SalesRegister.File_Creation_Programs.sales_present_quarter_file_creation import sales_present_quarter_file_creation
from SalesRegister.File_Creation_Programs.sales_previous_quarter_file_creation import \
    sales_previous_quarter_file_creation
from SalesRegister.File_Creation_Programs.hsn_codes_file_creation import hsn_codes_file_creation
from SalesRegister.File_Creation_Programs.mb51_file_creation import mb51_file_creation
from SalesRegister.File_Creation_Programs.sales_ledger_file_creation import sales_ledger_file_creation
from SalesRegister.File_Creation_Programs.open_po_file_creation import open_po_file_creation

from ReusableTasks.create_sheet_wise_config_dictionary import create_sheet_wise_config_dict


def process_execution(input_files,
                      present_quarter_sheet_name, previous_quarter_sheet_name,
                      hsn_codes_file_sheet_name, mb51_sheet_name,
                      sales_ledger_sheet_name, open_po_sheet_name,
                      present_quarter_column_name, previous_quarter_column_name,
                      company_name, statutory_audit_quarter, financial_year, config_main, request_id,
                      json_data_list,
                      env_file
                      ):
    print("Starting audit process for the input files")
    logging.info("Starting audit process for the input files")
    print(input_files)
    mb51_file_path = input_files[0]
    hsn_file_path = input_files[1]
    sales_register_present_quarter_file_path = input_files[2]
    sales_register_previous_quarter_file_path = input_files[3]
    sales_ledger_file_path = input_files[4]
    open_po_file_path = input_files[5]

    config_main['PresentQuarterColumnName'] = present_quarter_column_name
    config_main['PreviousQuarterColumnName'] = previous_quarter_column_name
    config_main['CompanyName'] = company_name
    config_main['StatutoryAuditQuarter'] = statutory_audit_quarter
    config_main['FinancialYear'] = financial_year

    print("*******************************************")
    print("Check if Output file exists")

    project_home_directory = os.getcwd()
    output_file_folder = os.path.join(project_home_directory, 'Data', 'Output', 'audit_requests', str(request_id))
    print("Output file folder is : ", output_file_folder)
    if not os.path.exists(output_file_folder):
        print("Output folder is not exist")
        print("Creating directory: ", output_file_folder)
        os.makedirs(output_file_folder)
        print("Directory" + output_file_folder + " is created")
    output_file_name = company_name.replace(' ', '_') + "_" + str(request_id) + "_Sales_Register_Output.xlsx"
    output_file_path = os.path.join(output_file_folder, output_file_name)
    print("Output file path is: " + output_file_path)
    config_main['Output_File_Path'] = output_file_path
    if os.path.exists(output_file_path):
        print("Output file exist")
        print("Deleting existing output file")
        os.remove(output_file_path)
        if os.path.exists(output_file_path):
            pass
        else:
            print("existing output file is deleted successfully")
        print("Creating a new output file")
        workbook = xlsxwriter.Workbook(output_file_path)
        workbook.close()
        if os.path.exists(output_file_path):
            print("New output file is created")
        else:
            print("New output file creation is failed")
    else:
        print("Output file not exist")
        print("Creating a new output file")
        workbook = xlsxwriter.Workbook(output_file_path)
        workbook.close()
        if os.path.exists(output_file_path):
            print("New output file is created")
        else:
            print("New output file creation is failed")

    print("*******************************************")

    try:
        print("Reading Sales register files is started")
        print("Reading Sales register present quarter sheet is started")
        print(sales_register_present_quarter_file_path)
        read_present_quarter_pd = pd.read_excel(sales_register_present_quarter_file_path,
                                                present_quarter_sheet_name)
        # print(read_present_quarter_pd.dtypes.to_list)
        read_present_quarter_pd = \
            read_present_quarter_pd.loc[:, ~read_present_quarter_pd.columns.duplicated(keep='first')]

        present_quarter_columns = read_present_quarter_pd.columns
        if config_main["sales_register_1st_column_name"] in present_quarter_columns and \
                config_main["sales_register_2nd_column_name"] in present_quarter_columns:
            print("Present Quarter file - The data is starting from first row only")
            pass

        else:
            print("Present Quarter file - The data is not starting from first row ")
            for index, row in read_present_quarter_pd.iterrows():
                if row[0] != config_main["sales_register_1st_column_name"]:
                    read_present_quarter_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = read_present_quarter_pd.iloc[0]
            read_present_quarter_pd = read_present_quarter_pd[1:]
            read_present_quarter_pd.columns = new_header
            read_present_quarter_pd.reset_index(drop=True, inplace=True)
            read_present_quarter_pd.columns.name = None
        read_present_quarter_pd = \
            read_present_quarter_pd.loc[:, ~read_present_quarter_pd.columns.duplicated(keep='first')]

        # print(read_present_quarter_pd)
        print(
            "Reading Sales register present quarter sheet is complete, creating new input file only with required columns")
        logging.info(
            "Reading Sales register present quarter sheet is complete, creating new input file only with required columns")
        sales_register_present_quarter_folder_path = os.path.dirname(sales_register_present_quarter_file_path)
        sales_register_present_quarter_file_name = os.path.basename(
            sales_register_present_quarter_file_path).lower()
        filtered_sales_present_file_name = "filtered_" + str(sales_register_present_quarter_file_name)
        filtered_sales_present_file_saving_path = os.path.join(sales_register_present_quarter_folder_path,
                                                               filtered_sales_present_file_name)
        filtered_sales_present_sheet_name = present_quarter_sheet_name
        sales_present_quarter_file_creation_output = sales_present_quarter_file_creation(config_main,
                                                                                         read_present_quarter_pd,
                                                                                         json_data_list,
                                                                                         filtered_sales_present_file_saving_path,
                                                                                         filtered_sales_present_sheet_name)

        sales_present_quarter_pd = sales_present_quarter_file_creation_output[0]
        config_main = sales_present_quarter_file_creation_output[1]

        logging.info("new sales register present quarter file is created in input folder in request ID folder")
        print("Reading Sales register present quarter sheet is completed")
        # -----------------------------------------------------------------------------------------------
        if env_file('VENDOR_AND_MATERIAL_COMPARISON') == 'YES':
            print("Reading Sales register previous quarter sheet is started")
            logging.info("Reading Sales register previous quarter sheet is started")
            # reading previous quarter sheet
            print("Reading previous quarter sheet")
            logging.info("Reading previous quarter sheet")
            read_previous_quarter_pd = pd.read_excel(sales_register_previous_quarter_file_path,
                                                     previous_quarter_sheet_name)
            # # print(read_previous_quarter_pd.dtypes.to_list)
            read_previous_quarter_pd = \
                read_previous_quarter_pd.loc[:, ~read_previous_quarter_pd.columns.duplicated(keep='first')]

            previous_quarter_columns = read_previous_quarter_pd.columns
            if config_main["sales_register_1st_column_name"] in previous_quarter_columns and \
                    config_main["sales_register_2nd_column_name"] in previous_quarter_columns:
                print("Previous Quarter file - The data is starting from first row only")
                pass

            else:
                print("Previous Quarter file - The data is not starting from first row ")
                for index, row in read_previous_quarter_pd.iterrows():
                    if row[0] != config_main["purchase_register_1st_column_name"]:
                        read_previous_quarter_pd.drop(index, axis=0, inplace=True)
                    else:
                        break
                new_header = read_previous_quarter_pd.iloc[0]
                read_previous_quarter_pd = read_previous_quarter_pd[1:]
                read_previous_quarter_pd.columns = new_header
                read_previous_quarter_pd.reset_index(drop=True, inplace=True)
                read_previous_quarter_pd.columns.name = None
            read_previous_quarter_pd = \
                read_previous_quarter_pd.loc[:, ~read_previous_quarter_pd.columns.duplicated(keep='first')]
            print(
                "Reading Sales register previous quarter sheet is complete, creating new input file only with required columns")
            logging.info(
                "Reading Sales register previous quarter sheet is complete, creating new input file only with required columns")
            sales_register_previous_quarter_folder_path = os.path.dirname(sales_register_previous_quarter_file_path)
            sales_register_previous_quarter_file_name = os.path.basename(
                sales_register_previous_quarter_file_path).lower()
            filtered_sales_previous_file_name = "filtered_" + str(sales_register_previous_quarter_file_name)
            filtered_sales_previous_file_saving_path = os.path.join(sales_register_previous_quarter_folder_path,
                                                                    filtered_sales_previous_file_name)
            filtered_sales_previous_sheet_name = previous_quarter_sheet_name

            sales_previous_quarter_file_creation_output = \
                sales_previous_quarter_file_creation(config_main,
                                                     read_previous_quarter_pd,
                                                     json_data_list,
                                                     filtered_sales_previous_file_saving_path,
                                                     filtered_sales_previous_sheet_name)
            read_previous_quarter_pd = sales_previous_quarter_file_creation_output[0]
            config_main = sales_previous_quarter_file_creation_output[1]

            logging.info("new sales register previous quarter file is created in input folder in request ID folder")
            print("new sales register previous quarter file is created in input folder in request ID folder")
            print("Reading Sales register previous quarter sheet is completed")
            print("Reading Sales register files is Completed")
        # --------------------------------------------------------------------------------------------------
        if env_file('GST_RATE_CHECK') == 'YES':
            print("Reading HSN Codes summary sheet is started")

            print(hsn_file_path)
            hsn_code_pd = pd.read_excel(hsn_file_path, hsn_codes_file_sheet_name)

            hsn_code_pd = \
                hsn_code_pd.loc[:, ~hsn_code_pd.columns.duplicated(keep='first')]

            hsn_code_pd_columns = hsn_code_pd.columns
            if config_main["HSN_sheet_1st_column_name"] in hsn_code_pd_columns and \
                    config_main["HSN_sheet_2nd_column_name"] in hsn_code_pd_columns:
                print("HSN Codes Sheet - The data is starting from first row only")
                pass

            else:
                print("HSN Codes Sheet - The data is not starting from first row ")
                for index, row in hsn_code_pd.iterrows():
                    if row[0] != config_main["HSN_sheet_1st_column_name"]:
                        hsn_code_pd.drop(index, axis=0, inplace=True)
                    else:
                        break
                new_header = hsn_code_pd.iloc[0]
                hsn_code_pd = hsn_code_pd[1:]
                hsn_code_pd.columns = new_header
                hsn_code_pd.reset_index(drop=True, inplace=True)
                hsn_code_pd.columns.name = None
            hsn_code_pd = \
                hsn_code_pd.loc[:, ~hsn_code_pd.columns.duplicated(keep='first')]

            print(
                "Reading HSN Codes sheet is complete, creating new input file only with required columns")
            logging.info(
                "Reading HSN Codes sheet is complete, creating new input file only with required columns")
            hsn_codes_file_folder_path = os.path.dirname(hsn_file_path)
            hsn_codes_file_name = os.path.basename(hsn_file_path).lower()
            filtered_hsn_codes_file_name = "filtered_" + str(hsn_codes_file_name)
            filtered_hsn_codes_file_saving_path = os.path.join(hsn_codes_file_folder_path, filtered_hsn_codes_file_name)

            filtered_hsn_codes_sheet_name = hsn_codes_file_sheet_name
            hsn_codes_file_creation_output = hsn_codes_file_creation(config_main,
                                                                     hsn_code_pd,
                                                                     json_data_list,
                                                                     filtered_hsn_codes_file_saving_path,
                                                                     filtered_hsn_codes_sheet_name)

            hsn_codes_new_dataframe = hsn_codes_file_creation_output[0]
            config_main = hsn_codes_file_creation_output[1]

            logging.info("HSN Codes file is created in input folder in request ID folder")
            print("Reading HSN Codes file is completed")
        else:
            hsn_codes_new_dataframe = pd.DataFrame()
        # ---------------------------------------------------------------------------------------------
        if env_file('SR_VS_MB51') == 'YES':
            print("Reading MB51 sheet is started")
            print(mb51_file_path)
            mb51_pd = pd.read_excel(mb51_file_path, mb51_sheet_name)

            mb51_pd = mb51_pd.loc[:, ~mb51_pd.columns.duplicated(keep='first')]

            mb51_pd_columns = mb51_pd.columns
            if config_main["MB51_first_column"] in mb51_pd_columns and \
                    config_main["MB51_second_column"] in mb51_pd_columns:
                print("MB51 Sheet- The data is starting from first row only")
                pass

            else:
                print("MB51 Sheet - The data is not starting from first row ")
                for index, row in mb51_pd.iterrows():
                    if row[0] != config_main["MB51_first_column"]:
                        mb51_pd.drop(index, axis=0, inplace=True)
                    else:
                        break
                new_header = mb51_pd.iloc[0]
                mb51_pd = mb51_pd[1:]
                mb51_pd.columns = new_header
                mb51_pd.reset_index(drop=True, inplace=True)
                mb51_pd.columns.name = None
            mb51_pd = \
                mb51_pd.loc[:, ~mb51_pd.columns.duplicated(keep='first')]

            print(
                "Reading MB51 sheet is complete, creating new input file only with required columns")
            logging.info(
                "Reading MB51 sheet is complete, creating new input file only with required columns")
            mb51_file_folder_path = os.path.dirname(mb51_file_path)
            mb51_file_name = os.path.basename(mb51_file_path).lower()
            filtered_mb51_file_name = "filtered_" + str(mb51_file_name)
            filtered_mb51_file_saving_path = os.path.join(mb51_file_folder_path, filtered_mb51_file_name)

            filtered_mb51_file_sheet_name = mb51_sheet_name
            mb51_file_creation_output = mb51_file_creation(config_main,
                                                           mb51_pd,
                                                           json_data_list,
                                                           filtered_mb51_file_saving_path,
                                                           filtered_mb51_file_sheet_name)

            mb51_new_dataframe = mb51_file_creation_output[0]
            config_main = mb51_file_creation_output[1]

            logging.info("Filtered MB51 file is created in input folder in request ID folder")
            print("Filtered MB51 Codes file is completed")
        else:
            mb51_new_dataframe = pd.DataFrame()
        # ---------------------------------------------------------------------------------------------
        # Sales Ledger
        if env_file('SR_VS_SL') == 'YES':
            print("Reading Sales Ledger sheet is started")
            logging.info("Reading Sales Ledger sheet is started")
            print(sales_ledger_file_path)
            sales_ledger_pd = pd.read_excel(sales_ledger_file_path, sales_ledger_sheet_name)

            sales_ledger_pd = \
                sales_ledger_pd.loc[:, ~sales_ledger_pd.columns.duplicated(keep='first')]

            sales_ledger_pd_columns = sales_ledger_pd.columns
            # print(sales_ledger_pd_columns)
            if "Credit" in sales_ledger_pd_columns and "Debit" in sales_ledger_pd_columns:
                print("Sales Ledger Sheet - The data is starting from first row only")
                pass

            else:
                print("Sales Ledger Sheet - The data is not starting from first row ")
                for index, row in sales_ledger_pd.iterrows():
                    if "Credit" not in row.values.tolist() and "Debit" not in row.values.tolist():
                        sales_ledger_pd.drop(index, axis=0, inplace=True)
                    else:
                        break
                new_header = sales_ledger_pd.iloc[0]
                sales_ledger_pd = sales_ledger_pd[1:]
                sales_ledger_pd.columns = new_header
                sales_ledger_pd.reset_index(drop=True, inplace=True)
                sales_ledger_pd.columns.name = None
            sales_ledger_pd = \
                sales_ledger_pd.loc[:, ~sales_ledger_pd.columns.duplicated(keep='first')]
            # print(sales_ledger_pd)
            print(
                "Reading Sales Ledger sheet is complete, creating new input file only with required columns")
            logging.info(
                "Reading Sales Ledger sheet is complete, creating new input file only with required columns")
            sales_ledger_file_folder_path = os.path.dirname(sales_ledger_file_path)
            sales_ledger_file_name = os.path.basename(sales_ledger_file_path).lower()
            filtered_sales_ledger_file_name = "filtered_" + str(sales_ledger_file_name)
            filtered_sales_ledger_file_saving_path = os.path.join(sales_ledger_file_folder_path,
                                                                  filtered_sales_ledger_file_name)

            filtered_sales_ledger_sheet_name = sales_ledger_sheet_name
            sales_ledger_file_creation_output = sales_ledger_file_creation(config_main,
                                                                           sales_ledger_pd,
                                                                           json_data_list,
                                                                           filtered_sales_ledger_file_saving_path,
                                                                           filtered_sales_ledger_sheet_name)

            sales_ledger_new_dataframe = sales_ledger_file_creation_output[0]
            config_main = sales_ledger_file_creation_output[1]

            logging.info("Sales Ledger file is created in input folder in request ID folder")
            print("Reading Sales Ledger file is completed")
        else:
            sales_ledger_new_dataframe = pd.DataFrame()

        # ---------------------------------------------------------------------------------------------
        # Open PO
        if env_file('OPEN_PO') == 'YES':
            print("Reading OPEN PO File is started")

            print(open_po_file_path)
            open_po_pd = pd.read_excel(open_po_file_path, open_po_sheet_name)

            open_po_pd = \
                open_po_pd.loc[:, ~open_po_pd.columns.duplicated(keep='first')]

            open_po_pd_columns = open_po_pd.columns
            if "Order Date" in open_po_pd_columns and "PO Date" in open_po_pd_columns:
                print("Open PO Sheet - The data is starting from first row only")
                pass

            else:
                print("Open PO Sheet - The data is not starting from first row ")
                for index, row in open_po_pd.iterrows():
                    if "Order Date" not in row.values.tolist() and "PO Date" not in row.values.tolist():
                        open_po_pd.drop(index, axis=0, inplace=True)
                    else:
                        break
                new_header = open_po_pd.iloc[0]
                open_po_pd = open_po_pd[1:]
                open_po_pd.columns = new_header
                open_po_pd.reset_index(drop=True, inplace=True)
                open_po_pd.columns.name = None
            open_po_pd = \
                open_po_pd.loc[:, ~open_po_pd.columns.duplicated(keep='first')]
            print(
                "Reading Open PO sheet is complete, creating new input file only with required columns")
            logging.info(
                "Reading Open PO sheet is complete, creating new input file only with required columns")
            open_po_file_folder_path = os.path.dirname(open_po_file_path)
            open_po_file_name = os.path.basename(open_po_file_path).lower()
            filtered_open_po_file_name = "filtered_" + str(open_po_file_name)
            filtered_open_po_file_saving_path = os.path.join(open_po_file_folder_path,
                                                             filtered_open_po_file_name)

            filtered_open_po_sheet_name = open_po_sheet_name
            open_po_file_creation_output = open_po_file_creation(config_main, open_po_pd, json_data_list,
                                                                 filtered_open_po_file_saving_path,
                                                                 filtered_open_po_sheet_name)

            open_po_new_dataframe = open_po_file_creation_output[0]
            config_main = open_po_file_creation_output[1]

            logging.info("Open PO file is created in input folder in request ID folder")
            print("Reading Open PO file is completed")
        else:
            open_po_new_dataframe = pd.DataFrame()

    # ---------------------------------------------------------------------------------------------
    except FileNotFoundError as notfound_error:
        send_mail(to=config_main["To_Mail_Address"], cc=config_main["CC_Mail_Address"],
                  subject=config_main["subject_file_not_found"],
                  body=config_main["body_file_not_found"])
        print(notfound_error)
        logging.error("file not found error occurred: \n\t {}".format(notfound_error))
        logging.exception(notfound_error)
        raise notfound_error
    except ValueError as sheetNotFound_error:
        send_mail(to=config_main["To_Mail_Address"], cc=config_main["CC_Mail_Address"],
                  subject=config_main["subject_sheet_not_found"],
                  body=config_main["body_sheet_not_found"])
        print(sheetNotFound_error)
        logging.error("sheet not found error occurred: \n\t {}".format(sheetNotFound_error))
        logging.exception(sheetNotFound_error)
        raise sheetNotFound_error
    except Exception as file_creation_exception:
        logging.exception(file_creation_exception)
        raise file_creation_exception

    print("*******************************************")
    # ------------------------------------------------------------------------------------
    try:
        if env_file('CONCENTRATION_CUSTOMER_WISE') == 'YES':
            print("Executing customer wise concentration program")
            config_concentration_customer_wise = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Concentrations_Customer_sheetname"])
            customer_wise_concentration(main_config=config_main, in_config=config_concentration_customer_wise,
                                        present_quarter_pd=sales_present_quarter_pd)

        elif env_file('CONCENTRATION_CUSTOMER_WISE') == 'NO':
            print("customer wise concentration program is skipped as per env file")
        else:
            print("select YES/NO for customer wise concentration program in env file")
            raise Exception("Error in Env file for 'customer wise concentration program' sheet")
    except Exception as e:
        print("Exception caught for Process: 'customer wise concentration program' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('CONCENTRATION_MONTH_WISE') == 'YES':
            print("Executing month wise concentration program")
            config_concentration_month_wise = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Concentrations_Month_sheetname"])
            month_wise_concentration(main_config=config_main, in_config=config_concentration_month_wise,
                                     present_quarter_pd=sales_present_quarter_pd)

        elif env_file('CONCENTRATION_MONTH_WISE') == 'NO':
            print("month wise concentration program is skipped as per env file")
        else:
            print("select YES/NO for month wise concentration program in env file")
            raise Exception("Error in Env file for 'month wise concentration program' sheet")
    except Exception as e:
        print("Exception caught for Process: 'customer wise concentration program' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('CONCENTRATION_PLANT_WISE') == 'YES':
            print("Executing plant wise concentration program")
            config_concentration_plant_wise = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Concentrations_Plant_sheetname"])
            plant_wise_concentration(main_config=config_main, in_config=config_concentration_plant_wise,
                                     present_quarter_pd=sales_present_quarter_pd)

        elif env_file('CONCENTRATION_PLANT_WISE') == 'NO':
            print("plant wise concentration program is skipped as per env file")
        else:
            print("select YES/NO for plant wise concentration program in env file")
            raise Exception("Error in Env file for 'plant wise concentration program' sheet")
    except Exception as e:
        print("Exception caught for Process: 'plant wise concentration program' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('CONCENTRATION_TYPE_OF_SALE_WISE') == 'YES':
            print("Executing 'type of sale' wise concentration program")
            config_concentration_type_of_sale_wise = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Concentrations_Type_of_sale_sheetname"])
            type_of_sale_wise_concentration(main_config=config_main, in_config=config_concentration_type_of_sale_wise,
                                            present_quarter_pd=sales_present_quarter_pd)

        elif env_file('CONCENTRATION_TYPE_OF_SALE_WISE') == 'NO':
            print("'type of sale' wise concentration program is skipped as per env file")
        else:
            print("select YES/NO for 'type of sale' wise concentration program in env file")
            raise Exception("Error in Env file for 'type of sale' wise concentration program sheet")
    except Exception as e:
        logging.exception(e)
        print("Exception caught for Process: 'type of sale wise concentration program' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('SR_VS_MB51') == 'YES':
            print("Executing SR vs MB51 sheet program")
            config_sr_vs_mb51 = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_SalesRegister_Vs_MB51_sheetname"])
            create_sales_register_vs_mb51_sheet(config_main, config_sr_vs_mb51, sales_present_quarter_pd,
                                                mb51_new_dataframe)

        elif env_file('SR_VS_MB51') == 'NO':
            print("Sales Register Vs MB51 program is skipped as per env file")
        else:
            print("select YES/NO for 'Sales Register Vs MB51' program in env file")
            raise Exception("Error in Env file for 'Sales Register Vs MB51 program' sheet")
    except Exception as e:
        print("Exception caught for Process: 'Sales Register Vs MB51 program' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('PRODUCT_MIX_COMPARISON') == 'YES':
            print("Executing 'product mix comparison' program")
            config_product_mix_comparison = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Product_Mix_Comparison_sheetname"])
            product_mix_comparison(main_config=config_main, in_config=config_product_mix_comparison,
                                   present_quarter_pd=sales_present_quarter_pd)

        elif env_file('PRODUCT_MIX_COMPARISON') == 'NO':
            print("Product mix comparison program is skipped as per env file")
        else:
            print("select YES/NO for 'product mix comparison' program in env file")
            raise Exception("Error in Env file for 'product mix comparison program' sheet")
    except Exception as e:
        print("Exception caught for Process: 'Product mix comparison' program' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('VENDOR_AND_MATERIAL_COMPARISON') == 'YES':
            print("Executing Vendor and Material Comparison program")
            config_vendor_material_comparison = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_VendorAndMaterial_Comparison_sheetname"])
            Vendor_And_Material(main_config=config_main, in_config=config_vendor_material_comparison,
                                present_quarter_pd=sales_present_quarter_pd,
                                previous_quarter_path=sales_register_previous_quarter_file_path)

        elif env_file('VENDOR_AND_MATERIAL_COMPARISON') == 'NO':
            print("Vendor and Material comparison program is skipped as per env file")
        else:
            print("select YES/NO for 'Vendor and Material Comparison' program in env file")
            raise Exception("Error in Env file for 'Vendor and Material Comparison' sheet")
    except Exception as e:
        print("Exception caught for Process: 'Vendor and Material Comparison' Sheet: ", e)
    # ------------------------------------------------------------------------------------

    try:
        if env_file('GST_RATE_CHECK') == 'YES':
            print("Executing GST Rate Check program")
            hsn_pd = hsn_codes_new_dataframe
            config_gst_rate_check = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_GST_Rate_Check_sheet_name"])
            gst_rate_check(config_main, config_gst_rate_check, sales_present_quarter_pd, hsn_pd)

        elif env_file('GST_RATE_CHECK') == 'NO':
            print("GST Rate Check process is skipped as per env file")
        else:
            print("select YES/NO for GST Rate Check process in env file")
            raise Exception("Error in Env file for 'GST Rate Check' sheet")
    except Exception as e:
        print("Exception caught for Process: 'GST Rate Check' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('TCS_RATE_CHECK') == 'YES':
            print("Executing 'TCS Rate Check' program")
            config_tcs_rate_check = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_TCS_Rate_Check_sheetname"])
            tcs_rate_check(main_config=config_main, in_config=config_tcs_rate_check,
                           present_quarter_pd=sales_present_quarter_pd)

        elif env_file('TCS_RATE_CHECK') == 'NO':
            print("TCS Rate Check program is skipped as per env file")
        else:
            print("select YES/NO for TCS Rate Check program in env file")
            raise Exception("Error in Env file for 'TCS Rate Check program' sheet")
    except Exception as e:
        print("Exception caught for Process: 'TCS Rate Check program' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('AVERAGE_DAY_SALES') == 'YES':
            print("Executing Average Day Sales program")
            average_day_sales(sales_register_df=sales_present_quarter_pd, main_config=config_main)

        elif env_file('AVERAGE_DAY_SALES') == 'NO':
            print("Average Day Sales program is skipped as per env file")
        else:
            print("select YES/NO for Average Day Sales program in env file")
            raise Exception("Error in Env file for 'Average Day Sales program' sheet")
    except Exception as e:
        print("Exception caught for Process: 'Average Day Sales program' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('RELATED_PARTY_TRANSACTION') == 'YES':
            print("Executing 'related party transaction' program")
            related_parties_transaction(rpt_df=sales_present_quarter_pd, dict_main_config=config_main)
        elif env_file('RELATED_PARTY_TRANSACTION') == 'NO':
            print("Related party transaction program is skipped as per env file")
        else:
            print("select YES/NO for Related party transaction program in env file")
            raise Exception("Error in Env file for 'Related party transaction program' sheet")
    except Exception as e:
        print("Exception caught for Process: 'Related party transaction program' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('SAME_MATERIAL_SCRAP') == 'YES':
            print("Executing 'same material scrap' program")
            config_same_material_scrap = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Same_Material_Scrap_sheetname"])
            same_material_scrap(dict_main_config=config_main, dict_in_config=config_same_material_scrap,
                                sales_present_quarter_pd=sales_present_quarter_pd)

        elif env_file('SAME_MATERIAL_SCRAP') == 'NO':
            print("'same material scrap' program is skipped as per env file")
        else:
            print("select YES/NO for 'same material scrap' program in env file")
            raise Exception("Error in Env file for ''same material scrap' program sheet")
    except Exception as e:
        print("Exception caught for Process: 'same material scrap' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('SAME_MATERIAL_DOMESTIC') == 'YES':
            print("Executing 'same material domestic' program")
            config_concentration_customer_wise = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Same_Material_Domestic_sheetname"])
            same_material_domestic(dict_main_config=config_main, dict_in_config=config_concentration_customer_wise,
                                   sales_present_quarter_pd=sales_present_quarter_pd)

        elif env_file('SAME_MATERIAL_DOMESTIC') == 'NO':
            print("same material domestic program is skipped as per env file")
        else:
            print("select YES/NO for same material domestic program in env file")
            raise Exception("Error in Env file for 'same material domestic' sheet")
    except Exception as e:
        print("Exception caught for Process: same material domestic Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('CUSTOMER_SPECIFIC') == 'YES':
            print("Executing customer specific program")
            customer_specific(sales_register_df=sales_present_quarter_pd, dict_config_main=config_main)

        elif env_file('CUSTOMER_SPECIFIC') == 'NO':
            print("customer specific program is skipped as per env file")
        else:
            print("select YES/NO for customer specific program in env file")
            raise Exception("Error in Env file for 'customer specific program' sheet")
    except Exception as e:
        print("Exception caught for Process: 'customer specific' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('SR_VS_SL') == 'YES':
            print("Executing Sales Register Vs Sales Ledger program")
            sr_vs_sl(data_of_sales_register=sales_present_quarter_pd, main_config=config_main,
                     sales_ledger_new_dataframe=sales_ledger_new_dataframe)

        elif env_file('SR_VS_SL') == 'NO':
            print("Sales Register vs Sales Ledger program is skipped as per env file")
        else:
            print("select YES/NO for Sales Register vs Sales Ledger program in env file")
            raise Exception("Error in Env file for 'Sales Register vs Sales Ledger' sheet")
    except Exception as e:
        logging.exception(e)
        print("Exception caught for Process: 'Sales Register vs Sales Ledger' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('OPEN_PO') == 'YES':
            print("Executing Open PO Reports program")
            open_po_reports(open_po_df=open_po_new_dataframe, main_config=config_main)

        elif env_file('OPEN_PO') == 'NO':
            print(" Open PO Reports program is skipped as per env file")
        else:
            print("select YES/NO for  Open PO Reports program in env file")
            raise Exception("Error in Env file for  Open PO Reports sheet")
    except Exception as e:
        logging.exception(e)
        print("Exception caught for Process:  Open PO Reports Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('SEQUENCE_CHECK') == 'YES':
            print("Executing Sequence Check program")
            config_sequence_check = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Sequence_Check_sheet_name"])
            sequence_check(config_main, config_sequence_check, sales_present_quarter_pd)

        elif env_file('SEQUENCE_CHECK') == 'NO':
            print("Sequence Check process is skipped as per env file")
        else:
            print("select YES/NO for Sequence Check process in env file")
            raise Exception("Error in Env file for 'Sequence Check' sheet")
    except Exception as e:
        print("Exception caught for Process: 'Sequence Check' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    try:
        if env_file('CASH_DISCOUNT') == 'YES':
            print("Executing Cash discount report program")
            config_cash_discount = create_sheet_wise_config_dict(
                sheet_name=config_main["Config_Cash_Discount_sheetname"])
            cash_discount(dict_main_config=config_main, dict_in_config=config_cash_discount,
                          sales_present_quarter_pd=sales_present_quarter_pd)

        elif env_file('CASH_DISCOUNT') == 'NO':
            print("Cash discount report process is skipped as per env file")
        else:
            print("select YES/NO for Cash discount report process in env file")
            raise Exception("Error in Env file for 'Cash discount report' sheet")
    except Exception as e:
        print("Exception caught for Process: 'Cash discount report' Sheet: ", e)
    # ------------------------------------------------------------------------------------
    print("*******************************************")

    final_output_file = openpyxl.load_workbook(output_file_path)
    if 'Sheet1' in final_output_file.sheetnames:
        final_output_file.remove(final_output_file['Sheet1'])

    final_output_file.save(output_file_path)

    # ------------------------------------------------------------------------------------

    config_saving_file_path_in_output = os.path.join(project_home_directory, 'Data', 'Output', 'audit_requests',
                                                     str(request_id), 'config.xlsx')
    config_saving_folder_path_in_output = os.path.join(project_home_directory, 'Data', 'Output', 'audit_requests',
                                                       str(request_id))
    if not os.path.exists(config_saving_folder_path_in_output):
        logging.warning("folder path {0} is not exist".format(config_saving_folder_path_in_output))
        os.makedirs(config_saving_folder_path_in_output)
        logging.warning("created the directory: {0}".format(config_saving_folder_path_in_output))
    print("Creating new Config file in output folder of Request from Config Dictionary")
    logging.info("Creating new Config file in output folder of Request from Config Dictionary")
    try:
        new_config_df = pd.DataFrame.from_dict(config_main, orient='index', columns=['Value'])
        new_config_df.index.name = 'Key'
        new_config_df.reset_index(level=0, inplace=True)
        new_config_df.to_excel(config_saving_file_path_in_output, index=False)
        print("Created new Config file in output folder of Request from Config Dictionary")
        logging.info("Created new Config file in Output folder of Request from Config Dictionary")
    except Exception as config_file_save_exception:
        logging.warning("Exception occurred while saving config file in Output folder of Request Folder directory")
        logging.exception(config_file_save_exception)
        exception_type, exception_object, exception_traceback = sys.exc_info()
        filename = exception_traceback.tb_frame.f_code.co_filename
        line_number = exception_traceback.tb_lineno
        logging.warning(str(exception_type))
        logging.warning("Exception occurred in file : {} at line number: {}".format(filename, line_number))
    else:
        print("Saved updated config data to an excel file in output folder in the request folder")
    # ------------------------------------------------------------------------------------
    print("Saved config data to an excel file")

    # Bot success mail notification
    end_to = config_main['To_Mail_Address']
    end_cc = config_main['CC_Mail_Address']
    end_subject = config_main['Success_Mail_Subject']
    end_body = config_main['Success_Mail_Body']
    send_mail_with_attachment(to=end_to, cc=end_cc, body=end_body, subject=end_subject,
                              attachment_path=output_file_path)
    print("Process complete mail notification is sent")

    print("Bot successfully finished Processing of the sheets")
    return output_file_path


if __name__ == '__main__':
    pass

    # sales_previous_quarter_file_path = None

    # input_files = [sales_present_quarter_file_path, sales_previous_quarter_file_path, hsn_file_path]
    #
    # present_quarter_sheet_name = 'Sheet1'
    # previous_quarter_sheet_name = ''
    # hsn_sheet_name = 'HSN Code summary'
    #
    # present_quarter_column_name = 'Q1 FY 2022-23'
    # previous_quarter_column_name = 'Q4 FY 2021-22'
    # company_name = 'Pitti Engineering Limited'
    # statutory_audit_quarter = ''
    # financial_year = '2022-23'
    #
    # present_working_directory = os.getcwd()
    # config_file_path = os.path.join(os.path.dirname(present_working_directory), 'Input', 'Config.xlsx')
    # config_sheet_name = 'Main'
    # config_main = create_config_dict_from_config_file(path=config_file_path, sheet_name=config_sheet_name)
    #
    # request_id = 20
    # process_execution(input_files,
    #                   present_quarter_sheet_name, previous_quarter_sheet_name,
    #                   present_quarter_column_name, previous_quarter_column_name,
    #                   company_name, statutory_audit_quarter, financial_year, config_main, request_id, hsn_sheet_name
    #                   )
